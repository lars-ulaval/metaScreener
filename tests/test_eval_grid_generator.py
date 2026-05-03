# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""test_eval_grid_generator.py - unit tests for tools/eval_grid_generator.py.

Covers:
  - Loaders reject empty / malformed inputs and missing required columns.
  - Stratification produces partitions of the correct size, with the
    overlap subset disjoint from each rater's unique allotment.
  - Partition is reproducible under a fixed seed and changes under a
    different seed.
  - Multi-stage partition: EL and IL stages run independently with
    derived seeds; partitions per stage are independent.
  - End-to-end: generated XLSX files open cleanly, contain the expected
    sheets (Read Me First + Decisions_EL + Decisions_IL + Reference),
    dropdown validations, and row counts.
  - Manifest CSV records every (stage, a_id, rater) assignment exactly once.
  - LLM decision columns are NOT exposed in any Decisions sheet.
"""
import csv
import importlib.util
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TOOLS_DIR = PROJECT_ROOT / "tools"
TESTS_DATA = PROJECT_ROOT / "tests" / "data"
GOLDEN = PROJECT_ROOT / "tests" / "golden"

# Import the script under test by file path (tools/ is not a package).
_spec = importlib.util.spec_from_file_location(
    "eval_grid_generator", str(TOOLS_DIR / "eval_grid_generator.py")
)
egg = importlib.util.module_from_spec(_spec)
sys.modules[_spec.name] = egg
_spec.loader.exec_module(egg)


# --------------------------------------------------------------------------
# Loader behaviour
# --------------------------------------------------------------------------

class TestLoaders:
    def test_load_stage_records_rejects_missing_file(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            egg.load_stage_records(tmp_path / "does_not_exist.csv", "IL")

    def test_load_stage_records_rejects_unknown_stage(self, tmp_path):
        with pytest.raises(ValueError, match="Unknown stage"):
            egg.load_stage_records(TESTS_DATA / "il_eval_fixture.csv", "ZZ")

    def test_load_stage_records_rejects_missing_required_columns(self, tmp_path):
        bad = tmp_path / "bad.csv"
        bad.write_text("local_id,title\nA001,Foo\n", encoding="utf-8")
        with pytest.raises(ValueError, match="missing required columns"):
            egg.load_stage_records(bad, "IL")

    def test_load_stage_records_il_fixture(self):
        rows = egg.load_stage_records(TESTS_DATA / "il_eval_fixture.csv", "IL")
        assert len(rows) == 30
        assert rows[0]["local_id"] == "A001"
        assert "il_outcome" in rows[0]

    def test_load_stage_records_el_fixture(self):
        rows = egg.load_stage_records(TESTS_DATA / "el_eval_fixture.csv", "EL")
        assert len(rows) == 30
        assert "el_outcome" in rows[0]

    def test_load_llm_criteria_returns_il_and_el_groups(self):
        crits = egg.load_llm_criteria(GOLDEN / "criteria_harmonized_v3.1.0.csv")
        # Published demonstration: IC-1 is IL/llm; EC-2 and EC-3 are EL/llm.
        assert "IL" in crits and "EL" in crits
        assert {c["id"] for c in crits["IL"]} == {"IC-1"}
        assert {c["id"] for c in crits["EL"]} == {"EC-2", "EC-3"}
        # All returned criteria must be llm-operator and enabled.
        for stage_crits in crits.values():
            for c in stage_crits:
                assert c["operator"] == "llm"
                assert str(c["enabled"]).strip() in ("1", "true", "True")

    def test_load_llm_criteria_rejects_when_no_llm_rows(self, tmp_path):
        bad = tmp_path / "no_llm.csv"
        bad.write_text(
            "stage,id,type,scope,label,operator,target,what,threshold,enabled,source_text\n"
            "EH,EC-1,exclude,metadata,foo,equals,lang,French,,1,foo\n",
            encoding="utf-8",
        )
        with pytest.raises(ValueError, match="No llm-operator criteria"):
            egg.load_llm_criteria(bad)


# --------------------------------------------------------------------------
# Single-stage stratification + partition logic
# --------------------------------------------------------------------------

class TestStratifyPartition:
    @pytest.fixture
    def il_records(self):
        return egg.load_stage_records(TESTS_DATA / "il_eval_fixture.csv", "IL")

    def test_partition_sizes_are_correct(self, il_records):
        overlap, per_rater = egg.stratify_and_partition(
            il_records, n_overlap=6, raters=["R1", "R2", "R3"], seed=42,
            outcome_col="il_outcome",
        )
        assert len(overlap) == 6
        total_disjoint = sum(len(v) for v in per_rater.values())
        assert total_disjoint == len(il_records) - 6
        sizes = sorted(len(v) for v in per_rater.values())
        assert sizes[-1] - sizes[0] <= 1

    def test_overlap_disjoint_from_unique_allotments(self, il_records):
        overlap, per_rater = egg.stratify_and_partition(
            il_records, n_overlap=9, raters=["R1", "R2", "R3"], seed=42,
            outcome_col="il_outcome",
        )
        overlap_ids = {r["local_id"] for r in overlap}
        for rater_recs in per_rater.values():
            unique_ids = {r["local_id"] for r in rater_recs}
            assert overlap_ids.isdisjoint(unique_ids)

    def test_partition_reproducible_under_same_seed(self, il_records):
        a_overlap, a_rater = egg.stratify_and_partition(
            il_records, n_overlap=6, raters=["R1", "R2", "R3"], seed=42,
            outcome_col="il_outcome",
        )
        b_overlap, b_rater = egg.stratify_and_partition(
            il_records, n_overlap=6, raters=["R1", "R2", "R3"], seed=42,
            outcome_col="il_outcome",
        )
        assert [r["local_id"] for r in a_overlap] == [r["local_id"] for r in b_overlap]
        for k in a_rater:
            assert [r["local_id"] for r in a_rater[k]] == \
                   [r["local_id"] for r in b_rater[k]]

    def test_partition_changes_under_different_seed(self, il_records):
        a_overlap, _ = egg.stratify_and_partition(
            il_records, n_overlap=6, raters=["R1", "R2", "R3"], seed=42,
            outcome_col="il_outcome",
        )
        b_overlap, _ = egg.stratify_and_partition(
            il_records, n_overlap=6, raters=["R1", "R2", "R3"], seed=99,
            outcome_col="il_outcome",
        )
        assert {r["local_id"] for r in a_overlap} != \
               {r["local_id"] for r in b_overlap}

    def test_overlap_proportional_stratification(self, il_records):
        # IL fixture has 25 REVIEW + 5 OUT. With n_overlap=6 the largest-remainder
        # apportionment yields ~5 REVIEW + ~1 OUT.
        overlap, _ = egg.stratify_and_partition(
            il_records, n_overlap=6, raters=["R1", "R2", "R3"], seed=42,
            outcome_col="il_outcome",
        )
        outcomes = [r["il_outcome"] for r in overlap]
        review_ct = sum(1 for o in outcomes if o == "REVIEW")
        out_ct = sum(1 for o in outcomes if o == "OUT")
        assert review_ct in (4, 5, 6)
        assert out_ct in (0, 1, 2)
        assert review_ct + out_ct == 6

    def test_rejects_invalid_inputs(self, il_records):
        with pytest.raises(ValueError, match="At least one rater"):
            egg.stratify_and_partition(
                il_records, n_overlap=5, raters=[], seed=1, outcome_col="il_outcome"
            )
        with pytest.raises(ValueError, match="n_overlap must be"):
            egg.stratify_and_partition(
                il_records, n_overlap=-1, raters=["R1"], seed=1, outcome_col="il_outcome"
            )
        with pytest.raises(ValueError, match="cannot exceed total records"):
            egg.stratify_and_partition(
                il_records, n_overlap=999, raters=["R1"], seed=1, outcome_col="il_outcome"
            )


# --------------------------------------------------------------------------
# Multi-stage partition
# --------------------------------------------------------------------------

class TestMultiStagePartition:
    @pytest.fixture
    def stage_records(self):
        return {
            "IL": egg.load_stage_records(TESTS_DATA / "il_eval_fixture.csv", "IL"),
            "EL": egg.load_stage_records(TESTS_DATA / "el_eval_fixture.csv", "EL"),
        }

    def test_partitions_returned_for_each_stage(self, stage_records):
        out = egg.partition_all_stages(
            stage_records=stage_records,
            n_overlap=6,
            raters=["R1", "R2", "R3"],
            seed=42,
        )
        assert set(out.keys()) == {"IL", "EL"}
        for stage, (overlap, per_rater) in out.items():
            assert len(overlap) == 6
            assert set(per_rater.keys()) == {"R1", "R2", "R3"}

    def test_il_and_el_partitions_are_independent(self, stage_records):
        # With derived seeds (42 vs 43), IL and EL overlap sets should differ
        # in record selection even though the underlying record sets are
        # almost identical (the synthetic fixtures share IDs A001-A030).
        out = egg.partition_all_stages(
            stage_records=stage_records,
            n_overlap=6,
            raters=["R1", "R2", "R3"],
            seed=42,
        )
        il_overlap_ids = {r["local_id"] for r in out["IL"][0]}
        el_overlap_ids = {r["local_id"] for r in out["EL"][0]}
        # Independent seeds -> different overlap sets (allowing for small
        # accidental intersection but not full equality).
        assert il_overlap_ids != el_overlap_ids


# --------------------------------------------------------------------------
# End-to-end: generate workbooks, manifest, and inspect them
# --------------------------------------------------------------------------

class TestEndToEnd:
    @pytest.fixture
    def generated(self, tmp_path):
        out_dir = tmp_path / "grids"
        rc = egg.main([
            "--el-filtered", str(TESTS_DATA / "el_eval_fixture.csv"),
            "--il-filtered", str(TESTS_DATA / "il_eval_fixture.csv"),
            "--criteria",    str(GOLDEN / "criteria_harmonized_v3.1.0.csv"),
            "--raters",      "AReyes", "JKiss", "JVoisin",
            "--output-dir",  str(out_dir),
            "--overlap",     "6",
            "--seed",        "42",
        ])
        assert rc == 0
        return out_dir

    def test_three_xlsx_files_produced(self, generated):
        files = sorted(p.name for p in generated.glob("*.xlsx"))
        assert files == [
            "eval_grid_AReyes.xlsx",
            "eval_grid_JKiss.xlsx",
            "eval_grid_JVoisin.xlsx",
        ]

    def test_each_workbook_has_expected_sheets(self, generated):
        for path in generated.glob("*.xlsx"):
            wb = load_workbook(path)
            assert wb.sheetnames == [
                "Read Me First", "Decisions_EL", "Decisions_IL", "Reference",
            ]

    def test_decisions_sheets_do_not_expose_llm_columns(self, generated):
        # Critical: blind adjudication. None of the il_*/el_* columns from
        # the input CSVs may appear as headers in any Decisions sheet.
        forbidden = {
            "il_outcome", "il_failed_ids", "il_met_ids", "il_uncertain_ids",
            "il_evidence_json", "il_reason_summary", "il_missing_ids",
            "el_outcome", "el_failed_ids", "el_met_ids", "el_uncertain_ids",
            "el_evidence_json", "el_reason_summary", "el_missing_ids",
        }
        for path in generated.glob("*.xlsx"):
            wb = load_workbook(path)
            for sheet_name in ("Decisions_EL", "Decisions_IL"):
                ws = wb[sheet_name]
                headers = [c.value for c in ws[1]]
                assert not (set(headers) & forbidden), \
                    f"LLM cols leaked into {path.name}/{sheet_name}: {headers}"

    def test_decisions_sheets_have_dropdown_validation(self, generated):
        for path in generated.glob("*.xlsx"):
            wb = load_workbook(path)
            for sheet_name in ("Decisions_EL", "Decisions_IL"):
                ws = wb[sheet_name]
                assert ws.data_validations.dataValidation
                formulas = {dv.formula1 for dv in ws.data_validations.dataValidation}
                assert any("include" in f and "exclude" in f and "uncertain" in f
                           for f in formulas)

    def test_decisions_per_stage_row_counts_sum_correctly(self, generated):
        # Both fixtures: 30 records, overlap=6 -> per-rater overlap=6, unique=8,
        # total per stage per rater = 14 (allowing +/-1 for stratum rounding).
        # Sum of UNIQUE rows across raters per stage == 30 - 6 = 24.
        for stage_sheet in ("Decisions_EL", "Decisions_IL"):
            n_unique_total = 0
            for path in generated.glob("*.xlsx"):
                wb = load_workbook(path)
                ws = wb[stage_sheet]
                n_rows = ws.max_row - 1
                n_unique_total += n_rows - 6
            assert n_unique_total == 24, f"{stage_sheet} unique sum mismatch"

    def test_decisions_per_criterion_columns_match_stage(self, generated):
        # EL has 2 criteria (EC-2, EC-3) -> 4 extra cols (decision+notes per).
        # IL has 1 criterion (IC-1) -> 2 extra cols.
        # Plus 7 record cols in either case.
        for path in generated.glob("*.xlsx"):
            wb = load_workbook(path)
            el_headers = [c.value for c in wb["Decisions_EL"][1]]
            il_headers = [c.value for c in wb["Decisions_IL"][1]]
            assert len(el_headers) == 7 + 2 * 2  # 7 record cols + 2 criteria * 2
            assert len(il_headers) == 7 + 1 * 2  # 7 record cols + 1 criterion * 2
            # And the decision columns are correctly labelled.
            assert "Decision: EC-2" in el_headers
            assert "Decision: EC-3" in el_headers
            assert "Decision: IC-1" in il_headers

    def test_partition_manifest_records_all_assignments(self, generated):
        manifest = generated / "partition_manifest.csv"
        assert manifest.exists()
        with manifest.open(encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
        # Per stage: 6 overlap * 3 raters + 24 disjoint * 1 rater = 42.
        # Two stages -> 84 rows total.
        assert len(rows) == 84
        # Per-stage overlap rows must total 18 each.
        from collections import Counter
        per_stage_overlap = Counter(
            r["stage"] for r in rows if r["set"] == "overlap"
        )
        per_stage_disjoint = Counter(
            r["stage"] for r in rows if r["set"] == "disjoint"
        )
        assert per_stage_overlap == {"IL": 18, "EL": 18}
        assert per_stage_disjoint == {"IL": 24, "EL": 24}
        # Every overlap (stage, a_id) pair appears for all three raters.
        per_pair = Counter(
            (r["stage"], r["a_id"]) for r in rows if r["set"] == "overlap"
        )
        assert all(v == 3 for v in per_pair.values())

    def test_partition_meta_sidecar_records_per_stage_criteria(self, generated):
        meta = generated / "partition_manifest.meta.txt"
        assert meta.exists()
        text = meta.read_text(encoding="utf-8")
        assert "seed=42" in text
        assert "EL_n_overlap=6" in text
        assert "IL_n_overlap=6" in text
        assert "EL_criteria=EC-2,EC-3" in text
        assert "IL_criteria=IC-1" in text

    def test_main_rejects_missing_stage_input_when_criteria_present(self, tmp_path):
        # If criteria has IL/llm rows but --il-filtered is omitted, error.
        out_dir = tmp_path / "grids"
        rc = egg.main([
            "--el-filtered", str(TESTS_DATA / "el_eval_fixture.csv"),
            # no --il-filtered
            "--criteria",    str(GOLDEN / "criteria_harmonized_v3.1.0.csv"),
            "--raters",      "AReyes", "JKiss", "JVoisin",
            "--output-dir",  str(out_dir),
            "--overlap",     "6",
            "--seed",        "42",
        ])
        assert rc == 2
