# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""test_eval_ingest.py - unit tests for tools/eval_ingest.py.

Covers:
  - Cohen's kappa pure-python implementation (perfect-agreement, perfect-
    disagreement, three-way reference values, NaN handling).
  - Fleiss' kappa pure-python implementation (perfect-agreement edge case,
    reference values, ValueError on inconsistent rater counts).
  - LLM status -> canonical decision mapping covers MET/FAILED/UNCERTAIN.
  - Filled-workbook reader: validates against manifest (rejects spurious
    rows, missing rows, empty cells, typos in dropdown values, malformed
    sheets), produces correctly-shaped long-format rows on valid input.
  - Join: missing LLM evidence rows get llm_status='MISSING' and
    agree=False; status mapping aligns the join key with human's
    include/exclude/uncertain vocabulary.
  - End-to-end: empty grid -> synthetic filler -> ingest -> verify Cohen
    kappa is high when agree_with_llm_rate is high, low when it is low.
  - Output files have expected schema and row counts.
"""
import csv
import importlib.util
import math
import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TOOLS_DIR = PROJECT_ROOT / "tools"
TESTS_DATA = PROJECT_ROOT / "tests" / "data"
GOLDEN = PROJECT_ROOT / "tests" / "golden"


def _import(modname: str):
    spec = importlib.util.spec_from_file_location(
        modname, str(TOOLS_DIR / f"{modname}.py")
    )
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    return mod


egg = _import("eval_grid_generator")
ein = _import("eval_ingest")
fil = _import("eval_grid_filler_synthetic")


# --------------------------------------------------------------------------
# Pure-math: Cohen's kappa
# --------------------------------------------------------------------------

class TestCohenKappa:
    def test_perfect_agreement_returns_one(self):
        pairs = [("yes", "yes")] * 10 + [("no", "no")] * 5
        result = ein.cohen_kappa(pairs)
        assert result["n"] == 15
        assert math.isclose(result["agreement_observed"], 1.0)
        assert math.isclose(result["kappa"], 1.0)

    def test_perfect_disagreement(self):
        # Every "yes" by A is "no" by B and vice versa.
        pairs = [("yes", "no")] * 5 + [("no", "yes")] * 5
        result = ein.cohen_kappa(pairs)
        assert result["n"] == 10
        assert math.isclose(result["agreement_observed"], 0.0)
        # P_e = 0.5 * 0.5 + 0.5 * 0.5 = 0.5; kappa = (0 - 0.5)/(1 - 0.5) = -1.
        assert math.isclose(result["kappa"], -1.0)

    def test_chance_agreement_zero(self):
        # 25 each of (yes,yes), (yes,no), (no,yes), (no,no).
        pairs = (
            [("yes", "yes")] * 25
            + [("yes", "no")] * 25
            + [("no", "yes")] * 25
            + [("no", "no")] * 25
        )
        result = ein.cohen_kappa(pairs)
        # P_o = 50/100 = 0.5; P_e = 0.5*0.5 + 0.5*0.5 = 0.5; kappa = 0.
        assert math.isclose(result["agreement_observed"], 0.5)
        assert math.isclose(result["kappa"], 0.0, abs_tol=1e-9)

    def test_empty_pairs_returns_nan(self):
        result = ein.cohen_kappa([])
        assert result["n"] == 0
        assert math.isnan(result["kappa"])

    def test_unanimous_single_class_returns_nan(self):
        # All raters always say "yes" -> P_e = 1 -> kappa undefined.
        pairs = [("yes", "yes")] * 20
        result = ein.cohen_kappa(pairs)
        assert math.isnan(result["kappa"])


# --------------------------------------------------------------------------
# Pure-math: Fleiss' kappa
# --------------------------------------------------------------------------

class TestFleissKappa:
    def test_perfect_agreement_returns_nan(self):
        # 5 items, 3 raters, all rating "include" (column 0) -> P_e = 1.
        matrix = [[3, 0, 0]] * 5
        result = ein.fleiss_kappa(matrix)
        assert result["n_items"] == 5
        assert result["n_raters"] == 3
        assert math.isclose(result["agreement_observed"], 1.0)
        assert math.isnan(result["kappa"])

    def test_perfect_disagreement_low_kappa(self):
        # 6 items, 3 raters, perfectly split (1 each in 3 categories per item).
        matrix = [[1, 1, 1]] * 6
        result = ein.fleiss_kappa(matrix)
        # P_i = 0 for every item -> P_o = 0; P_e = 1/3 -> kappa = -0.5.
        assert math.isclose(result["agreement_observed"], 0.0)
        assert math.isclose(result["kappa"], -0.5, abs_tol=1e-9)

    def test_reference_value_landis_koch(self):
        # Constructed: 3 items rated by 3 raters
        # Item 1: 3 include  -> P_1 = 1
        # Item 2: 2 include, 1 exclude -> P_2 = (2*1 + 1*0 + 0)/(3*2) = 2/6
        # Item 3: 1 each -> P_3 = 0
        matrix = [
            [3, 0, 0],
            [2, 1, 0],
            [1, 1, 1],
        ]
        result = ein.fleiss_kappa(matrix)
        assert result["n_items"] == 3
        assert result["n_raters"] == 3
        # P_o = (1 + 2/6 + 0)/3 = 4/9 = 0.4444
        assert math.isclose(result["agreement_observed"], 4 / 9, abs_tol=1e-9)
        # p_j: include=6/9=0.667, exclude=2/9=0.222, uncertain=1/9=0.111
        # P_e = 0.667^2 + 0.222^2 + 0.111^2 = 0.444 + 0.0494 + 0.0123 = 0.506
        # kappa = (0.4444 - 0.506) / (1 - 0.506) ~ -0.125
        assert -0.2 < result["kappa"] < 0.0

    def test_inconsistent_rater_counts_raises(self):
        # Items with different totals -> invalid.
        matrix = [[3, 0, 0], [2, 1, 1]]
        with pytest.raises(ValueError, match="same rater count"):
            ein.fleiss_kappa(matrix)

    def test_single_rater_raises(self):
        matrix = [[1, 0, 0]] * 3
        with pytest.raises(ValueError, match="at least 2 raters"):
            ein.fleiss_kappa(matrix)

    def test_empty_matrix_returns_nan_kappa(self):
        result = ein.fleiss_kappa([])
        assert result["n_items"] == 0
        assert math.isnan(result["kappa"])


# --------------------------------------------------------------------------
# Decision normalization
# --------------------------------------------------------------------------

class TestNormalizeDecision:
    def test_canonical_passthrough(self):
        # Bare canonical short codes (used by synthetic test inputs and
        # already-normalized data) pass through unchanged.
        assert ein.normalize_decision("yes") == "yes"
        assert ein.normalize_decision("no") == "no"
        assert ein.normalize_decision("unsure") == "unsure"

    def test_case_and_whitespace_tolerated(self):
        assert ein.normalize_decision("YES") == "yes"
        assert ein.normalize_decision("  NO  ") == "no"
        assert ein.normalize_decision("\tUnsure\n") == "unsure"

    def test_full_dropdown_sentences_canonicalize(self):
        # The actual values raters will produce: full sentences quoting
        # the criterion text. The prefix is what we match on.
        yes_sentence = (
            "YES - this is true: The paper considers immersive virtual "
            "reality OR a virtual simulation using a head-mounted display "
            "(HMD)."
        )
        no_sentence = (
            "NO - this is not true: The paper's primary focus is the "
            "rubber hand illusion paradigm."
        )
        unsure_sentence = "I cannot tell from the abstract alone."
        assert ein.normalize_decision(yes_sentence) == "yes"
        assert ein.normalize_decision(no_sentence) == "no"
        assert ein.normalize_decision(unsure_sentence) == "unsure"

    def test_empty_returns_none(self):
        assert ein.normalize_decision(None) is None
        assert ein.normalize_decision("") is None
        assert ein.normalize_decision("   ") is None

    def test_unrecognised_raises(self):
        with pytest.raises(ValueError, match="not recognised"):
            ein.normalize_decision("maybe")
        with pytest.raises(ValueError, match="not recognised"):
            ein.normalize_decision("include")  # old vocabulary, no longer valid


# --------------------------------------------------------------------------
# LLM status -> canonical decision mapping
# --------------------------------------------------------------------------

class TestLlmStatusMapping:
    """Polarity-aware LLM status -> canonical decision mapping.

    metaScreener's `status` field describes whether a record passes or fails
    the screening rule, not whether the criterion's claim holds. These
    coincide for inclusion criteria but invert for exclusion criteria.
    """

    def test_status_to_canonical_inclusion(self):
        assert ein.status_to_canonical("MET", "include") == "yes"
        assert ein.status_to_canonical("FAILED", "include") == "no"
        assert ein.status_to_canonical("UNCERTAIN", "include") == "unsure"

    def test_status_to_canonical_exclusion(self):
        # For exclusion criteria, status=MET means "passes the exclusion
        # check" = "not excluded" = "criterion's claim is NOT true."
        assert ein.status_to_canonical("MET", "exclude") == "no"
        assert ein.status_to_canonical("FAILED", "exclude") == "yes"
        assert ein.status_to_canonical("UNCERTAIN", "exclude") == "unsure"

    def test_status_to_canonical_case_insensitive(self):
        assert ein.status_to_canonical("met", "include") == "yes"
        assert ein.status_to_canonical("Failed", "Exclude") == "yes"
        assert ein.status_to_canonical(" UNCERTAIN ", "include") == "unsure"

    def test_status_to_canonical_unknown_status_defaults_unsure(self):
        assert ein.status_to_canonical("WEIRD", "include") == "unsure"
        assert ein.status_to_canonical("", "exclude") == "unsure"

    def test_status_to_canonical_unknown_polarity_defaults_inclusion(self):
        # Defensive default — if the criteria CSV has a typo, we still
        # produce a reasonable answer rather than crashing.
        assert ein.status_to_canonical("MET", "") == "yes"
        assert ein.status_to_canonical("MET", "unknown") == "yes"

    def test_join_with_llm_exclusion_criterion(self):
        # Reproduces the real-data semantics: EC-2 with status=MET means
        # "passes the exclusion check" = paper is NOT about navigation =
        # canonical "no". A human who said "no" should agree with the LLM.
        evidence = {
            "EL": {"A001": {"EC-2": {"status": "MET", "decision": "not_meet",
                                      "confidence": 0.9, "quote": "vr",
                                      "quote_valid": True}}}
        }
        humans = [{
            "stage": "EL", "a_id": "A001", "criterion_id": "EC-2",
            "rater_id": "R1", "set": "disjoint",
            "human_decision": "no", "human_notes": "",
        }]
        criteria_polarity = {"EC-2": "exclude"}
        out = ein.join_with_llm(humans, evidence, criteria_polarity)
        assert out[0]["llm_status"] == "MET"
        assert out[0]["llm_decision_canonical"] == "no"
        assert out[0]["agree"] == "True"

    def test_join_with_llm_inclusion_criterion(self):
        # IC-1 with status=MET means "passes the inclusion check" = paper IS
        # about HMD VR = canonical "yes".
        evidence = {
            "IL": {"A001": {"IC-1": {"status": "MET", "decision": "meet",
                                      "confidence": 0.9, "quote": "HMD",
                                      "quote_valid": True}}}
        }
        humans = [{
            "stage": "IL", "a_id": "A001", "criterion_id": "IC-1",
            "rater_id": "R1", "set": "disjoint",
            "human_decision": "yes", "human_notes": "",
        }]
        criteria_polarity = {"IC-1": "include"}
        out = ein.join_with_llm(humans, evidence, criteria_polarity)
        assert out[0]["llm_status"] == "MET"
        assert out[0]["llm_decision_canonical"] == "yes"
        assert out[0]["agree"] == "True"

    def test_join_missing_evidence_marked_disagreement(self):
        humans = [{
            "stage": "IL", "a_id": "A099", "criterion_id": "IC-1",
            "rater_id": "R1", "set": "disjoint",
            "human_decision": "yes", "human_notes": "",
        }]
        out = ein.join_with_llm(humans, {"IL": {}}, {"IC-1": "include"})
        assert out[0]["llm_status"] == "MISSING"
        assert out[0]["agree"] == "False"

    def test_load_criteria_polarity_from_csv(self):
        polarities = ein.load_criteria_polarity(
            GOLDEN / "criteria_harmonized_v3.1.0.csv"
        )
        # Sanity: published criteria all have known polarities.
        assert polarities.get("IC-1") == "include"
        assert polarities.get("EC-2") == "exclude"
        assert polarities.get("EC-3") == "exclude"


# --------------------------------------------------------------------------
# Workbook reader: schema validation
# --------------------------------------------------------------------------

class TestWorkbookReader:
    @pytest.fixture
    def empty_grids(self, tmp_path):
        out_dir = tmp_path / "empty"
        rc = egg.main([
            "--el-filtered", str(TESTS_DATA / "el_eval_fixture.csv"),
            "--il-filtered", str(TESTS_DATA / "il_eval_fixture.csv"),
            "--criteria",    str(GOLDEN / "criteria_harmonized_v3.1.0.csv"),
            "--raters",      "AReyes", "JKiss", "JVoisin",
            "--output-dir",  str(out_dir),
            "--overlap",     "6",
            "--seed",        "42",
        ])
        assert rc == 0
        return out_dir

    @pytest.fixture
    def manifest_by_rater(self, empty_grids):
        manifest = ein.load_partition_manifest(empty_grids / "partition_manifest.csv")
        return ein.index_manifest_by_rater(manifest)

    def test_unfilled_workbook_reports_empty_decisions(self, empty_grids, manifest_by_rater):
        path = empty_grids / "eval_grid_AReyes.xlsx"
        with pytest.raises(ValueError, match="decision is empty"):
            ein.read_filled_workbook(
                path=path, rater_id="AReyes",
                expected_per_stage=manifest_by_rater["AReyes"],
            )

    def test_filled_workbook_reads_cleanly(self, empty_grids, manifest_by_rater, tmp_path):
        # Use the synthetic filler.
        src = empty_grids / "eval_grid_AReyes.xlsx"
        filled = tmp_path / "filled" / "eval_grid_AReyes.xlsx"
        # Synthetic fixture has no real LLM decisions in its evidence_json,
        # so we inject canonical decisions manually for synthetic-fill.
        # All a_ids -> 'yes' for IC-1, EC-2, EC-3.
        manual_evidence = {
            "EL": {f"A{i:03d}": {"EC-2": "yes", "EC-3": "yes"} for i in range(1, 31)},
            "IL": {f"A{i:03d}": {"IC-1": "yes"} for i in range(1, 31)},
        }
        n = fil.fill_workbook(
            src_path=src, dst_path=filled,
            llm_canonical_by_stage=manual_evidence,
            agree_with_llm_rate=0.9, uncertain_rate=0.05, seed=1,
        )
        assert n > 0

        rows = ein.read_filled_workbook(
            path=filled, rater_id="AReyes",
            expected_per_stage=manifest_by_rater["AReyes"],
        )
        # Every row should have a valid decision in the new vocabulary.
        assert all(r["human_decision"] in ("yes", "no", "unsure")
                   for r in rows)
        # Manifest expects rater_AReyes to rate <expected> a_ids per stage,
        # times the number of criteria on that stage.
        n_expected = (
            len([r for r in manifest_by_rater["AReyes"]["IL"]]) * 1   # IL: IC-1
            + len([r for r in manifest_by_rater["AReyes"]["EL"]]) * 2  # EL: EC-2, EC-3
        )
        assert len(rows) == n_expected

    def test_workbook_with_typo_decision_raises(self, empty_grids, manifest_by_rater, tmp_path):
        # Write an unrecognised value into one decision cell. The value is
        # neither a canonical short code ("yes"/"no"/"unsure") nor a
        # YES/NO/I-cannot-tell sentence prefix, so normalize_decision
        # rejects it.
        src = empty_grids / "eval_grid_AReyes.xlsx"
        target = tmp_path / "typo.xlsx"
        shutil.copy(src, target)
        wb = load_workbook(target)
        ws = wb["Decisions_IL"]
        headers = [c.value for c in ws[1]]
        col = headers.index("Decision: IC-1") + 1
        # Fill all decisions with a valid canonical short code, then break one.
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).value = "yes"
        # Also fill EL columns.
        ws2 = wb["Decisions_EL"]
        hd = [c.value for c in ws2[1]]
        for cid in ("EC-2", "EC-3"):
            cidx = hd.index(f"Decision: {cid}") + 1
            for r in range(2, ws2.max_row + 1):
                ws2.cell(row=r, column=cidx).value = "yes"
        # The typo: an unrecognised value.
        ws.cell(row=2, column=col).value = "maybe"
        wb.save(target)
        with pytest.raises(ValueError, match="not recognised"):
            ein.read_filled_workbook(
                path=target, rater_id="AReyes",
                expected_per_stage=manifest_by_rater["AReyes"],
            )


# --------------------------------------------------------------------------
# End-to-end: empty grids -> synthetic fill -> ingest
# --------------------------------------------------------------------------

class TestEndToEnd:
    @pytest.fixture
    def pipeline(self, tmp_path):
        """Run generator, fill three grids with three different agree-rates,
        then ingest. Returns (out_dir, manifest_path, llm_evidence)."""
        empty_dir = tmp_path / "empty"
        rc = egg.main([
            "--el-filtered", str(TESTS_DATA / "el_eval_fixture.csv"),
            "--il-filtered", str(TESTS_DATA / "il_eval_fixture.csv"),
            "--criteria",    str(GOLDEN / "criteria_harmonized_v3.1.0.csv"),
            "--raters",      "R1", "R2", "R3",
            "--output-dir",  str(empty_dir),
            "--overlap",     "6",
            "--seed",        "42",
        ])
        assert rc == 0

        # Synthesize evidence for synthetic fixtures that lack evidence_json.
        # IL has only IC-1; EL has EC-2 and EC-3. Use new yes/no vocabulary.
        manual_evidence = {
            "EL": {f"A{i:03d}": {"EC-2": "yes", "EC-3": "no"}
                   for i in range(1, 31)},
            "IL": {f"A{i:03d}": {"IC-1": ("no" if i % 4 == 0 else "yes")}
                   for i in range(1, 31)},
        }

        filled_dir = tmp_path / "filled"
        # Three raters with different agree rates so the metric is exercised.
        for rater_id, agree, seed in [("R1", 0.95, 1), ("R2", 0.90, 2), ("R3", 0.88, 3)]:
            src = empty_dir / f"eval_grid_{rater_id}.xlsx"
            dst = filled_dir / f"eval_grid_{rater_id}.xlsx"
            fil.fill_workbook(
                src_path=src, dst_path=dst,
                llm_canonical_by_stage=manual_evidence,
                agree_with_llm_rate=agree, uncertain_rate=0.03, seed=seed,
            )

        # The ingestor reads evidence_json from the fixtures themselves.
        # The synthetic fixtures have empty il_evidence_json/el_evidence_json
        # ('{}'). We need to write modified fixtures so the ingest can
        # find LLM evidence to compare against. Polarity-aware so that the
        # canonical->status inverse mapping matches what the polarity-aware
        # ingestor will read back out.
        crit_polarity = {"IC-1": "include", "EC-2": "exclude", "EC-3": "exclude"}
        modified_il = tmp_path / "il_with_evidence.csv"
        modified_el = tmp_path / "el_with_evidence.csv"
        _patch_fixture_with_evidence(
            TESTS_DATA / "il_eval_fixture.csv", modified_il,
            ev_col="il_evidence_json", per_aid=manual_evidence["IL"],
            criteria_polarity=crit_polarity,
        )
        _patch_fixture_with_evidence(
            TESTS_DATA / "el_eval_fixture.csv", modified_el,
            ev_col="el_evidence_json", per_aid=manual_evidence["EL"],
            criteria_polarity=crit_polarity,
        )

        out_dir = tmp_path / "out"
        rc = ein.main([
            "--manifest",         str(empty_dir / "partition_manifest.csv"),
            "--criteria",         str(GOLDEN / "criteria_harmonized_v3.1.0.csv"),
            "--filled-grids-dir", str(filled_dir),
            "--el-filtered",      str(modified_el),
            "--il-filtered",      str(modified_il),
            "--output-dir",       str(out_dir),
        ])
        assert rc == 0
        return out_dir

    def test_evidence_files_produced(self, pipeline):
        for name in ("eval_decisions_v1.csv", "eval_results_v1.csv",
                     "eval_disagreements_v1.csv", "eval_summary_v1.txt"):
            assert (pipeline / name).exists(), f"missing {name}"

    def test_decisions_csv_schema(self, pipeline):
        with (pipeline / "eval_decisions_v1.csv").open(encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
        assert rows
        assert all(r["human_decision"] in ("yes", "no", "unsure")
                   for r in rows)

    def test_results_csv_has_agree_flag(self, pipeline):
        with (pipeline / "eval_results_v1.csv").open(encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
        assert all(r["agree"] in ("True", "False") for r in rows)
        # With ~90% agree rates on 3 raters, most rows should be agree=True.
        agree_count = sum(1 for r in rows if r["agree"] == "True")
        assert agree_count / len(rows) > 0.7

    def test_summary_text_lists_kappa_per_criterion(self, pipeline):
        text = (pipeline / "eval_summary_v1.txt").read_text(encoding="utf-8")
        # Expected criteria appear in the summary.
        for cid in ("IC-1", "EC-2", "EC-3"):
            assert cid in text
        assert "Cohen's kappa" in text
        assert "Fleiss' kappa" in text
        assert "Confusion matrix" in text


# --------------------------------------------------------------------------
# Test helpers
# --------------------------------------------------------------------------

def _patch_fixture_with_evidence(
    src_csv: Path,
    dst_csv: Path,
    ev_col: str,
    per_aid,
    criteria_polarity: dict = None,
) -> None:
    """Copy a fixture CSV, replacing the evidence_json column with synthesized
    JSON keyed by a_id and criterion. Polarity-aware: the inverse mapping
    from canonical decision back to LLM status depends on the criterion's
    polarity (mirror of tools/eval_ingest.py:status_to_canonical)."""
    import csv, json
    polarity = criteria_polarity or {}
    # status_map per polarity. For inclusion: canonical yes -> status MET
    # (paper passes inclusion = criterion's claim is true). For exclusion:
    # canonical yes -> status FAILED (criterion's claim is true = paper IS
    # about it = fails the exclusion check = excluded).
    status_map_include = {"yes": "MET", "no": "FAILED", "unsure": "UNCERTAIN"}
    status_map_exclude = {"yes": "FAILED", "no": "MET", "unsure": "UNCERTAIN"}
    with src_csv.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    fields = list(rows[0].keys())
    for r in rows:
        a_id = r.get("local_id", "").strip()
        if a_id in per_aid:
            criteria_dict = per_aid[a_id]
            ev = {}
            for cid, canonical in criteria_dict.items():
                pol = polarity.get(cid, "include")
                smap = status_map_exclude if pol == "exclude" else status_map_include
                ev[cid] = {
                    "status": smap[canonical],
                    "decision": "meet" if canonical == "yes" else "not_meet",
                    "confidence": 0.85,
                    "quote": "synthetic quote",
                    "quote_valid": True,
                }
            r[ev_col] = json.dumps(ev)
        else:
            r[ev_col] = "{}"
    dst_csv.parent.mkdir(parents=True, exist_ok=True)
    with dst_csv.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, lineterminator="\n")
        w.writeheader()
        for r in rows:
            w.writerow(r)
