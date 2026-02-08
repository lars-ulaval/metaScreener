# -*- coding: utf-8 -*-
"""
decisions_report.py — Screen A (metadata-only) reporting & exports — Foundation (v3)

Compatibility targets
- plugin.py (Foundation v3):
    • aggregate_final(meta_results=..., A_rows=...)
      (fallback: aggregate_decisions(meta_results, A_rows))
    • export_decisions_csv(path=..., meta_results=...)
    • export_decisions_xlsx(path=..., meta_results=...)
    • export_metadata_audit_csv(path=..., meta_results=..., A_rows=...)
    • export_metadata_audit_xlsx(path=..., meta_results=..., A_rows=...)
    • prisma_counts(final_rows=...)
    • save_metadata_charts(outdir=..., meta_results=...)

- metadata.py (Contract v2 engine):
    meta_results["caches"][stage]["rows"] contains per-record stage rows, including:
        a_id, title, stage_outcome, stage_reason_summary, criteria_results (list)
    caches[stage]["survivor_ids"] optional
    caches["FINAL"]["rows"] optional (if IL ran in engine)

Contract v2 recap (reporting assumptions)
- Stages: EH → IH → EL → IL
- Stage outcomes: OUT | PASS_CLEAN | PASS_FLAGGED | REVIEW
- FINAL rows should contain:
    a_id, title, year, venue, lang, doc_type,
    discarded_at_stage, final_outcome,
    outcome_EH/IH/EL/IL, reasons_EH/IH/EL/IL, history

This module is intentionally conservative and robust:
- Works even if only some stages ran (FINAL will be "in progress").
- Exports never crash due to missing keys (best-effort).
- XLSX uses openpyxl if available.
- Charts use matplotlib if available.
"""

from __future__ import annotations

from typing import Any, Dict, Iterable, List, Optional, Tuple
import csv
import json
import os
import time


STAGES: Tuple[str, ...] = ("EH", "IH", "EL", "IL")


# -----------------------------
# Small robust helpers
# -----------------------------
def _safe_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    try:
        return str(x)
    except Exception:
        return repr(x)


def _is_blank(x: Any) -> bool:
    s = _safe_str(x).strip()
    return s == "" or s.lower() in {"nan", "none", "null"}


def _ensure_dir(path: str) -> None:
    if not path:
        return
    os.makedirs(path, exist_ok=True)


def _as_dict(x: Any) -> Dict[str, Any]:
    return x if isinstance(x, dict) else {}


def _as_list(x: Any) -> List[Any]:
    return x if isinstance(x, list) else []


def _get_cache(meta_results: Dict[str, Any], stage: str) -> Dict[str, Any]:
    caches = _as_dict(meta_results.get("caches"))
    return _as_dict(caches.get(stage))


def _get_stage_rows(meta_results: Dict[str, Any], stage: str) -> List[Dict[str, Any]]:
    cache = _get_cache(meta_results, stage)
    rows = cache.get("rows")
    if isinstance(rows, list):
        return [r for r in rows if isinstance(r, dict)]
    return []


def _get_final_rows_if_present(meta_results: Dict[str, Any]) -> List[Dict[str, Any]]:
    # Prefer explicit meta_results["final_results"] (engine may set)
    fr = meta_results.get("final_results")
    if isinstance(fr, list) and fr:
        return [r for r in fr if isinstance(r, dict)]

    # Then caches["FINAL"]["rows"]
    cache = _get_cache(meta_results, "FINAL")
    rows = cache.get("rows")
    if isinstance(rows, list) and rows:
        return [r for r in rows if isinstance(r, dict)]

    return []


def _build_a_index(A_rows: Optional[List[Dict[str, Any]]]) -> Dict[str, Dict[str, Any]]:
    idx: Dict[str, Dict[str, Any]] = {}
    for r in (A_rows or []):
        if not isinstance(r, dict):
            continue
        a_id = _safe_str(r.get("a_id") or r.get("id") or r.get("ID")).strip()
        if a_id:
            idx[a_id] = r
    return idx


def _pick_first(d: Dict[str, Any], keys: Iterable[str]) -> str:
    for k in keys:
        v = d.get(k)
        if not _is_blank(v):
            return _safe_str(v).strip()
    return ""


def _history_from_outcomes(outcomes: Dict[str, str]) -> str:
    parts = []
    for st in STAGES:
        v = _safe_str(outcomes.get(st)).strip()
        if v:
            parts.append(f"{st}:{v}")
    return ";".join(parts)


def _final_outcome_from_stages(
    out_eh: str,
    out_ih: str,
    out_el: str,
    out_il: str,
) -> Tuple[str, str]:
    """
    Returns (discarded_at_stage, final_outcome).
    If IL hasn't run, final_outcome is "" (in progress) unless already discarded earlier.
    """
    seq = [("EH", out_eh), ("IH", out_ih), ("EL", out_el), ("IL", out_il)]
    for st, ov in seq:
        if ov == "OUT":
            return st, "OUT"

    # If IL ran, use IL outcome; else in-progress
    if out_il:
        if out_il == "PASS_FLAGGED":
            return "", "REVIEW"
        return "", out_il

    return "", ""


def _normalize_final_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """
    Ensure row has the recommended FINAL keys, without deleting extras.
    """
    r = dict(row or {})
    # Common normalizations
    if r.get("final_outcome") == "PASS_FLAGGED":
        r["final_outcome"] = "REVIEW"
    return r


def _collect_columns(rows: List[Dict[str, Any]], preferred: Optional[List[str]] = None) -> List[str]:
    preferred = preferred or []
    keys = []
    seen = set()
    for k in preferred:
        if k not in seen:
            keys.append(k)
            seen.add(k)
    for r in rows:
        for k in r.keys():
            if k not in seen:
                keys.append(k)
                seen.add(k)
    return keys


def _write_csv(path: str, rows: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> None:
    columns = columns or _collect_columns(rows)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: _safe_str(r.get(k, "")) for k in columns})


# -----------------------------
# Aggregation
# -----------------------------
def aggregate_final(*, meta_results: Dict[str, Any], A_rows: Optional[List[Dict[str, Any]]] = None, **kwargs) -> List[Dict[str, Any]]:
    """
    Build FINAL rows, even if the engine didn't create caches["FINAL"].

    If only some stages ran:
      - records OUT in any completed stage are marked OUT
      - otherwise final_outcome stays blank until IL is present
    """
    if not isinstance(meta_results, dict):
        return []

    # If engine already produced FINAL rows, normalize and return
    existing = _get_final_rows_if_present(meta_results)
    if existing:
        return [_normalize_final_row(r) for r in existing]

    # Otherwise compute from stage caches
    a_index = _build_a_index(A_rows)

    # Maps: a_id -> stage row
    stage_row_map: Dict[Tuple[str, str], Dict[str, Any]] = {}
    stage_out_map: Dict[Tuple[str, str], str] = {}
    for st in STAGES:
        for r in _get_stage_rows(meta_results, st):
            a_id = _safe_str(r.get("a_id")).strip()
            if not a_id:
                continue
            stage_row_map[(st, a_id)] = r
            stage_out_map[(st, a_id)] = _safe_str(r.get("stage_outcome")).strip()

    # Determine candidate universe:
    # - Prefer A_rows if given
    # - Else infer from any stage rows
    a_ids: List[str] = []
    if A_rows:
        for r in A_rows:
            a_id = _safe_str(r.get("a_id") or r.get("id") or r.get("ID")).strip()
            if a_id:
                a_ids.append(a_id)
    else:
        inferred = sorted({a_id for (_, a_id) in stage_row_map.keys()})
        a_ids = inferred

    finals: List[Dict[str, Any]] = []
    for a_id in a_ids:
        A = a_index.get(a_id, {})

        out_eh = stage_out_map.get(("EH", a_id), "")
        out_ih = stage_out_map.get(("IH", a_id), "")
        out_el = stage_out_map.get(("EL", a_id), "")
        out_il = stage_out_map.get(("IL", a_id), "")

        discarded_at, final_outcome = _final_outcome_from_stages(out_eh, out_ih, out_el, out_il)

        reasons_eh = _safe_str(stage_row_map.get(("EH", a_id), {}).get("stage_reason_summary")).strip()
        reasons_ih = _safe_str(stage_row_map.get(("IH", a_id), {}).get("stage_reason_summary")).strip()
        reasons_el = _safe_str(stage_row_map.get(("EL", a_id), {}).get("stage_reason_summary")).strip()
        reasons_il = _safe_str(stage_row_map.get(("IL", a_id), {}).get("stage_reason_summary")).strip()

        title = _pick_first(A, ("title", "ti", "article_title", "document_title", "paper_title")) or _safe_str(
            stage_row_map.get(("EH", a_id), {}).get("title")
            or stage_row_map.get(("IH", a_id), {}).get("title")
            or stage_row_map.get(("EL", a_id), {}).get("title")
            or stage_row_map.get(("IL", a_id), {}).get("title")
        ).strip()

        year = _pick_first(A, ("year", "py", "publication_year", "pub_year"))
        venue = _pick_first(A, ("venue", "journal", "source", "source_title", "publication"))
        lang = _pick_first(A, ("lang", "language", "langue"))
        doc_type = _pick_first(A, ("doc_type", "document_type", "type", "doctype"))

        history = _history_from_outcomes({"EH": out_eh, "IH": out_ih, "EL": out_el, "IL": out_il})

        finals.append(_normalize_final_row({
            "a_id": a_id,
            "title": title,
            "year": year,
            "venue": venue,
            "lang": lang,
            "doc_type": doc_type,
            "final_outcome": final_outcome,
            "discarded_at_stage": discarded_at,
            "outcome_EH": out_eh,
            "outcome_IH": out_ih,
            "outcome_EL": out_el,
            "outcome_IL": out_il,
            "reasons_EH": reasons_eh,
            "reasons_IH": reasons_ih,
            "reasons_EL": reasons_el,
            "reasons_IL": reasons_il,
            "history": history,
        }))

    return finals


def aggregate_decisions(meta_results: Dict[str, Any], A_rows: Optional[List[Dict[str, Any]]] = None, *args, **kwargs) -> List[Dict[str, Any]]:
    """
    Backward-friendly alias.
    Some callers pass positional args, so we accept both.
    """
    return aggregate_final(meta_results=meta_results, A_rows=A_rows)


# -----------------------------
# PRISMA-ish counts
# -----------------------------
def prisma_counts(*, final_rows: List[Dict[str, Any]], meta_results: Optional[Dict[str, Any]] = None, **kwargs) -> Dict[str, Any]:
    """
    Return a compact dict of counts for logs/quick sanity checks.
    """
    final_rows = [r for r in (final_rows or []) if isinstance(r, dict)]
    n_total = len(final_rows)

    def _is_out(r: Dict[str, Any]) -> bool:
        return _safe_str(r.get("final_outcome")).strip() == "OUT"

    def _is_review(r: Dict[str, Any]) -> bool:
        fo = _safe_str(r.get("final_outcome")).strip()
        return fo in {"REVIEW", "PASS_FLAGGED"}

    def _is_clean(r: Dict[str, Any]) -> bool:
        return _safe_str(r.get("final_outcome")).strip() == "PASS_CLEAN"

    def _is_in_progress(r: Dict[str, Any]) -> bool:
        fo = _safe_str(r.get("final_outcome")).strip()
        if fo:
            return False
        # If it has any outcomes recorded but no final, it's in progress
        for k in ("outcome_EH", "outcome_IH", "outcome_EL", "outcome_IL"):
            if _safe_str(r.get(k)).strip():
                return True
        return True

    n_out = sum(1 for r in final_rows if _is_out(r))
    n_clean = sum(1 for r in final_rows if _is_clean(r))
    n_review = sum(1 for r in final_rows if _is_review(r))
    n_in_progress = sum(1 for r in final_rows if _is_in_progress(r))

    out = {
        "n_total": n_total,
        "n_out": n_out,
        "n_clean": n_clean,
        "n_review": n_review,
        "n_in_progress": n_in_progress,
    }

    # Add stage survivor counts if meta_results provided
    if isinstance(meta_results, dict):
        caches = _as_dict(meta_results.get("caches"))
        for st in STAGES:
            cache = _as_dict(caches.get(st))
            surv = cache.get("survivor_ids")
            rows = cache.get("rows")
            if isinstance(surv, list):
                out[f"n_survivors_{st}"] = len(surv)
            if isinstance(rows, list):
                out[f"n_rows_{st}"] = len(rows)
                out[f"n_out_{st}"] = sum(
                    1 for r in rows if isinstance(r, dict) and _safe_str(r.get("stage_outcome")).strip() == "OUT"
                )
    return out


# -----------------------------
# Exports — Decisions
# -----------------------------
def export_decisions_csv(*, path: str, meta_results: Dict[str, Any], A_rows: Optional[List[Dict[str, Any]]] = None, **kwargs) -> str:
    """
    Export FINAL decisions to a single CSV.
    """
    if not path:
        raise ValueError("export_decisions_csv: empty path")
    if not isinstance(meta_results, dict):
        raise ValueError("export_decisions_csv: meta_results must be a dict")

    final_rows = aggregate_final(meta_results=meta_results, A_rows=A_rows)
    preferred = [
        "a_id", "title", "year", "venue", "lang", "doc_type",
        "final_outcome", "discarded_at_stage",
        "outcome_EH", "outcome_IH", "outcome_EL", "outcome_IL",
        "reasons_EH", "reasons_IH", "reasons_EL", "reasons_IL",
        "history",
    ]
    cols = _collect_columns(final_rows, preferred=preferred)
    _write_csv(path, final_rows, cols)
    return path


def export_decisions_xlsx(*, path: str, meta_results: Dict[str, Any], A_rows: Optional[List[Dict[str, Any]]] = None, **kwargs) -> str:
    """
    Export stage tabs + FINAL to an XLSX workbook.
    """
    if not path:
        raise ValueError("export_decisions_xlsx: empty path")
    if not isinstance(meta_results, dict):
        raise ValueError("export_decisions_xlsx: meta_results must be a dict")

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise RuntimeError(f"openpyxl is required for XLSX export: {e}")

    def write_sheet(ws, rows: List[Dict[str, Any]], preferred_cols: Optional[List[str]] = None) -> None:
        rows = [r for r in (rows or []) if isinstance(r, dict)]
        cols = _collect_columns(rows, preferred=preferred_cols or [])
        ws.append(cols)
        for r in rows:
            ws.append([_safe_str(r.get(c, "")) for c in cols])

        ws.freeze_panes = "A2"
        # basic column width
        for i, c in enumerate(cols, start=1):
            max_len = len(_safe_str(c))
            for r in rows[:2000]:
                v = _safe_str(r.get(c, ""))
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, max_len + 2))

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    # Stage tabs
    preferred_stage = [
        "stage", "a_id", "title", "stage_outcome", "passed_to_next", "hard_stop",
        "hard_stop_criterion_id", "hard_stop_criterion_label", "stage_reason_summary"
    ]
    for st in STAGES:
        ws = wb.create_sheet(st)
        write_sheet(ws, _get_stage_rows(meta_results, st), preferred_cols=preferred_stage)

    # FINAL
    ws_f = wb.create_sheet("FINAL")
    final_rows = aggregate_final(meta_results=meta_results, A_rows=A_rows)
    preferred_final = [
        "a_id", "title", "year", "venue", "lang", "doc_type",
        "final_outcome", "discarded_at_stage",
        "outcome_EH", "outcome_IH", "outcome_EL", "outcome_IL",
        "reasons_EH", "reasons_IH", "reasons_EL", "reasons_IL",
        "history",
    ]
    write_sheet(ws_f, final_rows, preferred_cols=preferred_final)

    # META (optional, small)
    meta = _as_dict(_get_cache(meta_results, "meta") or meta_results.get("meta"))
    if meta:
        ws_m = wb.create_sheet("META")
        ws_m.append(["key", "value"])
        for k in sorted(meta.keys()):
            v = meta.get(k)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, ensure_ascii=False)
            ws_m.append([_safe_str(k), _safe_str(v)])
        ws_m.freeze_panes = "A2"
        ws_m.column_dimensions["A"].width = 28
        ws_m.column_dimensions["B"].width = 80

    wb.save(path)
    return path


# -----------------------------
# Exports — Metadata audit (criterion-level)
# -----------------------------
def _audit_rows(meta_results: Dict[str, Any], A_rows: Optional[List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    """
    Long-form audit rows: one row per (a_id, stage, criterion_result).

    Uses:
      stage cache rows[*]["criteria_results"] which include:
        criterion_id, criterion_label, operator, status, evidence (dict)
    """
    a_index = _build_a_index(A_rows)
    out: List[Dict[str, Any]] = []

    for st in STAGES:
        for sr in _get_stage_rows(meta_results, st):
            a_id = _safe_str(sr.get("a_id")).strip()
            A = a_index.get(a_id, {})
            stage_outcome = _safe_str(sr.get("stage_outcome")).strip()
            stage_reason = _safe_str(sr.get("stage_reason_summary")).strip()
            title = _pick_first(A, ("title",)) or _safe_str(sr.get("title")).strip()

            # bibliographic extras (best-effort)
            year = _pick_first(A, ("year", "py", "publication_year", "pub_year"))
            doi = _pick_first(A, ("doi",))
            journal = _pick_first(A, ("journal", "venue", "source", "source_title"))
            authors = _pick_first(A, ("authors", "author", "creator", "creators"))
            lang = _pick_first(A, ("lang", "language", "langue"))
            doc_type = _pick_first(A, ("doc_type", "document_type", "type", "doctype"))
            availability = _pick_first(A, ("availability", "access", "open_access", "full_text", "fulltext"))

            crit_results = sr.get("criteria_results")
            crit_results = crit_results if isinstance(crit_results, list) else []

            if not crit_results:
                out.append({
                    "a_id": a_id,
                    "stage": st,
                    "stage_outcome": stage_outcome,
                    "stage_reason_summary": stage_reason,
                    "criterion_id": "",
                    "criterion_label": "",
                    "operator": "",
                    "status": "",
                    "evidence_json": "",
                    "title": title,
                    "year": year,
                    "doi": doi,
                    "journal": journal,
                    "authors": authors,
                    "lang": lang,
                    "doc_type": doc_type,
                    "availability": availability,
                })
                continue

            for cr in crit_results:
                if not isinstance(cr, dict):
                    continue
                ev = cr.get("evidence")
                ev = ev if isinstance(ev, dict) else {}
                # Pull common evidence keys (heuristic + LLM)
                row = {
                    "a_id": a_id,
                    "stage": st,
                    "stage_outcome": stage_outcome,
                    "stage_reason_summary": stage_reason,
                    "criterion_id": _safe_str(cr.get("criterion_id")).strip(),
                    "criterion_label": _safe_str(cr.get("criterion_label")).strip(),
                    "operator": _safe_str(cr.get("operator")).strip(),
                    "status": _safe_str(cr.get("status")).strip(),

                    # evidence (common)
                    "evidence_field": _safe_str(ev.get("field")).strip(),
                    "evidence_note": _safe_str(ev.get("note")).strip(),
                    "evidence_value": _safe_str(ev.get("value")).strip(),
                    "evidence_terms": _safe_str(ev.get("terms")).strip(),
                    "evidence_matched": _safe_str(ev.get("matched")).strip(),
                    "evidence_pattern": _safe_str(ev.get("pattern")).strip(),

                    # evidence (LLM)
                    "evidence_decision": _safe_str(ev.get("decision")).strip(),
                    "evidence_confidence": _safe_str(ev.get("confidence")).strip(),
                    "evidence_quote": _safe_str(ev.get("quote")).strip(),
                    "evidence_quote_valid": _safe_str(ev.get("quote_valid")).strip(),
                    "evidence_rationale": _safe_str(ev.get("rationale")).strip(),

                    "evidence_json": json.dumps(ev, ensure_ascii=False),

                    # record fields
                    "title": title,
                    "year": year,
                    "doi": doi,
                    "journal": journal,
                    "authors": authors,
                    "lang": lang,
                    "doc_type": doc_type,
                    "availability": availability,
                }
                out.append(row)

    return out


def export_metadata_audit_csv(*, path: str, meta_results: Dict[str, Any], A_rows: List[Dict[str, Any]], **kwargs) -> str:
    """
    Export criterion-level audit to CSV (long format).
    """
    if not path:
        raise ValueError("export_metadata_audit_csv: empty path")
    if not isinstance(meta_results, dict):
        raise ValueError("export_metadata_audit_csv: meta_results must be a dict")

    rows = _audit_rows(meta_results, A_rows)
    preferred = [
        "a_id", "title", "year", "doi", "journal", "authors", "lang", "doc_type", "availability",
        "stage", "stage_outcome", "stage_reason_summary",
        "criterion_id", "criterion_label", "operator", "status",
        "evidence_field", "evidence_note", "evidence_value", "evidence_terms", "evidence_matched", "evidence_pattern",
        "evidence_decision", "evidence_confidence", "evidence_quote", "evidence_quote_valid", "evidence_rationale",
        "evidence_json",
    ]
    cols = _collect_columns(rows, preferred=preferred)
    _write_csv(path, rows, cols)
    return path


def export_metadata_audit_xlsx(*, path: str, meta_results: Dict[str, Any], A_rows: List[Dict[str, Any]], **kwargs) -> str:
    """
    Export criterion-level audit to XLSX (single sheet: AUDIT).
    """
    if not path:
        raise ValueError("export_metadata_audit_xlsx: empty path")
    if not isinstance(meta_results, dict):
        raise ValueError("export_metadata_audit_xlsx: meta_results must be a dict")

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise RuntimeError(f"openpyxl is required for XLSX export: {e}")

    rows = _audit_rows(meta_results, A_rows)
    preferred = [
        "a_id", "title", "year", "doi", "journal", "authors", "lang", "doc_type", "availability",
        "stage", "stage_outcome", "stage_reason_summary",
        "criterion_id", "criterion_label", "operator", "status",
        "evidence_field", "evidence_note", "evidence_value", "evidence_terms", "evidence_matched", "evidence_pattern",
        "evidence_decision", "evidence_confidence", "evidence_quote", "evidence_quote_valid", "evidence_rationale",
        "evidence_json",
    ]
    cols = _collect_columns(rows, preferred=preferred)

    wb = Workbook()
    ws = wb.active
    ws.title = "AUDIT"
    ws.append(cols)

    for r in rows:
        ws.append([_safe_str(r.get(c, "")) for c in cols])

    ws.freeze_panes = "A2"
    for i, c in enumerate(cols, start=1):
        max_len = len(_safe_str(c))
        for r in rows[:2000]:
            v = _safe_str(r.get(c, ""))
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, max_len + 2))

    wb.save(path)
    return path


# -----------------------------
# Charts (best-effort)
# -----------------------------
def save_metadata_charts(*, outdir: str, meta_results: Dict[str, Any], final_rows: Optional[List[Dict[str, Any]]] = None, **kwargs) -> List[str]:
    """
    Save simple PNG charts to outdir.
    Best-effort: if matplotlib isn't available, it writes nothing and returns [].
    """
    if not outdir:
        raise ValueError("save_metadata_charts: empty outdir")
    if not isinstance(meta_results, dict):
        raise ValueError("save_metadata_charts: meta_results must be a dict")

    _ensure_dir(outdir)
    saved: List[str] = []

    try:
        import matplotlib.pyplot as plt  # type: ignore
    except Exception:
        return saved

    def count_outcomes(rows: List[Dict[str, Any]], key: str) -> Dict[str, int]:
        counts: Dict[str, int] = {}
        for r in rows:
            v = _safe_str(r.get(key)).strip()
            if not v:
                v = "(blank)"
            counts[v] = counts.get(v, 0) + 1
        return counts

    # Stage charts
    for st in STAGES:
        rows = _get_stage_rows(meta_results, st)
        if not rows:
            continue
        counts = count_outcomes(rows, "stage_outcome")
        labels = list(counts.keys())
        values = [counts[k] for k in labels]

        plt.figure()
        plt.title(f"Stage {st} outcomes")
        plt.bar(labels, values)
        plt.xlabel("Outcome")
        plt.ylabel("Count")
        plt.xticks(rotation=20, ha="right")

        p = os.path.join(outdir, f"stage_{st}_outcomes.png")
        plt.tight_layout()
        plt.savefig(p, dpi=160)
        plt.close()
        saved.append(p)

    # FINAL chart
    if final_rows is None:
        final_rows = aggregate_final(meta_results=meta_results, A_rows=None)

    if final_rows:
        counts = count_outcomes(final_rows, "final_outcome")
        labels = list(counts.keys())
        values = [counts[k] for k in labels]

        plt.figure()
        plt.title("FINAL outcomes")
        plt.bar(labels, values)
        plt.xlabel("Final outcome")
        plt.ylabel("Count")
        plt.xticks(rotation=20, ha="right")

        p = os.path.join(outdir, "final_outcomes.png")
        plt.tight_layout()
        plt.savefig(p, dpi=160)
        plt.close()
        saved.append(p)

    # Simple PRISMA flow (survivors after each stage) if available
    caches = _as_dict(meta_results.get("caches"))
    survivors = []
    labels = []
    for st in STAGES:
        cache = _as_dict(caches.get(st))
        surv = cache.get("survivor_ids")
        if isinstance(surv, list):
            labels.append(st)
            survivors.append(len(surv))
    if survivors:
        plt.figure()
        plt.title("Survivors after each stage")
        plt.bar(labels, survivors)
        plt.xlabel("Stage")
        plt.ylabel("Survivors")

        p = os.path.join(outdir, "prisma_survivors.png")
        plt.tight_layout()
        plt.savefig(p, dpi=160)
        plt.close()
        saved.append(p)

    return saved


# -----------------------------
# Small convenience for callers/tests
# -----------------------------
def _self_test_smoke(meta_results: Dict[str, Any], A_rows: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    """
    Internal smoke test utility (not used by plugin).
    """
    finals = aggregate_final(meta_results=meta_results, A_rows=A_rows)
    return {
        "final_rows": len(finals),
        "counts": prisma_counts(final_rows=finals, meta_results=meta_results),
        "ts": time.time(),
    }
