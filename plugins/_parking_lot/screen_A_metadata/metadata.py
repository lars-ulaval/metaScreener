# -*- coding: utf-8 -*-
"""
metadata.py — Screen A (metadata-only) engine (Contract v2)

Fresh, contract-first implementation intended to be the stable foundation for:
- plugin.py (staged buttons EH → IH → EL → IL; progressive reuse via prior_result)
- criteria.py (criteria authoring/harmonization)
- decisions_report.py (exports + FINAL aggregation)
- core.py (optional shared helpers later)

Contract v2 (frozen):
- Stages: EH -> IH -> EL -> IL
- Criterion statuses: MET | FAILED | MISSING | UNCERTAIN
  - UNCERTAIN is produced only by LLM evaluation (or when LLM is disabled/unavailable)
- Stage decision rule (uniform):
  - if any criterion is FAILED        -> OUT
  - else if all criteria are MET      -> PASS_CLEAN
  - else                              -> PASS_FLAGGED (or REVIEW if stage == IL)
  - if stage has no criteria          -> PASS_CLEAN

Standardized include/exclude semantics so "any FAILED -> OUT" is valid for every stage:
- Include criterion:
    condition true  -> MET
    condition false -> FAILED
    missing field   -> MISSING
- Exclude criterion:
    condition true  -> FAILED   (exclusion hit)
    condition false -> MET
    missing field   -> MISSING

Public API expected by plugin layer:
- parse_A_file(path) -> (rows, meta)
- parse_A_csv_xlsx(path) -> rows          (compat alias)
- screen_metadata(A, criteria, ..., subrun="EH|IH|EL|IL", prior_result=..., progress=..., cancel_token=...) -> dict

Notes on LLM:
- This module is provider-agnostic.
- If OpenAI SDK is available AND OPENAI_API_KEY is set, it can call OpenAI.
- Otherwise LLM criteria return UNCERTAIN (conservative) rather than crashing.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Callable, Dict, Iterable, List, Literal, Optional, Tuple, Union
import csv
import hashlib
import json
import os
import random
import re
import time


# -----------------------------
# Contract vocabulary
# -----------------------------
Status = Literal["MET", "FAILED", "MISSING", "UNCERTAIN"]
Stage = Literal["EH", "IH", "EL", "IL"]
Outcome = Literal["OUT", "PASS_CLEAN", "PASS_FLAGGED", "REVIEW"]

STAGES: Tuple[Stage, ...] = ("EH", "IH", "EL", "IL")
PREV_STAGE: Dict[Stage, Optional[Stage]] = {"EH": None, "IH": "EH", "EL": "IH", "IL": "EL"}


# -----------------------------
# Cancellation / progress
# -----------------------------
class _Cancelled(RuntimeError):
    pass


def _check_cancel(cancel_token: Optional[object]) -> None:
    if cancel_token is None:
        return
    try:
        if bool(getattr(cancel_token, "cancelled", False)):
            raise _Cancelled("cancelled")
    except Exception:
        # If token is weird, ignore rather than crashing.
        return


def _emit(progress: Optional[Callable[[Dict[str, Any]], None]], evt: Dict[str, Any]) -> None:
    if not progress:
        return
    try:
        if "ts" not in evt:
            evt["ts"] = time.time()
        progress(evt)
    except Exception:
        # UI callback errors must never break the engine.
        pass


def _log(log_fn: Optional[Callable[[str], None]], msg: str) -> None:
    if not log_fn:
        return
    try:
        log_fn(msg)
    except Exception:
        pass


# -----------------------------
# Small robust helpers
# -----------------------------
def _safe_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    try:
        return str(x)
    except Exception:
        return ""


def _is_blank(x: Any) -> bool:
    s = _safe_str(x).strip()
    return s == "" or s.lower() in {"nan", "none", "null"}


def _norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def _norm_text_for_match(s: str) -> str:
    # Deterministic, stable normalization for matching
    s = s.replace("\u00a0", " ")
    s = s.lower()
    s = _norm_space(s)
    return s


def _split_terms(v: Any) -> List[str]:
    """
    Accepts:
      - list/tuple
      - "a|b|c"
      - newline/semicolon/comma separated
    """
    if v is None:
        return []
    if isinstance(v, (list, tuple)):
        out = []
        for it in v:
            s = _safe_str(it).strip()
            if s:
                out.append(s)
        return out
    s = _safe_str(v)
    if not s.strip():
        return []
    # common separators
    parts = re.split(r"[|\n;,]+", s)
    return [p.strip() for p in parts if p.strip()]


def _canonical_key(k: str) -> str:
    k = _safe_str(k).strip()
    k = k.replace("\ufeff", "")
    k = k.lower()
    k = re.sub(r"\s+", "_", k)
    k = re.sub(r"[^a-z0-9_]+", "_", k).strip("_")
    return k


def _canonicalize_headers(row: Dict[str, Any], idx: int) -> Dict[str, Any]:
    """
    Convert arbitrary input columns into a stable internal set.
    We keep unknown columns too (canonized).
    """
    out: Dict[str, Any] = {}
    for k, v in (row or {}).items():
        ck = _canonical_key(k)
        out[ck] = v

    # Common synonyms
    synonym_map = {
        "title": {"ti", "article_title", "document_title", "paper_title"},
        "abstract": {"ab", "summary", "resume", "résumé"},
        "keywords": {"kw", "key_words", "author_keywords"},
        "year": {"py", "publication_year", "pub_year"},
        "authors": {"author", "author_s", "creator", "creators"},
        "journal": {"source", "source_title", "publication", "venue"},
        "doi": {"digital_object_identifier"},
        "lang": {"language", "langue"},
        "doc_type": {"document_type", "doctype", "type"},
        "availability": {"access", "accessibility", "full_text", "fulltext", "open_access"},
        "a_id": {"id", "record_id", "recordid", "uid", "unique_id"},
    }

    for target, syns in synonym_map.items():
        if target in out and not _is_blank(out.get(target)):
            continue
        for s in syns:
            if s in out and not _is_blank(out.get(s)):
                out[target] = out.get(s)
                break

    # Ensure a_id exists
    a_id = _safe_str(out.get("a_id")).strip()
    if not a_id:
        out["a_id"] = f"A{idx+1:06d}"

    return out


def _get_field_text(item: Dict[str, Any], field: str) -> str:
    """
    Fetch a field from an A item, with a few conveniences.
    """
    field = _canonical_key(field)
    if field in {"any", "any_text", "text", "blob"}:
        parts = [
            _safe_str(item.get("title")),
            _safe_str(item.get("abstract")),
            _safe_str(item.get("keywords")),
            _safe_str(item.get("journal")),
            _safe_str(item.get("authors")),
        ]
        return _norm_space(" | ".join([p for p in parts if p.strip()]))
    return _safe_str(item.get(field))


def _select_by_ids(A: List[Dict[str, Any]], ids: List[str]) -> List[Dict[str, Any]]:
    want = {str(x) for x in (ids or [])}
    out: List[Dict[str, Any]] = []
    for it in (A or []):
        a_id = str(it.get("a_id"))
        if a_id in want:
            out.append(it)
    return out


def _stage_pol(stage: Stage, crit: Dict[str, Any]) -> Literal["include", "exclude"]:
    """
    Primary polarity is derived from stage:
      EH/EL => exclude
      IH/IL => include
    If a criterion explicitly carries polarity, we accept it only if it matches the stage.
    """
    stage_pol: Literal["include", "exclude"] = "exclude" if stage.startswith("E") else "include"
    pol = _safe_str(crit.get("polarity")).strip().lower()
    if pol in {"include", "exclude"} and pol == stage_pol:
        return stage_pol
    return stage_pol


# -----------------------------
# A ingest
# -----------------------------
def parse_A_file(path: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Load A from CSV or Excel.

    Returns:
      (rows, meta)
    where rows is a list of dicts with canonical keys:
      a_id, title, abstract, keywords, year, authors, journal, doi, lang, doc_type, availability, ...
    """
    if not path:
        raise ValueError("parse_A_file: empty path")

    ext = os.path.splitext(path)[1].lower()
    raw_rows: List[Dict[str, Any]] = []
    meta: Dict[str, Any] = {"path": path, "ext": ext}

    if ext in {".csv", ".tsv"}:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            if ext == ".tsv":
                reader = csv.DictReader(f, delimiter="\t")
            else:
                reader = csv.DictReader(f)
            raw_rows = [dict(r) for r in reader]

    elif ext in {".xlsx", ".xls"}:
        # Prefer pandas if installed; fallback to openpyxl for .xlsx only.
        try:
            import pandas as pd  # type: ignore

            df = pd.read_excel(path)  # type: ignore
            raw_rows = df.to_dict(orient="records")  # type: ignore
            meta["excel_reader"] = "pandas"
        except Exception:
            if ext == ".xls":
                raise RuntimeError("Reading .xls requires pandas + appropriate engine. Convert to .xlsx or .csv.")
            try:
                from openpyxl import load_workbook  # type: ignore

                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    raw_rows = []
                else:
                    headers = [(_safe_str(h).strip() or f"col_{i}") for i, h in enumerate(rows[0])]
                    raw_rows = []
                    for r in rows[1:]:
                        d = {}
                        for i, h in enumerate(headers):
                            d[h] = r[i] if i < len(r) else None
                        raw_rows.append(d)
                meta["excel_reader"] = "openpyxl"
            except Exception as e:
                raise RuntimeError(f"Failed to read Excel file: {e}")

    else:
        raise ValueError(f"Unsupported file extension: {ext}")

    out_rows: List[Dict[str, Any]] = []
    for i, r in enumerate(raw_rows):
        if not isinstance(r, dict):
            continue
        out_rows.append(_canonicalize_headers(r, idx=i))

    meta["n_rows"] = len(out_rows)
    meta["columns_preview"] = sorted({k for rr in out_rows[:50] for k in rr.keys()})
    return out_rows, meta


# Backward-compat alias expected by some older plugin variants
def parse_A_csv_xlsx(path: str) -> List[Dict[str, Any]]:
    rows, _ = parse_A_file(path)
    return rows


# -----------------------------
# Criterion evaluation (heuristic)
# -----------------------------
def _parse_between(value: Any) -> Optional[Tuple[float, float]]:
    if value is None:
        return None
    if isinstance(value, (list, tuple)) and len(value) >= 2:
        try:
            return float(value[0]), float(value[1])
        except Exception:
            return None
    s = _safe_str(value).strip()
    if not s:
        return None
    m = re.match(r"^\s*(-?\d+(?:\.\d+)?)\s*[-–]\s*(-?\d+(?:\.\d+)?)\s*$", s)
    if not m:
        return None
    try:
        return float(m.group(1)), float(m.group(2))
    except Exception:
        return None


def _as_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = _safe_str(x).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        # Try to extract first numeric token
        m = re.search(r"(-?\d+(?:\.\d+)?)", s)
        if not m:
            return None
        try:
            return float(m.group(1))
        except Exception:
            return None


def _heuristic_condition_true(crit: Dict[str, Any], item: Dict[str, Any]) -> Tuple[bool, Dict[str, Any]]:
    """
    Returns: (condition_true, evidence)
    condition_true means: "this criterion condition holds" (independent from include/exclude).
    """
    op = _safe_str(crit.get("operator")).strip().lower()
    field = _safe_str(crit.get("field") or crit.get("target") or crit.get("column") or "any_text").strip()
    raw_val = crit.get("value")
    if raw_val is None and "pattern" in crit:
        raw_val = crit.get("pattern")
    if raw_val is None and "term" in crit:
        raw_val = crit.get("term")

    txt = _get_field_text(item, field)
    if _is_blank(txt):
        return False, {"field": field, "operator": op, "note": "field_blank"}

    # Text operators
    if op in {"contains", "contain"}:
        needle = _safe_str(raw_val).strip()
        if not needle:
            return False, {"field": field, "operator": op, "note": "empty_value"}
        hay = _norm_text_for_match(txt)
        ned = _norm_text_for_match(needle)
        ok = ned in hay
        ev = {"field": field, "operator": op, "value": needle}
        if ok:
            ev["matched"] = needle
        return ok, ev

    if op in {"any_of", "any", "in_keywords"}:
        terms = _split_terms(raw_val)
        if not terms:
            return False, {"field": field, "operator": op, "note": "empty_terms"}
        hay = _norm_text_for_match(txt)
        matched = [t for t in terms if _norm_text_for_match(t) in hay]
        return (len(matched) > 0), {"field": field, "operator": op, "terms": terms, "matched": matched}

    if op in {"all_of", "all"}:
        terms = _split_terms(raw_val)
        if not terms:
            return False, {"field": field, "operator": op, "note": "empty_terms"}
        hay = _norm_text_for_match(txt)
        matched = [t for t in terms if _norm_text_for_match(t) in hay]
        ok = len(matched) == len(terms)
        return ok, {"field": field, "operator": op, "terms": terms, "matched": matched}

    if op in {"regex", "re"}:
        pat = _safe_str(raw_val)
        if not pat.strip():
            return False, {"field": field, "operator": op, "note": "empty_pattern"}
        try:
            rx = re.compile(pat, flags=re.IGNORECASE)
            m = rx.search(txt or "")
            return bool(m), {"field": field, "operator": op, "pattern": pat, "matched": (m.group(0) if m else None)}
        except re.error as e:
            return False, {"field": field, "operator": op, "pattern": pat, "error": f"bad_regex:{e}"}

    if op in {"equals", "eq"}:
        v = _safe_str(raw_val).strip()
        if not v:
            return False, {"field": field, "operator": op, "note": "empty_value"}
        ok = _norm_text_for_match(txt) == _norm_text_for_match(v)
        return ok, {"field": field, "operator": op, "value": v}

    if op in {"exists", "present"}:
        return (not _is_blank(txt)), {"field": field, "operator": op}

    # Numeric operators (common: year)
    if op in {"gte", "ge"}:
        x = _as_float(txt)
        y = _as_float(raw_val)
        if x is None or y is None:
            return False, {"field": field, "operator": op, "note": "non_numeric"}
        return (x >= y), {"field": field, "operator": op, "value": y, "got": x}

    if op in {"lte", "le"}:
        x = _as_float(txt)
        y = _as_float(raw_val)
        if x is None or y is None:
            return False, {"field": field, "operator": op, "note": "non_numeric"}
        return (x <= y), {"field": field, "operator": op, "value": y, "got": x}

    if op in {"between", "range"}:
        x = _as_float(txt)
        rng = _parse_between(raw_val)
        if x is None or not rng:
            return False, {"field": field, "operator": op, "note": "non_numeric_or_bad_range"}
        lo, hi = rng
        return (lo <= x <= hi), {"field": field, "operator": op, "range": [lo, hi], "got": x}

    # List membership (for categorical fields)
    if op in {"in_list", "in"}:
        terms = _split_terms(raw_val)
        if not terms:
            return False, {"field": field, "operator": op, "note": "empty_terms"}
        got = _norm_text_for_match(txt)
        ok = got in {_norm_text_for_match(t) for t in terms}
        return ok, {"field": field, "operator": op, "terms": terms, "got": txt}

    if op in {"not_in", "not_in_list"}:
        terms = _split_terms(raw_val)
        if not terms:
            return False, {"field": field, "operator": op, "note": "empty_terms"}
        got = _norm_text_for_match(txt)
        ok = got not in {_norm_text_for_match(t) for t in terms}
        return ok, {"field": field, "operator": op, "terms": terms, "got": txt}

    # Unknown operator: default to False (conservative)
    return False, {"field": field, "operator": op, "note": "unknown_operator"}


def _eval_heuristic(crit: Dict[str, Any], item: Dict[str, Any], stage: Stage) -> Tuple[Status, Dict[str, Any]]:
    """
    Apply include/exclude semantics to a heuristic condition.
    """
    pol = _stage_pol(stage, crit)
    field = _safe_str(crit.get("field") or crit.get("target") or crit.get("column") or "any_text").strip()
    txt = _get_field_text(item, field)

    if _is_blank(txt) and _canonical_key(field) not in {"any", "any_text", "text", "blob"}:
        # If criterion points to a specific field and it's missing, treat as MISSING
        ev = {"field": field, "polarity": pol, "note": "missing_field"}
        return "MISSING", ev

    cond_true, ev = _heuristic_condition_true(crit, item)
    ev["polarity"] = pol

    if pol == "include":
        # condition true => MET ; false => FAILED
        return ("MET" if cond_true else "FAILED"), ev

    # exclude:
    # condition true => FAILED (exclusion hit) ; false => MET
    return ("FAILED" if cond_true else "MET"), ev


# -----------------------------
# Criterion evaluation (LLM)
# -----------------------------
@dataclass
class _LlmResult:
    decision: str
    confidence: float
    quote: str
    quote_valid: bool
    rationale: str
    raw: Dict[str, Any]


def _hash_payload(payload: Dict[str, Any]) -> str:
    b = json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(b).hexdigest()


def _llm_available() -> bool:
    # Minimal check; we only support OpenAI if installed + key set.
    if not os.getenv("OPENAI_API_KEY"):
        return False
    try:
        import openai  # noqa: F401
        return True
    except Exception:
        return False


def _call_openai_json(model: str, system: str, user: str, timeout_s: int = 60) -> Dict[str, Any]:
    """
    Best-effort OpenAI call with JSON response.
    Supports both classic 'openai' and new 'OpenAI' client styles if present.
    """
    # New style
    try:
        from openai import OpenAI  # type: ignore

        client = OpenAI()
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=0,
        )
        txt = resp.choices[0].message.content or ""
        return json.loads(txt)
    except Exception:
        pass

    # Classic style
    try:
        import openai  # type: ignore

        resp = openai.ChatCompletion.create(  # type: ignore[attr-defined]
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=0,
            request_timeout=timeout_s,
        )
        txt = resp["choices"][0]["message"]["content"] or ""
        return json.loads(txt)
    except Exception as e:
        raise RuntimeError(f"LLM call failed: {e}")


def _llm_prompt(
    crit: Dict[str, Any],
    item: Dict[str, Any],
    stage: Stage,
    trunc_chars: int
) -> Tuple[str, str]:
    """
    Build (system, user) messages. Provider-agnostic.
    """
    pol = _stage_pol(stage, crit)
    label = _safe_str(crit.get("label") or crit.get("name") or crit.get("id") or "").strip() or "Unnamed criterion"
    desc = _safe_str(crit.get("description") or crit.get("prompt") or crit.get("details") or "").strip()

    title = _safe_str(item.get("title")).strip()
    abstract = _safe_str(item.get("abstract")).strip()
    keywords = _safe_str(item.get("keywords")).strip()

    blob = _norm_space(f"TITLE: {title}\n\nABSTRACT: {abstract}\n\nKEYWORDS: {keywords}")
    if trunc_chars and len(blob) > trunc_chars:
        blob = blob[:trunc_chars] + "…"

    system = (
        "You are a cautious screening assistant for PRISMA-like metadata screening.\n"
        "Return ONLY valid JSON. Do not include extra text.\n"
        "You must be conservative: if evidence is weak, ambiguous, or absent, return decision='uncertain'.\n"
        "If the metadata does not contain enough information to decide, return decision='uncertain'.\n"
        "If the relevant field is empty, you may return decision='missing'.\n"
    )

    user = {
        "task": "Evaluate whether the record MEETS the criterion condition (condition true/false/uncertain).",
        "criterion": {
            "id": _safe_str(crit.get("id")),
            "label": label,
            "description": desc,
            "polarity": pol,  # include/exclude (for context only)
            "stage": stage,
        },
        "record": {
            "a_id": _safe_str(item.get("a_id")),
            "title": title,
            "abstract": abstract,
            "keywords": keywords,
            "blob": blob,
        },
        "output_schema": {
            "decision": "meet | not_meet | uncertain | missing",
            "confidence": "0.0 to 1.0",
            "quote": "a short quote from the blob supporting your decision (empty if uncertain/missing)",
            "quote_valid": "true if quote is a verbatim substring of blob; else false",
            "rationale": "short explanation (1-3 sentences)",
        }
    }
    return system, json.dumps(user, ensure_ascii=False)


def _parse_llm_result(d: Dict[str, Any]) -> _LlmResult:
    decision = _safe_str(d.get("decision")).strip().lower()
    conf = d.get("confidence")
    try:
        confidence = float(conf)
    except Exception:
        confidence = 0.0
    quote = _safe_str(d.get("quote")).strip()
    qv = bool(d.get("quote_valid", False))
    rationale = _safe_str(d.get("rationale")).strip()
    if decision not in {"meet", "not_meet", "uncertain", "missing"}:
        decision = "uncertain"
    if confidence < 0.0:
        confidence = 0.0
    if confidence > 1.0:
        confidence = 1.0
    return _LlmResult(decision=decision, confidence=confidence, quote=quote, quote_valid=qv, rationale=rationale, raw=d)


def _eval_llm(
    crit: Dict[str, Any],
    item: Dict[str, Any],
    stage: Stage,
    *,
    llm_model: Optional[str],
    trunc_chars: int,
    llm_cache: Dict[str, Any],
    log: Optional[Callable[[str], None]],
    progress: Optional[Callable[[Dict[str, Any]], None]],
    cancel_token: Optional[object],
    conf_threshold: float = 0.70,
) -> Tuple[Status, Dict[str, Any], Dict[str, Any]]:
    """
    Returns: (status, evidence, llm_cache)
    """
    pol = _stage_pol(stage, crit)
    field = _safe_str(crit.get("field") or crit.get("target") or crit.get("column") or "any_text").strip()

    # If there's no usable text at all, treat as MISSING
    blob = _get_field_text(item, "any_text")
    if _is_blank(blob):
        return "MISSING", {"polarity": pol, "note": "missing_all_text"}, llm_cache

    # If model/provider unavailable, be conservative (UNCERTAIN)
    if not llm_model or not _llm_available():
        return "UNCERTAIN", {"polarity": pol, "note": "llm_unavailable_or_disabled"}, llm_cache

    _check_cancel(cancel_token)

    sys_msg, user_msg = _llm_prompt(crit, item, stage, trunc_chars=trunc_chars)
    cache_key = _hash_payload({
        "model": llm_model,
        "stage": stage,
        "criterion_id": _safe_str(crit.get("id")),
        "criterion_label": _safe_str(crit.get("label") or crit.get("name")),
        "prompt_user": user_msg,
    })

    if cache_key in llm_cache:
        d = llm_cache[cache_key]
    else:
        _emit(progress, {"kind": "l_criterion_start", "stage": stage, "a_id": item.get("a_id"), "criterion_id": crit.get("id")})
        try:
            d = _call_openai_json(llm_model, sys_msg, user_msg)
        except Exception as e:
            _log(log, f"[LLM] error: {e}\n")
            return "UNCERTAIN", {"polarity": pol, "note": f"llm_error:{e}"}, llm_cache
        llm_cache[cache_key] = d
        _emit(progress, {"kind": "l_criterion_done", "stage": stage, "a_id": item.get("a_id"), "criterion_id": crit.get("id")})

    res = _parse_llm_result(d)

    ev: Dict[str, Any] = {
        "polarity": pol,
        "field": field,
        "decision": res.decision,
        "confidence": res.confidence,
        "quote": res.quote,
        "quote_valid": bool(res.quote_valid),
        "rationale": res.rationale,
    }

    # Conservative gating: only MET/FAILED if high confidence + quote_valid
    if res.decision == "missing":
        return "MISSING", ev, llm_cache

    if res.decision in {"meet", "not_meet"} and (res.confidence >= conf_threshold) and bool(res.quote_valid):
        cond_true = (res.decision == "meet")  # criterion condition true/false
        if pol == "include":
            # true => MET ; false => FAILED
            return ("MET" if cond_true else "FAILED"), ev, llm_cache
        # exclude: true => FAILED ; false => MET
        return ("FAILED" if cond_true else "MET"), ev, llm_cache

    return "UNCERTAIN", ev, llm_cache


# -----------------------------
# Stage runner
# -----------------------------
def _stage_criteria(criteria: List[Dict[str, Any]], stage: Stage) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for c in (criteria or []):
        if not isinstance(c, dict):
            continue
        if not bool(c.get("enabled", True)):
            continue
        st = _safe_str(c.get("stage")).upper().strip()
        if st != stage:
            continue
        out.append(c)
    # Stable order: preserve input order, but if "order" present, sort by it.
    if any("order" in c for c in out):
        def _k(x: Dict[str, Any]) -> Tuple[int, str]:
            try:
                return (int(x.get("order", 10_000)), _safe_str(x.get("id")))
            except Exception:
                return (10_000, _safe_str(x.get("id")))
        out = sorted(out, key=_k)
    return out


def _stage_name(stage: Stage) -> str:
    return {"EH": "E/H", "IH": "I/H", "EL": "E/L", "IL": "I/L"}[stage]


def _run_stage(
    stage: Stage,
    A_working: List[Dict[str, Any]],
    criteria: List[Dict[str, Any]],
    *,
    llm_model: Optional[str],
    llm_trunc_chars: int,
    llm_batch_size: int,  # accepted for compatibility; not required by contract
    llm_cache: Dict[str, Any],
    log: Optional[Callable[[str], None]],
    progress: Optional[Callable[[Dict[str, Any]], None]],
    cancel_token: Optional[object],
    hard_stop: bool,
) -> Dict[str, Any]:
    """
    Runs exactly one stage over A_working.
    Returns:
      {
        "rows": [...],
        "survivors": [...],
        "survivor_ids": [...],
        "outcome_map": {a_id: stage_outcome},
        "llm_cache": {...},
      }
    """
    _check_cancel(cancel_token)

    stage_criteria = _stage_criteria(criteria, stage)
    total_items = len(A_working)
    total_criteria = len(stage_criteria)

    _emit(progress, {"stage": stage, "msg": f"running {stage} ({_stage_name(stage)})", "done": 0, "total": total_items})

    rows: List[Dict[str, Any]] = []
    survivors: List[Dict[str, Any]] = []
    outcome_map: Dict[str, Any] = {}

    for idx, item in enumerate(A_working):
        _check_cancel(cancel_token)
        a_id = _safe_str(item.get("a_id"))
        title = _safe_str(item.get("title"))

        failed_ids: List[str] = []
        missing_ids: List[str] = []
        uncertain_ids: List[str] = []
        met_ids: List[str] = []

        crit_results: List[Dict[str, Any]] = []
        hard_stop_id = ""
        hard_stop_label = ""

        if total_criteria == 0:
            outcome: Outcome = "PASS_CLEAN"
            reason_summary = "no_criteria"
        else:
            for cidx, crit in enumerate(stage_criteria):
                _check_cancel(cancel_token)

                cid = _safe_str(crit.get("id") or f"{stage}_{cidx+1}")
                clabel = _safe_str(crit.get("label") or crit.get("name") or cid)

                op = _safe_str(crit.get("operator")).strip().lower()
                is_llm = (op == "llm")

                _emit(progress, {
                    "kind": "criterion_start",
                    "stage": stage,
                    "a_id": a_id,
                    "criterion_id": cid,
                    "criterion_label": clabel,
                    "done": idx,
                    "total": total_items,
                })

                if is_llm:
                    status, ev, llm_cache = _eval_llm(
                        crit, item, stage,
                        llm_model=llm_model,
                        trunc_chars=llm_trunc_chars,
                        llm_cache=llm_cache,
                        log=log,
                        progress=progress,
                        cancel_token=cancel_token,
                    )
                else:
                    status, ev = _eval_heuristic(crit, item, stage)

                crit_results.append({
                    "criterion_id": cid,
                    "criterion_label": clabel,
                    "operator": op,
                    "status": status,
                    "evidence": ev,
                })

                if status == "FAILED":
                    failed_ids.append(cid)
                    if hard_stop and not hard_stop_id:
                        hard_stop_id = cid
                        hard_stop_label = clabel
                        _emit(progress, {"kind": "hard_stop", "stage": stage, "a_id": a_id, "criterion_id": cid})
                        break
                elif status == "MISSING":
                    missing_ids.append(cid)
                elif status == "UNCERTAIN":
                    uncertain_ids.append(cid)
                elif status == "MET":
                    met_ids.append(cid)

                _emit(progress, {
                    "kind": "criterion_done",
                    "stage": stage,
                    "a_id": a_id,
                    "criterion_id": cid,
                })

            # Stage decision rule
            if failed_ids:
                outcome = "OUT"
            else:
                if total_criteria == 0:
                    outcome = "PASS_CLEAN"
                else:
                    all_met = (len(met_ids) == total_criteria)
                    if all_met:
                        outcome = "PASS_CLEAN"
                    else:
                        outcome = "REVIEW" if stage == "IL" else "PASS_FLAGGED"

            parts: List[str] = []
            if failed_ids:
                parts.append("FAILED=" + ",".join(failed_ids))
            if missing_ids:
                parts.append("MISSING=" + ",".join(missing_ids))
            if uncertain_ids:
                parts.append("UNCERTAIN=" + ",".join(uncertain_ids))
            if met_ids:
                parts.append("MET=" + ",".join(met_ids))
            reason_summary = " | ".join(parts) if parts else "no_criteria"

        row = {
            "a_id": a_id,
            "title": title,
            "stage": _stage_name(stage),
            "stage_outcome": outcome,
            "passed_to_next": bool(outcome != "OUT"),
            "hard_stop": bool(hard_stop_id),
            "hard_stop_criterion_id": hard_stop_id,
            "hard_stop_criterion_label": hard_stop_label,
            "failed_criteria_ids": failed_ids,
            "missing_criteria_ids": missing_ids,
            "uncertain_criteria_ids": uncertain_ids,
            "met_criteria_ids": met_ids,
            "criteria_results": crit_results,     # detailed per-criterion results (reporting will use this)
            "stage_reason_summary": reason_summary,
        }

        rows.append(row)
        outcome_map[a_id] = outcome

        if outcome != "OUT":
            survivors.append(item)

        # periodic progress
        if (idx + 1) % 10 == 0 or (idx + 1) == total_items:
            _emit(progress, {"stage": stage, "msg": "progress", "done": idx + 1, "total": total_items})

    survivor_ids = [str(it.get("a_id")) for it in survivors]
    return {
        "rows": rows,
        "survivors": survivors,
        "survivor_ids": survivor_ids,
        "outcome_map": outcome_map,
        "llm_cache": llm_cache,
    }


# -----------------------------
# FINAL builder (only after IL)
# -----------------------------
def _build_final_rows(A_all: List[Dict[str, Any]], caches: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Produce FINAL rows (one per A item):
      - outcome_EH/IH/EL/IL
      - discarded_at_stage
      - final_outcome
      - reasons_EH/IH/EL/IL (stage_reason_summary)
    """
    def _outcome_map(stage: Stage) -> Dict[str, str]:
        m = (caches.get(stage) or {}).get("outcome_map")
        if isinstance(m, dict) and m:
            return {str(k): str(v) for k, v in m.items()}
        # fallback from rows
        out: Dict[str, str] = {}
        for r in ((caches.get(stage) or {}).get("rows") or []):
            out[str(r.get("a_id"))] = str(r.get("stage_outcome"))
        return out

    oEH = _outcome_map("EH")
    oIH = _outcome_map("IH")
    oEL = _outcome_map("EL")
    oIL = _outcome_map("IL")

    # reasons lookup
    rows_map: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for st in STAGES:
        for r in ((caches.get(st) or {}).get("rows") or []):
            rows_map[(st, str(r.get("a_id")))] = r

    finals: List[Dict[str, Any]] = []
    for it in (A_all or []):
        a_id = str(it.get("a_id"))
        out_eh = oEH.get(a_id, "")
        out_ih = oIH.get(a_id, "")
        out_el = oEL.get(a_id, "")
        out_il = oIL.get(a_id, "")

        discarded_at = ""
        for st, ov in [("EH", out_eh), ("IH", out_ih), ("EL", out_el), ("IL", out_il)]:
            if ov == "OUT":
                discarded_at = st
                break

        final_outcome = "OUT" if discarded_at else (out_il or "")
        if final_outcome == "PASS_FLAGGED":
            final_outcome = "REVIEW"

        finals.append({
            "a_id": a_id,
            "title": _safe_str(it.get("title")),
            "outcome_EH": out_eh,
            "outcome_IH": out_ih,
            "outcome_EL": out_el,
            "outcome_IL": out_il,
            "discarded_at_stage": discarded_at,
            "final_outcome": final_outcome,
            "reasons_EH": rows_map.get(("EH", a_id), {}).get("stage_reason_summary", ""),
            "reasons_IH": rows_map.get(("IH", a_id), {}).get("stage_reason_summary", ""),
            "reasons_EL": rows_map.get(("EL", a_id), {}).get("stage_reason_summary", ""),
            "reasons_IL": rows_map.get(("IL", a_id), {}).get("stage_reason_summary", ""),
        })

    return finals


# -----------------------------
# Main entry (single-stage progressive engine)
# -----------------------------
def screen_metadata(
    A: List[Dict[str, Any]],
    criteria: List[Dict[str, Any]],
    *,
    # Backward-compatible aliases
    A_rows: Optional[List[Dict[str, Any]]] = None,
    criteria_rows: Optional[List[Dict[str, Any]]] = None,
    # Legacy parameters kept for compatibility (unused by contract rule)
    pass_thr: float = 0.60,
    border_thr: float = 0.40,
    missing_policy: str = "unknown",
    llm_model: Optional[str] = None,
    llm_trunc_chars: int = 1500,
    llm_batch_size: int = 8,
    stage_h_include_mode: str = "all",
    stage_l_include_mode: str = "all",
    randomize_within_blocks: bool = False,
    random_seed: Optional[Union[str, int]] = None,
    log: Optional[Callable[[str], None]] = None,
    hard_stop: bool = False,
    progress: Optional[Callable[[Dict[str, Any]], None]] = None,
    cancel_token: Optional[object] = None,
    subrun: Optional[Literal["EH", "IH", "EL", "IL"]] = None,
    return_stage_caches: bool = True,
    reuse_from_stage: Optional[Literal["EH", "IH", "EL"]] = None,
    initial_a_ids: Optional[List[str]] = None,
    prior_result: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Contract v2 expects staged execution:
      - subrun must be one of EH/IH/EL/IL
      - IH/EL/IL require prior_result so the working set is survivors from the previous stage

    Returns:
      {
        "mode": <subrun>,
        "caches": { "EH": {...}, "IH": {...}, "EL": {...}, "IL": {...}, "FINAL": {...}, "meta": {...} },
        "final_results": [...]   # only for IL
      }
    """
    _check_cancel(cancel_token)

    # Normalize aliases
    if A_rows is not None:
        A = A_rows
    if criteria_rows is not None:
        criteria = criteria_rows

    mode = _safe_str(subrun).upper().strip()
    if mode not in {"EH", "IH", "EL", "IL"}:
        raise ValueError("Contract v2 requires subrun one of: 'EH', 'IH', 'EL', 'IL'.")

    stage: Stage = mode  # type: ignore[assignment]

    # Seeded randomization (optional, deterministic)
    A_all = list(A or [])
    if randomize_within_blocks:
        try:
            seed_val = random_seed if random_seed is not None else "0"
            rng = random.Random(str(seed_val))
            rng.shuffle(A_all)
        except Exception:
            pass

    # Pull prior caches if present
    prior_caches: Dict[str, Any] = {}
    if isinstance(prior_result, dict):
        if isinstance(prior_result.get("caches"), dict):
            prior_caches = dict(prior_result.get("caches") or {})
        else:
            # Some callers pass caches directly
            prior_caches = dict(prior_result)

    # Carry LLM cache forward if present
    llm_cache: Dict[str, Any] = {}
    try:
        llm_cache = dict((prior_caches.get("meta") or {}).get("llm_cache") or {})
    except Exception:
        llm_cache = {}

    # Strict progressive working set
    working = A_all
    if stage != "EH":
        prev = PREV_STAGE[stage]
        if not prev:
            raise ValueError("Internal error: prev stage missing")
        prev_ids = (prior_caches.get(prev) or {}).get("survivor_ids")
        if not isinstance(prev_ids, list):
            raise ValueError(f"Progressive flow requires stage {prev} completed before running {stage}.")
        working = _select_by_ids(A_all, prev_ids)

        # IL requires EH/IH/EL rows for FINAL reasons
        if stage == "IL":
            for req in ("EH", "IH", "EL"):
                if not isinstance((prior_caches.get(req) or {}).get("survivor_ids"), list):
                    raise ValueError(f"Progressive flow requires stage {req} completed before running IL.")
                if not isinstance((prior_caches.get(req) or {}).get("rows"), list):
                    raise ValueError(f"Progressive flow requires stage {req} rows before running IL (needed for FINAL).")

    # Merge forward caches (keep earlier stage results intact)
    caches: Dict[str, Any] = {}
    if prior_caches:
        caches.update(prior_caches)

    # Run stage
    _log(log, f"[ENGINE] Running stage {stage} on {len(working)} record(s).\n")
    out = _run_stage(
        stage,
        working,
        criteria,
        llm_model=llm_model,
        llm_trunc_chars=llm_trunc_chars,
        llm_batch_size=llm_batch_size,
        llm_cache=llm_cache,
        log=log,
        progress=progress,
        cancel_token=cancel_token,
        hard_stop=hard_stop,
    )

    # Save stage cache
    caches[stage] = {
        "rows": out["rows"],
        "survivor_ids": out["survivor_ids"],
        "outcome_map": out["outcome_map"],
        # small conveniences (safe to ignore by reporters)
        "dropped_records": [r for r in out["rows"] if r.get("stage_outcome") == "OUT"],
        "criteria_count": len(_stage_criteria(criteria, stage)),
    }

    final_results: List[Dict[str, Any]] = []

    # Build FINAL only after IL
    if stage == "IL":
        final_results = _build_final_rows(A_all, caches)
        caches["FINAL"] = {
            "rows": final_results,
            "survivor_ids": [r["a_id"] for r in final_results if r.get("final_outcome") != "OUT"],
        }

    # meta cache (non-contract, diagnostic)
    caches["meta"] = {
        "contract": "v2",
        "ts": time.time(),
        "mode": stage,
        "n_A_total": len(A_all),
        "n_A_working": len(working),
        "llm_model": llm_model,
        "llm_trunc_chars": llm_trunc_chars,
        "llm_batch_size": llm_batch_size,
        "hard_stop": bool(hard_stop),
        # legacy fields preserved so plugin UI doesn't look "broken"
        "pass_thr": pass_thr,
        "border_thr": border_thr,
        "missing_policy": missing_policy,
        "stage_h_include_mode": stage_h_include_mode,
        "stage_l_include_mode": stage_l_include_mode,
        # carry forward cache
        "llm_cache": out.get("llm_cache", llm_cache),
    }

    return {
        "mode": stage,
        "caches": caches if return_stage_caches else {},
        "final_results": final_results,
    }
