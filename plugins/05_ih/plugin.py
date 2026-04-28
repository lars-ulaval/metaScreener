# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""
plugin.py — Screen A (IH-only) as a metaScreener tab plugin (Contract v2, IH stage)

UPDATED: Bundle-first input (Harmoniser ZIP)

✅ IH-only, self-contained (UI + engine)
✅ Expects a ScreenA bundle ZIP as input (produced by your harmoniser)
✅ Shows the IH criteria being applied (persistent left panel)
✅ Click-to-sort columns (full dataset; re-renders incrementally to keep UI responsive)
✅ Duplicate local_id: warn only (no enforcement)
✅ Row detail modal on double-click (recomputes per-criterion evaluation on demand)
✅ Criterion double-click filters reports to rows touched by that criterion (failed/missing/met/unknown)

IH semantics (same tags as EH; different assignment meaning)
- Per row, evaluate all IH criteria -> {MET, FAILED, MISSING, UNKNOWN}
- Outcome assignment:
    - OUT          if >=1 FAILED
    - PASS_FLAGGED if 0 FAILED and >=1 (MISSING or UNKNOWN)
    - PASS_CLEAN   if all MET
- Survivors = PASS_CLEAN + PASS_FLAGGED (forwarded to next stage bundle as data/current.csv)
- NOTE: In IH, OUT means "removed from pipeline after IH" (typically: confidently includable by heuristics).

Inputs (from bundle ZIP)
- data/current.csv (preferred)  [fallbacks supported: data/A.csv, data/aggregate.csv]
- criteria/criteria_harmonized.csv (preferred) [fallbacks supported: criteria/harmonized.csv, criteria/criteria.csv, criteria.csv]
- manifest.json (optional fields are OK; plugin will only warn if fields are missing)
Optional
- data/input_errors.csv (carried through; merged with any new parse skips)

Outputs
1) IH_FULL report (UI + export): all aggregate columns + outcome + reason columns
2) IH_SURVIVORS report (UI + export): survivors only (PASS_CLEAN + PASS_FLAGGED),
   in the exact same schema as the original aggregate input

Exports
- One XLSX with two sheets: IH_FULL, IH_SURVIVORS
- Optional input_errors.csv (skipped/invalid records)
- Next bundle ZIP (post-IH): survivors become new data/current.csv, manifest pipeline updated, and reports/ saved

Operator guardrails
- IH is heuristics-only. If a criterion uses operator "llm" (misconfig), it is treated as UNKNOWN (warn-only),
  and NEVER triggers any LLM call (same behavior spirit as EH).

Notes
- SHA256 checks are WARN-ONLY (never block the run).
"""

import csv
import io
import json
import re
import threading
import zipfile
from dataclasses import dataclass
from datetime import datetime
from hashlib import sha256
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from metascreener.plugin_api import BasePlugin, PluginMeta


TAB_TITLE = "Screen A — IH"

MAX_UI_ROWS_HINT = 2000  # only used for status text; we render all rows incrementally
RENDER_CHUNK = 300
PROGRESS_EVERY_N = 200

TRUTHY = {"1", "true", "yes", "y", "on", "enabled"}


# ----------------------------
# Normalization maps
# ----------------------------

DOC_TYPE_MAP = {
    # journal-like
    "journal": "journal",
    "journal article": "journal",
    "article": "journal",
    "research article": "journal",
    "original article": "journal",
    "journal-article": "journal",
    "peer reviewed article": "journal",
    # conference/proceedings-like
    "conference": "conference",
    "conference paper": "conference",
    "conference article": "conference",
    "proceedings": "conference",
    "proceedings article": "conference",
    "inproceedings": "conference",
    "conference proceedings": "conference",
    # misc common
    "preprint": "preprint",
    "arxiv": "preprint",
    "thesis": "thesis",
    "dissertation": "thesis",
    "book chapter": "book_chapter",
    "chapter": "book_chapter",
    "report": "report",
}

LANG_MAP = {
    "en": "en",
    "eng": "en",
    "english": "en",
    "en-us": "en",
    "en-gb": "en",
    "fr": "fr",
    "fra": "fr",
    "fre": "fr",
    "french": "fr",
    "fr-ca": "fr",
    "es": "es",
    "spa": "es",
    "spanish": "es",
}


# ----------------------------
# Data structures
# ----------------------------

@dataclass
class Criterion:
    stage: str
    cid: str
    ctype: str           # include / exclude
    scope: str
    label: str
    operator: str
    targets: List[str]   # parsed list of targets
    what_raw: str
    what_list: List[str]
    threshold: Optional[float]
    enabled: bool
    source_text: str


@dataclass
class ParseReport:
    header: List[str]
    rows: List[Dict[str, str]]           # integral rows only
    skipped: List[Tuple[int, str, str]]  # (record_index_1based_ex_header, reason, raw_record_text)


@dataclass
class CriteriaLoadReport:
    criteria: List[Criterion]
    warnings: List[str]


@dataclass
class BundleInfo:
    zip_path: str
    root: str                       # e.g. "ScreenA_Bundle/" or ""
    manifest: Dict[str, Any]
    members: List[str]


# ----------------------------
# Utilities
# ----------------------------

def _now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _iso_now() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _safe_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    try:
        return str(x)
    except Exception:
        return ""


def _norm_basic(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _truthy(x: Any) -> bool:
    return _safe_str(x).strip().lower() in TRUTHY


def _split_targets(target_cell: str) -> List[str]:
    t = _safe_str(target_cell)
    out: List[str] = []
    for p in t.split(","):
        p = p.strip().strip('"').strip("'")
        if p:
            out.append(p)
    return out


def _split_what_list(what_cell: str) -> List[str]:
    w = _safe_str(what_cell).replace("\n", ";").replace("\r", ";")
    out: List[str] = []
    for p in w.split(";"):
        p = p.strip().strip('"').strip("'")
        if p:
            out.append(p)
    return out


def _norm_doc_type(v: str) -> str:
    x = _norm_basic(v).lower()
    x = x.replace("_", " ").replace("-", " ")
    x = re.sub(r"\s+", " ", x).strip()
    return DOC_TYPE_MAP.get(x, x)


def _norm_lang(v: str) -> str:
    x = _norm_basic(v).lower()
    x = x.replace("_", "-")
    x = re.sub(r"\s+", "", x)
    return LANG_MAP.get(x, x)


def _norm_for_target(target: str, value: str) -> str:
    t = (target or "").strip().lower()
    v = _safe_str(value)
    if not v.strip():
        return ""
    if t == "doc_type":
        return _norm_doc_type(v)
    if t == "lang":
        return _norm_lang(v)
    return _norm_basic(v)


def _norm_what_for_target(target: str, what: str) -> str:
    return _norm_for_target(target, what)


def _sha256_hex(data: bytes) -> str:
    h = sha256()
    h.update(data)
    return h.hexdigest()


def _decode_bytes(b: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="ignore")


# ----------------------------
# Bundle IO
# ----------------------------

def _detect_bundle_root(members: Sequence[str]) -> str:
    """
    Determine bundle root prefix. Supports:
      - manifest.json at zip root
      - <folder>/manifest.json
    """
    if "manifest.json" in members:
        return ""
    for pref in ("ScreenA_Bundle/", "screenA_bundle/", "bundle/", "ScreenA/"):
        if pref + "manifest.json" in members:
            return pref
    for m in members:
        if m.endswith("/manifest.json"):
            return m[:-len("manifest.json")]
    return ""


def _read_zip_bytes(zf: zipfile.ZipFile, name: str) -> bytes:
    try:
        return zf.read(name)
    except KeyError:
        raise FileNotFoundError(f"Missing required file in bundle: {name}")


def _find_first_member(zf: zipfile.ZipFile, root: str, rel_candidates: Sequence[str]) -> Tuple[str, str]:
    """
    Returns (member_name_in_zip, rel_path_used)
      - member_name_in_zip includes root prefix
      - rel_path_used is the relative path without root
    """
    nameset = set(zf.namelist())
    for rel in rel_candidates:
        full = root + rel
        if full in nameset:
            return full, rel
    raise FileNotFoundError(f"None of these files were found in bundle: {', '.join(rel_candidates)}")


def _load_bundle(zip_path: str) -> BundleInfo:
    if not zip_path.lower().endswith(".zip"):
        raise ValueError("Bundle must be a .zip file.")
    with zipfile.ZipFile(zip_path, "r") as zf:
        members = zf.namelist()
        root = _detect_bundle_root(members)

        manifest_full = root + "manifest.json"
        if manifest_full not in members:
            raise FileNotFoundError("Bundle missing manifest.json")

        manifest_bytes = _read_zip_bytes(zf, manifest_full)
        try:
            manifest = json.loads(_decode_bytes(manifest_bytes))
        except Exception as e:
            raise ValueError(f"manifest.json is not valid JSON: {e}")

        if not isinstance(manifest, dict):
            raise ValueError("manifest.json must be a JSON object.")

        return BundleInfo(zip_path=zip_path, root=root, manifest=manifest, members=members)


# ----------------------------
# Robust CSV parsing (record splitter)
# ----------------------------

def _split_csv_records(text: str) -> List[str]:
    """
    Split CSV into record strings by scanning newlines not inside quotes.
    Preserves embedded newlines inside quoted fields.
    """
    if not text:
        return []
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    recs: List[str] = []
    buf: List[str] = []
    in_quote = False
    i = 0
    n = len(text)

    while i < n:
        ch = text[i]
        if ch == '"':
            if in_quote and (i + 1) < n and text[i + 1] == '"':
                buf.append('"')
                buf.append('"')
                i += 2
                continue
            in_quote = not in_quote
            buf.append(ch)
            i += 1
            continue

        if ch == "\n" and not in_quote:
            recs.append("".join(buf))
            buf = []
            i += 1
            continue

        buf.append(ch)
        i += 1

    if buf:
        recs.append("".join(buf))
    return recs


def _parse_csv_tolerant_text(text: str, required_id: str = "local_id") -> ParseReport:
    records = _split_csv_records(text)
    if not records:
        raise ValueError("CSV is empty.")

    try:
        header = next(csv.reader(io.StringIO(records[0] + "\n")))
    except Exception as e:
        raise ValueError(f"Failed to parse CSV header: {e}")

    header = [h.strip() for h in header]
    if not header or all(not h for h in header):
        raise ValueError("CSV header is empty.")

    expected_n = len(header)
    rows: List[Dict[str, str]] = []
    skipped: List[Tuple[int, str, str]] = []

    for rec_idx, rec in enumerate(records[1:], start=1):  # 1-based, excluding header
        raw = rec
        if not raw.strip():
            skipped.append((rec_idx, "empty_record", raw))
            continue
        try:
            parsed = next(csv.reader(io.StringIO(raw + "\n")))
        except csv.Error as e:
            skipped.append((rec_idx, f"csv_error:{e}", raw))
            continue
        except Exception as e:
            skipped.append((rec_idx, f"parse_error:{e}", raw))
            continue

        if len(parsed) != expected_n:
            skipped.append((rec_idx, f"bad_column_count:{len(parsed)}!=expected:{expected_n}", raw))
            continue

        d = {header[i]: _safe_str(parsed[i]) for i in range(expected_n)}

        lid = _safe_str(d.get(required_id, "")).strip()
        if not lid:
            skipped.append((rec_idx, "missing_local_id", raw))
            continue

        rows.append(d)

    return ParseReport(header=header, rows=rows, skipped=skipped)


def _load_input_errors_from_text(text: str) -> List[Tuple[int, str, str]]:
    """
    Parse a prior input_errors.csv (if present in bundle) and return skipped tuples.
    Expected columns: record_index_ex_header, reason, raw_record
    """
    out: List[Tuple[int, str, str]] = []
    try:
        rdr = csv.DictReader(io.StringIO(text))
        for row in rdr:
            idx_raw = (_safe_str(row.get("record_index_ex_header", "")) or "").strip()
            try:
                idx = int(idx_raw)
            except Exception:
                idx = 0
            reason = _safe_str(row.get("reason", "")).strip()
            raw = _safe_str(row.get("raw_record", "")).strip()
            if idx > 0 and reason:
                out.append((idx, reason, raw))
    except Exception:
        return []
    return out


# ----------------------------
# Criteria loading + contradiction warnings
# ----------------------------

def _load_criteria_ih_from_text(text: str) -> CriteriaLoadReport:
    warnings: List[str] = []
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return CriteriaLoadReport(criteria=[], warnings=["Criteria header not found."])

    has_stage = any((h or "").strip().lower() == "stage" for h in (reader.fieldnames or []))

    crits: List[Criterion] = []
    row_i = 0
    for row in reader:
        row_i += 1
        stage = _safe_str(row.get("stage", "")).strip()
        if not has_stage:
            stage = "IH"
        if stage.strip().upper() != "IH":
            continue

        enabled = _truthy(row.get("enabled", "1"))
        if not enabled:
            continue

        cid = _safe_str(row.get("id", "")).strip() or f"IH_ROW_{row_i}"
        ctype = (_safe_str(row.get("type", "")).strip().lower() or "include")
        scope = _safe_str(row.get("scope", "")).strip()
        label = _safe_str(row.get("label", "")).strip()
        operator = (_safe_str(row.get("operator", "")).strip().lower() or "equals")
        targets = _split_targets(_safe_str(row.get("target", "")).strip())
        what_raw = _safe_str(row.get("what", "")).strip()
        what_list = _split_what_list(what_raw)
        source_text = _safe_str(row.get("source_text", "")).strip()

        thr = None
        thr_cell = _safe_str(row.get("threshold", "")).strip()
        if thr_cell:
            try:
                thr = float(thr_cell)
            except Exception:
                warnings.append(f"[criteria] Row {row_i} ({cid}): invalid threshold '{thr_cell}' -> ignored.")

        if ctype not in ("include", "exclude"):
            warnings.append(f"[criteria] {cid}: unknown type '{ctype}' -> treating as include.")
            ctype = "include"

        if not targets:
            warnings.append(f"[criteria] {cid}: missing target -> criterion will be treated as MISSING (PASS_FLAGGED).")

        crits.append(
            Criterion(
                stage="IH",
                cid=cid,
                ctype=ctype,
                scope=scope,
                label=label,
                operator=operator,
                targets=targets,
                what_raw=what_raw,
                what_list=what_list,
                threshold=thr,
                enabled=True,
                source_text=source_text,
            )
        )

    warnings.extend(_detect_contradictions_simple(crits))
    return CriteriaLoadReport(criteria=crits, warnings=warnings)


def _detect_contradictions_simple(crits: Sequence[Criterion]) -> List[str]:
    warns: List[str] = []
    bucket: Dict[str, Dict[str, set]] = {}

    for c in crits:
        if not c.targets or len(c.targets) != 1:
            continue
        tgt = c.targets[0].strip()
        op = (c.operator or "").strip().lower()
        if op not in ("equals", "in_list"):
            continue
        vals = c.what_list[:] if c.what_list else ([_safe_str(c.what_raw)] if c.what_raw else [])
        vals_norm = {_norm_what_for_target(tgt, v) for v in vals if _norm_what_for_target(tgt, v)}
        if not vals_norm:
            continue
        bucket.setdefault(tgt, {"include": set(), "exclude": set()})
        bucket[tgt][c.ctype].update(vals_norm)

    for tgt, d in bucket.items():
        inc = d.get("include", set())
        exc = d.get("exclude", set())
        overlap = sorted(inc.intersection(exc))
        if overlap:
            warns.append(
                f"[criteria] Possible contradiction on target '{tgt}': values appear in BOTH include and exclude: {', '.join(overlap[:10])}"
                + (" ..." if len(overlap) > 10 else "")
            )
    return warns


# ----------------------------
# IH evaluation (per-criterion + aggregate)
# ----------------------------

def _get_first_nonempty(row: Dict[str, str], targets: Sequence[str]) -> Tuple[str, str]:
    if not targets:
        return "", ""
    for t in targets:
        if t in row:
            v = _safe_str(row.get(t, ""))
            if v.strip():
                return t, v
    t0 = targets[0]
    return t0, _safe_str(row.get(t0, "")) if t0 in row else ""


def _eval_criterion(row: Dict[str, str], header_set: set, c: Criterion) -> str:
    """
    Fast status only: returns one of {"MET","FAILED","MISSING","UNKNOWN"}.
    """
    if not c.targets:
        return "MISSING"

    if not any(t in header_set for t in c.targets):
        return "MISSING"

    target_used, raw_val = _get_first_nonempty(row, c.targets)
    val = _norm_for_target(target_used, raw_val)
    if not val:
        return "MISSING"

    op = (c.operator or "").strip().lower()

    what_list = c.what_list[:] if c.what_list else ([_safe_str(c.what_raw)] if c.what_raw else [])
    what_list_norm = [_norm_what_for_target(target_used, w) for w in what_list if _norm_what_for_target(target_used, w)]

    # Guardrail: heuristics stage never calls LLM
    if op in ("llm",):
        return "UNKNOWN"
    if op not in ("equals", "contains", "regex", "in_list", "not_in", "gte", "lte", "between"):
        return "UNKNOWN"

    def _as_float(x: str) -> Optional[float]:
        try:
            return float(x)
        except Exception:
            return None

    matched: Optional[bool] = None

    if op == "equals":
        if not what_list_norm:
            return "UNKNOWN"
        matched = (val.lower() == what_list_norm[0].lower())

    elif op == "contains":
        if not what_list_norm:
            return "UNKNOWN"
        hay = val.lower()
        matched = any((w.lower() in hay) for w in what_list_norm if w)

    elif op == "regex":
        if not what_list:
            return "UNKNOWN"
        pat = what_list[0]
        try:
            matched = bool(re.search(pat, raw_val, flags=re.IGNORECASE))
        except Exception:
            return "UNKNOWN"

    elif op == "in_list":
        if not what_list_norm:
            return "UNKNOWN"
        wset = {w.lower() for w in what_list_norm if w}
        matched = (val.lower() in wset)

    elif op == "not_in":
        if not what_list_norm:
            return "UNKNOWN"
        wset = {w.lower() for w in what_list_norm if w}
        matched = (val.lower() not in wset)

    elif op == "gte":
        if not what_list:
            return "UNKNOWN"
        a = _as_float(val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            return "UNKNOWN"
        matched = (a >= b)

    elif op == "lte":
        if not what_list:
            return "UNKNOWN"
        a = _as_float(val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            return "UNKNOWN"
        matched = (a <= b)

    elif op == "between":
        if len(what_list) < 2:
            return "UNKNOWN"
        a = _as_float(val)
        lo = _as_float(what_list[0])
        hi = _as_float(what_list[1])
        if a is None or lo is None or hi is None:
            return "UNKNOWN"
        if lo > hi:
            lo, hi = hi, lo
        matched = (lo <= a <= hi)

    if matched is None:
        return "UNKNOWN"

    # type semantics are preserved (same as EH)
    if c.ctype == "include":
        return "MET" if matched else "FAILED"
    else:
        return "FAILED" if matched else "MET"


def _eval_criterion_detail(row: Dict[str, str], header_set: set, c: Criterion) -> Dict[str, str]:
    """
    Slower, for UI detail modal.
    Returns dict with keys:
      status, note, target_used, raw_value, norm_value, operator, type, targets, what
    """
    out = {
        "cid": c.cid,
        "type": c.ctype,
        "operator": c.operator,
        "targets": ",".join(c.targets),
        "what": c.what_raw,
        "status": "",
        "note": "",
        "target_used": "",
        "raw_value": "",
        "norm_value": "",
    }

    if not c.targets:
        out["status"] = "MISSING"
        out["note"] = "missing_target"
        return out

    if not any(t in header_set for t in c.targets):
        out["status"] = "MISSING"
        out["note"] = "missing_column"
        return out

    target_used, raw_val = _get_first_nonempty(row, c.targets)
    out["target_used"] = target_used
    out["raw_value"] = raw_val

    norm_val = _norm_for_target(target_used, raw_val)
    out["norm_value"] = norm_val

    if not norm_val:
        out["status"] = "MISSING"
        out["note"] = "empty_value"
        return out

    op = (c.operator or "").strip().lower()

    what_list = c.what_list[:] if c.what_list else ([_safe_str(c.what_raw)] if c.what_raw else [])
    what_list_norm = [_norm_what_for_target(target_used, w) for w in what_list if _norm_what_for_target(target_used, w)]

    # Guardrail: heuristics stage never calls LLM
    if op in ("llm",):
        out["status"] = "UNKNOWN"
        out["note"] = "operator_llm_not_supported_in_IH"
        return out
    if op not in ("equals", "contains", "regex", "in_list", "not_in", "gte", "lte", "between"):
        out["status"] = "UNKNOWN"
        out["note"] = f"unknown_operator:{op}"
        return out

    def _as_float(x: str) -> Optional[float]:
        try:
            return float(x)
        except Exception:
            return None

    matched: Optional[bool] = None

    if op == "equals":
        if not what_list_norm:
            out["status"] = "UNKNOWN"
            out["note"] = "equals_missing_what"
            return out
        matched = (norm_val.lower() == what_list_norm[0].lower())
        out["note"] = "equals"

    elif op == "contains":
        if not what_list_norm:
            out["status"] = "UNKNOWN"
            out["note"] = "contains_missing_what"
            return out
        hay = norm_val.lower()
        matched = any((w.lower() in hay) for w in what_list_norm if w)
        out["note"] = "contains"

    elif op == "regex":
        if not what_list:
            out["status"] = "UNKNOWN"
            out["note"] = "regex_missing_pattern"
            return out
        pat = what_list[0]
        try:
            matched = bool(re.search(pat, raw_val, flags=re.IGNORECASE))
            out["note"] = f"regex:{pat}"
        except Exception:
            out["status"] = "UNKNOWN"
            out["note"] = f"bad_regex:{pat}"
            return out

    elif op == "in_list":
        if not what_list_norm:
            out["status"] = "UNKNOWN"
            out["note"] = "in_list_missing_what"
            return out
        wset = {w.lower() for w in what_list_norm if w}
        matched = (norm_val.lower() in wset)
        out["note"] = "in_list"

    elif op == "not_in":
        if not what_list_norm:
            out["status"] = "UNKNOWN"
            out["note"] = "not_in_missing_what"
            return out
        wset = {w.lower() for w in what_list_norm if w}
        matched = (norm_val.lower() not in wset)
        out["note"] = "not_in"

    elif op == "gte":
        if not what_list:
            out["status"] = "UNKNOWN"
            out["note"] = "gte_missing_what"
            return out
        a = _as_float(norm_val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            out["status"] = "UNKNOWN"
            out["note"] = "gte_non_numeric"
            return out
        matched = (a >= b)
        out["note"] = f"gte:{b}"

    elif op == "lte":
        if not what_list:
            out["status"] = "UNKNOWN"
            out["note"] = "lte_missing_what"
            return out
        a = _as_float(norm_val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            out["status"] = "UNKNOWN"
            out["note"] = "lte_non_numeric"
            return out
        matched = (a <= b)
        out["note"] = f"lte:{b}"

    elif op == "between":
        if len(what_list) < 2:
            out["status"] = "UNKNOWN"
            out["note"] = "between_requires_two_values"
            return out
        a = _as_float(norm_val)
        lo = _as_float(what_list[0])
        hi = _as_float(what_list[1])
        if a is None or lo is None or hi is None:
            out["status"] = "UNKNOWN"
            out["note"] = "between_non_numeric"
            return out
        if lo > hi:
            lo, hi = hi, lo
        matched = (lo <= a <= hi)
        out["note"] = f"between:{lo}..{hi}"

    if matched is None:
        out["status"] = "UNKNOWN"
        out["note"] = "no_match_result"
        return out

    if c.ctype == "include":
        out["status"] = "MET" if matched else "FAILED"
    else:
        out["status"] = "FAILED" if matched else "MET"

    return out


def _summarize_reason(outcome: str, failed: List[str], missing: List[str], met: List[str], unknown: List[str]) -> str:
    if outcome == "PASS_CLEAN":
        return "All IH criteria met."
    parts = []
    if failed:
        parts.append(f"Failed: {', '.join(failed[:6])}" + (" ..." if len(failed) > 6 else ""))
    if missing:
        parts.append(f"Missing: {', '.join(missing[:6])}" + (" ..." if len(missing) > 6 else ""))
    if unknown:
        parts.append(f"Unknown: {', '.join(unknown[:6])}" + (" ..." if len(unknown) > 6 else ""))
    if not parts:
        parts.append("Uncertain (no definitive failures).")
    return " | ".join(parts)


def run_ih_screen(
    parse: ParseReport,
    criteria_report: CriteriaLoadReport,
    cancel_event: threading.Event,
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], Dict[str, int], Dict[str, Dict[str, int]], List[Dict[str, List[str]]]]:
    """
    Returns:
      full_rows: aggregate row dicts + IH columns appended
      survivor_rows: original aggregate row dicts (no extra columns)
      counts: summary counts
      crit_impacts: {cid: {"failed":n, "missing":n, "met":n, "unknown":n}}
      row_eval_lists: list aligned with full_rows, each is {"failed":[...], "missing":[...], "met":[...], "unknown":[...]}
    """
    header_set = set(parse.header)
    rows = parse.rows
    crits = criteria_report.criteria

    counts = {
        "OUT": 0,
        "PASS_CLEAN": 0,
        "PASS_FLAGGED": 0,
        "SKIPPED_INVALID": len(parse.skipped),
        "TOTAL_INTEGRAL": len(rows),
        "TOTAL_INPUT_RECORDS_EX_HEADER": len(parse.rows) + len(parse.skipped),
    }

    crit_impacts: Dict[str, Dict[str, int]] = {c.cid: {"failed": 0, "missing": 0, "met": 0, "unknown": 0} for c in crits}

    full_rows: List[Dict[str, str]] = []
    survivors: List[Dict[str, str]] = []
    row_eval_lists: List[Dict[str, List[str]]] = []

    if not crits:
        for r in rows:
            if cancel_event.is_set():
                break
            fr = dict(r)
            fr["ih_outcome"] = "PASS_CLEAN"
            fr["ih_failed_ids"] = ""
            fr["ih_missing_ids"] = ""
            fr["ih_met_ids"] = ""
            fr["ih_reason_summary"] = "No active IH criteria: default PASS_CLEAN."
            full_rows.append(fr)
            survivors.append(dict(r))
            row_eval_lists.append({"failed": [], "missing": [], "met": [], "unknown": []})
        counts["PASS_CLEAN"] = len(survivors)
        if progress_cb:
            progress_cb(1.0)
        return full_rows, survivors, counts, crit_impacts, row_eval_lists

    n = len(rows)
    for idx, r in enumerate(rows, start=1):
        if cancel_event.is_set():
            break

        failed: List[str] = []
        missing: List[str] = []
        met: List[str] = []
        unknown: List[str] = []

        for c in crits:
            status = _eval_criterion(r, header_set, c)
            if status == "FAILED":
                failed.append(c.cid)
                crit_impacts[c.cid]["failed"] += 1
            elif status == "MISSING":
                missing.append(c.cid)
                crit_impacts[c.cid]["missing"] += 1
            elif status == "MET":
                met.append(c.cid)
                crit_impacts[c.cid]["met"] += 1
            else:
                unknown.append(c.cid)
                crit_impacts[c.cid]["unknown"] += 1

        # IH outcome assignment (as requested)
        if failed:
            outcome = "OUT"
            counts["OUT"] += 1
        else:
            if unknown or missing:
                outcome = "PASS_FLAGGED"
                counts["PASS_FLAGGED"] += 1
            else:
                outcome = "PASS_CLEAN"
                counts["PASS_CLEAN"] += 1

        fr = dict(r)
        fr["ih_outcome"] = outcome
        fr["ih_failed_ids"] = ";".join(failed)
        fr["ih_missing_ids"] = ";".join(missing + unknown)  # keep one column for "non-deterministic/absent"
        fr["ih_met_ids"] = ";".join(met)
        fr["ih_reason_summary"] = _summarize_reason(outcome, failed, missing, met, unknown)

        full_rows.append(fr)
        row_eval_lists.append({"failed": failed, "missing": missing, "met": met, "unknown": unknown})

        # Survivors are those that continue to next stage
        if outcome != "OUT":
            survivors.append(dict(r))

        if progress_cb and (idx % PROGRESS_EVERY_N == 0 or idx == n):
            progress_cb(idx / max(1, n))

    return full_rows, survivors, counts, crit_impacts, row_eval_lists


# ----------------------------
# Export helpers
# ----------------------------

def _export_input_errors_csv(path: str, skipped: Sequence[Tuple[int, str, str]]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["record_index_ex_header", "reason", "raw_record"])
        for rec_i, reason, raw in skipped:
            w.writerow([rec_i, reason, raw])


def _export_xlsx(path: str, full_rows: List[Dict[str, str]], survivors: List[Dict[str, str]], aggregate_header: List[str]) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError(f"openpyxl not available: {e}")

    full_cols = list(aggregate_header) + ["ih_outcome", "ih_failed_ids", "ih_missing_ids", "ih_met_ids", "ih_reason_summary"]
    surv_cols = list(aggregate_header)

    wb = Workbook(write_only=True)
    ws1 = wb.create_sheet("IH_FULL")
    ws1.append(full_cols)
    for r in full_rows:
        ws1.append([_safe_str(r.get(c, "")) for c in full_cols])

    ws2 = wb.create_sheet("IH_SURVIVORS")
    ws2.append(surv_cols)
    for r in survivors:
        ws2.append([_safe_str(r.get(c, "")) for c in surv_cols])

    wb.save(path)


def _write_csv_bytes(fieldnames: List[str], rows: Sequence[Dict[str, str]]) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
    w.writeheader()
    for r in rows:
        w.writerow({k: _safe_str(r.get(k, "")) for k in fieldnames})
    return buf.getvalue().encode("utf-8")


def _export_next_bundle_zip(
    out_zip_path: str,
    bundle: BundleInfo,
    data_rel: str,
    criteria_rel: str,
    input_errors_rel: Optional[str],
    parse_header: List[str],
    full_rows: List[Dict[str, str]],
    survivors: List[Dict[str, str]],
    skipped: List[Tuple[int, str, str]],
    counts: Dict[str, int],
) -> None:
    """
    Create a new bundle zip where data/current.csv becomes the IH survivors.
    Keeps other files from the input bundle, updates manifest pipeline + sha256 (warn-only downstream).
    Adds reports/IH_FULL.csv and reports/IH_SURVIVORS.csv.
    """
    root = bundle.root
    src_zip = bundle.zip_path

    manifest_rel = "manifest.json"
    rep_full_rel = "reports/IH_FULL.csv"
    rep_surv_rel = "reports/IH_SURVIVORS.csv"

    # Always write survivors to the canonical location (data/current.csv) for downstream stages
    out_data_rel = "data/current.csv"

    current_bytes = _write_csv_bytes(parse_header, survivors)
    rep_full_bytes = _write_csv_bytes(
        parse_header + ["ih_outcome", "ih_failed_ids", "ih_missing_ids", "ih_met_ids", "ih_reason_summary"],
        full_rows
    )
    rep_surv_bytes = current_bytes

    input_errors_bytes = None
    if skipped:
        buf = io.StringIO()
        w = csv.writer(buf, lineterminator="\n")
        w.writerow(["record_index_ex_header", "reason", "raw_record"])
        for rec_i, reason, raw in skipped:
            w.writerow([rec_i, reason, raw])
        input_errors_bytes = buf.getvalue().encode("utf-8")

    # Update manifest
    m = dict(bundle.manifest)
    pipeline = dict(m.get("pipeline", {}) or {})
    stages = dict(pipeline.get("stages", {}) or {})
    history = list(pipeline.get("history", []) or [])

    stages["IH"] = "done"
    history.append({
        "stage": "IH",
        "ran_at": _iso_now(),
        "counts": counts,
        "survivors_rows": len(survivors),
        "out_rows_full": len(full_rows),
    })
    pipeline["stages"] = stages
    pipeline["history"] = history
    m["pipeline"] = pipeline

    m["created_at"] = datetime.now().replace(microsecond=0).isoformat()
    m["created_by"] = "screen_a_ih_plugin"
    m.setdefault("derived_from", {})
    try:
        m["derived_from"]["zip_name"] = Path(src_zip).name
    except Exception:
        pass

    # Refresh sha256 map (only for files we overwrite/add)
    sha_map = dict(m.get("sha256", {}) or {})
    sha_map[out_data_rel] = _sha256_hex(current_bytes)
    sha_map[rep_full_rel] = _sha256_hex(rep_full_bytes)
    sha_map[rep_surv_rel] = _sha256_hex(rep_surv_bytes)
    if input_errors_bytes is not None:
        sha_map["data/input_errors.csv"] = _sha256_hex(input_errors_bytes)
    m["sha256"] = sha_map

    manifest_bytes = (json.dumps(m, ensure_ascii=False, indent=2) + "\n").encode("utf-8")

    # Copy everything except entries we overwrite
    overwrite_set = {
        root + manifest_rel,
        root + out_data_rel,
        root + rep_full_rel,
        root + rep_surv_rel,
        root + "data/input_errors.csv",
    }

    with zipfile.ZipFile(src_zip, "r") as zin, zipfile.ZipFile(out_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name in overwrite_set:
                continue
            data = zin.read(name)
            zout.writestr(name, data)

        zout.writestr(root + manifest_rel, manifest_bytes)
        zout.writestr(root + out_data_rel, current_bytes)
        zout.writestr(root + rep_full_rel, rep_full_bytes)
        zout.writestr(root + rep_surv_rel, rep_surv_bytes)
        if input_errors_bytes is not None:
            zout.writestr(root + "data/input_errors.csv", input_errors_bytes)


# ----------------------------
# UI Components
# ----------------------------

class DataTable(ttk.Frame):
    """
    Treeview wrapper with:
    - column setup
    - click-to-sort (requests sort callback)
    - incremental rendering to keep UI responsive
    - double-click callback to open details
    """
    def __init__(self, parent, on_sort: Callable[[str], None], on_row_activate: Optional[Callable[[Dict[str, str]], None]] = None):
        super().__init__(parent)
        self.on_sort = on_sort
        self.on_row_activate = on_row_activate

        self.tree = ttk.Treeview(self, show="headings", selectmode="browse")
        self.vs = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.hs = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.vs.set, xscrollcommand=self.hs.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        self.vs.grid(row=0, column=1, sticky="ns")
        self.hs.grid(row=1, column=0, sticky="ew")

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.columns: List[str] = []
        self._render_token = 0
        self._iid_to_row: Dict[str, Dict[str, str]] = {}

        self.tree.bind("<Double-1>", self._on_double_click)

    def set_columns(self, cols: List[str]):
        self.columns = cols
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols
        for c in cols:
            self.tree.heading(c, text=c, command=lambda col=c: self.on_sort(col))
            self.tree.column(c, width=140, minwidth=90, stretch=False)

    def clear(self):
        self._render_token += 1
        self._iid_to_row.clear()
        self.tree.delete(*self.tree.get_children())

    def render_rows_incremental(self, rows: List[Dict[str, str]]):
        self.clear()
        token = self._render_token

        def _insert_chunk(start: int):
            if token != self._render_token:
                return
            end = min(start + RENDER_CHUNK, len(rows))
            for i in range(start, end):
                r = rows[i]
                iid = f"r{i}"
                self._iid_to_row[iid] = r
                vals = [_safe_str(r.get(c, "")) for c in self.columns]
                self.tree.insert("", "end", iid=iid, values=vals)
            if end < len(rows):
                self.after(1, lambda: _insert_chunk(end))

        self.after(0, lambda: _insert_chunk(0))

    def _on_double_click(self, _evt):
        if not self.on_row_activate:
            return
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        r = self._iid_to_row.get(iid)
        if r:
            self.on_row_activate(r)


# ----------------------------
# Main View
# ----------------------------

class IHView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.bundle_zip_path: Optional[str] = None
        self.bundle: Optional[BundleInfo] = None

        self.data_full_member: Optional[str] = None
        self.data_rel_used: Optional[str] = None
        self.criteria_full_member: Optional[str] = None
        self.criteria_rel_used: Optional[str] = None
        self.input_errors_full_member: Optional[str] = None
        self.input_errors_rel_used: Optional[str] = None

        self.parse_report: Optional[ParseReport] = None
        self.criteria_report: Optional[CriteriaLoadReport] = None

        self.full_rows: List[Dict[str, str]] = []
        self.survivors: List[Dict[str, str]] = []
        self.counts: Dict[str, int] = {}

        self.crit_impacts: Dict[str, Dict[str, int]] = {}
        self.row_evals_full: List[Dict[str, List[str]]] = []

        self.active_criterion_id: Optional[str] = None

        self._worker: Optional[threading.Thread] = None
        self._cancel = threading.Event()

        self.sort_full: Tuple[Optional[str], bool] = (None, True)
        self.sort_surv: Tuple[Optional[str], bool] = (None, True)
        self.sort_crit: Tuple[Optional[str], bool] = (None, True)

        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        btn_b = ttk.Button(top, text="Load ScreenA bundle ZIP…", command=self._pick_bundle)
        btn_b.grid(row=0, column=0, padx=(0, 8), pady=2, sticky="w")

        self.lbl_bundle = ttk.Label(top, text="(no bundle loaded)")
        self.lbl_bundle.grid(row=0, column=1, sticky="w")

        self.lbl_bundle_meta = ttk.Label(top, text="")
        self.lbl_bundle_meta.grid(row=1, column=1, sticky="w")

        actions = ttk.Frame(top)
        actions.grid(row=0, column=2, rowspan=2, padx=(10, 0), sticky="e")

        self.btn_run = ttk.Button(actions, text="Run IH", command=self._run_clicked, state="disabled")
        self.btn_run.grid(row=0, column=0, padx=4, pady=2, sticky="e")

        self.btn_cancel = ttk.Button(actions, text="Cancel", command=self._cancel_run, state="disabled")
        self.btn_cancel.grid(row=1, column=0, padx=4, pady=2, sticky="e")

        self.btn_export = ttk.Button(actions, text="Export XLSX…", command=self._export_clicked, state="disabled")
        self.btn_export.grid(row=0, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_err = ttk.Button(actions, text="Export input_errors.csv…", command=self._export_errors_clicked, state="disabled")
        self.btn_export_err.grid(row=1, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_bundle = ttk.Button(actions, text="Export next bundle ZIP…", command=self._export_bundle_clicked, state="disabled")
        self.btn_export_bundle.grid(row=0, column=2, padx=4, pady=2, sticky="e")

        top.columnconfigure(1, weight=1)

        prog = ttk.Frame(self)
        prog.pack(fill="x", padx=10, pady=(0, 8))

        self.pbar = ttk.Progressbar(prog, orient="horizontal", mode="determinate")
        self.pbar.pack(fill="x", expand=True, side="left")

        self.lbl_status = ttk.Label(prog, text="Ready.")
        self.lbl_status.pack(side="left", padx=10)

        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        left = ttk.Frame(paned)
        right = ttk.Frame(paned)
        paned.add(left, weight=1)
        paned.add(right, weight=4)

        crit_box = ttk.Labelframe(left, text="IH Criteria (read-only)")
        crit_box.pack(fill="both", expand=True)

        self.criteria_table = DataTable(
            crit_box,
            on_sort=self._sort_criteria_table,
            on_row_activate=self._on_criterion_activated,
        )
        self.criteria_table.pack(fill="both", expand=True, padx=6, pady=6)

        cf = ttk.Frame(left)
        cf.pack(fill="x", pady=(6, 0))
        self.lbl_crit_filter = ttk.Label(cf, text="Criterion filter: (none)")
        self.lbl_crit_filter.pack(side="left")
        self.btn_clear_filter = ttk.Button(cf, text="Clear filter", command=self._clear_criterion_filter, state="disabled")
        self.btn_clear_filter.pack(side="right")

        warn_box = ttk.Labelframe(left, text="Notes / warnings")
        warn_box.pack(fill="both", expand=False, pady=(6, 0))

        self.txt_warn = tk.Text(warn_box, height=8, wrap="word")
        self.txt_warn.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_warn.configure(state="disabled")

        nb = ttk.Notebook(right)
        nb.pack(fill="both", expand=True)

        tab_full = ttk.Frame(nb)
        tab_surv = ttk.Frame(nb)
        nb.add(tab_full, text="IH Full report")
        nb.add(tab_surv, text="IH Survivors")

        self.full_table = DataTable(tab_full, on_sort=self._sort_full_table, on_row_activate=self._open_row_detail_modal)
        self.full_table.pack(fill="both", expand=True, padx=6, pady=6)

        self.surv_table = DataTable(tab_surv, on_sort=self._sort_surv_table, on_row_activate=self._open_row_detail_modal)
        self.surv_table.pack(fill="both", expand=True, padx=6, pady=6)

        self.lbl_counts = ttk.Label(self, text="")
        self.lbl_counts.pack(fill="x", padx=10, pady=(0, 10))

    # -------- helpers --------

    def _set_warnings(self, lines: Sequence[str]) -> None:
        self.txt_warn.configure(state="normal")
        self.txt_warn.delete("1.0", "end")
        self.txt_warn.insert("end", "\n".join(lines) if lines else "(none)")
        self.txt_warn.configure(state="disabled")

    def _refresh_counts_label(self):
        if not self.parse_report:
            self.lbl_counts.configure(text="")
            return
        pr = self.parse_report
        msg = f"Integral rows: {len(pr.rows)} | Skipped invalid: {len(pr.skipped)}"
        if self.counts:
            msg += (
                f" | OUT: {self.counts.get('OUT',0)}"
                f" | PASS_CLEAN: {self.counts.get('PASS_CLEAN',0)}"
                f" | PASS_FLAGGED: {self.counts.get('PASS_FLAGGED',0)}"
            )
        self.lbl_counts.configure(text=msg)

    def _detect_duplicate_local_ids(self, rows: List[Dict[str, str]]) -> Tuple[int, List[str]]:
        seen = set()
        dups = []
        for r in rows:
            lid = _safe_str(r.get("local_id", "")).strip()
            if not lid:
                continue
            if lid in seen:
                dups.append(lid)
            else:
                seen.add(lid)
        uniq = []
        s2 = set()
        for x in dups:
            if x not in s2:
                uniq.append(x)
                s2.add(x)
        return len(uniq), uniq[:10]

    # -------- bundle load --------

    def _pick_bundle(self):
        p = filedialog.askopenfilename(
            title="Select ScreenA bundle ZIP",
            filetypes=[("ZIP", "*.zip"), ("All files", "*.*")],
        )
        if not p:
            return
        self.bundle_zip_path = p
        self.lbl_bundle.configure(text=Path(p).name)
        self._load_bundle_inputs()

    def _load_bundle_inputs(self):
        if not self.bundle_zip_path:
            return

        warns: List[str] = []
        self.full_rows = []
        self.survivors = []
        self.counts = {}
        self.crit_impacts = {}
        self.row_evals_full = []
        self.active_criterion_id = None
        self.lbl_crit_filter.configure(text="Criterion filter: (none)")
        self.btn_clear_filter.configure(state="disabled")

        try:
            self.bundle = _load_bundle(self.bundle_zip_path)
        except Exception as e:
            self.bundle = None
            messagebox.showerror("Bundle load failed", str(e))
            return

        m = self.bundle.manifest
        schema = _safe_str(m.get("bundle_schema", m.get("schema", ""))).strip() or "unknown"
        created_at = _safe_str(m.get("created_at", "")).strip()
        created_by = _safe_str(m.get("created_by", "")).strip()
        stages = (m.get("pipeline", {}) or {}).get("stages", {}) or {}
        st_ih = _safe_str(stages.get("IH", "")).strip() or "unknown"
        self.lbl_bundle_meta.configure(text=f"schema={schema} | created_at={created_at} | created_by={created_by} | IH={st_ih}")

        root = self.bundle.root

        data_candidates = ["data/current.csv", "data/A.csv", "data/aggregate.csv"]
        crit_candidates = ["criteria/criteria_harmonized.csv", "criteria/harmonized.csv", "criteria/criteria.csv", "criteria.csv"]
        err_candidates = ["data/input_errors.csv", "input_errors.csv"]

        try:
            with zipfile.ZipFile(self.bundle_zip_path, "r") as zf:
                self.data_full_member, self.data_rel_used = _find_first_member(zf, root, data_candidates)
                self.criteria_full_member, self.criteria_rel_used = _find_first_member(zf, root, crit_candidates)

                data_bytes = zf.read(self.data_full_member)
                crit_bytes = zf.read(self.criteria_full_member)
                data_text = _decode_bytes(data_bytes)
                crit_text = _decode_bytes(crit_bytes)

                # Optional input_errors
                try:
                    self.input_errors_full_member, self.input_errors_rel_used = _find_first_member(zf, root, err_candidates)
                    err_text = _decode_bytes(zf.read(self.input_errors_full_member))
                    carried_skipped = _load_input_errors_from_text(err_text)
                except Exception:
                    self.input_errors_full_member = None
                    self.input_errors_rel_used = None
                    carried_skipped = []

                # SHA checks (warn only)
                sha_map = m.get("sha256", {}) or {}
                if isinstance(sha_map, dict):
                    if self.data_rel_used in sha_map:
                        exp = _safe_str(sha_map.get(self.data_rel_used, "")).strip()
                        got = _sha256_hex(data_bytes)
                        if exp and exp != got:
                            warns.append(f"[bundle] sha256 mismatch for {self.data_rel_used} (warn only).")
                    if self.criteria_rel_used in sha_map:
                        exp = _safe_str(sha_map.get(self.criteria_rel_used, "")).strip()
                        got = _sha256_hex(crit_bytes)
                        if exp and exp != got:
                            warns.append(f"[bundle] sha256 mismatch for {self.criteria_rel_used} (warn only).")

        except Exception as e:
            messagebox.showerror("Bundle read failed", str(e))
            return

        # Parse data CSV (still enforce local_id and handle malformed records)
        try:
            pr = _parse_csv_tolerant_text(data_text, required_id="local_id")
        except Exception as e:
            messagebox.showerror("Data CSV parse failed", str(e))
            return

        # Merge skipped: carried-forward first, then newly detected
        skipped_all: List[Tuple[int, str, str]] = []
        if carried_skipped:
            skipped_all.extend(carried_skipped)
        if pr.skipped:
            skipped_all.extend(pr.skipped)
        pr = ParseReport(header=pr.header, rows=pr.rows, skipped=skipped_all)
        self.parse_report = pr

        # Criteria (IH stage only)
        try:
            self.criteria_report = _load_criteria_ih_from_text(crit_text)
        except Exception as e:
            self.criteria_report = None
            messagebox.showerror("Criteria load failed", str(e))
            return

        # Manifest warnings passthrough
        for w in (m.get("warnings", []) or []):
            ws = _safe_str(w).strip()
            if ws:
                warns.append(f"[manifest] {ws}")

        if self.criteria_report:
            warns.extend(self.criteria_report.warnings)

        # Chosen file notes
        warns.append(f"[bundle] Using data file: {self.data_rel_used}")
        warns.append(f"[bundle] Using criteria file: {self.criteria_rel_used}")
        if self.input_errors_rel_used:
            warns.append(f"[bundle] Imported previous input_errors: {self.input_errors_rel_used} ({len(carried_skipped)} rows)")

        if "local_id" not in pr.header:
            warns.append("[data] Column 'local_id' not found in header (unexpected). Rows may be skipped.")

        ndup, examples = self._detect_duplicate_local_ids(pr.rows)
        if ndup > 0:
            warns.append(f"[data] Duplicate local_id detected: {ndup} unique duplicates (examples: {', '.join(examples)}). Policy: WARN ONLY.")

        self._set_warnings(warns)

        self.btn_run.configure(state="normal")
        self.btn_export_err.configure(state="normal" if pr.skipped else "disabled")
        self.btn_export.configure(state="disabled")
        self.btn_export_bundle.configure(state="disabled")

        self._refresh_criteria_table(pre_run=True)

        self.lbl_status.configure(text="Ready.")
        self._refresh_counts_label()

        self.full_table.clear()
        self.surv_table.clear()

    # -------- Criteria table --------

    def _refresh_criteria_table(self, pre_run: bool):
        crits = self.criteria_report.criteria if self.criteria_report else []

        cols = ["id", "type", "targets", "operator", "what", "status", "notes"]
        if not pre_run:
            cols += ["n_failed", "n_missing", "n_met", "n_unknown"]

        rows: List[Dict[str, str]] = []
        header_set = set(self.parse_report.header) if self.parse_report else set()

        for c in crits:
            status = "OK"
            notes = ""

            if not c.targets:
                status = "WARNING"
                notes = "missing target -> treated as MISSING (PASS_FLAGGED)"
            elif header_set and not any(t in header_set for t in c.targets):
                status = "WARNING"
                notes = f"missing column(s): {', '.join(c.targets)} -> treated as MISSING (PASS_FLAGGED)"

            op = (c.operator or "").strip().lower()
            if op in ("llm",):
                status = "WARNING"
                notes = (notes + " | " if notes else "") + "operator 'llm' not supported in IH -> UNKNOWN (PASS_FLAGGED)"
            elif op not in ("equals", "contains", "regex", "in_list", "not_in", "gte", "lte", "between"):
                status = "WARNING"
                notes = (notes + " | " if notes else "") + f"unknown operator '{op}' -> UNKNOWN (PASS_FLAGGED)"

            d: Dict[str, str] = {
                "id": c.cid,
                "type": c.ctype,
                "targets": ",".join(c.targets),
                "operator": c.operator,
                "what": c.what_raw,
                "status": status,
                "notes": notes,
            }

            if not pre_run:
                imp = self.crit_impacts.get(c.cid, {"failed": 0, "missing": 0, "met": 0, "unknown": 0})
                d["n_failed"] = str(imp.get("failed", 0))
                d["n_missing"] = str(imp.get("missing", 0))
                d["n_met"] = str(imp.get("met", 0))
                d["n_unknown"] = str(imp.get("unknown", 0))

            rows.append(d)

        self.criteria_table.set_columns(cols)
        col, asc = self.sort_crit
        if col:
            rows = self._sorted_rows(rows, col, asc)
        self.criteria_table.render_rows_incremental(rows)

    def _on_criterion_activated(self, row: Dict[str, str]):
        cid = _safe_str(row.get("id", "")).strip()
        if not cid:
            return
        self.active_criterion_id = cid
        self.lbl_crit_filter.configure(text=f"Criterion filter: {cid}")
        self.btn_clear_filter.configure(state="normal")
        self._refresh_reports_view()

    def _clear_criterion_filter(self):
        self.active_criterion_id = None
        self.lbl_crit_filter.configure(text="Criterion filter: (none)")
        self.btn_clear_filter.configure(state="disabled")
        self._refresh_reports_view()

    # -------- Sorting --------

    def _sorted_rows(self, rows: List[Dict[str, str]], col: str, asc: bool) -> List[Dict[str, str]]:
        sample = []
        for r in rows[:200]:
            v = _safe_str(r.get(col, "")).strip()
            if v != "":
                sample.append(v)
        num_hits = 0
        for v in sample[:50]:
            try:
                float(v)
                num_hits += 1
            except Exception:
                pass
        is_numeric = (len(sample) > 0 and num_hits >= max(3, int(0.7 * min(len(sample[:50]), 50))))

        def key_num(r: Dict[str, str]):
            v = _safe_str(r.get(col, "")).strip()
            try:
                return (0, float(v))
            except Exception:
                return (1, float("inf"))

        def key_txt(r: Dict[str, str]):
            return _safe_str(r.get(col, "")).strip().lower()

        key_fn = key_num if is_numeric else key_txt
        return sorted(rows, key=key_fn, reverse=not asc)

    def _toggle_sort(self, current: Tuple[Optional[str], bool], col: str) -> Tuple[str, bool]:
        cur_col, cur_asc = current
        if cur_col == col:
            return (col, not cur_asc)
        return (col, True)

    def _sort_criteria_table(self, col: str):
        self.sort_crit = self._toggle_sort(self.sort_crit, col)
        self._refresh_criteria_table(pre_run=(not bool(self.full_rows)))

    def _sort_full_table(self, col: str):
        self.sort_full = self._toggle_sort(self.sort_full, col)
        self._refresh_reports_view()

    def _sort_surv_table(self, col: str):
        self.sort_surv = self._toggle_sort(self.sort_surv, col)
        self._refresh_reports_view()

    # -------- Run / Cancel --------

    def _cancel_run(self):
        if self._worker and self._worker.is_alive():
            self._cancel.set()
            self.lbl_status.configure(text="Cancelling…")

    def _run_clicked(self):
        if self._worker and self._worker.is_alive():
            messagebox.showinfo("IH running", "A run is already in progress.")
            return
        if not self.parse_report:
            messagebox.showwarning("Missing input", "Load a ScreenA bundle ZIP first.")
            return
        if not self.criteria_report:
            messagebox.showwarning("Missing input", "Bundle criteria could not be loaded.")
            return

        self._cancel.clear()
        self.pbar["value"] = 0
        self.lbl_status.configure(text="Running IH…")
        self.btn_run.configure(state="disabled")
        self.btn_cancel.configure(state="normal")
        self.btn_export.configure(state="disabled")
        self.btn_export_bundle.configure(state="disabled")

        self.full_rows = []
        self.survivors = []
        self.counts = {}
        self.crit_impacts = {}
        self.row_evals_full = []

        self.full_table.clear()
        self.surv_table.clear()

        def progress_cb(frac: float):
            self.after(0, lambda: self._update_progress(frac))

        def worker():
            try:
                full, surv, counts, impacts, row_evals = run_ih_screen(
                    self.parse_report,
                    self.criteria_report,
                    self._cancel,
                    progress_cb=progress_cb,
                )
                self.after(0, lambda: self._finish_run(full, surv, counts, impacts, row_evals))
            except Exception as e:
                self.after(0, lambda msg=str(e): self._run_failed(msg))

        self._worker = threading.Thread(target=worker, daemon=True)
        self._worker.start()

    def _update_progress(self, frac: float):
        frac = max(0.0, min(1.0, float(frac)))
        self.pbar["value"] = frac * 100.0

    def _run_failed(self, err: str):
        self.lbl_status.configure(text="Run failed.")
        self.btn_run.configure(state="normal")
        self.btn_cancel.configure(state="disabled")
        messagebox.showerror("IH run failed", err)

    def _finish_run(
        self,
        full: List[Dict[str, str]],
        surv: List[Dict[str, str]],
        counts: Dict[str, int],
        impacts: Dict[str, Dict[str, int]],
        row_evals: List[Dict[str, List[str]]],
    ):
        self.full_rows = full
        self.survivors = surv
        self.counts = counts
        self.crit_impacts = impacts
        self.row_evals_full = row_evals

        self.btn_run.configure(state="normal")
        self.btn_cancel.configure(state="disabled")
        self.btn_export.configure(state="normal")
        self.btn_export_bundle.configure(state="normal")

        self._refresh_criteria_table(pre_run=False)
        self._refresh_reports_view()

        note = ""
        if len(full) > MAX_UI_ROWS_HINT or len(surv) > MAX_UI_ROWS_HINT:
            note = " (rendering incrementally; export contains all rows.)"

        self.lbl_status.configure(
            text=f"Done. OUT={counts.get('OUT',0)} CLEAN={counts.get('PASS_CLEAN',0)} FLAGGED={counts.get('PASS_FLAGGED',0)}{note}"
        )
        self._refresh_counts_label()

    # -------- Reports rendering --------

    def _refresh_reports_view(self):
        if not self.parse_report:
            return

        full_cols = list(self.parse_report.header) + ["ih_outcome", "ih_failed_ids", "ih_missing_ids", "ih_met_ids", "ih_reason_summary"]
        surv_cols = list(self.parse_report.header)

        full_view = self.full_rows
        if self.active_criterion_id:
            cid = self.active_criterion_id
            filtered = []
            for i, r in enumerate(self.full_rows):
                ev = self.row_evals_full[i] if i < len(self.row_evals_full) else {"failed": [], "missing": [], "met": [], "unknown": []}
                if (cid in ev["failed"]) or (cid in ev["missing"]) or (cid in ev["met"]) or (cid in ev["unknown"]):
                    filtered.append(r)
            full_view = filtered

        surv_view = self.survivors
        if self.active_criterion_id and self.full_rows:
            cid = self.active_criterion_id
            touched_survivor_ids = set()
            for i, r in enumerate(self.full_rows):
                if r.get("ih_outcome") == "OUT":
                    continue
                ev = self.row_evals_full[i]
                if (cid in ev["failed"]) or (cid in ev["missing"]) or (cid in ev["met"]) or (cid in ev["unknown"]):
                    touched_survivor_ids.add(_safe_str(r.get("local_id", "")).strip())
            surv_view = [r for r in self.survivors if _safe_str(r.get("local_id", "")).strip() in touched_survivor_ids]

        col, asc = self.sort_full
        if col and full_view:
            full_view = self._sorted_rows(full_view, col, asc)

        col2, asc2 = self.sort_surv
        if col2 and surv_view:
            surv_view = self._sorted_rows(surv_view, col2, asc2)

        self.full_table.set_columns(full_cols)
        self.full_table.render_rows_incremental(full_view)

        self.surv_table.set_columns(surv_cols)
        self.surv_table.render_rows_incremental(surv_view)

    # -------- Row detail modal --------

    def _open_row_detail_modal(self, row: Dict[str, str]):
        if not self.criteria_report or not self.parse_report:
            return

        win = tk.Toplevel(self)
        win.title("IH Row details")
        win.geometry("900x600")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        lid = _safe_str(row.get("local_id", "")).strip()
        title = _safe_str(row.get("title", "")).strip() or _safe_str(row.get("Title", "")).strip()

        top = ttk.Frame(win)
        top.pack(fill="x", padx=10, pady=10)

        ttk.Label(top, text=f"local_id: {lid}").pack(anchor="w")
        if title:
            ttk.Label(top, text=f"title: {title[:250]}").pack(anchor="w")

        outcome = _safe_str(row.get("ih_outcome", "")).strip()
        if outcome:
            ttk.Label(top, text=f"IH outcome: {outcome}").pack(anchor="w")
            rs = _safe_str(row.get("ih_reason_summary", "")).strip()
            if rs:
                ttk.Label(top, text=f"summary: {rs}").pack(anchor="w")

        box = ttk.Labelframe(win, text="Per-criterion evaluation (recomputed)")
        box.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ["cid", "type", "operator", "targets", "what", "status", "target_used", "norm_value", "note"]
        table = DataTable(box, on_sort=lambda _c: None, on_row_activate=None)
        table.pack(fill="both", expand=True, padx=6, pady=6)
        table.set_columns(cols)

        header_set = set(self.parse_report.header)
        crits = self.criteria_report.criteria

        detail_rows: List[Dict[str, str]] = []
        for c in crits:
            detail_rows.append(_eval_criterion_detail(row, header_set, c))

        table.render_rows_incremental(detail_rows)

        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(btns, text="Close", command=win.destroy).pack(side="right")
        win.bind("<Escape>", lambda _e: win.destroy())

    # -------- Export --------

    def _export_clicked(self):
        if not self.parse_report:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        if not self.full_rows:
            messagebox.showwarning("Nothing to export", "Run IH first.")
            return

        default_name = f"{_now_stamp()}_IH_reports.xlsx"
        p = filedialog.asksaveasfilename(
            title="Save IH reports",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not p:
            return

        try:
            _export_xlsx(p, self.full_rows, self.survivors, self.parse_report.header)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_errors_clicked(self):
        if not self.parse_report or not self.parse_report.skipped:
            messagebox.showinfo("No input errors", "No skipped/invalid records to export.")
            return

        default_name = f"{_now_stamp()}_input_errors.csv"
        p = filedialog.asksaveasfilename(
            title="Save input_errors.csv",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv")],
        )
        if not p:
            return
        try:
            _export_input_errors_csv(p, self.parse_report.skipped)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return
        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_bundle_clicked(self):
        if not self.bundle or not self.parse_report:
            messagebox.showwarning("Missing bundle", "Load a bundle first.")
            return
        if not self.full_rows:
            messagebox.showwarning("Nothing to export", "Run IH first.")
            return

        default_name = f"ScreenA_Bundle_IH_{_now_stamp()}.zip"
        p = filedialog.asksaveasfilename(
            title="Save next bundle (post-IH)",
            defaultextension=".zip",
            initialfile=default_name,
            filetypes=[("ZIP", "*.zip")],
        )
        if not p:
            return

        try:
            _export_next_bundle_zip(
                p,
                self.bundle,
                data_rel=(self.data_rel_used or "data/current.csv"),
                criteria_rel=(self.criteria_rel_used or "criteria/criteria_harmonized.csv"),
                input_errors_rel=self.input_errors_rel_used,
                parse_header=self.parse_report.header,
                full_rows=self.full_rows,
                survivors=self.survivors,
                skipped=self.parse_report.skipped,
                counts=self.counts,
            )
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported bundle: {Path(p).name}")


# ----------------------------
# Hub plugin wrapper
# ----------------------------

class Plugin(BasePlugin):
    def __init__(self, app=None, meta: Optional[PluginMeta] = None):
        if meta is None:
            meta = PluginMeta(id="screen_a_ih", title=TAB_TITLE, version="2.2.1")
        super().__init__(app, meta)
        self.view: Optional[IHView] = None

    def build_tab(self, parent: ttk.Notebook) -> tk.Frame:
        frame = ttk.Frame(parent)
        self.view = IHView(frame)
        self.view.pack(fill="both", expand=True)
        return frame

    def on_close(self):
        try:
            if self.view:
                self.view.destroy()
        except Exception:
            pass
        self.view = None
