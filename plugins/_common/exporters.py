# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""
plugins/_common/exporters.py — shared output formatters for EH/IH.

Owns:
  - _export_input_errors_csv(path, skipped): write the input_errors.csv
    artifact that the bundle carries forward when the upstream parser
    skips malformed records. Pure, no stage dependency.

  - _export_xlsx(path, full_rows, survivors, aggregate_header, stage):
    write a two-sheet workbook (FULL + SURVIVORS) using openpyxl. Sheet
    names and full-report column names are stage-prefixed (EH_FULL,
    EH_SURVIVORS, eh_outcome / eh_failed_ids / ... or the IH variants).

  - _write_csv_bytes(fieldnames, rows): serialise rows to UTF-8 CSV
    bytes with LF line terminators. Used by the FULL/SURVIVORS report
    writers AND by the EH/IH golden-file regressions, so its output is
    contractually byte-identity-locked.

Pulled-in from plugins._common.parser:
  - _safe_str (string coercion that the row writers rely on).

Consumed by:
  - plugins/04_eh/plugin.py  (called from _export_next_bundle_zip)
  - plugins/05_ih/plugin.py  (same)

Bodies are copied verbatim from plugins/04_eh/plugin.py. The only
deltas in _export_xlsx are: gain a `stage: str` parameter; sheet name
and column-prefix literals become f-strings over stage / stage.lower().
"""
import csv
import io
from typing import Dict, List, Sequence, Tuple

from plugins._common.parser import _safe_str
from plugins._common.input_errors import (
    from_tuple_skipped,
    merge_input_errors_csv,
    read_input_errors,
    write_input_errors_csv,
)


def _export_input_errors_csv(path: str, skipped: Sequence[Tuple[int, str, str]],
                             *, stage: str = "", prior_text: str = "") -> None:
    """Write the standalone "Export input_errors.csv…" file.

    F-03: emits the canonical schema via the single writer in
    plugins._common.input_errors instead of EH/IH's own three-column layout.

    F-71: the accumulated trail is now assembled here — ``prior_text`` is the
    bundle's incoming input_errors.csv, carried through verbatim, and
    ``skipped`` is this stage's own drops, stamped with ``stage``. The old
    contract (callers pre-merging carried rows into ``skipped``) re-stamped
    every prior row with the current stage; the same carried-copy subtraction
    the bundle writer applies is applied here for callers that still do.
    """
    prior_keys = {(e.record_number, e.reason, e.raw)
                  for e in read_input_errors(prior_text)}
    own_drops = [e for e in from_tuple_skipped(skipped, stage=stage)
                 if (e.record_number, e.reason, e.raw) not in prior_keys]
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(merge_input_errors_csv(prior_text, own_drops))

def _export_xlsx(path: str, full_rows: List[Dict[str, str]], survivors: List[Dict[str, str]], aggregate_header: List[str], stage: str) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError(f"openpyxl not available: {e}")

    sl = stage.lower()
    full_cols = list(aggregate_header) + [f"{sl}_outcome", f"{sl}_failed_ids", f"{sl}_missing_ids", f"{sl}_met_ids", f"{sl}_reason_summary"]
    surv_cols = list(aggregate_header)

    wb = Workbook(write_only=True)
    ws1 = wb.create_sheet(f"{stage}_FULL")
    ws1.append(full_cols)
    for r in full_rows:
        ws1.append([_safe_str(r.get(c, "")) for c in full_cols])

    ws2 = wb.create_sheet(f"{stage}_SURVIVORS")
    ws2.append(surv_cols)
    for r in survivors:
        ws2.append([_safe_str(r.get(c, "")) for c in surv_cols])

    wb.save(path)

def _write_csv_bytes(fieldnames: List[str], rows: Sequence[Dict[str, str]]) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
    w.writeheader()
    for r in rows:
        w.writerow({k: _safe_str(r.get(k, "")) for k in fieldnames})
    return buf.getvalue().encode("utf-8")
