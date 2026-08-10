
# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""ui.py - Plugin 06 EL (Screen A): Tkinter View and standalone shell.

After Conv 6 / Commit 3, this module owns the EL stage's UI surface:
  - DataTable: Treeview wrapper with click-to-sort and incremental rendering.
  - _now_stamp, _export_el_xlsx: small UI-side helpers for export actions.
  - ELView: the Tk Notebook tab widget for the EL stage.
  - StandaloneELPlugin: the standalone-app shell wrapper around ELView.

Engine logic stays in plugins/06_el/plugin.py for now; ui.py imports
the symbols it needs from there. Subsequent Conv 6 commits (4-5) will
finish moving engine code into plugins/_common/ and shrink
plugins/06_el/plugin.py to a thin shim that wires the View into the
plugin manager.

This file is GUI-only; its behaviour is verified manually after the
final thin-shim commit lands (load a bundle, click Run, double-click a
row to open the detail modal).
"""

import csv
import io
import json
import os
import re
import threading
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

# F-02: one shared export gate, so the rule that a cancelled run
# cannot be exported lives in one place for all four stages.
from plugins._common.bundle import (
    NOT_SCREENED,
    _export_block_reason,
    _export_confirm_reason,
    _run_summary_counts_text,
    _write_llm_stage_bundle,
)

from plugins._common.stage_state import (
    Outcome,
    Readiness,
    control_states,
    llm_readiness,
    parse_numeric_settings,
    run_outcome,
    tk_state,
)

from plugins._common.settings import load_settings
from plugins._common.provider_detect import last_known
from plugins._common.exporters import _export_input_errors_csv_from_dicts
from plugins._common.input_errors import (
    from_dict_skipped,
    merge_input_errors_csv,
    read_input_errors,
)

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Engine + dataclasses stay in plugin.py for now (Conv 6 / Commit 3).
# These imports break the circular dependency cleanly because plugin.py
# defines all of these BEFORE it does `from .ui import ELView, ...` near
# the bottom.
from .plugin import (
    BundleInfo,
    DEFAULT_BATCH_SIZE,
    DEFAULT_MODEL,
    DEFAULT_TRUNC_CHARS,
    DEFAULT_USE_CACHE,
    EL_CACHE_REL,
    RENDER_CHUNK,
    REPORTS_DIR_REL,
    _decode_bytes,
    _dump_cache_to_jsonl,
    _has_openai_key,
    _load_bundle,
    _load_cache_from_jsonl,
    _read_zip_bytes,
    _safe_str,
    _write_csv,
    run_el_screen,
)

# ------------------------------ UI helpers ------------------------------------

class DataTable(ttk.Frame):
    """
    Treeview wrapper with:
    - column setup
    - click-to-sort (optional)
    - incremental rendering to keep UI responsive
    - optional double-click callback

    Backward compatible:
      - DataTable(parent, on_sort_callable, on_row_activate=...)
      - DataTable(parent, ["col1","col2",...])  # legacy usage
    """
    def __init__(
        self,
        parent,
        on_sort=None,
        on_row_activate: Optional[Callable[[Dict[str, Any]], None]] = None
    ):
        super().__init__(parent)

        # Legacy signature: second arg is actually a list of columns
        cols_seed: Optional[List[str]] = None
        if on_sort is None:
            self.on_sort = lambda _c: None
        elif callable(on_sort):
            self.on_sort = on_sort
        else:
            # treat as columns list/tuple
            if isinstance(on_sort, (list, tuple)):
                cols_seed = list(on_sort)
            self.on_sort = lambda _c: None

        self.on_row_activate = on_row_activate

        self.tree = ttk.Treeview(self, show="headings", selectmode="browse")
        self.vs = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.hs = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.vs.set, xscrollcommand=self.hs.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        self.vs.grid(row=0, column=1, sticky="ns")
        self.hs.grid(row=1, column=0, sticky="ew")

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.columns: List[str] = []
        self._render_token = 0
        self._iid_to_row: Dict[str, Dict[str, Any]] = {}

        self.tree.bind("<Double-1>", self._on_double_click)

        if cols_seed:
            self.set_columns(cols_seed)

    def set_columns(self, cols: List[str]):
        self.columns = cols
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols
        for c in cols:
            self.tree.heading(c, text=c, command=lambda col=c: self.on_sort(col))
            self.tree.column(c, width=140, minwidth=90, stretch=False)

    def clear(self):
        self._render_token += 1
        self._iid_to_row.clear()
        self.tree.delete(*self.tree.get_children())

    def render_rows_incremental(self, rows: List[Dict[str, Any]]):
        self.clear()
        token = self._render_token

        def _insert_chunk(start: int):
            if token != self._render_token:
                return
            end = min(start + RENDER_CHUNK, len(rows))
            for i in range(start, end):
                r = rows[i]
                iid = f"r{i}"
                self._iid_to_row[iid] = r
                vals = [_safe_str(r.get(c, "")) for c in self.columns]
                self.tree.insert("", "end", iid=iid, values=vals)
            if end < len(rows):
                self.after(1, lambda: _insert_chunk(end))

        self.after(0, lambda: _insert_chunk(0))

    # ---- legacy helpers used by Plugin(ttk.Frame) ----
    def set_rows(self, rows: List[Dict[str, Any]]):
        if not rows:
            self.clear()
            return
        if not self.columns:
            self.set_columns(list(rows[0].keys()))
        self.render_rows_incremental(rows)

    def get_selected_row(self) -> Optional[Dict[str, Any]]:
        sel = self.tree.selection()
        if not sel:
            return None
        return self._iid_to_row.get(sel[0])

    def _on_double_click(self, _evt):
        if not self.on_row_activate:
            return
        r = self.get_selected_row()
        if r is not None:
            self.on_row_activate(r)

def _now_stamp() -> str:
    # Same convention as EH/IH
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _export_el_xlsx(
    path: str,
    full_rows: List[Dict[str, Any]],
    survivors: List[Dict[str, Any]],
    base_header: List[str],
) -> None:
    """
    Writes a 2-sheet XLSX:
      - EL_FULL
      - EL_SURVIVORS
    Uses openpyxl (same approach as your EH/IH plugins).
    """
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required for XLSX export. Install it (pip install openpyxl)."
        ) from e

    def cell(v: Any) -> Any:
        if v is None:
            return ""
        if isinstance(v, (dict, list)):
            return json.dumps(v, ensure_ascii=False)
        return v

    wb = Workbook(write_only=True)

    # ---- headers (stable + tolerant) ----
    base = list(base_header or [])
    if "local_id" in base:
        base = ["local_id"] + [h for h in base if h != "local_id"]
    else:
        base = ["local_id"] + base

    el_cols = [
        "el_outcome",
        "el_failed_ids",
        "el_missing_ids",
        "el_met_ids",
        "el_uncertain_ids",
        "el_reason_summary",
        "el_evidence_json",
    ]

    # Full sheet header = base + EL cols + any extras present in data
    full_header = base + [c for c in el_cols if c not in base]
    if full_rows:
        extras = [k for k in full_rows[0].keys() if k not in full_header]
        full_header += extras

    # Survivors sheet header = base + any extras present in survivors (rare)
    surv_header = list(base)
    if survivors:
        extras2 = [k for k in survivors[0].keys() if k not in surv_header]
        surv_header += extras2

    # ---- sheet: EL_FULL ----
    ws1 = wb.create_sheet(title="EL_FULL")
    ws1.append(full_header)
    for r in full_rows:
        ws1.append([cell(r.get(h, "")) for h in full_header])

    # ---- sheet: EL_SURVIVORS ----
    ws2 = wb.create_sheet(title="EL_SURVIVORS")
    ws2.append(surv_header)
    for r in survivors:
        ws2.append([cell(r.get(h, "")) for h in surv_header])

    # Remove default sheet if present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)


# ----------------------------
# Main View
# ----------------------------

class ELView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.bundle_zip_path: Optional[str] = None
        self.bundle: Optional[BundleInfo] = None

        self.full_rows: List[Dict[str, Any]] = []
        self.cancelled: bool = False   # F-02: last run stopped mid-corpus
        # Wave 8: what the last run learned, and what it failed to learn.
        # Derived by the engine from the evidence its decisions were made
        # from; empty until a run completes. A wholly failed run and a
        # wholly uncertain run are identical in every other field this
        # view holds, which is what made "EL done." the only thing the
        # interface could say for either.
        self.llm_report: Dict[str, Any] = {}
        # Wave 8 part 2: how the last run is classified, as data.
        self.outcome: Optional[Outcome] = None
        self.not_screened: bool = False  # F-34: last run had no criteria
        self.survivors: List[Dict[str, str]] = []
        self.counts: Dict[str, int] = {}

        self.crit_impacts: Dict[str, Dict[str, int]] = {}
        self.row_eval_lists: List[Dict[str, List[str]]] = []

        self.cache_map: Dict[str, Dict[str, Any]] = {}

        self.active_criterion_id: Optional[str] = None

        self._worker: Optional[threading.Thread] = None
        self._cancel = threading.Event()

        self.sort_full: Tuple[Optional[str], bool] = (None, True)
        self.sort_surv: Tuple[Optional[str], bool] = (None, True)
        self.sort_crit: Tuple[Optional[str], bool] = (None, True)

        # Settings
        self.var_model = tk.StringVar(value=DEFAULT_MODEL)
        self.var_temp = tk.DoubleVar(value=0.0)
        self.var_batch = tk.StringVar(value=str(DEFAULT_BATCH_SIZE))
        self.var_trunc = tk.StringVar(value=str(DEFAULT_TRUNC_CHARS))
        self.var_use_cache = tk.BooleanVar(value=DEFAULT_USE_CACHE)

        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        btn_b = ttk.Button(top, text="Load ScreenA bundle ZIP…", command=self._pick_bundle)
        btn_b.grid(row=0, column=0, padx=(0, 8), pady=2, sticky="w")

        self.lbl_bundle = ttk.Label(top, text="(no bundle loaded)")
        self.lbl_bundle.grid(row=0, column=1, sticky="w")

        self.lbl_bundle_meta = ttk.Label(top, text="")
        self.lbl_bundle_meta.grid(row=1, column=1, sticky="w")

        actions = ttk.Frame(top)
        actions.grid(row=0, column=2, rowspan=2, padx=(10, 0), sticky="e")

        self.btn_run = ttk.Button(actions, text="Run EL", command=self._run_clicked, state="disabled")
        self.btn_run.grid(row=0, column=0, padx=4, pady=2, sticky="e")

        self.btn_cancel = ttk.Button(actions, text="Cancel", command=self._cancel_run, state="disabled")
        self.btn_cancel.grid(row=1, column=0, padx=4, pady=2, sticky="e")

        self.btn_export = ttk.Button(actions, text="Export XLSX…", command=self._export_clicked, state="disabled")
        self.btn_export.grid(row=0, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_err = ttk.Button(actions, text="Export input_errors.csv…", command=self._export_errors_clicked, state="disabled")
        self.btn_export_err.grid(row=1, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_bundle = ttk.Button(actions, text="Export next bundle ZIP…", command=self._export_bundle_clicked, state="disabled")
        self.btn_export_bundle.grid(row=0, column=2, padx=4, pady=2, sticky="e")

        # API key indicator (EL requires OpenAI)
        self.lbl_key = ttk.Label(actions, text="")
        self.lbl_key.grid(row=1, column=2, padx=4, pady=2, sticky="e")
        self._refresh_readiness_label()
        # The indicator is only honest if it keeps up with the field it
        # reports on; without this it would go stale the moment the user
        # typed a model name.
        self.var_model.trace_add(
            "write", lambda *_a: self._refresh_readiness_label())

        top.columnconfigure(1, weight=1)

        prog = ttk.Frame(self)
        prog.pack(fill="x", padx=10, pady=(0, 8))

        self.pbar = ttk.Progressbar(prog, orient="horizontal", mode="determinate")
        self.pbar.pack(fill="x", expand=True, side="left")

        self.lbl_status = ttk.Label(prog, text="Ready.")
        self.lbl_status.pack(side="left", padx=10)

        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        left = ttk.Frame(paned)
        right = ttk.Frame(paned)
        paned.add(left, weight=1)
        paned.add(right, weight=4)

        crit_box = ttk.Labelframe(left, text="EL Criteria (read-only)")
        crit_box.pack(fill="both", expand=True)

        self.criteria_table = DataTable(
            crit_box,
            on_sort=self._sort_criteria_table,
            on_row_activate=self._on_criterion_activated,
        )
        self.criteria_table.pack(fill="both", expand=True, padx=6, pady=6)

        cf = ttk.Frame(left)
        cf.pack(fill="x", pady=(6, 0))
        self.lbl_crit_filter = ttk.Label(cf, text="Criterion filter: (none)")
        self.lbl_crit_filter.pack(side="left")
        self.btn_clear_filter = ttk.Button(cf, text="Clear filter", command=self._clear_criterion_filter, state="disabled")
        self.btn_clear_filter.pack(side="right")

        settings = ttk.Labelframe(left, text="EL Settings")
        settings.pack(fill="x", pady=(6, 0))

        ttk.Label(settings, text="Model").grid(row=0, column=0, sticky="w", padx=6, pady=2)
        ttk.Entry(settings, textvariable=self.var_model, width=24).grid(row=0, column=1, sticky="ew", padx=6, pady=2)

        ttk.Label(settings, text="Temperature").grid(row=1, column=0, sticky="w", padx=6, pady=2)
        ttk.Spinbox(settings, textvariable=self.var_temp, from_=0.0, to=2.0, increment=0.1, format="%.2f", width=10).grid(row=1, column=1, sticky="w", padx=6, pady=2)
        ttk.Label(settings, text="(0.0 = deterministic; non-zero invalidates cache)").grid(row=2, column=1, sticky="w", padx=6, pady=(0, 4))

        ttk.Label(settings, text="Batch size").grid(row=3, column=0, sticky="w", padx=6, pady=2)
        ttk.Entry(settings, textvariable=self.var_batch, width=10).grid(row=3, column=1, sticky="w", padx=6, pady=2)

        ttk.Label(settings, text="Trunc chars").grid(row=4, column=0, sticky="w", padx=6, pady=2)
        ttk.Entry(settings, textvariable=self.var_trunc, width=10).grid(row=4, column=1, sticky="w", padx=6, pady=2)

        ttk.Checkbutton(settings, text="Use cache (bundle cache/EL_cache.jsonl)", variable=self.var_use_cache).grid(row=5, column=0, columnspan=2, sticky="w", padx=6, pady=2)

        settings.columnconfigure(1, weight=1)

        warn_box = ttk.Labelframe(left, text="Notes / warnings")
        warn_box.pack(fill="both", expand=False, pady=(6, 0))

        self.txt_warn = tk.Text(warn_box, height=8, wrap="word")
        self.txt_warn.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_warn.configure(state="disabled")

        nb = ttk.Notebook(right)
        nb.pack(fill="both", expand=True)

        tab_full = ttk.Frame(nb)
        tab_surv = ttk.Frame(nb)
        tab_log = ttk.Frame(nb)
        nb.add(tab_full, text="EL Full report")
        nb.add(tab_surv, text="EL Survivors")
        nb.add(tab_log, text="Log")

        self.full_table = DataTable(tab_full, on_sort=self._sort_full_table, on_row_activate=self._open_row_detail_modal)
        self.full_table.pack(fill="both", expand=True, padx=6, pady=6)

        self.surv_table = DataTable(tab_surv, on_sort=self._sort_surv_table, on_row_activate=self._open_row_detail_modal)
        self.surv_table.pack(fill="both", expand=True, padx=6, pady=6)

        self.txt_log = tk.Text(tab_log, wrap="word")
        self.txt_log.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_log.configure(state="disabled")

        self.lbl_counts = ttk.Label(self, text="")
        self.lbl_counts.pack(fill="x", padx=10, pady=(0, 10))

    # -------- helpers --------

    def _readiness(self) -> Readiness:
        """F-111. One place decides whether this stage may run, and
        everything that asks — the indicator, the Run button, the Run
        handler — asks it. Three places deciding separately is how the
        gate came to be dropped (F-118) and how an empty model field
        came to start a run (F-93)."""
        # F-117. The provider decides whether a key is needed at all, so
        # readiness reads the persisted configuration rather than probing
        # the environment for one variable. A local server authenticates
        # nothing, and asking its user for a credential was the defect.
        # Review of this session: `load_settings` raises on a settings
        # file that exists and cannot be parsed, and this runs inside
        # `_build_ui` — where `main.py::resolve_plugin_entrypoint`
        # swallows every exception into a `print()`, so a JSON typo made
        # the EL and IL tabs silently absent, with nowhere for the
        # message to go in a windowed onefile build. Readiness degrades
        # to "unconfigured" instead, which blocks the run for a stated
        # reason rather than deleting the stage.
        try:
            cfg = load_settings()
        except Exception:
            cfg = {}
        # Wave 11 session B. `probe` is the last detection the app
        # deposited, never a call made here: this runs inside Tk
        # callbacks, and detection is a network operation. `None` means
        # "not checked yet", which readiness reports as its own state
        # rather than guessing either way.
        return llm_readiness(stage="EL",
                             has_bundle=bool(self.bundle_zip_path),
                             provider=cfg.get("provider", ""),
                             api_key=cfg.get("api_key", ""),
                             model=self.var_model.get(),
                             probe=last_known())

    def _refresh_readiness_label(self):
        """The widget used to read `OPENAI_API_KEY ✓` and nothing else:
        the startup modal cannot be passed without a non-empty key and
        nothing ever clears it, so the one provider-adjacent indicator
        in the application carried zero bits (F-111). It now names
        whichever thing is missing, and `Ready to run` when none is."""
        self.lbl_key.configure(text=self._readiness().label)

    def _log(self, msg: str) -> None:
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", msg)
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")

    def _set_warnings(self, lines: Sequence[str]) -> None:
        self.txt_warn.configure(state="normal")
        self.txt_warn.delete("1.0", "end")
        self.txt_warn.insert("end", "\n".join(lines) if lines else "(none)")
        self.txt_warn.configure(state="disabled")

    def _refresh_counts_label(self):
        if not self.bundle:
            self.lbl_counts.configure(text="")
            return
        pr = self.bundle.parse
        msg = f"Integral rows: {len(pr.rows)} | Skipped invalid: {len(pr.skipped)}"
        if self.counts:
            msg += (
                " | " + _run_summary_counts_text(
                    self.counts, stage="EL",
                    total_rows=len(pr.rows))
            )
        self.lbl_counts.configure(text=msg)

    def _set_controls_running(self, running: bool) -> None:
        # Wave 8 part 2: the five expressions that used to live here are
        # now plugins/_common/stage_state.py::control_states, so they can
        # be asserted on. This method reads state and writes widgets; it
        # decides nothing.
        st = control_states(
            running=running,
            readiness=self._readiness(),
            has_rows=bool(self.full_rows),
            has_input_errors=bool(self.bundle and self.bundle.parse.skipped),
        )
        self.btn_cancel.configure(state=tk_state(st.cancel))
        self.btn_run.configure(state=tk_state(st.run))
        self.btn_export.configure(state=tk_state(st.export))
        self.btn_export_err.configure(state=tk_state(st.export_errors))
        self.btn_export_bundle.configure(state=tk_state(st.export_bundle))

    def _sorted_rows(self, rows: List[Dict[str, Any]], col: str, asc: bool) -> List[Dict[str, Any]]:
        def _key(r: Dict[str, Any]):
            v = r.get(col, "")
            s = _safe_str(v)
            # numeric sort if it looks like a number
            try:
                if s.strip() == "":
                    return (1, 0.0, s)
                return (0, float(s), s)
            except Exception:
                return (0, 0.0, s.lower())
        return sorted(rows, key=_key, reverse=(not asc))

    # -------- bundle load --------

    def _pick_bundle(self):
        p = filedialog.askopenfilename(
            title="Select ScreenA bundle ZIP",
            filetypes=[("ZIP", "*.zip"), ("All files", "*.*")],
        )
        if not p:
            return
        self.bundle_zip_path = p
        self.lbl_bundle.configure(text=Path(p).name)
        self._load_bundle_inputs()

    def _load_bundle_inputs(self):
        if not self.bundle_zip_path:
            return

        warns: List[str] = []
        self.full_rows = []
        self.cancelled = False
        self.not_screened = False
        self.llm_report = {}
        self.outcome = None
        self.survivors = []
        self.counts = {}
        self.crit_impacts = {}
        self.row_eval_lists = []
        self.active_criterion_id = None
        self.lbl_crit_filter.configure(text="Criterion filter: (none)")
        self.btn_clear_filter.configure(state="disabled")

        self._refresh_readiness_label()

        try:
            self.bundle = _load_bundle(self.bundle_zip_path)
        except Exception as e:
            self.bundle = None
            messagebox.showerror("Bundle load failed", str(e))
            return

        # load cache if present
        self.cache_map = {}
        try:
            with zipfile.ZipFile(self.bundle_zip_path, "r") as zf:
                cache_member = self.bundle.root + EL_CACHE_REL
                if cache_member in zf.namelist():
                    self.cache_map = _load_cache_from_jsonl(_decode_bytes(_read_zip_bytes(zf, cache_member)))
        except Exception:
            self.cache_map = {}

        m = self.bundle.manifest
        schema = _safe_str(m.get("bundle_schema", m.get("schema", ""))).strip() or "unknown"
        created_at = _safe_str(m.get("created_at", "")).strip()
        created_by = _safe_str(m.get("created_by", "")).strip()
        stages = ((m.get("pipeline", {}) or {}).get("stages", None)) or ((m.get("pipeline_state", {}) or {}).get("stages", None)) or {}
        st_el = _safe_str(stages.get("EL", "")).strip() or "unknown"
        self.lbl_bundle_meta.configure(text=f"schema={schema} | created_at={created_at} | created_by={created_by} | EL={st_el}")

        # warnings: duplicates (they were moved to parse.skipped during load)
        dup_ids = []
        for e in (self.bundle.parse.skipped or []):
            if _safe_str(e.get("reason", "")).strip() == "duplicate local_id":
                lid = _safe_str((e.get("row") or {}).get("local_id", "")).strip()
                if lid:
                    dup_ids.append(lid)
        # unique sample
        sample = []
        seen = set()
        for x in dup_ids:
            if x not in seen:
                sample.append(x); seen.add(x)
            if len(sample) >= 10:
                break
        if dup_ids:
            warns.append(f"[data] duplicate local_id detected (n={len(set(dup_ids))}), sample={', '.join(sample)}")

        # refresh criteria table + notes
        self._refresh_criteria_table(pre_run=True)
        if self.bundle.criteria and self.bundle.criteria.warnings:
            warns.extend(self.bundle.criteria.warnings)
        self._set_warnings(warns)

        # clear report tables
        self.full_table.clear()
        self.surv_table.clear()
        self._refresh_counts_label()

        # F-118: the load path used to set four buttons itself, and its
        # Run predicate (`_has_openai_key()`) disagreed with the one in
        # `_set_controls_running` (`self.bundle_zip_path`). Since the
        # latter runs in the `finally` of every run, the gate applied
        # here survived only until the first run of a session ended.
        # There is one predicate now, so there is nothing to disagree.
        # `full_rows` was cleared above, so this yields exactly what the
        # four lines did — and it also resets IL's sixth button, which
        # they forgot.
        self._set_controls_running(False)

        self.lbl_status.configure(text="Ready.")

    def _detect_duplicate_local_ids(self, rows: List[Dict[str, str]]) -> Tuple[int, List[str]]:
        seen = set()
        dups = []
        for r in rows:
            lid = _safe_str(r.get("local_id", "")).strip()
            if not lid:
                continue
            if lid in seen:
                dups.append(lid)
            else:
                seen.add(lid)
        uniq = []
        s2 = set()
        for x in dups:
            if x not in s2:
                uniq.append(x)
                s2.add(x)
        return len(uniq), uniq[:10]

    # -------- criteria table --------

    def _refresh_criteria_table(self, pre_run: bool):
        crits = self.bundle.criteria.criteria if (self.bundle and self.bundle.criteria) else []

        cols = ["id", "type", "targets", "operator", "what", "threshold", "status", "notes"]
        if not pre_run:
            cols += ["n_failed", "n_missing", "n_met", "n_uncertain"]

        rows: List[Dict[str, Any]] = []
        header_set = set(self.bundle.parse.header) if (self.bundle and self.bundle.parse) else set()
        header_set_l = {h.lower() for h in header_set}

        for c in crits:
            status = "OK"
            notes = ""

            if not c.targets:
                status = "WARNING"
                notes = "missing target -> treated as MISSING (PASS_FLAGGED)"
            elif header_set and not any((t or "").lower() in header_set_l for t in c.targets):
                status = "WARNING"
                notes = f"missing column(s): {', '.join(c.targets)} -> treated as MISSING (PASS_FLAGGED)"

            op = (c.operator or "").strip().lower()
            if op != "llm":
                status = "WARNING"
                notes = (notes + " | " if notes else "") + f"operator '{op}' treated as UNCERTAIN in EL"

            d: Dict[str, Any] = {
                "id": c.id,
                "type": c.ctype,
                "targets": ",".join(c.targets),
                "operator": c.operator,
                "what": c.what_raw,
                "threshold": _safe_str(c.threshold),
                "status": status,
                "notes": notes,
            }

            if not pre_run:
                imp = self.crit_impacts.get(c.id, {"failed": 0, "missing": 0, "met": 0, "uncertain": 0})
                d["n_failed"] = str(imp.get("failed", 0))
                d["n_missing"] = str(imp.get("missing", 0))
                d["n_met"] = str(imp.get("met", 0))
                d["n_uncertain"] = str(imp.get("uncertain", 0))

            rows.append(d)

        self.criteria_table.set_columns(cols)
        col, asc = self.sort_crit
        if col:
            rows = self._sorted_rows(rows, col, asc)
        self.criteria_table.render_rows_incremental(rows)

    def _on_criterion_activated(self, row: Dict[str, str]):
        cid = _safe_str(row.get("id", "")).strip()
        if not cid:
            return
        self.active_criterion_id = cid
        self.lbl_crit_filter.configure(text=f"Criterion filter: {cid}")
        self.btn_clear_filter.configure(state="normal")
        self._refresh_reports_view()

    def _clear_criterion_filter(self):
        self.active_criterion_id = None
        self.lbl_crit_filter.configure(text="Criterion filter: (none)")
        self.btn_clear_filter.configure(state="disabled")
        self._refresh_reports_view()

    def _sort_criteria_table(self, col: str):
        cur_col, asc = self.sort_crit
        if cur_col == col:
            asc = not asc
        else:
            asc = True
        self.sort_crit = (col, asc)
        self._refresh_criteria_table(pre_run=(not bool(self.full_rows)))

    # -------- reports --------

    def _refresh_reports_view(self):
        if not self.bundle:
            return

        full_cols = list(self.bundle.parse.header) + [
            "el_outcome", "el_failed_ids", "el_missing_ids", "el_met_ids", "el_uncertain_ids", "el_reason_summary"
        ]
        surv_cols = list(self.bundle.parse.header)

        full_view = self.full_rows
        if self.active_criterion_id:
            cid = self.active_criterion_id
            filtered = []
            for i, r in enumerate(self.full_rows):
                ev = self.row_eval_lists[i] if i < len(self.row_eval_lists) else {"failed": [], "missing": [], "met": [], "uncertain": []}
                if (cid in ev.get("failed", [])) or (cid in ev.get("missing", [])) or (cid in ev.get("met", [])) or (cid in ev.get("uncertain", [])):
                    filtered.append(r)
            full_view = filtered

        surv_view = self.survivors
        if self.active_criterion_id and self.full_rows:
            cid = self.active_criterion_id
            touched_survivor_ids = set()
            for i, r in enumerate(self.full_rows):
                if r.get("el_outcome") == "OUT":
                    continue
                ev = self.row_eval_lists[i]
                if (cid in ev.get("failed", [])) or (cid in ev.get("missing", [])) or (cid in ev.get("met", [])) or (cid in ev.get("uncertain", [])):
                    touched_survivor_ids.add(_safe_str(r.get("local_id", "")).strip())
            surv_view = [r for r in self.survivors if _safe_str(r.get("local_id", "")).strip() in touched_survivor_ids]

        col, asc = self.sort_full
        if col and full_view:
            full_view = self._sorted_rows(full_view, col, asc)

        col2, asc2 = self.sort_surv
        if col2 and surv_view:
            surv_view = self._sorted_rows(surv_view, col2, asc2)

        self.full_table.set_columns(full_cols)
        self.full_table.render_rows_incremental(full_view)

        self.surv_table.set_columns(surv_cols)
        self.surv_table.render_rows_incremental(surv_view)

    def _sort_full_table(self, col: str):
        cur_col, asc = self.sort_full
        if cur_col == col:
            asc = not asc
        else:
            asc = True
        self.sort_full = (col, asc)
        self._refresh_reports_view()

    def _sort_surv_table(self, col: str):
        cur_col, asc = self.sort_surv
        if cur_col == col:
            asc = not asc
        else:
            asc = True
        self.sort_surv = (col, asc)
        self._refresh_reports_view()

    # -------- Row detail modal --------

    def _open_row_detail_modal(self, row: Dict[str, Any]):
        if not self.bundle:
            return

        win = tk.Toplevel(self)
        win.title("EL Row details")
        win.geometry("950x650")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        lid = _safe_str(row.get("local_id", "")).strip()
        title = _safe_str(row.get("title", "")).strip() or _safe_str(row.get("Title", "")).strip()

        top = ttk.Frame(win)
        top.pack(fill="x", padx=10, pady=10)

        ttk.Label(top, text=f"local_id: {lid}").pack(anchor="w")
        if title:
            ttk.Label(top, text=f"title: {title[:250]}").pack(anchor="w")

        outcome = _safe_str(row.get("el_outcome", "")).strip()
        if outcome:
            ttk.Label(top, text=f"EL outcome: {outcome}").pack(anchor="w")
            rs = _safe_str(row.get("el_reason_summary", "")).strip()
            if rs:
                ttk.Label(top, text=f"summary: {rs}").pack(anchor="w")

        box = ttk.Labelframe(win, text="Per-criterion evidence (from el_evidence_json)")
        box.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ["cid", "type", "operator", "targets", "what", "threshold", "status", "decision", "confidence", "field", "quote_valid", "quote"]
        table = DataTable(box, on_sort=lambda _c: None, on_row_activate=None)
        table.pack(fill="both", expand=True, padx=6, pady=6)
        table.set_columns(cols)

        # parse evidence dict
        ev_raw = _safe_str(row.get("el_evidence_json", "{}"))
        try:
            evj = json.loads(ev_raw) if ev_raw.strip() else {}
        except Exception:
            evj = {}

        detail_rows: List[Dict[str, Any]] = []
        for c in self.bundle.criteria.criteria:
            obj = evj.get(c.id, {}) if isinstance(evj, dict) else {}
            status = _safe_str(obj.get("status", ""))
            decision = _safe_str(obj.get("decision", ""))
            confidence = _safe_str(obj.get("confidence", ""))
            field = _safe_str(obj.get("field", ""))
            quote = _safe_str(obj.get("quote", ""))
            qv = _safe_str(obj.get("quote_valid", obj.get("quote_valid", "")))
            if isinstance(obj.get("quote_valid", None), bool):
                qv = "True" if obj.get("quote_valid") else "False"

            detail_rows.append({
                "cid": c.id,
                "type": c.ctype,
                "operator": c.operator,
                "targets": ",".join(c.targets),
                "what": c.what_raw,
                "threshold": _safe_str(c.threshold),
                "status": status,
                "decision": decision,
                "confidence": confidence,
                "field": field,
                "quote_valid": qv,
                "quote": quote,
            })

        table.render_rows_incremental(detail_rows)

        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(btns, text="Close", command=win.destroy).pack(side="right")
        win.bind("<Escape>", lambda _e: win.destroy())

    # -------- run --------

    def _run_clicked(self):
        if not self.bundle or self._worker:
            return
        # F-93/F-111: one readiness check, and it is the same one the
        # indicator and the Run button consult. The model field is part of
        # it. `(self.var_model.get() or DEFAULT_MODEL).strip()` put the
        # strip OUTSIDE the `or`, so a whitespace-only field was truthy,
        # survived the fallback, and reached the engine as "" — which the
        # engine skips silently, leaving a full corpus of unscreened records
        # under a status line reading "done".
        #
        # It refuses rather than substituting DEFAULT_MODEL. A blanked field
        # is a mistake, and quietly screening a whole corpus against a model
        # the user did not type costs real money and produces results
        # attributable to the wrong model. F-119's wording for the missing
        # key now lives in stage_state::llm_readiness with the other cases.
        ready = self._readiness()
        if not ready.can_run:
            messagebox.showerror(f"EL cannot start", ready.detail)
            return

        # parse settings
        model = ready.model
        try:
            temperature = float(self.var_temp.get())
        except Exception:
            temperature = 0.0
        # F-118: `try: int(...) except: <default>` rescued 'abc' and
        # accepted '-100'. A negative truncation reaches the prompt
        # builder as a negative slice: it removes the LAST |n|
        # characters of every field and empties any field shorter than
        # that — in practice the title and keywords of every record.
        numeric = parse_numeric_settings(
            batch_raw=self.var_batch.get(),
            trunc_raw=self.var_trunc.get(),
            batch_default=DEFAULT_BATCH_SIZE,
            trunc_default=DEFAULT_TRUNC_CHARS)
        batch_size = numeric.batch_size
        trunc_chars = numeric.trunc_chars
        for _problem in numeric.problems:
            self._log("[EL] setting corrected: " + _problem + "\n")
        if numeric.problems:
            messagebox.showwarning("EL settings corrected",
                                   "\n\n".join(numeric.problems))
        use_cache = bool(self.var_use_cache.get())

        self._cancel.clear()
        self.pbar.configure(value=0.0, maximum=100.0)
        self.lbl_status.configure(text="Running EL…")
        self._set_controls_running(True)

        def progress_cb(frac: float):
            self.after(0, lambda: self.pbar.configure(value=max(0.0, min(100.0, frac * 100.0))))

        def progress_evt(evt: Dict[str, Any]):
            kind = _safe_str(evt.get("kind", ""))
            sub = _safe_str(evt.get("sub", ""))
            if kind == "l_batch" and sub == "sending":
                bi = evt.get("batch_idx"); bt = evt.get("batch_total")
                ci = evt.get("crit_idx"); ct = evt.get("crit_total")
                self.after(0, lambda: self.lbl_status.configure(text=f"LLM: criterion {ci}/{ct} batch {bi}/{bt}…"))

        def work():
            try:
                (full_rows, survivors, counts, crit_impacts,
                 row_eval_lists, cache_out, cancelled,
                 run_report) = run_el_screen(
                    self.bundle.parse,
                    self.bundle.criteria,
                    model=model,
                    trunc_chars=trunc_chars,
                    batch_size=batch_size,
                    temperature=temperature,
                    use_cache=use_cache,
                    cache_in=self.cache_map if use_cache else {},
                    cancel_event=self._cancel,  # ✅ ELView uses self._cancel
                    log_cb=lambda s: self.after(0, (lambda ss=s: self._log(ss))),
                    progress_cb=progress_cb,
                    progress_evt=progress_evt,
                )

                self.llm_report = run_report

                if cancelled:
                    # F-02: the engine reports that it stopped mid-corpus.
                    # Trust that rather than re-reading the event, drop the
                    # partial results instead of leaving them on screen, and
                    # latch the flag so the export handlers refuse even if an
                    # earlier complete run left rows behind.
                    self.cancelled = True
                    self.full_rows = []
                    self.survivors = []
                    self.after(0, lambda: self._log(
                        "\n[CANCELLED] Partial results discarded; export stays "
                        "disabled until EL is re-run to completion.\n"))
                    self.outcome = run_outcome(
                        stage="EL", counts=counts,
                        llm_report=run_report, cancelled=True,
                        not_screened=False, total_rows=len(full_rows))
                    self.after(0, lambda t=self.outcome.label:
                               self.lbl_status.configure(text=t))
                    return

                self.full_rows = full_rows
                self.survivors = survivors
                self.counts = counts
                # F-34: the engine reports the no-op through the counts.
                self.not_screened = bool(counts.get(NOT_SCREENED, 0))
                self.crit_impacts = crit_impacts
                self.row_eval_lists = row_eval_lists
                self.cache_map = cache_out

                # ✅ refresh views + counts + criteria impacts
                self.after(0, self._refresh_reports_view)
                self.after(0, lambda: self._refresh_criteria_table(pre_run=False))
                self.after(0, self._refresh_counts_label)
                # F-34: a stage that evaluated no criteria is not "done"
                # in any sense the user means by the word. The
                # classification is stage_state::run_outcome's now.
                self.outcome = run_outcome(
                    stage="EL", counts=counts,
                    llm_report=run_report, cancelled=False,
                    not_screened=self.not_screened,
                    total_rows=len(full_rows))
                self.after(0, lambda t=self.outcome.label:
                           self.lbl_status.configure(text=t))
                # F-93: the status line is one line in a bar the user
                # may not be looking at; the counts belong in the trail
                # too, next to the per-batch failures that explain them.
                self.after(0, lambda t=self.outcome.label:
                           self._log("\n[EL] " + t + "\n"))

            except Exception as e:
                self.after(0, lambda m=str(e): messagebox.showerror("EL run failed", m))
                self.after(0, lambda: self.lbl_status.configure(text="EL failed."))
            finally:
                self.after(0, lambda: self._set_controls_running(False))
                # clear worker flag so you can run again
                self._worker = None

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    def _cancel_run(self):
        self._cancel.set()
        self._log("\n[Cancel requested]\n")
        self.lbl_status.configure(text="Cancelling…")

    # -------- Export --------

    def _export_clicked(self):
        if not self.bundle:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        blocked = _export_block_reason(has_rows=bool(self.full_rows),
                                       cancelled=self.cancelled)
        if blocked:
            messagebox.showwarning("Cannot export", blocked)
            return
        # F-34: a stage that screened nothing may still be exported,
        # but not without the user saying so out loud.
        # F-93 extends F-34's gate rather than adding a second one:
        # `not_screened` covers a stage with no criteria, and the
        # outcome covers a stage that had criteria and still learned
        # nothing. One gate, two diagnoses.
        confirm = _export_confirm_reason(
            not_screened=self.not_screened,
            stage="EL",
            outcome_reason=(self.outcome.ack_reason if self.outcome
                            else None))
        if confirm:
            title = ("Nothing was screened" if self.not_screened
                     else "Check this before exporting")
            if not messagebox.askyesno(title, confirm):
                return

        default_name = f"{_now_stamp()}_EL_reports.xlsx"
        p = filedialog.asksaveasfilename(
            title="Save EL reports",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not p:
            return

        try:
            _export_el_xlsx(p, self.full_rows, self.survivors, self.bundle.parse.header)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_errors_clicked(self):
        if not self.bundle:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        if not self.bundle.parse.skipped:
            messagebox.showinfo("No input errors", "No skipped/invalid rows were recorded for this bundle.")
            return

        default_name = f"{_now_stamp()}_input_errors.csv"
        p = filedialog.asksaveasfilename(
            title="Save input_errors.csv",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv")],
        )
        if not p:
            return

        try:
            # F-74: the canonical six-column schema through the shared
            # writer, not the legacy reason,row_json layout this handler
            # used to emit inline.
            _export_input_errors_csv_from_dicts(
                p, self.bundle.parse.skipped, stage="EL")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_bundle_clicked(self):
        if not self.bundle:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        blocked = _export_block_reason(has_rows=bool(self.full_rows),
                                       cancelled=self.cancelled)
        if blocked:
            messagebox.showwarning("Cannot export", blocked)
            return
        # F-34: a stage that screened nothing may still be exported,
        # but not without the user saying so out loud.
        # F-93 extends F-34's gate rather than adding a second one:
        # `not_screened` covers a stage with no criteria, and the
        # outcome covers a stage that had criteria and still learned
        # nothing. One gate, two diagnoses.
        confirm = _export_confirm_reason(
            not_screened=self.not_screened,
            stage="EL",
            outcome_reason=(self.outcome.ack_reason if self.outcome
                            else None))
        if confirm:
            title = ("Nothing was screened" if self.not_screened
                     else "Check this before exporting")
            if not messagebox.askyesno(title, confirm):
                return

        default_name = f"{_now_stamp()}_post_EL_bundle.zip"
        out_zip = filedialog.asksaveasfilename(
            title="Save next bundle ZIP (post-EL)",
            defaultextension=".zip",
            initialfile=default_name,
            filetypes=[("ZIP", "*.zip")],
        )
        if not out_zip:
            return

        try:
            self._build_next_bundle_zip(out_zip)
        except Exception as e:
            messagebox.showerror("Bundle export failed", str(e))
            return

        self.lbl_status.configure(text=f"Bundle exported: {Path(out_zip).name}")

    def _build_next_bundle_zip(self, out_zip: str) -> None:
        """Create a new bundle zip where data/current.csv becomes the EL survivors.

        Keeps other files from the input bundle, updates manifest pipeline stage, adds reports, carries input_errors,
        and writes cache/EL_cache.jsonl when enabled.
        """
        assert self.bundle is not None

        with zipfile.ZipFile(self.bundle.zip_path, "r") as zf_in:
            members = zf_in.namelist()
            root = self.bundle.root

            manifest = dict(self.bundle.manifest)

            def _set_stage(m: Dict[str, Any], key: str, value: str):
                if "pipeline" in m and isinstance(m.get("pipeline"), dict):
                    m["pipeline"].setdefault("stages", {})
                    m["pipeline"]["stages"][key] = value
                if "pipeline_state" in m and isinstance(m.get("pipeline_state"), dict):
                    m["pipeline_state"].setdefault("stages", {})
                    m["pipeline_state"]["stages"][key] = value

            _set_stage(manifest, "EL", "done")
            manifest["updated_at"] = datetime.utcnow().isoformat() + "Z"

            header = list(self.bundle.parse.header)
            if "local_id" not in header:
                header = ["local_id"] + header

            header_full = [
                *header,
                "el_outcome", "el_failed_ids", "el_missing_ids", "el_met_ids", "el_uncertain_ids", "el_reason_summary", "el_evidence_json",
            ]

            # F-05: the writing itself lives in plugins/_common/bundle.py so
            # that this method, its standalone.py twin and IL's two copies
            # share one implementation — and, in particular, one that
            # refreshes the manifest's sha256 map. None of the four did.
            _write_llm_stage_bundle(
                out_zip, zf_in,
                root=root,
                manifest=manifest,
                stage="EL",
                reports_dir_rel=REPORTS_DIR_REL,
                cache_rel=EL_CACHE_REL,
                parse_header=header,
                survivors=self.survivors,
                full_rows=self.full_rows,
                full_header=header_full,
                skipped=self.bundle.parse.skipped,
                counts=self.counts,
                not_screened=self.not_screened,
                cancelled=self.cancelled,
                llm_report=self.llm_report,
                cache_text=(_dump_cache_to_jsonl(self.cache_map)
                            if self.var_use_cache.get() else None),
            )


# StandaloneELPlugin moved to plugins/06_el/standalone.py in
# Conv 6 / Commit 4. The plugin module re-exports it from there.
