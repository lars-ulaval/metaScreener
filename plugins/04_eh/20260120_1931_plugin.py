# -*- coding: utf-8 -*-
"""
plugin.py — Screen A (EH-only) as a PRISMA Hub tab plugin (Contract v2, EH stage)

Self-contained (UI + engine).

Inputs
- Aggregate A vector: *_aggregate.csv (CSV). May contain malformed rows.
- Harmonized criteria: criteria_harmonized.csv (CSV). May contain non-EH stages too.

EH-only behavior (frozen defaults)
- Process only "integral" rows: rows that parse correctly AND match header column count AND have non-empty local_id.
- Skip malformed rows; show counts in UI; export input_errors.csv (raw record text).
- Load criteria where stage == "EH" (case-insensitive) and enabled truthy.
- If stage column missing, assume all criteria are EH.
- If a criterion targets a missing column: treat as MISSING => PASS_FLAGGED (+ warning). Never OUT.
- Normalize: case-insensitive matching; trim; empty => MISSING. doc_type + lang normalization maps.
- Decision logic:
    any FAILED => OUT
    else if all MET => PASS_CLEAN
    else => PASS_FLAGGED
- Contradiction handling: STRICT APPLY + WARN (no auto-fix, no stop).

Outputs
1) EH_FULL report (UI + export): all aggregate columns + EH outcome + reasons
2) EH_SURVIVORS report (UI + export): only survivors (PASS_CLEAN + PASS_FLAGGED),
   in the exact same schema/order as original aggregate (no extra columns).

Exports
- One XLSX with two sheets: EH_FULL, EH_SURVIVORS
- Optional input_errors.csv for skipped records

Notes
- This plugin is EH-only; it does NOT perform IH/EL/IL or LLM checks.
- Unknown operators (including llm) => MISSING/UNCERTAIN => PASS_FLAGGED (+ warning).

"""

import csv
import io
import os
import re
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from prisma_hub.plugin_api import BasePlugin, PluginMeta


TAB_TITLE = "Screen A — EH"


# ----------------------------
# Config
# ----------------------------

MAX_UI_ROWS = 2000
PROGRESS_EVERY_N = 200

TRUTHY = {"1", "true", "yes", "y", "on", "enabled"}


# ----------------------------
# Normalization maps
# ----------------------------

DOC_TYPE_MAP = {
    # journal-like
    "journal": "journal",
    "journal article": "journal",
    "article": "journal",
    "research article": "journal",
    "original article": "journal",
    "journal-article": "journal",
    "peer reviewed article": "journal",
    # conference/proceedings-like
    "conference": "conference",
    "conference paper": "conference",
    "conference article": "conference",
    "proceedings": "conference",
    "proceedings article": "conference",
    "inproceedings": "conference",
    "conference proceedings": "conference",
    # misc common
    "preprint": "preprint",
    "arxiv": "preprint",
    "thesis": "thesis",
    "dissertation": "thesis",
    "book chapter": "book_chapter",
    "chapter": "book_chapter",
    "report": "report",
}

LANG_MAP = {
    "en": "en",
    "eng": "en",
    "english": "en",
    "en-us": "en",
    "en-gb": "en",
    "fr": "fr",
    "fra": "fr",
    "fre": "fr",
    "french": "fr",
    "fr-ca": "fr",
    "es": "es",
    "spa": "es",
    "spanish": "es",
}


# ----------------------------
# Data structures
# ----------------------------

@dataclass
class Criterion:
    stage: str
    cid: str
    ctype: str           # include / exclude
    scope: str
    label: str
    operator: str
    targets: List[str]   # parsed list of targets, canonicalized minimally
    what_raw: str
    what_list: List[str]
    threshold: Optional[float]
    enabled: bool
    source_text: str


@dataclass
class ParseReport:
    header: List[str]
    rows: List[Dict[str, str]]           # integral rows only
    skipped: List[Tuple[int, str, str]]  # (record_index_1based_in_file_ex_header, reason, raw_record_text)


@dataclass
class CriteriaLoadReport:
    criteria: List[Criterion]
    warnings: List[str]


# ----------------------------
# Utilities
# ----------------------------

def _now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    try:
        return str(x)
    except Exception:
        return ""


def _norm_basic(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _truthy(x: Any) -> bool:
    s = _safe_str(x).strip().lower()
    return s in TRUTHY


def _split_targets(target_cell: str) -> List[str]:
    # Accept comma-separated targets. Keep original token names trimmed.
    t = _safe_str(target_cell)
    parts = []
    for p in t.split(","):
        p = p.strip().strip('"').strip("'")
        if p:
            parts.append(p)
    return parts


def _split_what_list(what_cell: str) -> List[str]:
    # Semicolon-delimited list (also normalize newlines -> ;)
    w = _safe_str(what_cell).replace("\n", ";").replace("\r", ";")
    parts = []
    for p in w.split(";"):
        p = p.strip().strip('"').strip("'")
        if p:
            parts.append(p)
    return parts


def _norm_doc_type(v: str) -> str:
    x = _norm_basic(v).lower()
    x = x.replace("_", " ").replace("-", " ")
    x = re.sub(r"\s+", " ", x).strip()
    return DOC_TYPE_MAP.get(x, x)


def _norm_lang(v: str) -> str:
    x = _norm_basic(v).lower()
    x = x.replace("_", "-")
    x = re.sub(r"\s+", "", x)
    return LANG_MAP.get(x, x)


def _norm_for_target(target: str, value: str) -> str:
    t = (target or "").strip().lower()
    v = _safe_str(value)
    if not v.strip():
        return ""
    if t == "doc_type":
        return _norm_doc_type(v)
    if t == "lang":
        return _norm_lang(v)
    return _norm_basic(v)


def _norm_what_for_target(target: str, what: str) -> str:
    # Apply same normalization to criteria "what" values when target is doc_type/lang
    return _norm_for_target(target, what)


# ----------------------------
# Robust CSV parsing (record splitter)
# ----------------------------

def _decode_bytes(b: bytes) -> str:
    # Best-effort decoding
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="ignore")


def _split_csv_records(text: str) -> List[str]:
    """
    Split CSV into record strings by scanning newlines not inside quotes.
    Preserves embedded newlines inside quoted fields.
    """
    if not text:
        return []
    # Normalize newlines to \n (keep content otherwise)
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    recs: List[str] = []
    buf: List[str] = []
    in_quote = False
    i = 0
    n = len(text)

    while i < n:
        ch = text[i]
        if ch == '"':
            # Handle escaped quotes "" within a quoted field
            if in_quote and (i + 1) < n and text[i + 1] == '"':
                buf.append('"')
                buf.append('"')
                i += 2
                continue
            in_quote = not in_quote
            buf.append(ch)
            i += 1
            continue

        if ch == "\n" and not in_quote:
            recs.append("".join(buf))
            buf = []
            i += 1
            continue

        buf.append(ch)
        i += 1

    if buf:
        recs.append("".join(buf))
    return recs


def _parse_csv_tolerant(path: str, required_id: str = "local_id") -> ParseReport:
    b = Path(path).read_bytes()
    text = _decode_bytes(b)

    records = _split_csv_records(text)
    if not records:
        raise ValueError("CSV is empty.")

    # Parse header
    try:
        header = next(csv.reader(io.StringIO(records[0] + "\n")))
    except Exception as e:
        raise ValueError(f"Failed to parse CSV header: {e}")

    header = [h.strip() for h in header]
    if not header or all(not h for h in header):
        raise ValueError("CSV header is empty.")

    # Validate id column presence (EH contract default)
    if required_id not in header:
        # We keep going, but then A4 rule will skip all rows (no id) unless you rename.
        # For robustness: allow aliasing common id names to local_id by duplicating.
        pass

    rows: List[Dict[str, str]] = []
    skipped: List[Tuple[int, str, str]] = []
    expected_n = len(header)

    for rec_idx, rec in enumerate(records[1:], start=1):  # 1-based after header
        raw = rec
        if not raw.strip():
            skipped.append((rec_idx, "empty_record", raw))
            continue
        try:
            parsed = next(csv.reader(io.StringIO(raw + "\n")))
        except csv.Error as e:
            skipped.append((rec_idx, f"csv_error:{e}", raw))
            continue
        except Exception as e:
            skipped.append((rec_idx, f"parse_error:{e}", raw))
            continue

        if len(parsed) != expected_n:
            skipped.append((rec_idx, f"bad_column_count:{len(parsed)}!=expected:{expected_n}", raw))
            continue

        d = {header[i]: _safe_str(parsed[i]) for i in range(expected_n)}

        # A4: missing/empty local_id => skip
        lid = _safe_str(d.get(required_id, "")).strip()
        if not lid:
            skipped.append((rec_idx, "missing_local_id", raw))
            continue

        rows.append(d)

    return ParseReport(header=header, rows=rows, skipped=skipped)


# ----------------------------
# Criteria loading + contradiction warnings
# ----------------------------

def _load_criteria_eh(path: str) -> CriteriaLoadReport:
    warnings: List[str] = []

    # Read with DictReader (criteria files should be valid; still robust)
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        # Peek header
        sniffer = csv.reader(f)
        try:
            hdr = next(sniffer)
        except StopIteration:
            return CriteriaLoadReport(criteria=[], warnings=["Criteria file is empty."])
        hdr_norm = [h.strip() for h in hdr]
        f.seek(0)
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return CriteriaLoadReport(criteria=[], warnings=["Criteria header not found."])

        has_stage = any((h or "").strip().lower() == "stage" for h in reader.fieldnames)
        # if DictReader fieldnames differ from peek (rare), trust DictReader.

        crits: List[Criterion] = []
        row_i = 0
        for row in reader:
            row_i += 1
            stage = _safe_str(row.get("stage", "")).strip()
            if not has_stage:
                stage = "EH"  # default if stage column absent
            if stage.strip().upper() != "EH":
                continue

            enabled = _truthy(row.get("enabled", "1"))
            if not enabled:
                continue

            cid = _safe_str(row.get("id", "")).strip() or f"EH_ROW_{row_i}"
            ctype = (_safe_str(row.get("type", "")).strip().lower() or "include")
            scope = _safe_str(row.get("scope", "")).strip()
            label = _safe_str(row.get("label", "")).strip()
            operator = (_safe_str(row.get("operator", "")).strip().lower() or "equals")
            target_cell = _safe_str(row.get("target", "")).strip()
            targets = _split_targets(target_cell) or []
            what_raw = _safe_str(row.get("what", "")).strip()
            what_list = _split_what_list(what_raw)
            source_text = _safe_str(row.get("source_text", "")).strip()

            thr = None
            thr_cell = _safe_str(row.get("threshold", "")).strip()
            if thr_cell:
                try:
                    thr = float(thr_cell)
                except Exception:
                    warnings.append(f"[criteria] Row {row_i} ({cid}): invalid threshold '{thr_cell}' -> ignored.")

            if ctype not in ("include", "exclude"):
                warnings.append(f"[criteria] {cid}: unknown type '{ctype}' -> treating as include.")
                ctype = "include"

            if not targets:
                warnings.append(f"[criteria] {cid}: missing target -> criterion will be treated as MISSING (PASS_FLAGGED).")

            crits.append(
                Criterion(
                    stage="EH",
                    cid=cid,
                    ctype=ctype,
                    scope=scope,
                    label=label,
                    operator=operator,
                    targets=targets,
                    what_raw=what_raw,
                    what_list=what_list,
                    threshold=thr,
                    enabled=True,
                    source_text=source_text,
                )
            )

    # Contradiction warnings (G1-A): only warn, no auto-fix
    warnings.extend(_detect_contradictions_simple(crits))
    return CriteriaLoadReport(criteria=crits, warnings=warnings)


def _detect_contradictions_simple(crits: Sequence[Criterion]) -> List[str]:
    """
    Heuristic contradiction detection:
    - For same single-target column, intersect include allowed set (equals/in_list)
      with exclude forbidden set (equals/in_list). If overlap => warn.
    """
    warns: List[str] = []
    # target -> {"include": set(values), "exclude": set(values)}
    bucket: Dict[str, Dict[str, set]] = {}

    for c in crits:
        if not c.targets or len(c.targets) != 1:
            continue
        tgt = c.targets[0].strip()
        op = (c.operator or "").strip().lower()
        if op not in ("equals", "in_list"):
            continue
        vals = c.what_list[:] if c.what_list else ([_safe_str(c.what_raw)] if c.what_raw else [])
        # normalize against target
        vals_norm = {_norm_what_for_target(tgt, v) for v in vals if _norm_what_for_target(tgt, v)}
        if not vals_norm:
            continue
        bucket.setdefault(tgt, {"include": set(), "exclude": set()})
        bucket[tgt][c.ctype].update(vals_norm)

    for tgt, d in bucket.items():
        inc = d.get("include", set())
        exc = d.get("exclude", set())
        overlap = sorted(inc.intersection(exc))
        if overlap:
            warns.append(
                f"[criteria] Possible contradiction on target '{tgt}': values appear in BOTH include and exclude sets: {', '.join(overlap[:10])}"
                + (" ..." if len(overlap) > 10 else "")
            )
    return warns


# ----------------------------
# EH evaluation engine
# ----------------------------

def _get_first_nonempty(row: Dict[str, str], targets: Sequence[str]) -> Tuple[str, str]:
    """
    Returns (target_used, value) using the first non-empty target in order.
    If all empty or missing, returns (first_target_or_empty, "").
    """
    if not targets:
        return "", ""
    for t in targets:
        if t in row:
            v = _safe_str(row.get(t, ""))
            if v.strip():
                return t, v
    # if none non-empty, still return first for normalization context
    t0 = targets[0]
    return t0, _safe_str(row.get(t0, "")) if t0 in row else ""


def _eval_criterion(row: Dict[str, str], header_set: set, c: Criterion) -> Tuple[str, str]:
    """
    Evaluate one criterion against a row.

    Returns (status, note):
      status in {"MET","FAILED","MISSING","UNKNOWN"}
      note is a compact human string for debugging.
    """
    # Missing target => MISSING (PASS_FLAGGED)
    if not c.targets:
        return "MISSING", "missing_target"

    # If all targets absent from header => MISSING (PASS_FLAGGED)
    if not any(t in header_set for t in c.targets):
        return "MISSING", f"missing_column:{','.join(c.targets)}"

    target_used, raw_val = _get_first_nonempty(row, c.targets)
    val = _norm_for_target(target_used, raw_val)

    if not val:
        return "MISSING", f"empty_value:{target_used}"

    op = (c.operator or "").strip().lower()

    # Normalize "what" values according to the chosen target
    what_list = c.what_list[:] if c.what_list else ([_safe_str(c.what_raw)] if c.what_raw else [])
    what_list_norm = [_norm_what_for_target(target_used, w) for w in what_list if _norm_what_for_target(target_used, w)]

    # Unknown operators -> UNKNOWN (PASS_FLAGGED)
    if op in ("llm",):
        return "UNKNOWN", "operator_llm_not_supported_in_EH"
    if op not in ("equals", "contains", "regex", "in_list", "not_in", "gte", "lte", "between"):
        return "UNKNOWN", f"unknown_operator:{op}"

    # Helper: determine match truth for include/exclude semantics
    def _as_float(x: str) -> Optional[float]:
        try:
            return float(x)
        except Exception:
            return None

    matched: Optional[bool] = None
    note = ""

    if op == "equals":
        if not what_list_norm:
            return "UNKNOWN", "equals_missing_what"
        matched = (val.lower() == what_list_norm[0].lower())
        note = f"equals({target_used})"

    elif op == "contains":
        if not what_list_norm:
            return "UNKNOWN", "contains_missing_what"
        hay = val.lower()
        matched = any((w.lower() in hay) for w in what_list_norm if w)
        note = f"contains({target_used})"

    elif op == "regex":
        if not what_list:
            return "UNKNOWN", "regex_missing_pattern"
        pat = what_list[0]
        try:
            matched = bool(re.search(pat, raw_val, flags=re.IGNORECASE))
            note = f"regex({target_used})"
        except Exception:
            return "UNKNOWN", f"bad_regex:{pat}"

    elif op == "in_list":
        if not what_list_norm:
            return "UNKNOWN", "in_list_missing_what"
        wset = {w.lower() for w in what_list_norm if w}
        matched = (val.lower() in wset)
        note = f"in_list({target_used})"

    elif op == "not_in":
        if not what_list_norm:
            return "UNKNOWN", "not_in_missing_what"
        wset = {w.lower() for w in what_list_norm if w}
        matched = (val.lower() not in wset)
        note = f"not_in({target_used})"

    elif op == "gte":
        if not what_list:
            return "UNKNOWN", "gte_missing_what"
        a = _as_float(val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            return "UNKNOWN", "gte_non_numeric"
        matched = (a >= b)
        note = f"gte({target_used})"

    elif op == "lte":
        if not what_list:
            return "UNKNOWN", "lte_missing_what"
        a = _as_float(val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            return "UNKNOWN", "lte_non_numeric"
        matched = (a <= b)
        note = f"lte({target_used})"

    elif op == "between":
        if len(what_list) < 2:
            return "UNKNOWN", "between_requires_two_values"
        a = _as_float(val)
        lo = _as_float(what_list[0])
        hi = _as_float(what_list[1])
        if a is None or lo is None or hi is None:
            return "UNKNOWN", "between_non_numeric"
        if lo > hi:
            lo, hi = hi, lo
        matched = (lo <= a <= hi)
        note = f"between({target_used})"

    # Apply include/exclude semantics
    if matched is None:
        return "UNKNOWN", "no_match_result"

    if c.ctype == "include":
        # Must match; otherwise FAILED
        return ("MET" if matched else "FAILED"), note
    else:
        # exclude: if it matches, FAILED; else MET
        return ("FAILED" if matched else "MET"), note


def _summarize_reason(outcome: str, failed: List[str], missing: List[str], met: List[str], unknown: List[str]) -> str:
    if outcome == "PASS_CLEAN":
        return "All EH criteria met."
    parts = []
    if failed:
        parts.append(f"Failed: {', '.join(failed[:6])}" + (" ..." if len(failed) > 6 else ""))
    if missing:
        parts.append(f"Missing: {', '.join(missing[:6])}" + (" ..." if len(missing) > 6 else ""))
    if unknown:
        parts.append(f"Unknown: {', '.join(unknown[:6])}" + (" ..." if len(unknown) > 6 else ""))
    if not parts:
        parts.append("Uncertain (no definitive failures).")
    return " | ".join(parts)


def run_eh_screen(
    parse: ParseReport,
    criteria_report: CriteriaLoadReport,
    cancel_event: threading.Event,
    progress_cb=None,
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], Dict[str, int]]:
    """
    Returns (full_rows, survivor_rows, counts)
    full_rows: aggregate row dicts + EH columns appended
    survivor_rows: original aggregate row dicts (no extra columns)
    """
    header = parse.header
    header_set = set(header)
    rows = parse.rows
    crits = criteria_report.criteria

    counts = {
        "OUT": 0,
        "PASS_CLEAN": 0,
        "PASS_FLAGGED": 0,
        "SKIPPED_INVALID": len(parse.skipped),
        "TOTAL_INTEGRAL": len(rows),
        "TOTAL_INPUT_RECORDS_EX_HEADER": len(parse.rows) + len(parse.skipped),
    }

    # E3: if no active EH criteria, all integral rows PASS_CLEAN
    if not crits:
        full = []
        survivors = []
        for i, r in enumerate(rows, start=1):
            if cancel_event.is_set():
                break
            fr = dict(r)
            fr["eh_outcome"] = "PASS_CLEAN"
            fr["eh_failed_ids"] = ""
            fr["eh_missing_ids"] = ""
            fr["eh_met_ids"] = ""
            fr["eh_reason_summary"] = "No active EH criteria: default PASS_CLEAN."
            full.append(fr)
            survivors.append(dict(r))
        counts["PASS_CLEAN"] = len(survivors)
        if progress_cb:
            progress_cb(1.0)
        return full, survivors, counts

    full_rows: List[Dict[str, str]] = []
    survivors: List[Dict[str, str]] = []

    n = len(rows)
    for idx, r in enumerate(rows, start=1):
        if cancel_event.is_set():
            break

        failed: List[str] = []
        missing: List[str] = []
        met: List[str] = []
        unknown: List[str] = []

        for c in crits:
            status, _note = _eval_criterion(r, header_set, c)
            if status == "FAILED":
                failed.append(c.cid)
            elif status == "MISSING":
                missing.append(c.cid)
            elif status == "MET":
                met.append(c.cid)
            else:
                unknown.append(c.cid)

        if failed:
            outcome = "OUT"
            counts["OUT"] += 1
        else:
            # All criteria met => clean; otherwise flagged
            if len(met) == len(crits) and not missing and not unknown:
                outcome = "PASS_CLEAN"
                counts["PASS_CLEAN"] += 1
            else:
                outcome = "PASS_FLAGGED"
                counts["PASS_FLAGGED"] += 1

        fr = dict(r)
        fr["eh_outcome"] = outcome
        fr["eh_failed_ids"] = ";".join(failed)
        fr["eh_missing_ids"] = ";".join(missing)
        fr["eh_met_ids"] = ";".join(met)
        # Unknown/unsupported operator ids aren’t requested as a column; include in summary only.
        fr["eh_reason_summary"] = _summarize_reason(outcome, failed, missing, met, unknown)

        full_rows.append(fr)
        if outcome != "OUT":
            survivors.append(dict(r))

        if progress_cb and (idx % PROGRESS_EVERY_N == 0 or idx == n):
            progress_cb(idx / max(1, n))

    return full_rows, survivors, counts


# ----------------------------
# Export helpers
# ----------------------------

def _export_input_errors_csv(path: str, skipped: Sequence[Tuple[int, str, str]]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["record_index_ex_header", "reason", "raw_record"])
        for rec_i, reason, raw in skipped:
            w.writerow([rec_i, reason, raw])


def _export_xlsx(path: str, full_rows: List[Dict[str, str]], survivors: List[Dict[str, str]], aggregate_header: List[str]) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError(f"openpyxl not available: {e}")

    # Sheet columns
    full_cols = list(aggregate_header) + ["eh_outcome", "eh_failed_ids", "eh_missing_ids", "eh_met_ids", "eh_reason_summary"]
    surv_cols = list(aggregate_header)

    wb = Workbook(write_only=True)

    ws1 = wb.create_sheet("EH_FULL")
    ws1.append(full_cols)
    for r in full_rows:
        ws1.append([_safe_str(r.get(c, "")) for c in full_cols])

    ws2 = wb.create_sheet("EH_SURVIVORS")
    ws2.append(surv_cols)
    for r in survivors:
        ws2.append([_safe_str(r.get(c, "")) for c in surv_cols])

    # Remove default sheet if present (write_only typically starts empty, but safe)
    try:
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 2:
            pass
    except Exception:
        pass

    wb.save(path)


# ----------------------------
# UI
# ----------------------------

class EHView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.aggregate_path: Optional[str] = None
        self.criteria_path: Optional[str] = None

        self.parse_report: Optional[ParseReport] = None
        self.criteria_report: Optional[CriteriaLoadReport] = None

        self.full_rows: List[Dict[str, str]] = []
        self.survivors: List[Dict[str, str]] = []
        self.counts: Dict[str, int] = {}

        self._worker: Optional[threading.Thread] = None
        self._cancel = threading.Event()

        self._build_ui()

    def _build_ui(self):
        # Top controls
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        btn_a = ttk.Button(top, text="Load A (aggregate CSV)…", command=self._pick_aggregate)
        btn_a.grid(row=0, column=0, padx=(0, 8), pady=2, sticky="w")

        self.lbl_a = ttk.Label(top, text="(no aggregate loaded)")
        self.lbl_a.grid(row=0, column=1, sticky="w")

        btn_c = ttk.Button(top, text="Load Criteria (harmonized CSV)…", command=self._pick_criteria)
        btn_c.grid(row=1, column=0, padx=(0, 8), pady=2, sticky="w")

        self.lbl_c = ttk.Label(top, text="(no criteria loaded)")
        self.lbl_c.grid(row=1, column=1, sticky="w")

        # Run / Cancel / Export
        actions = ttk.Frame(top)
        actions.grid(row=0, column=2, rowspan=2, padx=(10, 0), sticky="e")
        actions.columnconfigure(0, weight=1)

        self.btn_run = ttk.Button(actions, text="Run EH", command=self._run_clicked)
        self.btn_run.grid(row=0, column=0, padx=4, pady=2, sticky="e")

        self.btn_cancel = ttk.Button(actions, text="Cancel", command=self._cancel_run, state="disabled")
        self.btn_cancel.grid(row=1, column=0, padx=4, pady=2, sticky="e")

        self.btn_export = ttk.Button(actions, text="Export XLSX…", command=self._export_clicked, state="disabled")
        self.btn_export.grid(row=0, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_err = ttk.Button(actions, text="Export input_errors.csv…", command=self._export_errors_clicked, state="disabled")
        self.btn_export_err.grid(row=1, column=1, padx=4, pady=2, sticky="e")

        top.columnconfigure(1, weight=1)

        # Progress + summary
        prog = ttk.Frame(self)
        prog.pack(fill="x", padx=10, pady=(0, 8))

        self.pbar = ttk.Progressbar(prog, orient="horizontal", mode="determinate")
        self.pbar.pack(fill="x", expand=True, side="left")

        self.lbl_status = ttk.Label(prog, text="Ready.")
        self.lbl_status.pack(side="left", padx=10)

        # Warnings box
        warn_box = ttk.Labelframe(self, text="Criteria warnings / notes")
        warn_box.pack(fill="both", padx=10, pady=(0, 8))

        self.txt_warn = tk.Text(warn_box, height=6, wrap="word")
        self.txt_warn.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_warn.configure(state="disabled")

        # Reports notebook
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.tab_full = ttk.Frame(nb)
        self.tab_surv = ttk.Frame(nb)
        nb.add(self.tab_full, text="EH Full report")
        nb.add(self.tab_surv, text="EH Survivors")

        self.tree_full = self._make_tree(self.tab_full)
        self.tree_surv = self._make_tree(self.tab_surv)

        # Footer summary
        self.lbl_counts = ttk.Label(self, text="")
        self.lbl_counts.pack(fill="x", padx=10, pady=(0, 10))

    def _make_tree(self, parent) -> ttk.Treeview:
        wrap = ttk.Frame(parent)
        wrap.pack(fill="both", expand=True)

        tree = ttk.Treeview(wrap, show="headings")
        vs = ttk.Scrollbar(wrap, orient="vertical", command=tree.yview)
        hs = ttk.Scrollbar(wrap, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vs.set, xscrollcommand=hs.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vs.grid(row=0, column=1, sticky="ns")
        hs.grid(row=1, column=0, sticky="ew")

        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)
        return tree

    def _set_warnings(self, lines: Sequence[str]) -> None:
        self.txt_warn.configure(state="normal")
        self.txt_warn.delete("1.0", "end")
        if lines:
            self.txt_warn.insert("end", "\n".join(lines))
        else:
            self.txt_warn.insert("end", "(none)")
        self.txt_warn.configure(state="disabled")

    def _pick_aggregate(self):
        p = filedialog.askopenfilename(
            title="Select aggregate A CSV",
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not p:
            return
        self.aggregate_path = p
        self.lbl_a.configure(text=Path(p).name)
        self._try_load_inputs()

    def _pick_criteria(self):
        p = filedialog.askopenfilename(
            title="Select harmonized criteria CSV",
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not p:
            return
        self.criteria_path = p
        self.lbl_c.configure(text=Path(p).name)
        self._try_load_inputs()

    def _try_load_inputs(self):
        # Load aggregate
        if self.aggregate_path:
            try:
                self.parse_report = _parse_csv_tolerant(self.aggregate_path, required_id="local_id")
            except Exception as e:
                self.parse_report = None
                messagebox.showerror("Aggregate load failed", str(e))
                return

        # Load criteria
        if self.criteria_path:
            try:
                self.criteria_report = _load_criteria_eh(self.criteria_path)
            except Exception as e:
                self.criteria_report = None
                messagebox.showerror("Criteria load failed", str(e))
                return

        # Update warnings display
        warns = []
        if self.criteria_report:
            warns.extend(self.criteria_report.warnings)

        # Also warn if aggregate header lacks local_id
        if self.parse_report and "local_id" not in self.parse_report.header:
            warns.insert(0, "[aggregate] Column 'local_id' not found in header. All rows may be skipped by the 'missing local_id' rule.")

        self._set_warnings(warns)

        # Enable exporting input_errors if available
        if self.parse_report and self.parse_report.skipped:
            self.btn_export_err.configure(state="normal")
        else:
            self.btn_export_err.configure(state="disabled")

        self.lbl_status.configure(text="Ready.")
        self._refresh_counts_label()

    def _refresh_counts_label(self):
        if not self.parse_report:
            self.lbl_counts.configure(text="")
            return
        pr = self.parse_report
        msg = f"Integral rows: {len(pr.rows)} | Skipped invalid: {len(pr.skipped)}"
        if self.counts:
            msg += (
                f" | OUT: {self.counts.get('OUT',0)}"
                f" | PASS_CLEAN: {self.counts.get('PASS_CLEAN',0)}"
                f" | PASS_FLAGGED: {self.counts.get('PASS_FLAGGED',0)}"
            )
        self.lbl_counts.configure(text=msg)

    def _cancel_run(self):
        if self._worker and self._worker.is_alive():
            self._cancel.set()
            self.lbl_status.configure(text="Cancelling…")

    def _run_clicked(self):
        if self._worker and self._worker.is_alive():
            messagebox.showinfo("EH running", "A run is already in progress.")
            return
        if not self.parse_report or not self.aggregate_path:
            messagebox.showwarning("Missing input", "Load an aggregate A CSV first.")
            return
        if not self.criteria_report or not self.criteria_path:
            messagebox.showwarning("Missing input", "Load a harmonized criteria CSV first.")
            return

        self._cancel.clear()
        self.pbar["value"] = 0
        self.lbl_status.configure(text="Running EH…")
        self.btn_run.configure(state="disabled")
        self.btn_cancel.configure(state="normal")
        self.btn_export.configure(state="disabled")

        # Clear trees
        self._populate_tree(self.tree_full, [], [])
        self._populate_tree(self.tree_surv, [], [])
        self.full_rows = []
        self.survivors = []
        self.counts = {}

        def progress_cb(frac: float):
            self.after(0, lambda: self._update_progress(frac))

        def worker():
            try:
                full, surv, counts = run_eh_screen(
                    self.parse_report,
                    self.criteria_report,
                    self._cancel,
                    progress_cb=progress_cb,
                )
                self.after(0, lambda: self._finish_run(full, surv, counts))
            except Exception as e:
                self.after(0, lambda: self._run_failed(str(e)))

        self._worker = threading.Thread(target=worker, daemon=True)
        self._worker.start()

    def _update_progress(self, frac: float):
        frac = max(0.0, min(1.0, float(frac)))
        self.pbar["value"] = frac * 100.0

    def _run_failed(self, err: str):
        self.lbl_status.configure(text="Run failed.")
        self.btn_run.configure(state="normal")
        self.btn_cancel.configure(state="disabled")
        messagebox.showerror("EH run failed", err)

    def _finish_run(self, full: List[Dict[str, str]], surv: List[Dict[str, str]], counts: Dict[str, int]):
        self.full_rows = full
        self.survivors = surv
        self.counts = counts

        self.lbl_status.configure(text="Done.")
        self.btn_run.configure(state="normal")
        self.btn_cancel.configure(state="disabled")
        self.btn_export.configure(state="normal")

        # Populate UI (truncate for responsiveness)
        full_cols = list(self.parse_report.header) + ["eh_outcome", "eh_failed_ids", "eh_missing_ids", "eh_met_ids", "eh_reason_summary"]
        surv_cols = list(self.parse_report.header)

        full_ui = full[:MAX_UI_ROWS]
        surv_ui = surv[:MAX_UI_ROWS]

        self._populate_tree(self.tree_full, full_cols, full_ui)
        self._populate_tree(self.tree_surv, surv_cols, surv_ui)

        note = ""
        if len(full) > MAX_UI_ROWS or len(surv) > MAX_UI_ROWS:
            note = f" (UI truncated to {MAX_UI_ROWS} rows; export contains all rows.)"
        self.lbl_status.configure(
            text=f"Done. OUT={counts.get('OUT',0)} CLEAN={counts.get('PASS_CLEAN',0)} FLAGGED={counts.get('PASS_FLAGGED',0)}{note}"
        )
        self._refresh_counts_label()

    def _populate_tree(self, tree: ttk.Treeview, cols: List[str], rows: List[Dict[str, str]]):
        # Reset
        tree.delete(*tree.get_children())
        tree["columns"] = cols

        for c in cols:
            tree.heading(c, text=c)
            # modest widths; allow horizontal scroll
            tree.column(c, width=140, minwidth=90, stretch=False)

        for r in rows:
            vals = [_safe_str(r.get(c, "")) for c in cols]
            tree.insert("", "end", values=vals)

    def _export_clicked(self):
        if not self.parse_report:
            messagebox.showwarning("Nothing to export", "Load aggregate first.")
            return
        if not self.full_rows:
            messagebox.showwarning("Nothing to export", "Run EH first.")
            return

        default_name = f"{_now_stamp()}_EH_reports.xlsx"
        p = filedialog.asksaveasfilename(
            title="Save EH reports",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not p:
            return

        try:
            _export_xlsx(p, self.full_rows, self.survivors, self.parse_report.header)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_errors_clicked(self):
        if not self.parse_report or not self.parse_report.skipped:
            messagebox.showinfo("No input errors", "No skipped/invalid records to export.")
            return

        default_name = f"{_now_stamp()}_input_errors.csv"
        p = filedialog.asksaveasfilename(
            title="Save input_errors.csv",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv")],
        )
        if not p:
            return
        try:
            _export_input_errors_csv(p, self.parse_report.skipped)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return
        self.lbl_status.configure(text=f"Exported: {Path(p).name}")


# ----------------------------
# Hub plugin wrapper
# ----------------------------

class Plugin(BasePlugin):
    def __init__(self, app=None, meta: Optional[PluginMeta] = None):
        if meta is None:
            meta = PluginMeta(id="screen_a_eh", title=TAB_TITLE, version="2.0.0")
        super().__init__(app, meta)
        self.view: Optional[EHView] = None

    def build_tab(self, parent: ttk.Notebook) -> tk.Frame:
        frame = ttk.Frame(parent)
        self.view = EHView(frame)
        self.view.pack(fill="both", expand=True)
        return frame

    def on_close(self):
        try:
            if self.view:
                self.view.destroy()
        except Exception:
            pass
        self.view = None
