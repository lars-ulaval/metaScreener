# -*- coding: utf-8 -*-
"""
plugin.py — Screen A (EH-only) as a PRISMA Hub tab plugin (Contract v2, EH stage)

✅ EH-only, self-contained (UI + engine)
✅ Shows the EH criteria being applied (persistent left panel)
✅ Click-to-sort columns (full dataset; re-renders incrementally to keep UI responsive)
✅ Duplicate local_id: warn only (no enforcement)
✅ Row detail modal on double-click (recomputes per-criterion evaluation on demand)
✅ Criterion click filters reports to rows touched by that criterion (failed/missing/met/unknown)

Inputs
- Aggregate A CSV (may contain malformed rows)
- Harmonized criteria CSV (may contain other stages; EH-only plugin loads EH)

Outputs
1) EH_FULL report (UI + export): all aggregate columns + outcome + reason columns
2) EH_SURVIVORS report (UI + export): survivors only (PASS_CLEAN + PASS_FLAGGED),
   in the exact same schema as the original aggregate input

Exports
- One XLSX with two sheets: EH_FULL, EH_SURVIVORS
- Optional input_errors.csv for skipped records

"""

import csv
import io
import re
import threading
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from prisma_hub.plugin_api import BasePlugin, PluginMeta


TAB_TITLE = "Screen A — EH"

MAX_UI_ROWS_HINT = 2000  # only used for status text; we render all rows incrementally
RENDER_CHUNK = 300
PROGRESS_EVERY_N = 200

TRUTHY = {"1", "true", "yes", "y", "on", "enabled"}


# ----------------------------
# Normalization maps
# ----------------------------

DOC_TYPE_MAP = {
    # journal-like
    "journal": "journal",
    "journal article": "journal",
    "article": "journal",
    "research article": "journal",
    "original article": "journal",
    "journal-article": "journal",
    "peer reviewed article": "journal",
    # conference/proceedings-like
    "conference": "conference",
    "conference paper": "conference",
    "conference article": "conference",
    "proceedings": "conference",
    "proceedings article": "conference",
    "inproceedings": "conference",
    "conference proceedings": "conference",
    # misc common
    "preprint": "preprint",
    "arxiv": "preprint",
    "thesis": "thesis",
    "dissertation": "thesis",
    "book chapter": "book_chapter",
    "chapter": "book_chapter",
    "report": "report",
}

LANG_MAP = {
    "en": "en",
    "eng": "en",
    "english": "en",
    "en-us": "en",
    "en-gb": "en",
    "fr": "fr",
    "fra": "fr",
    "fre": "fr",
    "french": "fr",
    "fr-ca": "fr",
    "es": "es",
    "spa": "es",
    "spanish": "es",
}


# ----------------------------
# Data structures
# ----------------------------

@dataclass
class Criterion:
    stage: str
    cid: str
    ctype: str           # include / exclude
    scope: str
    label: str
    operator: str
    targets: List[str]   # parsed list of targets
    what_raw: str
    what_list: List[str]
    threshold: Optional[float]
    enabled: bool
    source_text: str


@dataclass
class ParseReport:
    header: List[str]
    rows: List[Dict[str, str]]           # integral rows only
    skipped: List[Tuple[int, str, str]]  # (record_index_1based_ex_header, reason, raw_record_text)


@dataclass
class CriteriaLoadReport:
    criteria: List[Criterion]
    warnings: List[str]


# ----------------------------
# Utilities
# ----------------------------

def _now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    try:
        return str(x)
    except Exception:
        return ""


def _norm_basic(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _truthy(x: Any) -> bool:
    return _safe_str(x).strip().lower() in TRUTHY


def _split_targets(target_cell: str) -> List[str]:
    t = _safe_str(target_cell)
    out: List[str] = []
    for p in t.split(","):
        p = p.strip().strip('"').strip("'")
        if p:
            out.append(p)
    return out


def _split_what_list(what_cell: str) -> List[str]:
    w = _safe_str(what_cell).replace("\n", ";").replace("\r", ";")
    out: List[str] = []
    for p in w.split(";"):
        p = p.strip().strip('"').strip("'")
        if p:
            out.append(p)
    return out


def _norm_doc_type(v: str) -> str:
    x = _norm_basic(v).lower()
    x = x.replace("_", " ").replace("-", " ")
    x = re.sub(r"\s+", " ", x).strip()
    return DOC_TYPE_MAP.get(x, x)


def _norm_lang(v: str) -> str:
    x = _norm_basic(v).lower()
    x = x.replace("_", "-")
    x = re.sub(r"\s+", "", x)
    return LANG_MAP.get(x, x)


def _norm_for_target(target: str, value: str) -> str:
    t = (target or "").strip().lower()
    v = _safe_str(value)
    if not v.strip():
        return ""
    if t == "doc_type":
        return _norm_doc_type(v)
    if t == "lang":
        return _norm_lang(v)
    return _norm_basic(v)


def _norm_what_for_target(target: str, what: str) -> str:
    return _norm_for_target(target, what)


# ----------------------------
# Robust CSV parsing (record splitter)
# ----------------------------

def _decode_bytes(b: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="ignore")


def _split_csv_records(text: str) -> List[str]:
    """
    Split CSV into record strings by scanning newlines not inside quotes.
    Preserves embedded newlines inside quoted fields.
    """
    if not text:
        return []
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    recs: List[str] = []
    buf: List[str] = []
    in_quote = False
    i = 0
    n = len(text)

    while i < n:
        ch = text[i]
        if ch == '"':
            if in_quote and (i + 1) < n and text[i + 1] == '"':
                buf.append('"')
                buf.append('"')
                i += 2
                continue
            in_quote = not in_quote
            buf.append(ch)
            i += 1
            continue

        if ch == "\n" and not in_quote:
            recs.append("".join(buf))
            buf = []
            i += 1
            continue

        buf.append(ch)
        i += 1

    if buf:
        recs.append("".join(buf))
    return recs


def _parse_csv_tolerant(path: str, required_id: str = "local_id") -> ParseReport:
    b = Path(path).read_bytes()
    text = _decode_bytes(b)

    records = _split_csv_records(text)
    if not records:
        raise ValueError("CSV is empty.")

    try:
        header = next(csv.reader(io.StringIO(records[0] + "\n")))
    except Exception as e:
        raise ValueError(f"Failed to parse CSV header: {e}")

    header = [h.strip() for h in header]
    if not header or all(not h for h in header):
        raise ValueError("CSV header is empty.")

    expected_n = len(header)
    rows: List[Dict[str, str]] = []
    skipped: List[Tuple[int, str, str]] = []

    for rec_idx, rec in enumerate(records[1:], start=1):  # 1-based, excluding header
        raw = rec
        if not raw.strip():
            skipped.append((rec_idx, "empty_record", raw))
            continue
        try:
            parsed = next(csv.reader(io.StringIO(raw + "\n")))
        except csv.Error as e:
            skipped.append((rec_idx, f"csv_error:{e}", raw))
            continue
        except Exception as e:
            skipped.append((rec_idx, f"parse_error:{e}", raw))
            continue

        if len(parsed) != expected_n:
            skipped.append((rec_idx, f"bad_column_count:{len(parsed)}!=expected:{expected_n}", raw))
            continue

        d = {header[i]: _safe_str(parsed[i]) for i in range(expected_n)}

        lid = _safe_str(d.get(required_id, "")).strip()
        if not lid:
            skipped.append((rec_idx, "missing_local_id", raw))
            continue

        rows.append(d)

    return ParseReport(header=header, rows=rows, skipped=skipped)


# ----------------------------
# Criteria loading + contradiction warnings
# ----------------------------

def _load_criteria_eh(path: str) -> CriteriaLoadReport:
    warnings: List[str] = []

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return CriteriaLoadReport(criteria=[], warnings=["Criteria header not found."])

        has_stage = any((h or "").strip().lower() == "stage" for h in (reader.fieldnames or []))

        crits: List[Criterion] = []
        row_i = 0
        for row in reader:
            row_i += 1
            stage = _safe_str(row.get("stage", "")).strip()
            if not has_stage:
                stage = "EH"
            if stage.strip().upper() != "EH":
                continue

            enabled = _truthy(row.get("enabled", "1"))
            if not enabled:
                continue

            cid = _safe_str(row.get("id", "")).strip() or f"EH_ROW_{row_i}"
            ctype = (_safe_str(row.get("type", "")).strip().lower() or "include")
            scope = _safe_str(row.get("scope", "")).strip()
            label = _safe_str(row.get("label", "")).strip()
            operator = (_safe_str(row.get("operator", "")).strip().lower() or "equals")
            targets = _split_targets(_safe_str(row.get("target", "")).strip())
            what_raw = _safe_str(row.get("what", "")).strip()
            what_list = _split_what_list(what_raw)
            source_text = _safe_str(row.get("source_text", "")).strip()

            thr = None
            thr_cell = _safe_str(row.get("threshold", "")).strip()
            if thr_cell:
                try:
                    thr = float(thr_cell)
                except Exception:
                    warnings.append(f"[criteria] Row {row_i} ({cid}): invalid threshold '{thr_cell}' -> ignored.")

            if ctype not in ("include", "exclude"):
                warnings.append(f"[criteria] {cid}: unknown type '{ctype}' -> treating as include.")
                ctype = "include"

            if not targets:
                warnings.append(f"[criteria] {cid}: missing target -> criterion will be treated as MISSING (PASS_FLAGGED).")

            crits.append(
                Criterion(
                    stage="EH",
                    cid=cid,
                    ctype=ctype,
                    scope=scope,
                    label=label,
                    operator=operator,
                    targets=targets,
                    what_raw=what_raw,
                    what_list=what_list,
                    threshold=thr,
                    enabled=True,
                    source_text=source_text,
                )
            )

    warnings.extend(_detect_contradictions_simple(crits))
    return CriteriaLoadReport(criteria=crits, warnings=warnings)


def _detect_contradictions_simple(crits: Sequence[Criterion]) -> List[str]:
    warns: List[str] = []
    bucket: Dict[str, Dict[str, set]] = {}

    for c in crits:
        if not c.targets or len(c.targets) != 1:
            continue
        tgt = c.targets[0].strip()
        op = (c.operator or "").strip().lower()
        if op not in ("equals", "in_list"):
            continue
        vals = c.what_list[:] if c.what_list else ([_safe_str(c.what_raw)] if c.what_raw else [])
        vals_norm = {_norm_what_for_target(tgt, v) for v in vals if _norm_what_for_target(tgt, v)}
        if not vals_norm:
            continue
        bucket.setdefault(tgt, {"include": set(), "exclude": set()})
        bucket[tgt][c.ctype].update(vals_norm)

    for tgt, d in bucket.items():
        inc = d.get("include", set())
        exc = d.get("exclude", set())
        overlap = sorted(inc.intersection(exc))
        if overlap:
            warns.append(
                f"[criteria] Possible contradiction on target '{tgt}': values appear in BOTH include and exclude: {', '.join(overlap[:10])}"
                + (" ..." if len(overlap) > 10 else "")
            )
    return warns


# ----------------------------
# EH evaluation
# ----------------------------

def _get_first_nonempty(row: Dict[str, str], targets: Sequence[str]) -> Tuple[str, str]:
    if not targets:
        return "", ""
    for t in targets:
        if t in row:
            v = _safe_str(row.get(t, ""))
            if v.strip():
                return t, v
    t0 = targets[0]
    return t0, _safe_str(row.get(t0, "")) if t0 in row else ""


def _eval_criterion(row: Dict[str, str], header_set: set, c: Criterion) -> str:
    """
    Fast status only: returns one of {"MET","FAILED","MISSING","UNKNOWN"}.
    """
    if not c.targets:
        return "MISSING"

    if not any(t in header_set for t in c.targets):
        return "MISSING"

    target_used, raw_val = _get_first_nonempty(row, c.targets)
    val = _norm_for_target(target_used, raw_val)
    if not val:
        return "MISSING"

    op = (c.operator or "").strip().lower()

    what_list = c.what_list[:] if c.what_list else ([_safe_str(c.what_raw)] if c.what_raw else [])
    what_list_norm = [_norm_what_for_target(target_used, w) for w in what_list if _norm_what_for_target(target_used, w)]

    if op in ("llm",):
        return "UNKNOWN"
    if op not in ("equals", "contains", "regex", "in_list", "not_in", "gte", "lte", "between"):
        return "UNKNOWN"

    def _as_float(x: str) -> Optional[float]:
        try:
            return float(x)
        except Exception:
            return None

    matched: Optional[bool] = None

    if op == "equals":
        if not what_list_norm:
            return "UNKNOWN"
        matched = (val.lower() == what_list_norm[0].lower())

    elif op == "contains":
        if not what_list_norm:
            return "UNKNOWN"
        hay = val.lower()
        matched = any((w.lower() in hay) for w in what_list_norm if w)

    elif op == "regex":
        if not what_list:
            return "UNKNOWN"
        pat = what_list[0]
        try:
            matched = bool(re.search(pat, raw_val, flags=re.IGNORECASE))
        except Exception:
            return "UNKNOWN"

    elif op == "in_list":
        if not what_list_norm:
            return "UNKNOWN"
        wset = {w.lower() for w in what_list_norm if w}
        matched = (val.lower() in wset)

    elif op == "not_in":
        if not what_list_norm:
            return "UNKNOWN"
        wset = {w.lower() for w in what_list_norm if w}
        matched = (val.lower() not in wset)

    elif op == "gte":
        if not what_list:
            return "UNKNOWN"
        a = _as_float(val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            return "UNKNOWN"
        matched = (a >= b)

    elif op == "lte":
        if not what_list:
            return "UNKNOWN"
        a = _as_float(val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            return "UNKNOWN"
        matched = (a <= b)

    elif op == "between":
        if len(what_list) < 2:
            return "UNKNOWN"
        a = _as_float(val)
        lo = _as_float(what_list[0])
        hi = _as_float(what_list[1])
        if a is None or lo is None or hi is None:
            return "UNKNOWN"
        if lo > hi:
            lo, hi = hi, lo
        matched = (lo <= a <= hi)

    if matched is None:
        return "UNKNOWN"

    if c.ctype == "include":
        return "MET" if matched else "FAILED"
    else:
        return "FAILED" if matched else "MET"


def _eval_criterion_detail(row: Dict[str, str], header_set: set, c: Criterion) -> Dict[str, str]:
    """
    Slower, for UI detail modal.
    Returns dict with keys:
      status, note, target_used, raw_value, norm_value, operator, type, targets, what
    """
    out = {
        "cid": c.cid,
        "type": c.ctype,
        "operator": c.operator,
        "targets": ",".join(c.targets),
        "what": c.what_raw,
        "status": "",
        "note": "",
        "target_used": "",
        "raw_value": "",
        "norm_value": "",
    }

    if not c.targets:
        out["status"] = "MISSING"
        out["note"] = "missing_target"
        return out

    if not any(t in header_set for t in c.targets):
        out["status"] = "MISSING"
        out["note"] = "missing_column"
        return out

    target_used, raw_val = _get_first_nonempty(row, c.targets)
    out["target_used"] = target_used
    out["raw_value"] = raw_val

    norm_val = _norm_for_target(target_used, raw_val)
    out["norm_value"] = norm_val

    if not norm_val:
        out["status"] = "MISSING"
        out["note"] = "empty_value"
        return out

    op = (c.operator or "").strip().lower()

    what_list = c.what_list[:] if c.what_list else ([_safe_str(c.what_raw)] if c.what_raw else [])
    what_list_norm = [_norm_what_for_target(target_used, w) for w in what_list if _norm_what_for_target(target_used, w)]

    if op in ("llm",):
        out["status"] = "UNKNOWN"
        out["note"] = "operator_llm_not_supported_in_EH"
        return out
    if op not in ("equals", "contains", "regex", "in_list", "not_in", "gte", "lte", "between"):
        out["status"] = "UNKNOWN"
        out["note"] = f"unknown_operator:{op}"
        return out

    def _as_float(x: str) -> Optional[float]:
        try:
            return float(x)
        except Exception:
            return None

    matched: Optional[bool] = None

    if op == "equals":
        if not what_list_norm:
            out["status"] = "UNKNOWN"
            out["note"] = "equals_missing_what"
            return out
        matched = (norm_val.lower() == what_list_norm[0].lower())
        out["note"] = "equals"

    elif op == "contains":
        if not what_list_norm:
            out["status"] = "UNKNOWN"
            out["note"] = "contains_missing_what"
            return out
        hay = norm_val.lower()
        matched = any((w.lower() in hay) for w in what_list_norm if w)
        out["note"] = "contains"

    elif op == "regex":
        if not what_list:
            out["status"] = "UNKNOWN"
            out["note"] = "regex_missing_pattern"
            return out
        pat = what_list[0]
        try:
            matched = bool(re.search(pat, raw_val, flags=re.IGNORECASE))
            out["note"] = f"regex:{pat}"
        except Exception:
            out["status"] = "UNKNOWN"
            out["note"] = f"bad_regex:{pat}"
            return out

    elif op == "in_list":
        if not what_list_norm:
            out["status"] = "UNKNOWN"
            out["note"] = "in_list_missing_what"
            return out
        wset = {w.lower() for w in what_list_norm if w}
        matched = (norm_val.lower() in wset)
        out["note"] = "in_list"

    elif op == "not_in":
        if not what_list_norm:
            out["status"] = "UNKNOWN"
            out["note"] = "not_in_missing_what"
            return out
        wset = {w.lower() for w in what_list_norm if w}
        matched = (norm_val.lower() not in wset)
        out["note"] = "not_in"

    elif op == "gte":
        if not what_list:
            out["status"] = "UNKNOWN"
            out["note"] = "gte_missing_what"
            return out
        a = _as_float(norm_val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            out["status"] = "UNKNOWN"
            out["note"] = "gte_non_numeric"
            return out
        matched = (a >= b)
        out["note"] = f"gte:{b}"

    elif op == "lte":
        if not what_list:
            out["status"] = "UNKNOWN"
            out["note"] = "lte_missing_what"
            return out
        a = _as_float(norm_val)
        b = _as_float(what_list[0])
        if a is None or b is None:
            out["status"] = "UNKNOWN"
            out["note"] = "lte_non_numeric"
            return out
        matched = (a <= b)
        out["note"] = f"lte:{b}"

    elif op == "between":
        if len(what_list) < 2:
            out["status"] = "UNKNOWN"
            out["note"] = "between_requires_two_values"
            return out
        a = _as_float(norm_val)
        lo = _as_float(what_list[0])
        hi = _as_float(what_list[1])
        if a is None or lo is None or hi is None:
            out["status"] = "UNKNOWN"
            out["note"] = "between_non_numeric"
            return out
        if lo > hi:
            lo, hi = hi, lo
        matched = (lo <= a <= hi)
        out["note"] = f"between:{lo}..{hi}"

    if matched is None:
        out["status"] = "UNKNOWN"
        out["note"] = "no_match_result"
        return out

    if c.ctype == "include":
        out["status"] = "MET" if matched else "FAILED"
    else:
        out["status"] = "FAILED" if matched else "MET"

    return out


def _summarize_reason(outcome: str, failed: List[str], missing: List[str], met: List[str], unknown: List[str]) -> str:
    if outcome == "PASS_CLEAN":
        return "All EH criteria met."
    parts = []
    if failed:
        parts.append(f"Failed: {', '.join(failed[:6])}" + (" ..." if len(failed) > 6 else ""))
    if missing:
        parts.append(f"Missing: {', '.join(missing[:6])}" + (" ..." if len(missing) > 6 else ""))
    if unknown:
        parts.append(f"Unknown: {', '.join(unknown[:6])}" + (" ..." if len(unknown) > 6 else ""))
    if not parts:
        parts.append("Uncertain (no definitive failures).")
    return " | ".join(parts)


def run_eh_screen(
    parse: ParseReport,
    criteria_report: CriteriaLoadReport,
    cancel_event: threading.Event,
    progress_cb: Optional[Callable[[float], None]] = None,
) -> Tuple[List[Dict[str, str]], List[Dict[str, str]], Dict[str, int], Dict[str, Dict[str, int]], List[Dict[str, List[str]]]]:
    """
    Returns:
      full_rows: aggregate row dicts + EH columns appended
      survivor_rows: original aggregate row dicts (no extra columns)
      counts: summary counts
      crit_impacts: {cid: {"failed":n, "missing":n, "met":n, "unknown":n}}
      row_eval_lists: list aligned with full_rows, each is {"failed":[...], "missing":[...], "met":[...], "unknown":[...]}
                     (used for criterion filtering and detail views)
    """
    header = parse.header
    header_set = set(header)
    rows = parse.rows
    crits = criteria_report.criteria

    counts = {
        "OUT": 0,
        "PASS_CLEAN": 0,
        "PASS_FLAGGED": 0,
        "SKIPPED_INVALID": len(parse.skipped),
        "TOTAL_INTEGRAL": len(rows),
        "TOTAL_INPUT_RECORDS_EX_HEADER": len(parse.rows) + len(parse.skipped),
    }

    crit_impacts: Dict[str, Dict[str, int]] = {c.cid: {"failed": 0, "missing": 0, "met": 0, "unknown": 0} for c in crits}

    full_rows: List[Dict[str, str]] = []
    survivors: List[Dict[str, str]] = []
    row_eval_lists: List[Dict[str, List[str]]] = []

    # No active EH criteria => default PASS_CLEAN for all integral rows
    if not crits:
        for i, r in enumerate(rows, start=1):
            if cancel_event.is_set():
                break
            fr = dict(r)
            fr["eh_outcome"] = "PASS_CLEAN"
            fr["eh_failed_ids"] = ""
            fr["eh_missing_ids"] = ""
            fr["eh_met_ids"] = ""
            fr["eh_reason_summary"] = "No active EH criteria: default PASS_CLEAN."
            full_rows.append(fr)
            survivors.append(dict(r))
            row_eval_lists.append({"failed": [], "missing": [], "met": [], "unknown": []})

        counts["PASS_CLEAN"] = len(survivors)
        if progress_cb:
            progress_cb(1.0)
        return full_rows, survivors, counts, crit_impacts, row_eval_lists

    n = len(rows)
    for idx, r in enumerate(rows, start=1):
        if cancel_event.is_set():
            break

        failed: List[str] = []
        missing: List[str] = []
        met: List[str] = []
        unknown: List[str] = []

        for c in crits:
            status = _eval_criterion(r, header_set, c)
            if status == "FAILED":
                failed.append(c.cid)
                crit_impacts[c.cid]["failed"] += 1
            elif status == "MISSING":
                missing.append(c.cid)
                crit_impacts[c.cid]["missing"] += 1
            elif status == "MET":
                met.append(c.cid)
                crit_impacts[c.cid]["met"] += 1
            else:
                unknown.append(c.cid)
                crit_impacts[c.cid]["unknown"] += 1

        if failed:
            outcome = "OUT"
            counts["OUT"] += 1
        else:
            if len(met) == len(crits) and not missing and not unknown:
                outcome = "PASS_CLEAN"
                counts["PASS_CLEAN"] += 1
            else:
                outcome = "PASS_FLAGGED"
                counts["PASS_FLAGGED"] += 1

        fr = dict(r)
        fr["eh_outcome"] = outcome
        fr["eh_failed_ids"] = ";".join(failed)
        fr["eh_missing_ids"] = ";".join(missing)
        fr["eh_met_ids"] = ";".join(met)
        fr["eh_reason_summary"] = _summarize_reason(outcome, failed, missing, met, unknown)

        full_rows.append(fr)
        row_eval_lists.append({"failed": failed, "missing": missing, "met": met, "unknown": unknown})

        if outcome != "OUT":
            survivors.append(dict(r))

        if progress_cb and (idx % PROGRESS_EVERY_N == 0 or idx == n):
            progress_cb(idx / max(1, n))

    return full_rows, survivors, counts, crit_impacts, row_eval_lists


# ----------------------------
# Export helpers
# ----------------------------

def _export_input_errors_csv(path: str, skipped: Sequence[Tuple[int, str, str]]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["record_index_ex_header", "reason", "raw_record"])
        for rec_i, reason, raw in skipped:
            w.writerow([rec_i, reason, raw])


def _export_xlsx(path: str, full_rows: List[Dict[str, str]], survivors: List[Dict[str, str]], aggregate_header: List[str]) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError(f"openpyxl not available: {e}")

    full_cols = list(aggregate_header) + ["eh_outcome", "eh_failed_ids", "eh_missing_ids", "eh_met_ids", "eh_reason_summary"]
    surv_cols = list(aggregate_header)

    wb = Workbook(write_only=True)
    ws1 = wb.create_sheet("EH_FULL")
    ws1.append(full_cols)
    for r in full_rows:
        ws1.append([_safe_str(r.get(c, "")) for c in full_cols])

    ws2 = wb.create_sheet("EH_SURVIVORS")
    ws2.append(surv_cols)
    for r in survivors:
        ws2.append([_safe_str(r.get(c, "")) for c in surv_cols])

    wb.save(path)


# ----------------------------
# UI Components
# ----------------------------

class DataTable(ttk.Frame):
    """
    Treeview wrapper with:
    - column setup
    - click-to-sort (requests sort callback)
    - incremental rendering to keep UI responsive
    - double-click callback to open details
    """
    def __init__(self, parent, on_sort: Callable[[str], None], on_row_activate: Optional[Callable[[Dict[str, str]], None]] = None):
        super().__init__(parent)
        self.on_sort = on_sort
        self.on_row_activate = on_row_activate

        self.tree = ttk.Treeview(self, show="headings", selectmode="browse")
        self.vs = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.hs = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.vs.set, xscrollcommand=self.hs.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        self.vs.grid(row=0, column=1, sticky="ns")
        self.hs.grid(row=1, column=0, sticky="ew")

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.columns: List[str] = []
        self._render_token = 0
        self._iid_to_row: Dict[str, Dict[str, str]] = {}

        # Double-click row => activate
        self.tree.bind("<Double-1>", self._on_double_click)

    def set_columns(self, cols: List[str]):
        self.columns = cols
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols
        for c in cols:
            self.tree.heading(c, text=c, command=lambda col=c: self.on_sort(col))
            self.tree.column(c, width=140, minwidth=90, stretch=False)

    def clear(self):
        self._render_token += 1
        self._iid_to_row.clear()
        self.tree.delete(*self.tree.get_children())

    def render_rows_incremental(self, rows: List[Dict[str, str]]):
        """
        Render all rows, but insert them in chunks via after().
        Cancels any prior render in progress.
        """
        self.clear()
        token = self._render_token

        # Generate stable iids for this render: r0, r1, ...
        def _insert_chunk(start: int):
            if token != self._render_token:
                return  # cancelled
            end = min(start + RENDER_CHUNK, len(rows))
            for i in range(start, end):
                r = rows[i]
                iid = f"r{i}"
                self._iid_to_row[iid] = r
                vals = [_safe_str(r.get(c, "")) for c in self.columns]
                self.tree.insert("", "end", iid=iid, values=vals)
            if end < len(rows):
                self.after(1, lambda: _insert_chunk(end))

        self.after(0, lambda: _insert_chunk(0))

    def _on_double_click(self, _evt):
        if not self.on_row_activate:
            return
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        r = self._iid_to_row.get(iid)
        if r:
            self.on_row_activate(r)


# ----------------------------
# Main View
# ----------------------------

class EHView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.aggregate_path: Optional[str] = None
        self.criteria_path: Optional[str] = None

        self.parse_report: Optional[ParseReport] = None
        self.criteria_report: Optional[CriteriaLoadReport] = None

        self.full_rows: List[Dict[str, str]] = []
        self.survivors: List[Dict[str, str]] = []
        self.counts: Dict[str, int] = {}

        self.crit_impacts: Dict[str, Dict[str, int]] = {}
        self.row_evals_full: List[Dict[str, List[str]]] = []  # aligned with full_rows

        # criterion filter state
        self.active_criterion_id: Optional[str] = None  # if set, filter reports by this criterion id

        self._worker: Optional[threading.Thread] = None
        self._cancel = threading.Event()

        # sorting state per table
        self.sort_full: Tuple[Optional[str], bool] = (None, True)      # (col, asc)
        self.sort_surv: Tuple[Optional[str], bool] = (None, True)
        self.sort_crit: Tuple[Optional[str], bool] = (None, True)

        self._build_ui()

    # -------- UI layout --------

    def _build_ui(self):
        # Top controls
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        btn_a = ttk.Button(top, text="Load A (aggregate CSV)…", command=self._pick_aggregate)
        btn_a.grid(row=0, column=0, padx=(0, 8), pady=2, sticky="w")

        self.lbl_a = ttk.Label(top, text="(no aggregate loaded)")
        self.lbl_a.grid(row=0, column=1, sticky="w")

        btn_c = ttk.Button(top, text="Load Criteria (harmonized CSV)…", command=self._pick_criteria)
        btn_c.grid(row=1, column=0, padx=(0, 8), pady=2, sticky="w")

        self.lbl_c = ttk.Label(top, text="(no criteria loaded)")
        self.lbl_c.grid(row=1, column=1, sticky="w")

        actions = ttk.Frame(top)
        actions.grid(row=0, column=2, rowspan=2, padx=(10, 0), sticky="e")

        self.btn_run = ttk.Button(actions, text="Run EH", command=self._run_clicked)
        self.btn_run.grid(row=0, column=0, padx=4, pady=2, sticky="e")

        self.btn_cancel = ttk.Button(actions, text="Cancel", command=self._cancel_run, state="disabled")
        self.btn_cancel.grid(row=1, column=0, padx=4, pady=2, sticky="e")

        self.btn_export = ttk.Button(actions, text="Export XLSX…", command=self._export_clicked, state="disabled")
        self.btn_export.grid(row=0, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_err = ttk.Button(actions, text="Export input_errors.csv…", command=self._export_errors_clicked, state="disabled")
        self.btn_export_err.grid(row=1, column=1, padx=4, pady=2, sticky="e")

        top.columnconfigure(1, weight=1)

        # Progress + status
        prog = ttk.Frame(self)
        prog.pack(fill="x", padx=10, pady=(0, 8))

        self.pbar = ttk.Progressbar(prog, orient="horizontal", mode="determinate")
        self.pbar.pack(fill="x", expand=True, side="left")

        self.lbl_status = ttk.Label(prog, text="Ready.")
        self.lbl_status.pack(side="left", padx=10)

        # Main paned area: left criteria, right reports
        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        left = ttk.Frame(paned)
        right = ttk.Frame(paned)
        paned.add(left, weight=1)
        paned.add(right, weight=4)

        # LEFT: Criteria + warnings + criterion filter status
        crit_box = ttk.Labelframe(left, text="EH Criteria (read-only)")
        crit_box.pack(fill="both", expand=True)

        self.criteria_table = DataTable(
            crit_box,
            on_sort=self._sort_criteria_table,
            on_row_activate=self._on_criterion_activated,
        )
        self.criteria_table.pack(fill="both", expand=True, padx=6, pady=6)

        # Criterion filter control
        cf = ttk.Frame(left)
        cf.pack(fill="x", pady=(6, 0))
        self.lbl_crit_filter = ttk.Label(cf, text="Criterion filter: (none)")
        self.lbl_crit_filter.pack(side="left")
        self.btn_clear_filter = ttk.Button(cf, text="Clear filter", command=self._clear_criterion_filter, state="disabled")
        self.btn_clear_filter.pack(side="right")

        warn_box = ttk.Labelframe(left, text="Notes / warnings")
        warn_box.pack(fill="both", expand=False, pady=(6, 0))

        self.txt_warn = tk.Text(warn_box, height=7, wrap="word")
        self.txt_warn.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_warn.configure(state="disabled")

        # RIGHT: Reports notebook
        nb = ttk.Notebook(right)
        nb.pack(fill="both", expand=True)

        tab_full = ttk.Frame(nb)
        tab_surv = ttk.Frame(nb)
        nb.add(tab_full, text="EH Full report")
        nb.add(tab_surv, text="EH Survivors")

        self.full_table = DataTable(tab_full, on_sort=self._sort_full_table, on_row_activate=self._open_row_detail_modal)
        self.full_table.pack(fill="both", expand=True, padx=6, pady=6)

        self.surv_table = DataTable(tab_surv, on_sort=self._sort_surv_table, on_row_activate=self._open_row_detail_modal)
        self.surv_table.pack(fill="both", expand=True, padx=6, pady=6)

        # Footer summary
        self.lbl_counts = ttk.Label(self, text="")
        self.lbl_counts.pack(fill="x", padx=10, pady=(0, 10))

    # -------- helpers --------

    def _set_warnings(self, lines: Sequence[str]) -> None:
        self.txt_warn.configure(state="normal")
        self.txt_warn.delete("1.0", "end")
        self.txt_warn.insert("end", "\n".join(lines) if lines else "(none)")
        self.txt_warn.configure(state="disabled")

    def _refresh_counts_label(self):
        if not self.parse_report:
            self.lbl_counts.configure(text="")
            return
        pr = self.parse_report
        msg = f"Integral rows: {len(pr.rows)} | Skipped invalid: {len(pr.skipped)}"
        if self.counts:
            msg += (
                f" | OUT: {self.counts.get('OUT',0)}"
                f" | PASS_CLEAN: {self.counts.get('PASS_CLEAN',0)}"
                f" | PASS_FLAGGED: {self.counts.get('PASS_FLAGGED',0)}"
            )
        self.lbl_counts.configure(text=msg)

    def _detect_duplicate_local_ids(self, rows: List[Dict[str, str]]) -> Tuple[int, List[str]]:
        seen = set()
        dups = []
        for r in rows:
            lid = _safe_str(r.get("local_id", "")).strip()
            if not lid:
                continue
            if lid in seen:
                dups.append(lid)
            else:
                seen.add(lid)
        # Unique duplicates for warning
        uniq = []
        s2 = set()
        for x in dups:
            if x not in s2:
                uniq.append(x)
                s2.add(x)
        return len(uniq), uniq[:10]

    # -------- file pickers & preload --------

    def _pick_aggregate(self):
        p = filedialog.askopenfilename(
            title="Select aggregate A CSV",
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not p:
            return
        self.aggregate_path = p
        self.lbl_a.configure(text=Path(p).name)
        self._try_load_inputs()

    def _pick_criteria(self):
        p = filedialog.askopenfilename(
            title="Select harmonized criteria CSV",
            filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
        )
        if not p:
            return
        self.criteria_path = p
        self.lbl_c.configure(text=Path(p).name)
        self._try_load_inputs()

    def _try_load_inputs(self):
        warns: List[str] = []

        # Load aggregate
        if self.aggregate_path:
            try:
                self.parse_report = _parse_csv_tolerant(self.aggregate_path, required_id="local_id")
            except Exception as e:
                self.parse_report = None
                messagebox.showerror("Aggregate load failed", str(e))
                return

            if self.parse_report and "local_id" not in self.parse_report.header:
                warns.append("[aggregate] Column 'local_id' not found in header. The 'missing local_id' rule may skip rows unexpectedly.")

            # Duplicate local_id warning (warn only)
            if self.parse_report:
                ndup, examples = self._detect_duplicate_local_ids(self.parse_report.rows)
                if ndup > 0:
                    warns.append(f"[aggregate] Duplicate local_id detected: {ndup} unique duplicates (examples: {', '.join(examples)}). Policy: WARN ONLY.")

        # Load criteria
        if self.criteria_path:
            try:
                self.criteria_report = _load_criteria_eh(self.criteria_path)
            except Exception as e:
                self.criteria_report = None
                messagebox.showerror("Criteria load failed", str(e))
                return

            if self.criteria_report:
                warns.extend(self.criteria_report.warnings)

        # Update warnings display
        self._set_warnings(warns)

        # Enable exporting input_errors if available
        if self.parse_report and self.parse_report.skipped:
            self.btn_export_err.configure(state="normal")
        else:
            self.btn_export_err.configure(state="disabled")

        # Populate criteria table if available
        self._refresh_criteria_table(pre_run=True)

        self.lbl_status.configure(text="Ready.")
        self._refresh_counts_label()

    # -------- Criteria table --------

    def _refresh_criteria_table(self, pre_run: bool):
        """
        Show criteria + status. After run, include impact columns.
        """
        crits = self.criteria_report.criteria if self.criteria_report else []

        cols = ["id", "type", "targets", "operator", "what", "status", "notes"]
        if not pre_run:
            cols += ["n_failed", "n_missing", "n_met", "n_unknown"]

        rows: List[Dict[str, str]] = []
        header_set = set(self.parse_report.header) if self.parse_report else set()

        for c in crits:
            # status/notes: missing target? missing column? unknown operator?
            status = "OK"
            notes = ""

            if not c.targets:
                status = "WARNING"
                notes = "missing target -> treated as MISSING (PASS_FLAGGED)"
            elif header_set and not any(t in header_set for t in c.targets):
                status = "WARNING"
                notes = f"missing column(s): {', '.join(c.targets)} -> treated as MISSING (PASS_FLAGGED)"

            op = (c.operator or "").strip().lower()
            if op in ("llm",):
                status = "WARNING"
                notes = (notes + " | " if notes else "") + "operator 'llm' not supported in EH -> UNKNOWN (PASS_FLAGGED)"
            elif op not in ("equals", "contains", "regex", "in_list", "not_in", "gte", "lte", "between"):
                status = "WARNING"
                notes = (notes + " | " if notes else "") + f"unknown operator '{op}' -> UNKNOWN (PASS_FLAGGED)"

            d: Dict[str, str] = {
                "id": c.cid,
                "type": c.ctype,
                "targets": ",".join(c.targets),
                "operator": c.operator,
                "what": c.what_raw,
                "status": status,
                "notes": notes,
            }

            if not pre_run:
                imp = self.crit_impacts.get(c.cid, {"failed": 0, "missing": 0, "met": 0, "unknown": 0})
                d["n_failed"] = str(imp.get("failed", 0))
                d["n_missing"] = str(imp.get("missing", 0))
                d["n_met"] = str(imp.get("met", 0))
                d["n_unknown"] = str(imp.get("unknown", 0))

            rows.append(d)

        self.criteria_table.set_columns(cols)
        # Apply sorting if set
        col, asc = self.sort_crit
        if col:
            rows = self._sorted_rows(rows, col, asc)
        self.criteria_table.render_rows_incremental(rows)

    def _on_criterion_activated(self, row: Dict[str, str]):
        """
        Double-click criterion row => filter reports to rows touched by that criterion.
        """
        cid = _safe_str(row.get("id", "")).strip()
        if not cid:
            return
        self.active_criterion_id = cid
        self.lbl_crit_filter.configure(text=f"Criterion filter: {cid}")
        self.btn_clear_filter.configure(state="normal")
        self._refresh_reports_view()

    def _clear_criterion_filter(self):
        self.active_criterion_id = None
        self.lbl_crit_filter.configure(text="Criterion filter: (none)")
        self.btn_clear_filter.configure(state="disabled")
        self._refresh_reports_view()

    # -------- Sorting --------

    def _sorted_rows(self, rows: List[Dict[str, str]], col: str, asc: bool) -> List[Dict[str, str]]:
        # infer numeric vs text from sample
        sample = []
        for r in rows[:200]:
            v = _safe_str(r.get(col, "")).strip()
            if v != "":
                sample.append(v)
        num_hits = 0
        for v in sample[:50]:
            try:
                float(v)
                num_hits += 1
            except Exception:
                pass
        is_numeric = (len(sample) > 0 and num_hits >= max(3, int(0.7 * min(len(sample[:50]), 50))))

        def key_num(r: Dict[str, str]):
            v = _safe_str(r.get(col, "")).strip()
            try:
                return (0, float(v))
            except Exception:
                # put non-numeric at end
                return (1, float("inf"))

        def key_txt(r: Dict[str, str]):
            v = _safe_str(r.get(col, "")).strip().lower()
            return v

        key_fn = key_num if is_numeric else key_txt
        return sorted(rows, key=key_fn, reverse=not asc)

    def _toggle_sort(self, current: Tuple[Optional[str], bool], col: str) -> Tuple[str, bool]:
        cur_col, cur_asc = current
        if cur_col == col:
            return (col, not cur_asc)
        return (col, True)

    def _sort_criteria_table(self, col: str):
        self.sort_crit = self._toggle_sort(self.sort_crit, col)
        self._refresh_criteria_table(pre_run=(not bool(self.full_rows)))

    def _sort_full_table(self, col: str):
        self.sort_full = self._toggle_sort(self.sort_full, col)
        self._refresh_reports_view()

    def _sort_surv_table(self, col: str):
        self.sort_surv = self._toggle_sort(self.sort_surv, col)
        self._refresh_reports_view()

    # -------- Run / Cancel --------

    def _cancel_run(self):
        if self._worker and self._worker.is_alive():
            self._cancel.set()
            self.lbl_status.configure(text="Cancelling…")

    def _run_clicked(self):
        if self._worker and self._worker.is_alive():
            messagebox.showinfo("EH running", "A run is already in progress.")
            return
        if not self.parse_report or not self.aggregate_path:
            messagebox.showwarning("Missing input", "Load an aggregate A CSV first.")
            return
        if not self.criteria_report or not self.criteria_path:
            messagebox.showwarning("Missing input", "Load a harmonized criteria CSV first.")
            return

        self._cancel.clear()
        self.pbar["value"] = 0
        self.lbl_status.configure(text="Running EH…")
        self.btn_run.configure(state="disabled")
        self.btn_cancel.configure(state="normal")
        self.btn_export.configure(state="disabled")

        self.full_rows = []
        self.survivors = []
        self.counts = {}
        self.crit_impacts = {}
        self.row_evals_full = []

        self.full_table.clear()
        self.surv_table.clear()

        # keep criterion filter but it will apply after run
        def progress_cb(frac: float):
            self.after(0, lambda: self._update_progress(frac))

        def worker():
            try:
                full, surv, counts, impacts, row_evals = run_eh_screen(
                    self.parse_report,
                    self.criteria_report,
                    self._cancel,
                    progress_cb=progress_cb,
                )
                self.after(0, lambda: self._finish_run(full, surv, counts, impacts, row_evals))
            except Exception as e:
                self.after(0, lambda: self._run_failed(str(e)))

        self._worker = threading.Thread(target=worker, daemon=True)
        self._worker.start()

    def _update_progress(self, frac: float):
        frac = max(0.0, min(1.0, float(frac)))
        self.pbar["value"] = frac * 100.0

    def _run_failed(self, err: str):
        self.lbl_status.configure(text="Run failed.")
        self.btn_run.configure(state="normal")
        self.btn_cancel.configure(state="disabled")
        messagebox.showerror("EH run failed", err)

    def _finish_run(
        self,
        full: List[Dict[str, str]],
        surv: List[Dict[str, str]],
        counts: Dict[str, int],
        impacts: Dict[str, Dict[str, int]],
        row_evals: List[Dict[str, List[str]]],
    ):
        self.full_rows = full
        self.survivors = surv
        self.counts = counts
        self.crit_impacts = impacts
        self.row_evals_full = row_evals

        self.btn_run.configure(state="normal")
        self.btn_cancel.configure(state="disabled")
        self.btn_export.configure(state="normal")

        # Refresh criteria with impact columns
        self._refresh_criteria_table(pre_run=False)

        # Refresh reports (applies criterion filter + sorting)
        self._refresh_reports_view()

        # Status message
        note = ""
        if len(full) > MAX_UI_ROWS_HINT or len(surv) > MAX_UI_ROWS_HINT:
            note = " (rendering incrementally; export contains all rows.)"

        self.lbl_status.configure(
            text=f"Done. OUT={counts.get('OUT',0)} CLEAN={counts.get('PASS_CLEAN',0)} FLAGGED={counts.get('PASS_FLAGGED',0)}{note}"
        )
        self._refresh_counts_label()

    # -------- Reports rendering (filter + sort + incremental) --------

    def _refresh_reports_view(self):
        if not self.parse_report:
            return

        full_cols = list(self.parse_report.header) + ["eh_outcome", "eh_failed_ids", "eh_missing_ids", "eh_met_ids", "eh_reason_summary"]
        surv_cols = list(self.parse_report.header)

        # Build filtered full view
        full_view = self.full_rows
        if self.active_criterion_id:
            cid = self.active_criterion_id
            # Use row evals for accurate membership
            filtered = []
            for i, r in enumerate(self.full_rows):
                ev = self.row_evals_full[i] if i < len(self.row_evals_full) else {"failed": [], "missing": [], "met": [], "unknown": []}
                if (cid in ev["failed"]) or (cid in ev["missing"]) or (cid in ev["met"]) or (cid in ev["unknown"]):
                    filtered.append(r)
            full_view = filtered

        # survivors view can’t use row_evals directly (we didn’t store aligned list),
        # but we can filter survivors by evaluating against full_rows local_id membership
        surv_view = self.survivors
        if self.active_criterion_id and self.full_rows:
            cid = self.active_criterion_id
            # Determine local_ids in full_rows that are "touched" by cid AND survive
            touched_survivor_ids = set()
            for i, r in enumerate(self.full_rows):
                if r.get("eh_outcome") == "OUT":
                    continue
                ev = self.row_evals_full[i]
                if (cid in ev["failed"]) or (cid in ev["missing"]) or (cid in ev["met"]) or (cid in ev["unknown"]):
                    touched_survivor_ids.add(_safe_str(r.get("local_id", "")).strip())
            surv_view = [r for r in self.survivors if _safe_str(r.get("local_id", "")).strip() in touched_survivor_ids]

        # Apply sorting
        col, asc = self.sort_full
        if col and full_view:
            full_view = self._sorted_rows(full_view, col, asc)

        col2, asc2 = self.sort_surv
        if col2 and surv_view:
            surv_view = self._sorted_rows(surv_view, col2, asc2)

        # Render
        self.full_table.set_columns(full_cols)
        self.full_table.render_rows_incremental(full_view)

        self.surv_table.set_columns(surv_cols)
        self.surv_table.render_rows_incremental(surv_view)

    # -------- Row detail modal --------

    def _open_row_detail_modal(self, row: Dict[str, str]):
        if not self.criteria_report or not self.parse_report:
            return

        win = tk.Toplevel(self)
        win.title("EH Row details")
        win.geometry("900x600")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        lid = _safe_str(row.get("local_id", "")).strip()
        title = _safe_str(row.get("title", "")).strip() or _safe_str(row.get("Title", "")).strip()

        top = ttk.Frame(win)
        top.pack(fill="x", padx=10, pady=10)

        ttk.Label(top, text=f"local_id: {lid}").pack(anchor="w")
        if title:
            ttk.Label(top, text=f"title: {title[:250]}").pack(anchor="w")

        outcome = _safe_str(row.get("eh_outcome", "")).strip()
        if outcome:
            ttk.Label(top, text=f"EH outcome: {outcome}").pack(anchor="w")
            rs = _safe_str(row.get("eh_reason_summary", "")).strip()
            if rs:
                ttk.Label(top, text=f"summary: {rs}").pack(anchor="w")

        # Criteria evaluations
        box = ttk.Labelframe(win, text="Per-criterion evaluation (recomputed)")
        box.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ["cid", "type", "operator", "targets", "what", "status", "target_used", "norm_value", "note"]
        table = DataTable(box, on_sort=lambda _c: None, on_row_activate=None)
        table.pack(fill="both", expand=True, padx=6, pady=6)
        table.set_columns(cols)

        header_set = set(self.parse_report.header)
        crits = self.criteria_report.criteria

        detail_rows: List[Dict[str, str]] = []
        for c in crits:
            det = _eval_criterion_detail(row, header_set, c)
            detail_rows.append(det)

        table.render_rows_incremental(detail_rows)

        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(btns, text="Close", command=win.destroy).pack(side="right")

        # ESC to close
        win.bind("<Escape>", lambda _e: win.destroy())

    # -------- Export --------

    def _export_clicked(self):
        if not self.parse_report:
            messagebox.showwarning("Nothing to export", "Load aggregate first.")
            return
        if not self.full_rows:
            messagebox.showwarning("Nothing to export", "Run EH first.")
            return

        default_name = f"{_now_stamp()}_EH_reports.xlsx"
        p = filedialog.asksaveasfilename(
            title="Save EH reports",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not p:
            return

        try:
            _export_xlsx(p, self.full_rows, self.survivors, self.parse_report.header)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_errors_clicked(self):
        if not self.parse_report or not self.parse_report.skipped:
            messagebox.showinfo("No input errors", "No skipped/invalid records to export.")
            return

        default_name = f"{_now_stamp()}_input_errors.csv"
        p = filedialog.asksaveasfilename(
            title="Save input_errors.csv",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv")],
        )
        if not p:
            return
        try:
            _export_input_errors_csv(p, self.parse_report.skipped)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return
        self.lbl_status.configure(text=f"Exported: {Path(p).name}")


# ----------------------------
# Hub plugin wrapper
# ----------------------------

class Plugin(BasePlugin):
    def __init__(self, app=None, meta: Optional[PluginMeta] = None):
        if meta is None:
            meta = PluginMeta(id="screen_a_eh", title=TAB_TITLE, version="2.1.0")
        super().__init__(app, meta)
        self.view: Optional[EHView] = None

    def build_tab(self, parent: ttk.Notebook) -> tk.Frame:
        frame = ttk.Frame(parent)
        self.view = EHView(frame)
        self.view.pack(fill="both", expand=True)
        return frame

    def on_close(self):
        try:
            if self.view:
                self.view.destroy()
        except Exception:
            pass
        self.view = None
