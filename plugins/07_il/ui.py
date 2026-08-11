
# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""ui.py - Plugin 07 IL (Screen A): Tkinter View and standalone shell.

After Conv 6 / Commit 3, this module owns the IL stage's UI surface:
  - DataTable: Treeview wrapper with click-to-sort and incremental rendering.
  - _now_stamp, _export_il_xlsx: small UI-side helpers for export actions.
  - Final-report aggregation helpers (_find_bundle_member,
    _load_csv_rows_from_zip, _load_master_rows, _stage_prefix,
    _extract_contract_stage_row, _compute_final_outcome,
    _build_final_report_xlsx_bytes): IL-specific because IL is the
    terminal stage that produces the cross-stage final report.
  - ILView: the Tk Notebook tab widget for the IL stage.
  - StandaloneILPlugin: the standalone-app shell wrapper around ILView.

Engine logic stays in plugins/07_il/plugin.py for now; ui.py imports
the symbols it needs from there. Subsequent Conv 6 commits will finish
moving engine code into plugins/_common/ and shrink
plugins/07_il/plugin.py to a thin shim.

This file is GUI-only; its behaviour is verified manually after the
final thin-shim commit lands.
"""

import csv
import io
import json
import os
import re
import threading
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

# F-02: one shared export gate, so the rule that a cancelled run
# cannot be exported lives in one place for all four stages.
from plugins._common.bundle import (
    NOT_SCREENED,
    _export_block_reason,
    _export_confirm_reason,
    _run_summary_counts_text,
    _write_llm_stage_bundle,
)

from plugins._common.stage_state import (
    Outcome,
    Readiness,
    batch_size_tooltip,
    control_states,
    llm_readiness,
    parse_numeric_settings,
    run_outcome,
    tk_state,
)
from plugins._common.widgets import RecheckButton, Tooltip

from plugins._common.settings import (
    apply_stage_fields,
    load_settings,
    resolve_stage,
)
from plugins._common import provider_detect as _pd
from plugins._common.provider_detect import (
    last_known,
    model_choices,
    refresh as pd_refresh,
)
from plugins._common.exporters import _export_input_errors_csv_from_dicts
from plugins._common.input_errors import (
    from_dict_skipped,
    merge_input_errors_csv,
    read_input_errors,
)

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Engine + dataclasses + IL-specific constants stay in plugin.py for now
# (Conv 6 / Commit 3). These imports break the circular dependency
# cleanly because plugin.py defines all of these BEFORE it does
# `from .ui import ILView, ...` near the bottom.
from .plugin import (
    BundleInfo,
    CONTRACT_STAGE_SHEET_COLS,
    DEFAULT_BATCH_SIZE,
    DEFAULT_MODEL,
    DEFAULT_TRUNC_CHARS,
    DEFAULT_USE_CACHE,
    FINAL_REPORT_NAME,
    FINAL_REPORT_REL,
    IL_CACHE_REL,
    RENDER_CHUNK,
    REPORTS_DIR_REL,
    _csv_read,
    _decode_bytes,
    _dump_cache_to_jsonl,
    _has_openai_key,
    _load_bundle,
    _load_cache_from_jsonl,
    _read_zip_bytes,
    _safe_str,
    _write_csv,
    run_il_screen,
)

# ------------------------------ UI helpers ------------------------------------

class DataTable(ttk.Frame):
    """
    Treeview wrapper with:
    - column setup
    - click-to-sort (optional)
    - incremental rendering to keep UI responsive
    - optional double-click callback

    Backward compatible:
      - DataTable(parent, on_sort_callable, on_row_activate=...)
      - DataTable(parent, ["col1","col2",...])  # legacy usage
    """
    def __init__(
        self,
        parent,
        on_sort=None,
        on_row_activate: Optional[Callable[[Dict[str, Any]], None]] = None
    ):
        super().__init__(parent)

        # Legacy signature: second arg is actually a list of columns
        cols_seed: Optional[List[str]] = None
        if on_sort is None:
            self.on_sort = lambda _c: None
        elif callable(on_sort):
            self.on_sort = on_sort
        else:
            # treat as columns list/tuple
            if isinstance(on_sort, (list, tuple)):
                cols_seed = list(on_sort)
            self.on_sort = lambda _c: None

        self.on_row_activate = on_row_activate

        self.tree = ttk.Treeview(self, show="headings", selectmode="browse")
        self.vs = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.hs = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.vs.set, xscrollcommand=self.hs.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        self.vs.grid(row=0, column=1, sticky="ns")
        self.hs.grid(row=1, column=0, sticky="ew")

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.columns: List[str] = []
        self._render_token = 0
        self._iid_to_row: Dict[str, Dict[str, Any]] = {}

        self.tree.bind("<Double-1>", self._on_double_click)

        if cols_seed:
            self.set_columns(cols_seed)

    def set_columns(self, cols: List[str]):
        self.columns = cols
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols
        for c in cols:
            self.tree.heading(c, text=c, command=lambda col=c: self.on_sort(col))
            self.tree.column(c, width=140, minwidth=90, stretch=False)

    def clear(self):
        self._render_token += 1
        self._iid_to_row.clear()
        self.tree.delete(*self.tree.get_children())

    def render_rows_incremental(self, rows: List[Dict[str, Any]]):
        self.clear()
        token = self._render_token

        def _insert_chunk(start: int):
            if token != self._render_token:
                return
            end = min(start + RENDER_CHUNK, len(rows))
            for i in range(start, end):
                r = rows[i]
                iid = f"r{i}"
                self._iid_to_row[iid] = r
                vals = [_safe_str(r.get(c, "")) for c in self.columns]
                self.tree.insert("", "end", iid=iid, values=vals)
            if end < len(rows):
                self.after(1, lambda: _insert_chunk(end))

        self.after(0, lambda: _insert_chunk(0))

    # ---- legacy helpers used by Plugin(ttk.Frame) ----
    def set_rows(self, rows: List[Dict[str, Any]]):
        if not rows:
            self.clear()
            return
        if not self.columns:
            self.set_columns(list(rows[0].keys()))
        self.render_rows_incremental(rows)

    def get_selected_row(self) -> Optional[Dict[str, Any]]:
        sel = self.tree.selection()
        if not sel:
            return None
        return self._iid_to_row.get(sel[0])

    def _on_double_click(self, _evt):
        if not self.on_row_activate:
            return
        r = self.get_selected_row()
        if r is not None:
            self.on_row_activate(r)

def _now_stamp() -> str:
    # Same convention as EH/IH
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _export_il_xlsx(
    path: str,
    full_rows: List[Dict[str, Any]],
    survivors: List[Dict[str, Any]],
    base_header: List[str],
) -> None:
    """
    Writes a 2-sheet XLSX:
      - IL_FULL
      - IL_SURVIVORS
    Uses openpyxl (same approach as your EH/IH plugins).
    """
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required for XLSX export. Install it (pip install openpyxl)."
        ) from e

    def cell(v: Any) -> Any:
        if v is None:
            return ""
        if isinstance(v, (dict, list)):
            return json.dumps(v, ensure_ascii=False)
        return v

    wb = Workbook(write_only=True)

    # ---- headers (stable + tolerant) ----
    base = list(base_header or [])
    if "local_id" in base:
        base = ["local_id"] + [h for h in base if h != "local_id"]
    else:
        base = ["local_id"] + base

    el_cols = [
        "il_outcome",
        "il_failed_ids",
        "il_missing_ids",
        "il_met_ids",
        "il_uncertain_ids",
        "il_reason_summary",
        "il_evidence_json",
    ]

    # Full sheet header = base + IL cols + any extras present in data
    full_header = base + [c for c in el_cols if c not in base]
    if full_rows:
        extras = [k for k in full_rows[0].keys() if k not in full_header]
        full_header += extras

    # Survivors sheet header = base + any extras present in survivors (rare)
    surv_header = list(base)
    if survivors:
        extras2 = [k for k in survivors[0].keys() if k not in surv_header]
        surv_header += extras2

    # ---- sheet: IL_FULL ----
    ws1 = wb.create_sheet(title="IL_FULL")
    ws1.append(full_header)
    for r in full_rows:
        ws1.append([cell(r.get(h, "")) for h in full_header])

    # ---- sheet: IL_SURVIVORS ----
    ws2 = wb.create_sheet(title="IL_SURVIVORS")
    ws2.append(surv_header)
    for r in survivors:
        ws2.append([cell(r.get(h, "")) for h in surv_header])

    # Remove default sheet if present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)



def _find_bundle_member(zf: zipfile.ZipFile, root: str, rel: str) -> Optional[str]:
    """Return the member name for root+rel if present, else try suffix match case-insensitive."""
    exact = root + rel
    if exact in zf.namelist():
        return exact
    suffix = ("/" + rel).lower()
    for name in zf.namelist():
        if name.lower().endswith(suffix):
            return name
    return None


def _load_csv_rows_from_zip(zf: zipfile.ZipFile, member: str) -> Tuple[List[str], List[Dict[str, str]]]:
    b = _read_zip_bytes(zf, member)
    txt = _decode_bytes(b)
    return _csv_read(txt)


def _load_master_rows(zf: zipfile.ZipFile, root: str) -> Tuple[List[str], List[Dict[str, str]]]:
    """
    Best-effort: try to load the *original* A (all items) if present; else fall back to current.
    """
    candidates = [
        "data/A.csv",
        "data/original.csv",
        "data/aggregate.csv",
        "data/current.csv",
    ]
    for rel in candidates:
        mem = _find_bundle_member(zf, root, rel)
        if mem:
            try:
                return _load_csv_rows_from_zip(zf, mem)
            except Exception:
                continue
    return ([], [])


def _stage_prefix(stage: str) -> str:
    return stage.lower() + "_"


def _extract_contract_stage_row(stage: str, r: Dict[str, Any]) -> Dict[str, Any]:
    """
    Map a stage FULL row (plugin convention: {prefix}outcome, {prefix}*_ids, {prefix}reason_summary, {prefix}evidence_json)
    to contract v2 standardized columns.
    """
    pref = _stage_prefix(stage)
    a_id = _safe_str(r.get("a_id") or r.get("local_id") or r.get("id") or r.get("doi") or "")
    stage_outcome = _safe_str(r.get(pref + "outcome") or r.get("stage_outcome") or "")
    failed = _safe_str(r.get(pref + "failed_ids") or "")
    missing = _safe_str(r.get(pref + "missing_ids") or "")
    uncertain = _safe_str(r.get(pref + "uncertain_ids") or "")
    met = _safe_str(r.get(pref + "met_ids") or "")
    matched_evidence = _safe_str(r.get(pref + "evidence_json") or r.get(pref + "matched_evidence") or "")
    reason = _safe_str(r.get(pref + "reason_summary") or r.get("stage_reason_summary") or "")
    passed_to_next = "true" if stage_outcome and stage_outcome != "OUT" else "false"
    # F-69: this key set IS the stage-sheet header (CONTRACT_STAGE_SHEET_COLS
    # in plugin.py); the two must stay identical or the sheets go blank
    # again. ``history`` is gone — it was written as "" unconditionally.
    return {
        "a_id": a_id,
        "stage": stage,
        "stage_outcome": stage_outcome,
        "passed_to_next": passed_to_next,
        "failed_criteria_ids": failed,
        "missing_criteria_ids": missing,
        "uncertain_criteria_ids": uncertain,
        "met_criteria_ids": met,
        "matched_evidence": matched_evidence,
        "stage_reason_summary": reason,
    }


def _compute_final_outcome(outcomes: Dict[str, str]) -> str:
    """
    Contract v2: FINAL outcome:
      - OUT if any stage outcome == OUT
      - PASS_CLEAN if ALL non-empty stage outcomes are PASS_CLEAN AND IL is PASS_CLEAN
      - REVIEW otherwise
    """
    for st in ("EH", "IH", "EL", "IL"):
        if outcomes.get(st) == "OUT":
            return "OUT"
    il_out = outcomes.get("IL", "")
    if il_out == "PASS_CLEAN" and all((outcomes.get(st, "") in {"", "PASS_CLEAN"} for st in ("EH", "IH", "EL", "IL"))):
        return "PASS_CLEAN"
    return "REVIEW"


def _build_final_report_xlsx_bytes(
    zf_in: zipfile.ZipFile,
    root: str,
    il_full_rows: List[Dict[str, Any]],
) -> bytes:
    """
    Build the contract v2 5-sheet Excel workbook:
      EH / IH / EL / IL / FINAL
    Uses stage FULL CSVs from the input bundle when available.
    """
    # Load prior stage FULL rows (best-effort)
    stage_full: Dict[str, List[Dict[str, str]]] = {}
    for st in ("EH", "IH", "EL"):
        mem = _find_bundle_member(zf_in, root, f"{REPORTS_DIR_REL}/{st}_FULL.csv")
        if mem:
            try:
                _, rows = _load_csv_rows_from_zip(zf_in, mem)
                stage_full[st] = rows
            except Exception:
                stage_full[st] = []
        else:
            stage_full[st] = []

    stage_full["IL"] = [{k: (_safe_str(v) if not isinstance(v, str) else v) for k, v in r.items()} for r in il_full_rows]

    # Master list (all items) for FINAL
    master_header, master_rows = _load_master_rows(zf_in, root)
    # Build meta lookup by a_id
    meta_by_id: Dict[str, Dict[str, str]] = {}
    for r in master_rows:
        a_id = _safe_str(r.get("a_id") or r.get("local_id") or r.get("id") or "")
        if a_id:
            meta_by_id[a_id] = r

    # Outcomes + reasons per stage
    stage_by_id: Dict[str, Dict[str, Dict[str, Any]]] = {st: {} for st in ("EH", "IH", "EL", "IL")}
    for st in ("EH", "IH", "EL", "IL"):
        for r in stage_full.get(st, []):
            a_id = _safe_str(r.get("a_id") or r.get("local_id") or r.get("id") or "")
            if not a_id:
                continue
            pref = _stage_prefix(st)
            stage_by_id[st][a_id] = {
                "outcome": _safe_str(r.get(pref + "outcome") or r.get("stage_outcome") or ""),
                "reason": _safe_str(r.get(pref + "reason_summary") or r.get("stage_reason_summary") or ""),
            }

    # Universe of ids
    all_ids = set(meta_by_id.keys())
    for st in ("EH", "IH", "EL", "IL"):
        all_ids.update(stage_by_id[st].keys())
    all_ids = sorted(all_ids, key=lambda x: (len(x), x))

    # Meta columns for FINAL (keep master header order, minus id columns)
    meta_cols = [c for c in master_header if c not in {"a_id", "local_id"}] if master_header else []

    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError("openpyxl is required to build the final report workbook.") from e

    wb = Workbook(write_only=True)

    # Stage sheets
    for st in ("EH", "IH", "EL", "IL"):
        ws = wb.create_sheet(title=st)
        ws.append(CONTRACT_STAGE_SHEET_COLS)
        for r in stage_full.get(st, []):
            row_obj = _extract_contract_stage_row(st, r)
            ws.append([row_obj.get(c, "") for c in CONTRACT_STAGE_SHEET_COLS])

    # FINAL sheet
    ws = wb.create_sheet(title="FINAL")
    # F-69: no ``history`` column — it was emitted as "" for every record
    # and had no data source. Do not fabricate one.
    final_cols = ["a_id"] + meta_cols + [
        "outcome_EH", "reason_EH",
        "outcome_IH", "reason_IH",
        "outcome_EL", "reason_EL",
        "outcome_IL", "reason_IL",
        "final_outcome", "final_reason_summary",
    ]
    ws.append(final_cols)

    for a_id in all_ids:
        meta = meta_by_id.get(a_id, {})
        outcomes = {st: stage_by_id[st].get(a_id, {}).get("outcome", "") for st in ("EH", "IH", "EL", "IL")}
        reasons = {st: stage_by_id[st].get(a_id, {}).get("reason", "") for st in ("EH", "IH", "EL", "IL")}
        final_out = _compute_final_outcome(outcomes)
        if final_out == "OUT":
            # first stage that OUTed (in pipeline order)
            out_stage = next((st for st in ("EH", "IH", "EL", "IL") if outcomes.get(st) == "OUT"), "")
            final_reason = f"OUT at {out_stage}" if out_stage else "OUT"
        elif final_out == "PASS_CLEAN":
            final_reason = "PASS_CLEAN across all stages"
        else:
            final_reason = "Needs review after staged screening"

        row = [a_id]
        row.extend([meta.get(c, "") for c in meta_cols])
        row.extend([outcomes["EH"], reasons["EH"], outcomes["IH"], reasons["IH"], outcomes["EL"], reasons["EL"], outcomes["IL"], reasons["IL"]])
        row.extend([final_out, final_reason])
        ws.append(row)

    # Ensure workbook has proper first sheet order (write_only creates default Sheet)
    # openpyxl write_only starts with a default sheet named "Sheet" sometimes; remove it if empty.
    # (In write_only mode, default may still exist.)
    try:
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 5:
            wb.remove(wb["Sheet"])
    except Exception:
        pass

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------------------
# Main View
# ----------------------------

class ILView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.bundle_zip_path: Optional[str] = None
        self.bundle: Optional[BundleInfo] = None

        self.full_rows: List[Dict[str, Any]] = []
        self.cancelled: bool = False   # F-02: last run stopped mid-corpus
        # Wave 8: what the last run learned, and what it failed to learn.
        # Derived by the engine from the evidence its decisions were made
        # from; empty until a run completes. A wholly failed run and a
        # wholly uncertain run are identical in every other field this
        # view holds, which is what made "IL done." the only thing the
        # interface could say for either.
        self.llm_report: Dict[str, Any] = {}
        # Wave 8 part 2: how the last run is classified, as data.
        self.outcome: Optional[Outcome] = None
        self.not_screened: bool = False  # F-34: last run had no criteria
        self.survivors: List[Dict[str, str]] = []
        self.counts: Dict[str, int] = {}

        self.crit_impacts: Dict[str, Dict[str, int]] = {}
        self.row_eval_lists: List[Dict[str, List[str]]] = []

        self.cache_map: Dict[str, Dict[str, Any]] = {}

        self.active_criterion_id: Optional[str] = None

        self._worker: Optional[threading.Thread] = None
        self._cancel = threading.Event()

        self.sort_full: Tuple[Optional[str], bool] = (None, True)
        self.sort_surv: Tuple[Optional[str], bool] = (None, True)
        self.sort_crit: Tuple[Optional[str], bool] = (None, True)

        # Settings. Wave 11 session C: seeded from this stage's effective
        # configuration rather than from the module defaults, so a choice
        # made once survives a launch. An unconfigured install resolves to
        # exactly the old defaults, which is why nothing about a fresh
        # install changes.
        seed = self._stored_config()
        self.var_model = tk.StringVar(value=seed.model or DEFAULT_MODEL)
        self.var_endpoint = tk.StringVar(value=seed.endpoint)
        self.var_temp = tk.DoubleVar(value=0.0)
        self.var_batch = tk.StringVar(
            value=str(seed.batch_size if seed.batch_size is not None
                      else DEFAULT_BATCH_SIZE))
        self.var_trunc = tk.StringVar(value=str(DEFAULT_TRUNC_CHARS))
        self.var_use_cache = tk.BooleanVar(value=DEFAULT_USE_CACHE)

        self._build_ui()
        self._refresh_discovery()

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        btn_b = ttk.Button(top, text="Load ScreenA bundle ZIP…", command=self._pick_bundle)
        btn_b.grid(row=0, column=0, padx=(0, 8), pady=2, sticky="w")

        self.lbl_bundle = ttk.Label(top, text="(no bundle loaded)")
        self.lbl_bundle.grid(row=0, column=1, sticky="w")

        self.lbl_bundle_meta = ttk.Label(top, text="")
        self.lbl_bundle_meta.grid(row=1, column=1, sticky="w")

        actions = ttk.Frame(top)
        actions.grid(row=0, column=2, rowspan=2, padx=(10, 0), sticky="e")

        self.btn_run = ttk.Button(actions, text="Run IL", command=self._run_clicked, state="disabled")
        self.btn_run.grid(row=0, column=0, padx=4, pady=2, sticky="e")

        self.btn_cancel = ttk.Button(actions, text="Cancel", command=self._cancel_run, state="disabled")
        self.btn_cancel.grid(row=1, column=0, padx=4, pady=2, sticky="e")

        self.btn_export = ttk.Button(actions, text="Export XLSX…", command=self._export_clicked, state="disabled")
        self.btn_export.grid(row=0, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_err = ttk.Button(actions, text="Export input_errors.csv…", command=self._export_errors_clicked, state="disabled")
        self.btn_export_err.grid(row=1, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_bundle = ttk.Button(actions, text="Export next bundle ZIP…", command=self._export_bundle_clicked, state="disabled")
        self.btn_export_bundle.grid(row=0, column=2, padx=4, pady=2, sticky="e")

        self.btn_export_final = ttk.Button(actions, text="Export ScreenA_Report.xlsx…", command=self._export_final_clicked, state="disabled")
        self.btn_export_final.grid(row=0, column=3, padx=4, pady=2, sticky="e")

        # API key indicator (IL requires OpenAI)
        self.lbl_key = ttk.Label(actions, text="")
        self.lbl_key.grid(row=1, column=3, padx=4, pady=2, sticky="e")

        # F-149. Beside the indicator that says "Unreachable", because
        # that is where the user is looking when they need it.
        self.btn_recheck = RecheckButton(
            actions, prepare=self._reprobe_job,
            on_done=self.on_provider_changed)
        self.btn_recheck.grid(row=1, column=2, padx=4, pady=2, sticky="e")
        Tooltip(self.btn_recheck,
                "Ask this stage's endpoint again. The provider is checked "
                "once at startup, so a server started, restarted or woken "
                "since then is not noticed until you ask.")
        self._refresh_readiness_label()
        # The indicator is only honest if it keeps up with the fields it
        # reports on; without this it would go stale the moment the user
        # typed a model name. Session C adds the endpoint for the same
        # reason: readiness now asks where the stage points, so an
        # endpoint edited without the label following would leave the one
        # provider-adjacent indicator in the tab answering about the
        # previous server.
        for _var in (self.var_model, self.var_endpoint):
            _var.trace_add(
                "write", lambda *_a: self._refresh_readiness_label())

        top.columnconfigure(1, weight=1)

        prog = ttk.Frame(self)
        prog.pack(fill="x", padx=10, pady=(0, 8))

        self.pbar = ttk.Progressbar(prog, orient="horizontal", mode="determinate")
        self.pbar.pack(fill="x", expand=True, side="left")

        self.lbl_status = ttk.Label(prog, text="Ready.")
        self.lbl_status.pack(side="left", padx=10)

        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        left = ttk.Frame(paned)
        right = ttk.Frame(paned)
        paned.add(left, weight=1)
        paned.add(right, weight=4)

        crit_box = ttk.Labelframe(left, text="IL Criteria (read-only)")
        crit_box.pack(fill="both", expand=True)

        self.criteria_table = DataTable(
            crit_box,
            on_sort=self._sort_criteria_table,
            on_row_activate=self._on_criterion_activated,
        )
        self.criteria_table.pack(fill="both", expand=True, padx=6, pady=6)

        cf = ttk.Frame(left)
        cf.pack(fill="x", pady=(6, 0))
        self.lbl_crit_filter = ttk.Label(cf, text="Criterion filter: (none)")
        self.lbl_crit_filter.pack(side="left")
        self.btn_clear_filter = ttk.Button(cf, text="Clear filter", command=self._clear_criterion_filter, state="disabled")
        self.btn_clear_filter.pack(side="right")

        settings = ttk.Labelframe(left, text="IL Settings")
        settings.pack(fill="x", pady=(6, 0))

        # The model control is an EDITABLE combobox, and that is the
        # point rather than a detail. llama.cpp ignores the model field
        # entirely, so a readonly dropdown fed from discovery would
        # rebuild the enumeration problem this project keeps removing: a
        # user whose server will not enumerate could not name a model at
        # all. Nothing anywhere may set this widget's state - asserted by
        # AST in tests/test_model_discovery.py. Discovery is an aid,
        # never a gate.
        ttk.Label(settings, text="Model").grid(row=0, column=0, sticky="w", padx=6, pady=2)
        self.cmb_model = ttk.Combobox(settings, textvariable=self.var_model,
                                      width=22, values=())
        self.cmb_model.grid(row=0, column=1, sticky="ew", padx=6, pady=2)

        self.lbl_models = ttk.Label(settings, text="", wraplength=260,
                                    justify="left", foreground="#555")
        self.lbl_models.grid(row=1, column=0, columnspan=2, sticky="w",
                             padx=6, pady=(0, 4))

        # F-91's per-stage surface. What it shows is the endpoint this
        # stage would actually use - resolved, not the raw stored key,
        # because those differ whenever a default is doing the work and a
        # box showing "" while the run goes to the vendor is the whole of
        # F-91.
        ttk.Label(settings, text="Endpoint").grid(row=2, column=0, sticky="w", padx=6, pady=2)
        self.ent_endpoint = ttk.Entry(settings, textvariable=self.var_endpoint,
                                      width=24)
        self.ent_endpoint.grid(row=2, column=1, sticky="ew", padx=6, pady=2)

        self.lbl_endpoint_src = ttk.Label(settings, text="", wraplength=260,
                                          justify="left", foreground="#555")
        self.lbl_endpoint_src.grid(row=3, column=0, columnspan=2, sticky="w",
                                   padx=6, pady=(0, 4))

        for widget in (self.ent_endpoint, self.cmb_model):
            widget.bind("<Return>", lambda _e: self._stage_fields_edited())
            widget.bind("<FocusOut>", lambda _e: self._stage_fields_edited())
        self.cmb_model.bind("<<ComboboxSelected>>",
                            lambda _e: self._stage_fields_edited())

        ttk.Label(settings, text="Temperature").grid(row=4, column=0, sticky="w", padx=6, pady=2)
        ttk.Spinbox(settings, textvariable=self.var_temp, from_=0.0, to=2.0, increment=0.1, format="%.2f", width=10).grid(row=4, column=1, sticky="w", padx=6, pady=2)
        ttk.Label(settings, text="(0.0 = deterministic; non-zero invalidates cache)").grid(row=5, column=1, sticky="w", padx=6, pady=(0, 4))

        ttk.Label(settings, text="Batch size").grid(row=6, column=0, sticky="w", padx=6, pady=2)
        self.ent_batch = ttk.Entry(settings, textvariable=self.var_batch,
                                   width=10)
        self.ent_batch.grid(row=6, column=1, sticky="w", padx=6, pady=2)
        self.ent_batch.bind("<Return>", lambda _e: self._stage_fields_edited())
        self.ent_batch.bind("<FocusOut>",
                            lambda _e: self._stage_fields_edited())
        # D6. The reason lives beside the number. The wording is decided
        # in stage_state -- pure, and therefore assertable -- because what
        # it must NOT say is the load-bearing part: batch size is a
        # quality knob, not a safety one, and changing it does not
        # invalidate a cache.
        _cfg = self._stored_config()
        self.tip_batch = Tooltip(
            self.ent_batch,
            batch_size_tooltip(_cfg.provider, _cfg.endpoint))

        ttk.Label(settings, text="Trunc chars").grid(row=7, column=0, sticky="w", padx=6, pady=2)
        ttk.Entry(settings, textvariable=self.var_trunc, width=10).grid(row=7, column=1, sticky="w", padx=6, pady=2)

        ttk.Checkbutton(settings, text="Use cache (bundle cache/IL_cache.jsonl)", variable=self.var_use_cache).grid(row=8, column=0, columnspan=2, sticky="w", padx=6, pady=2)

        settings.columnconfigure(1, weight=1)

        warn_box = ttk.Labelframe(left, text="Notes / warnings")
        warn_box.pack(fill="both", expand=False, pady=(6, 0))

        self.txt_warn = tk.Text(warn_box, height=8, wrap="word")
        self.txt_warn.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_warn.configure(state="disabled")

        nb = ttk.Notebook(right)
        nb.pack(fill="both", expand=True)

        tab_full = ttk.Frame(nb)
        tab_surv = ttk.Frame(nb)
        tab_log = ttk.Frame(nb)
        nb.add(tab_full, text="IL Full report")
        nb.add(tab_surv, text="IL Survivors")
        nb.add(tab_log, text="Log")

        self.full_table = DataTable(tab_full, on_sort=self._sort_full_table, on_row_activate=self._open_row_detail_modal)
        self.full_table.pack(fill="both", expand=True, padx=6, pady=6)

        self.surv_table = DataTable(tab_surv, on_sort=self._sort_surv_table, on_row_activate=self._open_row_detail_modal)
        self.surv_table.pack(fill="both", expand=True, padx=6, pady=6)

        self.txt_log = tk.Text(tab_log, wrap="word")
        self.txt_log.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_log.configure(state="disabled")

        self.lbl_counts = ttk.Label(self, text="")
        self.lbl_counts.pack(fill="x", padx=10, pady=(0, 10))

    # -------- helpers --------

    STAGE = "IL"

    def _stored_config(self):
        """This stage's effective configuration, from the store.

        Guarded for the same reason ``_readiness`` is: ``load_settings``
        raises on a settings file that exists and cannot be parsed, this
        runs during construction, and ``main.py::resolve_plugin_entrypoint``
        swallows every exception into a ``print()`` -- so an unguarded read
        makes the tab silently absent. Degrading to defaults blocks the
        run for a stated reason instead of deleting the stage.
        """
        try:
            cfg = load_settings()
        except Exception:
            cfg = {}
        return resolve_stage(cfg, self.STAGE,
                             env_endpoint=os.environ.get("OPENAI_BASE_URL", ""),
                             env_api_key=os.environ.get("OPENAI_API_KEY", ""))

    def _reprobe_job(self):
        """Build the probe to run off the GUI thread (F-149).

        Every widget read happens **here**, on the GUI thread, and the
        returned closure touches nothing but the network. `RecheckButton`
        calls this at click and the closure on its worker.

        `provider_detect.refresh` files its answer under the endpoint it
        asked about, so this re-probes only the server this tab points
        at. That is deliberate: the application-wide
        `_refresh_provider_status` begins with `pd.forget()`, which drops
        every probe and would send the other tabs to NOT_CHECKED -- label
        "Checking...", Run disabled -- because this one asked a question.
        """
        cfg = self._stored_config()
        endpoint = self._effective_endpoint()
        api_key, provider = cfg.api_key, cfg.provider
        return lambda: pd_refresh(endpoint, api_key=api_key,
                                  provider=provider)

    def _effective_endpoint(self) -> str:
        """Where this stage would send a request right now.

        The live widget value when there is one, so the indicator answers
        for what the user is looking at rather than for what was last
        saved; the resolved configuration otherwise.
        """
        return self.var_endpoint.get().strip() or self._stored_config().endpoint

    def _stage_fields_edited(self) -> None:
        """Persist what the tab now says, so the engine cannot disagree.

        The engine resolves from the store. If the widgets held values the
        store had never seen, a run would go somewhere other than what the
        tab shows -- which is the defect family this whole wave is about,
        arriving through the controls added to fix it. So the write
        happens when the field is left, and again before a run starts.

        A field equal to what the stage would resolve to *without* an
        override stores nothing; see ``settings::stage_overrides_for``.
        """
        try:
            apply_stage_fields(
                self.STAGE,
                # The values these widgets show when the store says
                # nothing. Without them every seeded default looks like a
                # deliberate edit and gets pinned as a permanent override
                # — which defeated D6 and stopped the user's own model
                # choice reaching any stage they had opened. Session C's
                # review, measured.
                fallbacks={"model": DEFAULT_MODEL, "batch_size": DEFAULT_BATCH_SIZE},
                model=self.var_model.get(),
                endpoint=self.var_endpoint.get(),
                batch_size=self.var_batch.get(),
            )
        except Exception as e:
            # Never fatal: a store that cannot be written must not stop
            # the user working. The values still apply to this session
            # through the widgets the run reads.
            self._log(f"[IL] settings not saved: {e}\n")
        self._refresh_readiness_label()
        self._refresh_endpoint_source()
        self._set_controls_running(False)

    def _refresh_endpoint_source(self) -> None:
        """Say which source the endpoint came from.

        F-119's lesson: ``endpoint=https://api.openai.com/v1`` alone does
        not distinguish *I chose the public API* from *my configuration
        was not read*.
        """
        self.lbl_endpoint_src.configure(
            text=f"Source: {self._stored_config().endpoint_source}")

    def _refresh_discovery(self) -> None:
        """Fill the model combobox from the last detection.

        Never probes: this runs inside Tk callbacks, and detection is a
        network call. It reads the answer the application deposited, the
        same way ``_readiness`` does -- so the two cannot describe
        different servers.
        """
        # Session C's review: the probe is keyed by endpoint now, and this
        # asks for THIS stage's. It used to read one global answer, so a
        # stage with an endpoint override was offered the models of a
        # server it would never call — and reported "Ready to run" while
        # its own endpoint had nothing listening.
        choices = model_choices(last_known(self._effective_endpoint()))
        try:
            self.cmb_model.configure(values=list(choices.values))
        except Exception:
            pass
        self.lbl_models.configure(text=choices.note)
        self._refresh_endpoint_source()

    def on_provider_changed(self) -> None:
        """The application settled or re-probed a provider.

        ``main.py`` has notified this hook since session B and no plugin
        implemented it, so the call was a silent no-op and the tabs kept
        reporting the previous answer until something else happened to
        refresh them. Re-seeding here is what makes a provider change
        visible in the stage the user is not looking at.
        """
        seed = self._stored_config()
        if seed.endpoint:
            self.var_endpoint.set(seed.endpoint)
        if seed.model:
            self.var_model.set(seed.model)
        if seed.batch_size is not None:
            self.var_batch.set(str(seed.batch_size))
        # D6: the reason beside the number is provider-specific, so a
        # provider change that left the old wording in place would be a
        # tooltip explaining a number the box no longer shows.
        try:
            self.tip_batch.set_text(
                batch_size_tooltip(seed.provider, seed.endpoint))
        except Exception:
            pass
        self._refresh_discovery()
        self._refresh_readiness_label()
        self._set_controls_running(False)

    def _readiness(self) -> Readiness:
        """F-111. One place decides whether this stage may run, and
        everything that asks — the indicator, the Run button, the Run
        handler — asks it. Three places deciding separately is how the
        gate came to be dropped (F-118) and how an empty model field
        came to start a run (F-93)."""
        # F-117. The provider decides whether a key is needed at all, so
        # readiness reads the persisted configuration rather than probing
        # the environment for one variable. A local server authenticates
        # nothing, and asking its user for a credential was the defect.
        # Review of this session: `load_settings` raises on a settings
        # file that exists and cannot be parsed, and this runs inside
        # `_build_ui` — where `main.py::resolve_plugin_entrypoint`
        # swallows every exception into a `print()`, so a JSON typo made
        # the EL and IL tabs silently absent, with nowhere for the
        # message to go in a windowed onefile build. Readiness degrades
        # to "unconfigured" instead, which blocks the run for a stated
        # reason rather than deleting the stage.
        cfg = self._stored_config()
        # Wave 11 session B. `probe` is the last detection the app
        # deposited, never a call made here: this runs inside Tk
        # callbacks, and detection is a network operation. `None` means
        # "not checked yet", which readiness reports as its own state
        # rather than guessing either way.
        #
        # Session C passes the ENDPOINT this stage would use - the live
        # widget value, not the stored one, so the indicator answers for
        # what the user is looking at. `key_required_for` then decides the
        # key question on the resolved pair, which is what stops a stage
        # pointed at the paid vendor from being waved through because its
        # provider field says local.
        return llm_readiness(stage=self.STAGE,
                             has_bundle=bool(self.bundle_zip_path),
                             provider=cfg.provider,
                             api_key=cfg.api_key,
                             model=self.var_model.get(),
                             endpoint=self._effective_endpoint(),
                             probe=last_known(self._effective_endpoint()))

    def _refresh_readiness_label(self):
        """The widget used to read `OPENAI_API_KEY ✓` and nothing else:
        the startup modal cannot be passed without a non-empty key and
        nothing ever clears it, so the one provider-adjacent indicator
        in the application carried zero bits (F-111). It now names
        whichever thing is missing, and `Ready to run` when none is."""
        self.lbl_key.configure(text=self._readiness().label)

    def _log(self, msg: str) -> None:
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", msg)
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")

    def _set_warnings(self, lines: Sequence[str]) -> None:
        self.txt_warn.configure(state="normal")
        self.txt_warn.delete("1.0", "end")
        self.txt_warn.insert("end", "\n".join(lines) if lines else "(none)")
        self.txt_warn.configure(state="disabled")

    def _refresh_counts_label(self):
        if not self.bundle:
            self.lbl_counts.configure(text="")
            return
        pr = self.bundle.parse
        msg = f"Integral rows: {len(pr.rows)} | Skipped invalid: {len(pr.skipped)}"
        if self.counts:
            msg += (
                " | " + _run_summary_counts_text(
                    self.counts, stage="IL",
                    total_rows=len(pr.rows))
            )
        self.lbl_counts.configure(text=msg)

    def _set_controls_running(self, running: bool) -> None:
        # Wave 8 part 2: the five expressions that used to live here are
        # now plugins/_common/stage_state.py::control_states, so they can
        # be asserted on. This method reads state and writes widgets; it
        # decides nothing.
        st = control_states(
            running=running,
            readiness=self._readiness(),
            has_rows=bool(self.full_rows),
            has_input_errors=bool(self.bundle and self.bundle.parse.skipped),
        )
        self.btn_cancel.configure(state=tk_state(st.cancel))
        self.btn_run.configure(state=tk_state(st.run))
        self.btn_export.configure(state=tk_state(st.export))
        self.btn_export_err.configure(state=tk_state(st.export_errors))
        self.btn_export_bundle.configure(state=tk_state(st.export_bundle))
        # IL's sixth button follows the same rule as `export` and reads
        # it directly; a ControlStates field always equal to another
        # field would be the F-69 shape in miniature.
        self.btn_export_final.configure(state=tk_state(st.export))

    def _sorted_rows(self, rows: List[Dict[str, Any]], col: str, asc: bool) -> List[Dict[str, Any]]:
        def _key(r: Dict[str, Any]):
            v = r.get(col, "")
            s = _safe_str(v)
            # numeric sort if it looks like a number
            try:
                if s.strip() == "":
                    return (1, 0.0, s)
                return (0, float(s), s)
            except Exception:
                return (0, 0.0, s.lower())
        return sorted(rows, key=_key, reverse=(not asc))

    # -------- bundle load --------

    def _pick_bundle(self):
        p = filedialog.askopenfilename(
            title="Select ScreenA bundle ZIP",
            filetypes=[("ZIP", "*.zip"), ("All files", "*.*")],
        )
        if not p:
            return
        self.bundle_zip_path = p
        self.lbl_bundle.configure(text=Path(p).name)
        self._load_bundle_inputs()

    def _load_bundle_inputs(self):
        if not self.bundle_zip_path:
            return

        warns: List[str] = []
        self.full_rows = []
        self.cancelled = False
        self.not_screened = False
        self.llm_report = {}
        self.outcome = None
        self.survivors = []
        self.counts = {}
        self.crit_impacts = {}
        self.row_eval_lists = []
        self.active_criterion_id = None
        self.lbl_crit_filter.configure(text="Criterion filter: (none)")
        self.btn_clear_filter.configure(state="disabled")

        self._refresh_readiness_label()

        try:
            self.bundle = _load_bundle(self.bundle_zip_path)
        except Exception as e:
            self.bundle = None
            messagebox.showerror("Bundle load failed", str(e))
            return

        # load cache if present
        self.cache_map = {}
        try:
            with zipfile.ZipFile(self.bundle_zip_path, "r") as zf:
                cache_member = self.bundle.root + IL_CACHE_REL
                if cache_member in zf.namelist():
                    self.cache_map = _load_cache_from_jsonl(_decode_bytes(_read_zip_bytes(zf, cache_member)))
        except Exception:
            self.cache_map = {}

        m = self.bundle.manifest
        schema = _safe_str(m.get("bundle_schema", m.get("schema", ""))).strip() or "unknown"
        created_at = _safe_str(m.get("created_at", "")).strip()
        created_by = _safe_str(m.get("created_by", "")).strip()
        stages = ((m.get("pipeline", {}) or {}).get("stages", None)) or ((m.get("pipeline_state", {}) or {}).get("stages", None)) or {}
        st_il = _safe_str(stages.get("IL", "")).strip() or "unknown"
        self.lbl_bundle_meta.configure(text=f"schema={schema} | created_at={created_at} | created_by={created_by} | IL={st_il}")

        # warnings: duplicates (they were moved to parse.skipped during load)
        dup_ids = []
        for e in (self.bundle.parse.skipped or []):
            if _safe_str(e.get("reason", "")).strip() == "duplicate local_id":
                lid = _safe_str((e.get("row") or {}).get("local_id", "")).strip()
                if lid:
                    dup_ids.append(lid)
        # unique sample
        sample = []
        seen = set()
        for x in dup_ids:
            if x not in seen:
                sample.append(x); seen.add(x)
            if len(sample) >= 10:
                break
        if dup_ids:
            warns.append(f"[data] duplicate local_id detected (n={len(set(dup_ids))}), sample={', '.join(sample)}")

        # refresh criteria table + notes
        self._refresh_criteria_table(pre_run=True)
        if self.bundle.criteria and self.bundle.criteria.warnings:
            warns.extend(self.bundle.criteria.warnings)
        self._set_warnings(warns)

        # clear report tables
        self.full_table.clear()
        self.surv_table.clear()
        self._refresh_counts_label()

        # F-118: the load path used to set four buttons itself, and its
        # Run predicate (`_has_openai_key()`) disagreed with the one in
        # `_set_controls_running` (`self.bundle_zip_path`). Since the
        # latter runs in the `finally` of every run, the gate applied
        # here survived only until the first run of a session ended.
        # There is one predicate now, so there is nothing to disagree.
        # `full_rows` was cleared above, so this yields exactly what the
        # four lines did — and it also resets IL's sixth button, which
        # they forgot.
        self._set_controls_running(False)

        self.lbl_status.configure(text="Ready.")

    def _detect_duplicate_local_ids(self, rows: List[Dict[str, str]]) -> Tuple[int, List[str]]:
        seen = set()
        dups = []
        for r in rows:
            lid = _safe_str(r.get("local_id", "")).strip()
            if not lid:
                continue
            if lid in seen:
                dups.append(lid)
            else:
                seen.add(lid)
        uniq = []
        s2 = set()
        for x in dups:
            if x not in s2:
                uniq.append(x)
                s2.add(x)
        return len(uniq), uniq[:10]

    # -------- criteria table --------

    def _refresh_criteria_table(self, pre_run: bool):
        crits = self.bundle.criteria.criteria if (self.bundle and self.bundle.criteria) else []

        cols = ["id", "type", "targets", "operator", "what", "threshold", "status", "notes"]
        if not pre_run:
            cols += ["n_failed", "n_missing", "n_met", "n_uncertain"]

        rows: List[Dict[str, Any]] = []
        header_set = set(self.bundle.parse.header) if (self.bundle and self.bundle.parse) else set()
        header_set_l = {h.lower() for h in header_set}

        for c in crits:
            status = "OK"
            notes = ""

            if not c.targets:
                status = "WARNING"
                notes = "missing target -> treated as MISSING (REVIEW)"
            elif header_set and not any((t or "").lower() in header_set_l for t in c.targets):
                status = "WARNING"
                notes = f"missing column(s): {', '.join(c.targets)} -> treated as MISSING (REVIEW)"

            op = (c.operator or "").strip().lower()
            if op != "llm":
                status = "WARNING"
                notes = (notes + " | " if notes else "") + f"operator '{op}' treated as UNCERTAIN in IL"

            d: Dict[str, Any] = {
                "id": c.id,
                "type": c.ctype,
                "targets": ",".join(c.targets),
                "operator": c.operator,
                "what": c.what_raw,
                "threshold": _safe_str(c.threshold),
                "status": status,
                "notes": notes,
            }

            if not pre_run:
                imp = self.crit_impacts.get(c.id, {"failed": 0, "missing": 0, "met": 0, "uncertain": 0})
                d["n_failed"] = str(imp.get("failed", 0))
                d["n_missing"] = str(imp.get("missing", 0))
                d["n_met"] = str(imp.get("met", 0))
                d["n_uncertain"] = str(imp.get("uncertain", 0))

            rows.append(d)

        self.criteria_table.set_columns(cols)
        col, asc = self.sort_crit
        if col:
            rows = self._sorted_rows(rows, col, asc)
        self.criteria_table.render_rows_incremental(rows)

    def _on_criterion_activated(self, row: Dict[str, str]):
        cid = _safe_str(row.get("id", "")).strip()
        if not cid:
            return
        self.active_criterion_id = cid
        self.lbl_crit_filter.configure(text=f"Criterion filter: {cid}")
        self.btn_clear_filter.configure(state="normal")
        self._refresh_reports_view()

    def _clear_criterion_filter(self):
        self.active_criterion_id = None
        self.lbl_crit_filter.configure(text="Criterion filter: (none)")
        self.btn_clear_filter.configure(state="disabled")
        self._refresh_reports_view()

    def _sort_criteria_table(self, col: str):
        cur_col, asc = self.sort_crit
        if cur_col == col:
            asc = not asc
        else:
            asc = True
        self.sort_crit = (col, asc)
        self._refresh_criteria_table(pre_run=(not bool(self.full_rows)))

    # -------- reports --------

    def _refresh_reports_view(self):
        if not self.bundle:
            return

        full_cols = list(self.bundle.parse.header) + [
            "il_outcome", "il_failed_ids", "il_missing_ids", "il_met_ids", "il_uncertain_ids", "il_reason_summary"
        ]
        surv_cols = list(self.bundle.parse.header)

        full_view = self.full_rows
        if self.active_criterion_id:
            cid = self.active_criterion_id
            filtered = []
            for i, r in enumerate(self.full_rows):
                ev = self.row_eval_lists[i] if i < len(self.row_eval_lists) else {"failed": [], "missing": [], "met": [], "uncertain": [], "suppressed": []}
                if (cid in ev.get("failed", [])) or (cid in ev.get("missing", [])) or (cid in ev.get("met", [])) or (cid in ev.get("uncertain", [])):
                    filtered.append(r)
            full_view = filtered

        surv_view = self.survivors
        if self.active_criterion_id and self.full_rows:
            cid = self.active_criterion_id
            touched_survivor_ids = set()
            for i, r in enumerate(self.full_rows):
                if r.get("il_outcome") == "OUT":
                    continue
                ev = self.row_eval_lists[i]
                if (cid in ev.get("failed", [])) or (cid in ev.get("missing", [])) or (cid in ev.get("met", [])) or (cid in ev.get("uncertain", [])):
                    touched_survivor_ids.add(_safe_str(r.get("local_id", "")).strip())
            surv_view = [r for r in self.survivors if _safe_str(r.get("local_id", "")).strip() in touched_survivor_ids]

        col, asc = self.sort_full
        if col and full_view:
            full_view = self._sorted_rows(full_view, col, asc)

        col2, asc2 = self.sort_surv
        if col2 and surv_view:
            surv_view = self._sorted_rows(surv_view, col2, asc2)

        self.full_table.set_columns(full_cols)
        self.full_table.render_rows_incremental(full_view)

        self.surv_table.set_columns(surv_cols)
        self.surv_table.render_rows_incremental(surv_view)

    def _sort_full_table(self, col: str):
        cur_col, asc = self.sort_full
        if cur_col == col:
            asc = not asc
        else:
            asc = True
        self.sort_full = (col, asc)
        self._refresh_reports_view()

    def _sort_surv_table(self, col: str):
        cur_col, asc = self.sort_surv
        if cur_col == col:
            asc = not asc
        else:
            asc = True
        self.sort_surv = (col, asc)
        self._refresh_reports_view()

    # -------- Row detail modal --------

    def _open_row_detail_modal(self, row: Dict[str, Any]):
        if not self.bundle:
            return

        win = tk.Toplevel(self)
        win.title("IL Row details")
        win.geometry("950x650")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        lid = _safe_str(row.get("local_id", "")).strip()
        title = _safe_str(row.get("title", "")).strip() or _safe_str(row.get("Title", "")).strip()

        top = ttk.Frame(win)
        top.pack(fill="x", padx=10, pady=10)

        ttk.Label(top, text=f"local_id: {lid}").pack(anchor="w")
        if title:
            ttk.Label(top, text=f"title: {title[:250]}").pack(anchor="w")

        outcome = _safe_str(row.get("il_outcome", "")).strip()
        if outcome:
            ttk.Label(top, text=f"IL outcome: {outcome}").pack(anchor="w")
            rs = _safe_str(row.get("il_reason_summary", "")).strip()
            if rs:
                ttk.Label(top, text=f"summary: {rs}").pack(anchor="w")

        box = ttk.Labelframe(win, text="Per-criterion evidence (from il_evidence_json)")
        box.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ["cid", "type", "operator", "targets", "what", "threshold", "status", "decision", "confidence", "field", "quote_valid", "quote"]
        table = DataTable(box, on_sort=lambda _c: None, on_row_activate=None)
        table.pack(fill="both", expand=True, padx=6, pady=6)
        table.set_columns(cols)

        # parse evidence dict
        ev_raw = _safe_str(row.get("il_evidence_json", "{}"))
        try:
            evj = json.loads(ev_raw) if ev_raw.strip() else {}
        except Exception:
            evj = {}

        detail_rows: List[Dict[str, Any]] = []
        for c in self.bundle.criteria.criteria:
            obj = evj.get(c.id, {}) if isinstance(evj, dict) else {}
            status = _safe_str(obj.get("status", ""))
            decision = _safe_str(obj.get("decision", ""))
            confidence = _safe_str(obj.get("confidence", ""))
            field = _safe_str(obj.get("field", ""))
            quote = _safe_str(obj.get("quote", ""))
            qv = _safe_str(obj.get("quote_valid", obj.get("quote_valid", "")))
            if isinstance(obj.get("quote_valid", None), bool):
                qv = "True" if obj.get("quote_valid") else "False"

            detail_rows.append({
                "cid": c.id,
                "type": c.ctype,
                "operator": c.operator,
                "targets": ",".join(c.targets),
                "what": c.what_raw,
                "threshold": _safe_str(c.threshold),
                "status": status,
                "decision": decision,
                "confidence": confidence,
                "field": field,
                "quote_valid": qv,
                "quote": quote,
            })

        table.render_rows_incremental(detail_rows)

        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(btns, text="Close", command=win.destroy).pack(side="right")
        win.bind("<Escape>", lambda _e: win.destroy())

    # -------- run --------

    def _run_clicked(self):
        if not self.bundle or self._worker:
            return
        # F-93/F-111: one readiness check, and it is the same one the
        # indicator and the Run button consult. The model field is part of
        # it. `(self.var_model.get() or DEFAULT_MODEL).strip()` put the
        # strip OUTSIDE the `or`, so a whitespace-only field was truthy,
        # survived the fallback, and reached the engine as "" — which the
        # engine skips silently, leaving a full corpus of unscreened records
        # under a status line reading "done".
        #
        # It refuses rather than substituting DEFAULT_MODEL. A blanked field
        # is a mistake, and quietly screening a whole corpus against a model
        # the user did not type costs real money and produces results
        # attributable to the wrong model. F-119's wording for the missing
        # key now lives in stage_state::llm_readiness with the other cases.
        # Session C: the tab's fields reach the store BEFORE the readiness
        # check and before the worker starts. The engine resolves the
        # endpoint from the store, so a value living only in a widget
        # would send the run somewhere other than what this tab shows -
        # and the readiness check would be answering about a different
        # configuration from the one that ran.
        self._stage_fields_edited()

        ready = self._readiness()
        if not ready.can_run:
            messagebox.showerror(f"IL cannot start", ready.detail)
            return

        # parse settings
        model = ready.model
        try:
            temperature = float(self.var_temp.get())
        except Exception:
            temperature = 0.0
        # F-118: `try: int(...) except: <default>` rescued 'abc' and
        # accepted '-100'. A negative truncation reaches the prompt
        # builder as a negative slice: it removes the LAST |n|
        # characters of every field and empties any field shorter than
        # that — in practice the title and keywords of every record.
        numeric = parse_numeric_settings(
            batch_raw=self.var_batch.get(),
            trunc_raw=self.var_trunc.get(),
            batch_default=DEFAULT_BATCH_SIZE,
            trunc_default=DEFAULT_TRUNC_CHARS)
        batch_size = numeric.batch_size
        trunc_chars = numeric.trunc_chars
        for _problem in numeric.problems:
            self._log("[IL] setting corrected: " + _problem + "\n")
        if numeric.problems:
            messagebox.showwarning("IL settings corrected",
                                   "\n\n".join(numeric.problems))
        use_cache = bool(self.var_use_cache.get())

        self._cancel.clear()
        self.pbar.configure(value=0.0, maximum=100.0)
        self.lbl_status.configure(text="Running IL…")
        self._set_controls_running(True)

        def progress_cb(frac: float):
            self.after(0, lambda: self.pbar.configure(value=max(0.0, min(100.0, frac * 100.0))))

        def progress_evt(evt: Dict[str, Any]):
            kind = _safe_str(evt.get("kind", ""))
            sub = _safe_str(evt.get("sub", ""))
            if kind == "l_batch" and sub == "sending":
                bi = evt.get("batch_idx"); bt = evt.get("batch_total")
                ci = evt.get("crit_idx"); ct = evt.get("crit_total")
                self.after(0, lambda: self.lbl_status.configure(text=f"LLM: criterion {ci}/{ct} batch {bi}/{bt}…"))

        def work():
            try:
                (full_rows, survivors, counts, crit_impacts,
                 row_eval_lists, cache_out, cancelled,
                 run_report) = run_il_screen(
                    self.bundle.parse,
                    self.bundle.criteria,
                    model=model,
                    trunc_chars=trunc_chars,
                    batch_size=batch_size,
                    temperature=temperature,
                    use_cache=use_cache,
                    cache_in=self.cache_map if use_cache else {},
                    cancel_event=self._cancel,  # ✅ ILView uses self._cancel
                    log_cb=lambda s: self.after(0, (lambda ss=s: self._log(ss))),
                    progress_cb=progress_cb,
                    progress_evt=progress_evt,
                )

                self.llm_report = run_report

                if cancelled:
                    # F-02: the engine reports that it stopped mid-corpus.
                    # Trust that rather than re-reading the event, drop the
                    # partial results instead of leaving them on screen, and
                    # latch the flag so the export handlers refuse even if an
                    # earlier complete run left rows behind.
                    self.cancelled = True
                    self.full_rows = []
                    self.survivors = []
                    self.after(0, lambda: self._log(
                        "\n[CANCELLED] Partial results discarded; export stays "
                        "disabled until IL is re-run to completion.\n"))
                    self.outcome = run_outcome(
                        stage="IL", counts=counts,
                        llm_report=run_report, cancelled=True,
                        not_screened=False, total_rows=len(full_rows))
                    self.after(0, lambda t=self.outcome.label:
                               self.lbl_status.configure(text=t))
                    return

                self.full_rows = full_rows
                self.survivors = survivors
                self.counts = counts
                # F-34: the engine reports the no-op through the counts.
                self.not_screened = bool(counts.get(NOT_SCREENED, 0))
                self.crit_impacts = crit_impacts
                self.row_eval_lists = row_eval_lists
                self.cache_map = cache_out

                # ✅ refresh views + counts + criteria impacts
                self.after(0, self._refresh_reports_view)
                self.after(0, lambda: self._refresh_criteria_table(pre_run=False))
                self.after(0, self._refresh_counts_label)
                # F-34: a stage that evaluated no criteria is not "done"
                # in any sense the user means by the word. The
                # classification is stage_state::run_outcome's now.
                self.outcome = run_outcome(
                    stage="IL", counts=counts,
                    llm_report=run_report, cancelled=False,
                    not_screened=self.not_screened,
                    total_rows=len(full_rows))
                self.after(0, lambda t=self.outcome.label:
                           self.lbl_status.configure(text=t))
                # F-93: the status line is one line in a bar the user
                # may not be looking at; the counts belong in the trail
                # too, next to the per-batch failures that explain them.
                self.after(0, lambda t=self.outcome.label:
                           self._log("\n[IL] " + t + "\n"))

            except Exception as e:
                self.after(0, lambda m=str(e): messagebox.showerror("IL run failed", m))
                self.after(0, lambda: self.lbl_status.configure(text="IL failed."))
            finally:
                self.after(0, lambda: self._set_controls_running(False))
                # clear worker flag so you can run again
                self._worker = None

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    def _cancel_run(self):
        self._cancel.set()
        self._log("\n[Cancel requested]\n")
        self.lbl_status.configure(text="Cancelling…")

    # -------- Export --------

    def _export_clicked(self):
        if not self.bundle:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        blocked = _export_block_reason(has_rows=bool(self.full_rows),
                                       cancelled=self.cancelled)
        if blocked:
            messagebox.showwarning("Cannot export", blocked)
            return
        # F-34: a stage that screened nothing may still be exported,
        # but not without the user saying so out loud.
        # F-93 extends F-34's gate rather than adding a second one:
        # `not_screened` covers a stage with no criteria, and the
        # outcome covers a stage that had criteria and still learned
        # nothing. One gate, two diagnoses.
        confirm = _export_confirm_reason(
            not_screened=self.not_screened,
            stage="IL",
            outcome_reason=(self.outcome.ack_reason if self.outcome
                            else None))
        if confirm:
            title = ("Nothing was screened" if self.not_screened
                     else "Check this before exporting")
            if not messagebox.askyesno(title, confirm):
                return

        default_name = f"{_now_stamp()}_IL_reports.xlsx"
        p = filedialog.asksaveasfilename(
            title="Save IL reports",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not p:
            return

        try:
            _export_il_xlsx(p, self.full_rows, self.survivors, self.bundle.parse.header)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_final_clicked(self):
        blocked = (
            (None if self.bundle else "Load a bundle first.")
            or _export_block_reason(has_rows=bool(self.full_rows),
                                    cancelled=self.cancelled)
        )
        if blocked:
            messagebox.showinfo("Not ready", blocked)
            return

        # F-93. This handler had no acknowledgement gate at all, unlike its
        # two siblings — a pre-existing asymmetry, and the worst place for
        # one: ScreenA_Report.xlsx is the terminal deliverable, the artefact
        # a reviewer actually reads. Gating the other two doors and leaving
        # this one open would not be a gate.
        confirm = _export_confirm_reason(
            not_screened=self.not_screened,
            stage="IL",
            outcome_reason=(self.outcome.ack_reason if self.outcome
                            else None))
        if confirm:
            title = ("Nothing was screened" if self.not_screened
                     else "Check this before exporting")
            if not messagebox.askyesno(title, confirm):
                return

        default_name = f"{_now_stamp()}_{FINAL_REPORT_NAME}"
        p = filedialog.asksaveasfilename(
            title="Save ScreenA_Report.xlsx",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not p:
            return

        try:
            with zipfile.ZipFile(self.bundle.zip_path, "r") as zf:
                wb_bytes = _build_final_report_xlsx_bytes(zf, self.bundle.root, self.full_rows)
            Path(p).write_bytes(wb_bytes)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_errors_clicked(self):
        if not self.bundle:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        if not self.bundle.parse.skipped:
            messagebox.showinfo("No input errors", "No skipped/invalid rows were recorded for this bundle.")
            return

        default_name = f"{_now_stamp()}_input_errors.csv"
        p = filedialog.asksaveasfilename(
            title="Save input_errors.csv",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv")],
        )
        if not p:
            return

        try:
            # F-74: the canonical six-column schema through the shared
            # writer, not the legacy reason,row_json layout this handler
            # used to emit inline.
            _export_input_errors_csv_from_dicts(
                p, self.bundle.parse.skipped, stage="IL")
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_bundle_clicked(self):
        if not self.bundle:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        blocked = _export_block_reason(has_rows=bool(self.full_rows),
                                       cancelled=self.cancelled)
        if blocked:
            messagebox.showwarning("Cannot export", blocked)
            return
        # F-34: a stage that screened nothing may still be exported,
        # but not without the user saying so out loud.
        # F-93 extends F-34's gate rather than adding a second one:
        # `not_screened` covers a stage with no criteria, and the
        # outcome covers a stage that had criteria and still learned
        # nothing. One gate, two diagnoses.
        confirm = _export_confirm_reason(
            not_screened=self.not_screened,
            stage="IL",
            outcome_reason=(self.outcome.ack_reason if self.outcome
                            else None))
        if confirm:
            title = ("Nothing was screened" if self.not_screened
                     else "Check this before exporting")
            if not messagebox.askyesno(title, confirm):
                return

        default_name = f"{_now_stamp()}_post_IL_bundle.zip"
        out_zip = filedialog.asksaveasfilename(
            title="Save next bundle ZIP (post-IL)",
            defaultextension=".zip",
            initialfile=default_name,
            filetypes=[("ZIP", "*.zip")],
        )
        if not out_zip:
            return

        try:
            self._build_next_bundle_zip(out_zip)
        except Exception as e:
            messagebox.showerror("Bundle export failed", str(e))
            return

        self.lbl_status.configure(text=f"Bundle exported: {Path(out_zip).name}")

    def _build_next_bundle_zip(self, out_zip: str) -> None:
        """Create a new bundle zip where data/current.csv becomes the IL survivors.

        Keeps other files from the input bundle, updates manifest pipeline stage, adds reports, carries input_errors,
        and writes cache/IL_cache.jsonl when enabled.
        """
        assert self.bundle is not None

        with zipfile.ZipFile(self.bundle.zip_path, "r") as zf_in:
            members = zf_in.namelist()
            root = self.bundle.root

            manifest = dict(self.bundle.manifest)

            def _set_stage(m: Dict[str, Any], key: str, value: str):
                if "pipeline" in m and isinstance(m.get("pipeline"), dict):
                    m["pipeline"].setdefault("stages", {})
                    m["pipeline"]["stages"][key] = value
                if "pipeline_state" in m and isinstance(m.get("pipeline_state"), dict):
                    m["pipeline_state"].setdefault("stages", {})
                    m["pipeline_state"]["stages"][key] = value

            _set_stage(manifest, "IL", "done")
            manifest["updated_at"] = datetime.utcnow().isoformat() + "Z"

            header = list(self.bundle.parse.header)
            if "local_id" not in header:
                header = ["local_id"] + header

            header_full = [
                *header,
                "il_outcome", "il_failed_ids", "il_missing_ids", "il_met_ids", "il_uncertain_ids", "il_reason_summary", "il_evidence_json",
            ]

            # F-05: shared writer, which refreshes the manifest digests. See
            # the EL twin. IL additionally emits the cross-stage FINAL
            # workbook, passed through as an extra member so it is digested
            # alongside everything else rather than written behind the
            # manifest's back.
            _write_llm_stage_bundle(
                out_zip, zf_in,
                root=root,
                manifest=manifest,
                stage="IL",
                reports_dir_rel=REPORTS_DIR_REL,
                cache_rel=IL_CACHE_REL,
                parse_header=header,
                survivors=self.survivors,
                full_rows=self.full_rows,
                full_header=header_full,
                skipped=self.bundle.parse.skipped,
                counts=self.counts,
                not_screened=self.not_screened,
                cancelled=self.cancelled,
                llm_report=self.llm_report,
                cache_text=(_dump_cache_to_jsonl(self.cache_map)
                            if self.var_use_cache.get() else None),
                extra_members={
                    FINAL_REPORT_REL: _build_final_report_xlsx_bytes(
                        zf_in, root, self.full_rows),
                },
            )


# StandaloneILPlugin moved to plugins/07_il/standalone.py in
# Conv 6 / Commit 4. The plugin module re-exports it from there.
