# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""
plugin.py — Screen A (IL-only) as a metaScreener tab plugin (Contract v2, IL stage)

Design goals (aligned with your EH/IH conventions)
- Single-file plugin: UI + engine + LLM interaction (no local imports except metascreener.plugin_api)
- Bundle-first input (Harmoniser ZIP): reads data/current.csv + criteria/criteria_harmonized.csv
- IL semantics (contract v2):
    - Per criterion status: MET / FAILED / MISSING / UNCERTAIN
    - Per row outcome:
        any FAILED           -> OUT
        else all MET         -> PASS_CLEAN
        else                 -> REVIEW
- LLM logic (recovered + compatible with your legacy metadata.py approach):
    - System: JSON-only list; per item: a_id, decision(meet|not_meet|uncertain), confidence(0..1), field, quote, span
    - Evidence gating: decision counts only if confidence>=threshold AND quote_valid=True
    - Batching: chunk items; adaptive shrink on 429/oversize; optional truncation reduction
- Persistent cache (default ON): cache/IL_cache.jsonl stored inside output bundle.
  Key includes model + criterion_id + a_id + hash(truncated target text) + prompt_version.

Notes
- Requires metascreener.plugin_api.BasePlugin / PluginMeta
- Requires OPENAI_API_KEY in environment to run LLM
- Uses OpenAI Chat Completions via `from openai import OpenAI` (imported inside function)
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
import threading
import time
import zipfile
from dataclasses import dataclass
from datetime import datetime
from hashlib import sha256
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from metascreener.plugin_api import BasePlugin, PluginMeta


# ------------------------------ constants -------------------------------------

TAB_TITLE = "Screen A — IL"
PLUGIN_ID = "screen_a_il"
PLUGIN_VERSION = "2.0.0"
# Defaults (safe; overridable in UI + env)
DEFAULT_MODEL = os.environ.get("SCREENA_IL_MODEL", "gpt-4o-mini")
DEFAULT_TRUNC_CHARS = int(os.environ.get("SCREENA_IL_TRUNC_CHARS", "1500"))
DEFAULT_BATCH_SIZE = int(os.environ.get("SCREENA_IL_BATCH_SIZE", "50"))
DEFAULT_USE_CACHE = os.environ.get("SCREENA_IL_USE_CACHE", "1").strip() not in {"0", "false", "False", "no", "NO"}

RENDER_CHUNK = 400

PROMPT_VERSION = "IL_v1_jsonlist"

IL_CACHE_REL = "cache/IL_cache.jsonl"
REPORTS_DIR_REL = "reports"
FINAL_REPORT_NAME = "ScreenA_Report.xlsx"
FINAL_REPORT_REL = f"{REPORTS_DIR_REL}/{FINAL_REPORT_NAME}"

# Contract v2: standardized stage sheet columns
CONTRACT_STAGE_SHEET_COLS = [
    "a_id",
    "stage",
    "stage_outcome",
    "passed_to_next",
    "failed_criteria_ids",
    "missing_criteria_ids",
    "uncertain_criteria_ids",
    "met_criteria_ids",
    "matched_evidence",
    "stage_reason_summary",
    "history",
]
OUTCOMES = ("OUT", "PASS_CLEAN", "REVIEW")

# ------------------------------ dataclasses -----------------------------------

@dataclass
class Criterion:
    id: str
    stage: str                  # EH / IH / EL / IL
    ctype: str                  # include / exclude
    enabled: bool
    operator: str               # contains / regex / llm / ...
    targets: List[str]          # ["title","abstract"] etc
    what_raw: str               # raw "what" cell
    what_list: List[str]        # parsed list
    threshold: float            # for llm
    source_text: str            # human-readable criterion text (for UI)
    label: str = ""             # optional

@dataclass
class ParseReport:
    header: List[str]
    rows: List[Dict[str, str]]
    skipped: List[Dict[str, Any]]  # rows skipped due to parse issues

@dataclass
class CriteriaLoadReport:
    criteria: List[Criterion]
    warnings: List[str]

@dataclass
class BundleInfo:
    zip_path: str
    root: str
    manifest: Dict[str, Any]
    parse: ParseReport
    criteria: CriteriaLoadReport


# ------------------------------ small utils ----------------------------------

def _safe_str(x: Any) -> str:
    return "" if x is None else str(x)

def _decode_bytes(b: bytes) -> str:
    # BOM-safe decode
    return b.decode("utf-8-sig", errors="replace")

def _read_zip_bytes(zf: zipfile.ZipFile, member: str) -> bytes:
    with zf.open(member, "r") as fp:
        return fp.read()

def _detect_bundle_root(members: Sequence[str]) -> str:
    """
    Accept:
      manifest.json at root
      OR inside a single top folder, e.g. ScreenA_Bundle/manifest.json
    Return root prefix ("" or "ScreenA_Bundle/").
    """
    if "manifest.json" in members:
        return ""
    # find any */manifest.json
    for m in members:
        if m.endswith("/manifest.json"):
            return m[:-len("manifest.json")]
    # fallback: try first segment
    tops = {m.split("/", 1)[0] for m in members if "/" in m}
    for t in sorted(tops):
        if f"{t}/manifest.json" in members:
            return f"{t}/"
    return ""

def _csv_read(text: str) -> Tuple[List[str], List[Dict[str, str]]]:
    # Robust CSV reading with UTF-8 and tolerant rows
    f = io.StringIO(text)
    reader = csv.reader(f)
    rows = list(reader)
    if not rows:
        return [], []
    header = [h.strip() for h in rows[0]]
    out: List[Dict[str, str]] = []
    for i, r in enumerate(rows[1:], start=2):
        if not any((c or "").strip() for c in r):
            continue
        d: Dict[str, str] = {}
        for j, h in enumerate(header):
            d[h] = (r[j] if j < len(r) else "")
        out.append(d)
    return header, out

def _write_csv(path: str, header: List[str], rows: List[Dict[str, Any]]):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in header})

def _sha_text(s: str) -> str:
    return sha256(s.encode("utf-8", errors="ignore")).hexdigest()

def _normalize_space(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()

def _quote_in_text(quote: str, text: str) -> bool:
    """
    Quote validity check:
      - exact substring OR
      - whitespace-normalized substring (prevents false invalid due to newlines/multiple spaces)
    """
    if not quote or not text:
        return False
    if quote in text:
        return True
    qn = _normalize_space(quote)
    tn = _normalize_space(text)
    return bool(qn) and (qn in tn)

def _has_openai_key() -> bool:
    return bool(os.environ.get("OPENAI_API_KEY", "").strip())

def chunked(seq: Sequence[Any], n: int):
    n = max(1, int(n))
    for i in range(0, len(seq), n):
        yield seq[i:i+n]


# -------------------------- bundle parsing ------------------------------------

def _load_bundle(zip_path: str) -> BundleInfo:
    if not zip_path.lower().endswith(".zip"):
        raise ValueError("Bundle must be a .zip file.")
    with zipfile.ZipFile(zip_path, "r") as zf:
        members = zf.namelist()
        root = _detect_bundle_root(members)

        manifest_full = root + "manifest.json"
        if manifest_full not in members:
            raise FileNotFoundError("Bundle missing manifest.json")

        manifest_bytes = _read_zip_bytes(zf, manifest_full)
        try:
            manifest = json.loads(_decode_bytes(manifest_bytes))
        except Exception as e:
            raise ValueError(f"manifest.json is not valid JSON: {e}")

        # data/current.csv
        current_full = root + "data/current.csv"
        if current_full not in members:
            raise FileNotFoundError("Bundle missing data/current.csv")
        current_txt = _decode_bytes(_read_zip_bytes(zf, current_full))
        header, rows = _csv_read(current_txt)

        skipped: List[Dict[str, Any]] = []
        # Minimal sanity: require local_id
        if "local_id" not in header:
            # Try common fallbacks
            for cand in ("id", "ID", "LocalID", "localId"):
                if cand in header:
                    # rename in-memory
                    for r in rows:
                        r["local_id"] = r.get(cand, "")
                    header.append("local_id")
                    break
        # mark rows with empty or duplicate local_id as skipped
        filtered: List[Dict[str, str]] = []
        seen_ids: set = set()
        for r in rows:
            lid = (r.get("local_id") or "").strip()
            if not lid:
                skipped.append({"reason": "missing local_id", "row": r})
                continue
            if lid in seen_ids:
                skipped.append({"reason": "duplicate local_id", "row": r})
                continue
            seen_ids.add(lid)
            filtered.append(r)

        parse = ParseReport(header=header, rows=filtered, skipped=skipped)

        # criteria/criteria_harmonized.csv
        crit_full = root + "criteria/criteria_harmonized.csv"
        if crit_full not in members:
            raise FileNotFoundError("Bundle missing criteria/criteria_harmonized.csv")
        crit_txt = _decode_bytes(_read_zip_bytes(zf, crit_full))
        criteria = _parse_criteria_harmonized_csv(crit_txt, stage_filter="IL")

        return BundleInfo(zip_path=zip_path, root=root, manifest=manifest, parse=parse, criteria=criteria)

def _parse_criteria_harmonized_csv(csv_text: str, stage_filter: str) -> CriteriaLoadReport:
    header, rows = _csv_read(csv_text)
    # columns we expect (tolerate variations)
    def get(d: Dict[str,str], *keys: str) -> str:
        for k in keys:
            if k in d:
                return d.get(k, "")
        # case-insensitive
        kl = {kk.lower(): kk for kk in d.keys()}
        for k in keys:
            if k.lower() in kl:
                return d.get(kl[k.lower()], "")
        return ""

    crits: List[Criterion] = []
    warnings: List[str] = []

    for r in rows:
        stage = _safe_str(get(r, "stage")).strip().upper()
        if stage != stage_filter.upper():
            continue

        enabled_s = _safe_str(get(r, "enabled")).strip().lower()
        enabled = enabled_s not in {"0", "false", "no"}

        cid = _safe_str(get(r, "id", "criterion_id")).strip()
        if not cid:
            continue

        ctype = _safe_str(get(r, "type", "ctype")).strip().lower() or "exclude"
        operator = _safe_str(get(r, "operator")).strip().lower() or "llm"
        target_raw = _safe_str(get(r, "target", "targets")).strip()
        targets = [t.strip().lower() for t in re.split(r"[,+;]", target_raw) if t.strip()] or ["abstract"]

        what_raw = _safe_str(get(r, "what")).strip()
        # "what" might be JSON list, or newline-separated
        what_list: List[str] = []
        if what_raw:
            try:
                val = json.loads(what_raw)
                if isinstance(val, list):
                    what_list = [str(x) for x in val]
                else:
                    what_list = [str(val)]
            except Exception:
                what_list = [w.strip() for w in re.split(r"[\n;|]+", what_raw) if w.strip()]
        thr_s = _safe_str(get(r, "threshold", "thr")).strip()
        try:
            thr = float(thr_s) if thr_s else 0.6
        except Exception:
            thr = 0.6

        label = _safe_str(get(r, "label")).strip()
        source_text = _safe_str(get(r, "source_text", "text", "criterion_text")).strip()
        if not source_text:
            # synthesize from what/label
            source_text = label or (what_list[0] if what_list else "")

        crits.append(
            Criterion(
                id=cid,
                stage=stage,
                ctype=ctype,
                enabled=enabled,
                operator=operator,
                targets=targets,
                what_raw=what_raw,
                what_list=what_list,
                threshold=thr,
                source_text=source_text,
                label=label,
            )
        )

    if not crits:
        warnings.append(f"No IL criteria found (stage={stage_filter}).")

    return CriteriaLoadReport(criteria=crits, warnings=warnings)


# ------------------------ LLM utilities (legacy-compatible) -------------------

def _parse_llm_json_array(text: str) -> List[Dict[str, Any]]:
    """
    Robust parse of a JSON array. Accepts:
      - pure JSON list
      - fenced code block
      - extra text before/after (extract first [...] block)
    """
    t = (text or "").strip()
    # strip fenced code blocks
    if t.startswith("```"):
        t = re.sub(r"^```[a-zA-Z0-9]*\s*", "", t)
        t = re.sub(r"\s*```$", "", t).strip()

    # direct
    try:
        val = json.loads(t)
        if isinstance(val, list):
            return [v for v in val if isinstance(v, dict)]
    except Exception:
        pass

    # extract first bracketed list
    m = re.search(r"\[\s*\{.*\}\s*\]", t, flags=re.S)
    if m:
        try:
            val = json.loads(m.group(0))
            if isinstance(val, list):
                return [v for v in val if isinstance(v, dict)]
        except Exception:
            pass
    return []

def _build_llm_messages_for_criterion(criterion: Dict[str, Any], items: List[Dict[str, Any]], trunc_chars: int) -> List[Dict[str, str]]:
    sys = (
        "You are scoring research items against ONE screening criterion. "
        "For each item, answer with JSON only. Keys per item: "
        "a_id, decision ('meet'|'not_meet'|'uncertain'), confidence (0..1), "
        "field ('title'|'abstract'|'keywords'), quote (exact substring from that field), span [start,end]. "
        "Return a JSON list of objects, nothing else."
    )

    c_pack = {
        "id": criterion["id"],
        "type": criterion.get("type", "exclude"),
        "operator": criterion.get("operator", "llm"),
        "target": criterion.get("target", "abstract"),
        "what": criterion.get("what", []),
        "how": criterion.get("how", "llm"),
        "label": criterion.get("label", ""),
        "threshold": criterion.get("threshold", 0.6),
    }

    def trunc(s: str) -> str:
        s = s or ""
        if trunc_chars and len(s) > trunc_chars:
            return s[:trunc_chars]
        return s

    items_pack = []
    for it in items:
        items_pack.append({
            "a_id": it.get("a_id", ""),
            "title": trunc(_safe_str(it.get("title",""))),
            "abstract": trunc(_safe_str(it.get("abstract",""))),
            "keywords": trunc(_safe_str(it.get("keywords",""))),
        })

    user = json.dumps({"criterion": c_pack, "items": items_pack}, ensure_ascii=False)
    return [{"role":"system","content":sys}, {"role":"user","content":user}]

def run_m1_llm_for_criterion(
    criterion: Dict[str, Any],
    items: List[Dict[str, Any]],
    *,
    model: Optional[str],
    trunc_chars: int = DEFAULT_TRUNC_CHARS,
    batch_size: int = DEFAULT_BATCH_SIZE,
    log: Optional[Callable[[str], None]] = None,
    progress: Optional[Callable[[Dict[str, Any]], None]] = None,
    cancel_token: Optional[object] = None,
    crit_idx: Optional[int] = None,
    crit_total: Optional[int] = None,
    block_tag: str = "exclude",
) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """
    Returns (a_id, criterion_id) -> llm_decision_dict (used/decision/confidence/field/quote/span/valid_quote).
    Implements adaptive batching on errors (429/oversize).
    """
    if not model:
        if log: log("[IL-LLM] model=None; skipping.\n")
        return {}
    if not _has_openai_key():
        if log: log("[IL-LLM] OPENAI_API_KEY not visible in environment; skipping.\n")
        return {}

    try:
        from openai import OpenAI  # type: ignore
        client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
    except Exception as e:
        if log: log(f"[IL-LLM] OpenAI client import/init failed: {e}\n")
        return {}

    def _check_cancel():
        if cancel_token is not None and bool(getattr(cancel_token, "cancelled", False)):
            raise RuntimeError("Cancelled")
        if cancel_token is not None and bool(getattr(cancel_token, "is_set", lambda: False)()):
            raise RuntimeError("Cancelled")

    def _call_once(batch: List[Dict[str, Any]], cur_trunc: int):
        msgs = _build_llm_messages_for_criterion(criterion, batch, cur_trunc)
        return client.chat.completions.create(
            model=model,
            messages=msgs,
            temperature=0.0,
        )

    out: Dict[Tuple[str, str], Dict[str, Any]] = {}
    cid = criterion["id"]
    
    # Build a_id -> field texts map (used for quote validation + ignoring unknown a_id)
    idx_map: Dict[str, Dict[str, str]] = {}
    for it in items:
        a_id = _safe_str(it.get("a_id", "")).strip()
        if not a_id:
            continue
        idx_map[a_id] = {
            "title": _safe_str(it.get("title", "")),
            "abstract": _safe_str(it.get("abstract", "")),
            "keywords": _safe_str(it.get("keywords", "")),
        }

    # initial batches
    batches: List[List[Dict[str, Any]]] = [list(b) for b in chunked(items, max(1, int(batch_size)))]

    bi = 0
    while bi < len(batches):
        _check_cancel()

        cur_batch = list(batches[bi])
        cur_trunc = int(trunc_chars)
        attempts = 0

        while True:
            attempts += 1
            try:
                total_batches = len(batches)
                batch_num = bi + 1

                if progress:
                    progress({
                        "kind": "l_batch",
                        "stage": "IL",
                        "block": block_tag,
                        "crit_idx": crit_idx,
                        "crit_total": crit_total,
                        "batch_idx": batch_num,
                        "batch_total": total_batches,
                        "sub": "sending",
                    })

                resp = _call_once(cur_batch, cur_trunc)
                _check_cancel()

                if progress:
                    progress({
                        "kind": "l_batch",
                        "stage": "IL",
                        "block": block_tag,
                        "crit_idx": crit_idx,
                        "crit_total": crit_total,
                        "batch_idx": batch_num,
                        "batch_total": total_batches,
                        "sub": "parsing",
                    })

                txt = (resp.choices[0].message.content or "[]")
                arr = _parse_llm_json_array(txt)

                # parse response objects
                for obj in arr:
                    a_id = _safe_str(obj.get("a_id", "")).strip()
                    if not a_id or a_id not in idx_map:
                        continue

                    decision = _safe_str(obj.get("decision", "uncertain")).strip()
                    if decision not in {"meet", "not_meet", "uncertain"}:
                        decision = "uncertain"

                    try:
                        confidence = float(obj.get("confidence", 0.0))
                    except Exception:
                        confidence = 0.0
                    confidence = min(1.0, max(0.0, confidence))

                    field = _safe_str(obj.get("field", "")).strip().lower()
                    if field not in {"title", "abstract", "keywords"}:
                        field = "abstract"

                    quote = _safe_str(obj.get("quote", ""))
                    span = obj.get("span", None)
                    if not (isinstance(span, list) and len(span) == 2 and all(isinstance(x, int) for x in span)):
                        span = None

                    fld_txt = (idx_map.get(a_id) or {}).get(field) or ""
                    fld_txt = (idx_map.get(a_id) or {}).get(field) or ""
                    # Validate against the SAME truncated text that was sent to the model for this call
                    fld_txt_prompt = (fld_txt[:cur_trunc] if cur_trunc and len(fld_txt) > cur_trunc else fld_txt)
                    valid_quote = _quote_in_text(quote, fld_txt_prompt)

                    out[(a_id, cid)] = {
                        "used": True,
                        "decision": decision,
                        "confidence": confidence,
                        "field": field,
                        "quote": quote,
                        "span": span,
                        "valid_quote": valid_quote,
                    }

                # ensure every item in THIS cur_batch has an entry
                for it in cur_batch:
                    a_id = _safe_str(it.get("a_id", "")).strip()
                    if not a_id:
                        continue
                    if (a_id, cid) not in out:
                        out[(a_id, cid)] = {
                            "used": False,
                            "decision": "uncertain",
                            "confidence": 0.0,
                            "field": "abstract",
                            "quote": "",
                            "span": None,
                            "valid_quote": False,
                        }

                if progress:
                    progress({
                        "kind": "l_batch",
                        "stage": "IL",
                        "block": block_tag,
                        "crit_idx": crit_idx,
                        "crit_total": crit_total,
                        "batch_idx": batch_num,
                        "batch_total": total_batches,
                        "sub": "batch_done",
                    })

                break  # ✅ batch success

            except Exception as e:
                msg = str(e).lower()
                is_rate = ("429" in msg) or ("too many requests" in msg) or ("rate" in msg and "limit" in msg)
                is_big = ("too large" in msg) or ("context" in msg and "length" in msg) or ("max tokens" in msg)

                # ✅ split WITHOUT losing items (requeue remainder right after this batch)
                if (is_rate or is_big) and len(cur_batch) > 1:
                    new_n = max(1, len(cur_batch) // 2)
                    remainder = cur_batch[new_n:]
                    cur_batch = cur_batch[:new_n]

                    # replace current batch, insert remainder as next batch
                    batches[bi] = cur_batch
                    if remainder:
                        batches.insert(bi + 1, remainder)

                    if log:
                        log(
                            f"[IL-LLM] batch {bi+1}/{len(batches)} error ({e}); "
                            f"split into {len(cur_batch)} + {len(remainder)}\n"
                        )

                    time.sleep(min(4.0, 0.4 * attempts))
                    continue

                # reduce truncation if still big/rate and trunc is high
                if (is_rate or is_big) and cur_trunc > 600:
                    new_trunc = max(600, int(cur_trunc * 0.75))
                    if log:
                        log(f"[IL-LLM] batch {bi+1}/{len(batches)} error ({e}); trunc {cur_trunc} -> {new_trunc}\n")
                    cur_trunc = new_trunc
                    time.sleep(min(4.0, 0.4 * attempts))
                    continue

                # final failure for this batch: mark all items in THIS cur_batch as uncertain
                if log:
                    log(f"[IL-LLM] batch {bi+1}/{len(batches)} failed: {e}\n")

                for it in cur_batch:
                    a_id = _safe_str(it.get("a_id", "")).strip()
                    if not a_id:
                        continue
                    out[(a_id, cid)] = {
                        "used": False,
                        "decision": "uncertain",
                        "confidence": 0.0,
                        "field": "abstract",
                        "quote": "",
                        "span": None,
                        "valid_quote": False,
                        "error": str(e),
                    }
                break  # stop retrying this batch

        bi += 1

    return out


# ------------------------ IL engine (self-contained) --------------------------

def _make_item_for_llm(row: Dict[str, str]) -> Dict[str, Any]:
    return {
        "a_id": _safe_str(row.get("local_id","")).strip(),
        "title": _safe_str(row.get("title","")),
        "abstract": _safe_str(row.get("abstract","")),
        "keywords": _safe_str(row.get("keywords","")),
    }

def _row_target_text_hash(row: Dict[str, str], targets: List[str], trunc_chars: int) -> str:
    parts: List[str] = []
    for t in targets:
        v = _safe_str(row.get(t, ""))
        if trunc_chars and len(v) > trunc_chars:
            v = v[:trunc_chars]
        parts.append(v)
    return _sha_text("|".join(parts))

def _cache_key(*, model: str, cid: str, a_id: str, text_hash: str, trunc_chars: int) -> str:
    base = f"{PROMPT_VERSION}|{model}|{cid}|{a_id}|{text_hash}|{trunc_chars}"
    return _sha_text(base)

def _load_cache_from_jsonl(text: str) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for ln in (text or "").splitlines():
        ln = ln.strip()
        if not ln:
            continue
        try:
            obj = json.loads(ln)
            k = _safe_str(obj.get("key",""))
            v = obj.get("val", None)
            if k and isinstance(v, dict):
                out[k] = v
        except Exception:
            continue
    return out

def _dump_cache_to_jsonl(cache: Dict[str, Dict[str, Any]]) -> str:
    lines: List[str] = []
    for k, v in cache.items():
        lines.append(json.dumps({"key": k, "val": v}, ensure_ascii=False))
    return "\n".join(lines) + ("\n" if lines else "")

def run_il_screen(
    parse: ParseReport,
    criteria_report: CriteriaLoadReport,
    *,
    model: str,
    trunc_chars: int,
    batch_size: int,
    use_cache: bool,
    cache_in: Optional[Dict[str, Dict[str, Any]]],
    cancel_event: threading.Event,
    log_cb: Optional[Callable[[str], None]] = None,
    progress_cb: Optional[Callable[[float], None]] = None,
    progress_evt: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> Tuple[
    List[Dict[str, Any]],   # full_rows with IL columns
    List[Dict[str, str]],   # survivors (original schema)
    Dict[str, int],         # counts
    Dict[str, Dict[str, int]],  # crit_impacts
    List[Dict[str, List[str]]], # row_eval_lists (aligned with full_rows)
    Dict[str, Dict[str, Any]],  # cache_out
]:
    rows = parse.rows
    crits = [c for c in criteria_report.criteria if c.enabled]
    counts = {k: 0 for k in OUTCOMES}
    crit_impacts: Dict[str, Dict[str, int]] = {c.id: {"failed":0,"missing":0,"met":0,"uncertain":0} for c in crits}

    full_rows: List[Dict[str, Any]] = []
    survivors: List[Dict[str, str]] = []
    row_eval_lists: List[Dict[str, List[str]]] = []

    cache_out: Dict[str, Dict[str, Any]] = dict(cache_in or {})

    if not crits:
        for r in rows:
            if cancel_event.is_set():
                break
            fr = dict(r)
            fr["il_outcome"] = "PASS_CLEAN"
            fr["il_failed_ids"] = ""
            fr["il_missing_ids"] = ""
            fr["il_met_ids"] = ""
            fr["il_uncertain_ids"] = ""
            fr["il_reason_summary"] = "No active IL criteria: default PASS_CLEAN."
            fr["il_evidence_json"] = "{}"
            full_rows.append(fr)
            survivors.append(dict(r))
            row_eval_lists.append({"failed": [], "missing": [], "met": [], "uncertain": []})
        counts["PASS_CLEAN"] = len(survivors)
        if progress_cb:
            progress_cb(1.0)
        return full_rows, survivors, counts, crit_impacts, row_eval_lists, cache_out

    # Build items for LLM (same base shape as legacy), but robust to header casing
    header_map: Dict[str, str] = {h.lower(): h for h in (parse.header or [])}

    def getv(row: Dict[str, str], key: str) -> str:
        k = header_map.get(key.lower(), key)
        return _safe_str(row.get(k, ""))

    def row_text_hash(row: Dict[str, str], targets: List[str]) -> str:
        parts: List[str] = []
        for t in targets:
            v = getv(row, t)
            if trunc_chars and len(v) > trunc_chars:
                v = v[:trunc_chars]
            parts.append(v)
        return _sha_text("|".join(parts))

    items = [{
        "a_id": getv(r, "local_id").strip(),
        "title": getv(r, "title"),
        "abstract": getv(r, "abstract"),
        "keywords": getv(r, "keywords"),
    } for r in rows]

    id_to_row: Dict[str, Dict[str, str]] = {}
    for r in rows:
        lid = getv(r, "local_id").strip()
        if lid:
            id_to_row[lid] = r

    # Precompute per-row text hashes per criterion for caching
    per_row_hash: Dict[Tuple[str,str], str] = {}  # (a_id,cid)->text_hash
    for c in crits:
        for r in rows:
            a_id = getv(r, "local_id").strip()
            if not a_id:
                continue
            per_row_hash[(a_id, c.id)] = row_text_hash(r, c.targets)

    # Run criterion by criterion (legacy-style)
    llm_results: Dict[Tuple[str,str], Dict[str, Any]] = {}

    for ci, c in enumerate(crits, start=1):
        if cancel_event.is_set():
            break

        # build criterion pack
        crit_pack = {
            "id": c.id,
            "type": c.ctype,
            "operator": c.operator,
            "target": ",".join(c.targets),
            "what": c.what_list,
            "how": "llm" if c.operator == "llm" else c.operator,
            "label": c.label or c.source_text,
            "threshold": c.threshold,
        }

        if log_cb:
            log_cb(f"\n[IL] Criterion {ci}/{len(crits)} {c.id} ({c.operator})\n")

        # Separate cached vs to-call items
        to_call: List[Dict[str, Any]] = []
        cached_pairs: List[Tuple[str,str]] = []

        for it in items:
            a_id = _safe_str(it.get("a_id","")).strip()
            if not a_id:
                continue
            th = per_row_hash.get((a_id, c.id), "")
            k = _cache_key(model=model, cid=c.id, a_id=a_id, text_hash=th, trunc_chars=trunc_chars)
            if use_cache and k in cache_out:
                # reuse cached evidence
                ev = dict(cache_out[k])
                ev.setdefault("used", True)
                llm_results[(a_id, c.id)] = ev
                cached_pairs.append((a_id, c.id))
            else:
                to_call.append(it)

        if log_cb:
            log_cb(f"[IL] cache_hits={len(cached_pairs)} | to_call={len(to_call)}\n")

        if c.operator == "llm" and to_call:
            res = run_m1_llm_for_criterion(
                crit_pack,
                to_call,
                model=model,
                trunc_chars=trunc_chars,
                batch_size=batch_size,
                log=log_cb,
                progress=progress_evt,
                cancel_token=cancel_event,
                crit_idx=ci,
                crit_total=len(crits),
                block_tag="exclude",
            )
            # merge + write to cache
            for (a_id, cid), ev in res.items():
                llm_results[(a_id, cid)] = ev
                r = id_to_row.get(a_id)
                if not r:
                    continue
                th = per_row_hash.get((a_id, cid), "")
                k = _cache_key(model=model, cid=cid, a_id=a_id, text_hash=th, trunc_chars=trunc_chars)
                if use_cache:
                    cache_out[k] = dict(ev)

        elif c.operator != "llm":
            # deterministic operator (rare in IL). Mark uncertain; real deterministic logic can be added later.
            for it in to_call:
                a_id = _safe_str(it.get("a_id","")).strip()
                llm_results[(a_id, c.id)] = {"used": False, "decision":"uncertain","confidence":0.0,"field":"abstract","quote":"","span":None,"valid_quote":False}

        if progress_cb:
            progress_cb(ci / max(1, len(crits)) * 0.7)

    # Now compute per-row statuses
    for idx, r in enumerate(rows, start=1):
        if cancel_event.is_set():
            break

        a_id = getv(r, "local_id").strip()
        failed: List[str] = []
        missing: List[str] = []
        met: List[str] = []
        uncertain: List[str] = []
        evidence: Dict[str, Any] = {}

        for c in crits:
            # missing if all target fields are empty
            all_empty = True
            for t in c.targets:
                if getv(r, t).strip():
                    all_empty = False
                    break
            if all_empty:
                missing.append(c.id)
                crit_impacts[c.id]["missing"] += 1
                evidence[c.id] = {"status":"MISSING"}
                continue

            if c.operator != "llm":
                uncertain.append(c.id)
                crit_impacts[c.id]["uncertain"] += 1
                evidence[c.id] = {"status":"UNCERTAIN", "note":"non-llm operator in IL stage"}
                continue

            ev = llm_results.get((a_id, c.id), None) or {}
            decision = _safe_str(ev.get("decision","uncertain")).strip()
            try:
                confidence = float(ev.get("confidence", 0.0))
            except Exception:
                confidence = 0.0
            valid_quote = bool(ev.get("valid_quote", False))
            usable = valid_quote and (confidence >= float(c.threshold)) and (decision in {"meet","not_meet"})

            status = "UNCERTAIN"
            if usable:
                if c.ctype == "exclude":
                    if decision == "meet":
                        status = "FAILED"
                        failed.append(c.id)
                        crit_impacts[c.id]["failed"] += 1
                    elif decision == "not_meet":
                        status = "MET"
                        met.append(c.id)
                        crit_impacts[c.id]["met"] += 1
                else:
                    # (not expected in IL) treat as include
                    if decision == "meet":
                        status = "MET"
                        met.append(c.id)
                        crit_impacts[c.id]["met"] += 1
                    elif decision == "not_meet":
                        status = "FAILED"
                        failed.append(c.id)
                        crit_impacts[c.id]["failed"] += 1
            else:
                uncertain.append(c.id)
                crit_impacts[c.id]["uncertain"] += 1

            evidence[c.id] = {
                "status": status,
                "decision": decision,
                "confidence": confidence,
                "threshold": c.threshold,
                "field": _safe_str(ev.get("field","")),
                "quote": _safe_str(ev.get("quote","")),
                "quote_valid": bool(valid_quote),
                "span": ev.get("span", None),
                "used": bool(ev.get("used", False)),
            }

        if failed:
            outcome = "OUT"
        elif (len(met) == len(crits)) and not missing and not uncertain:
            outcome = "PASS_CLEAN"
        else:
            outcome = "REVIEW"

        fr = dict(r)
        fr["il_outcome"] = outcome
        fr["il_failed_ids"] = ",".join(failed)
        fr["il_missing_ids"] = ",".join(missing)
        fr["il_met_ids"] = ",".join(met)
        fr["il_uncertain_ids"] = ",".join(uncertain)
        fr["il_evidence_json"] = json.dumps(evidence, ensure_ascii=False)
        fr["il_reason_summary"] = _summarize_el_reason(outcome, failed, missing, uncertain)

        full_rows.append(fr)
        row_eval_lists.append({"failed": failed, "missing": missing, "met": met, "uncertain": uncertain})

        if outcome != "OUT":
            survivors.append(dict(r))

        counts[outcome] = counts.get(outcome, 0) + 1

        if progress_cb and idx % 25 == 0:
            # remaining 30% of progress
            progress_cb(0.7 + (idx / max(1, len(rows))) * 0.3)

    if progress_cb:
        progress_cb(1.0)

    return full_rows, survivors, counts, crit_impacts, row_eval_lists, cache_out

def _summarize_el_reason(outcome: str, failed: List[str], missing: List[str], uncertain: List[str]) -> str:
    if outcome == "OUT":
        return f"OUT: failed {', '.join(failed)}"
    if outcome == "PASS_CLEAN":
        return "PASS_CLEAN: all IL criteria MET."
    bits: List[str] = ["REVIEW:"]
    if missing:
        bits.append(f"missing {', '.join(missing)}")
    if uncertain:
        bits.append(f"uncertain {', '.join(uncertain)}")
    if not missing and not uncertain:
        bits.append("no failures")
    return " ".join(bits)


# ------------------------------ UI helpers ------------------------------------

class DataTable(ttk.Frame):
    """
    Treeview wrapper with:
    - column setup
    - click-to-sort (optional)
    - incremental rendering to keep UI responsive
    - optional double-click callback

    Backward compatible:
      - DataTable(parent, on_sort_callable, on_row_activate=...)
      - DataTable(parent, ["col1","col2",...])  # legacy usage
    """
    def __init__(
        self,
        parent,
        on_sort=None,
        on_row_activate: Optional[Callable[[Dict[str, Any]], None]] = None
    ):
        super().__init__(parent)

        # Legacy signature: second arg is actually a list of columns
        cols_seed: Optional[List[str]] = None
        if on_sort is None:
            self.on_sort = lambda _c: None
        elif callable(on_sort):
            self.on_sort = on_sort
        else:
            # treat as columns list/tuple
            if isinstance(on_sort, (list, tuple)):
                cols_seed = list(on_sort)
            self.on_sort = lambda _c: None

        self.on_row_activate = on_row_activate

        self.tree = ttk.Treeview(self, show="headings", selectmode="browse")
        self.vs = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        self.hs = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.vs.set, xscrollcommand=self.hs.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        self.vs.grid(row=0, column=1, sticky="ns")
        self.hs.grid(row=1, column=0, sticky="ew")

        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.columns: List[str] = []
        self._render_token = 0
        self._iid_to_row: Dict[str, Dict[str, Any]] = {}

        self.tree.bind("<Double-1>", self._on_double_click)

        if cols_seed:
            self.set_columns(cols_seed)

    def set_columns(self, cols: List[str]):
        self.columns = cols
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cols
        for c in cols:
            self.tree.heading(c, text=c, command=lambda col=c: self.on_sort(col))
            self.tree.column(c, width=140, minwidth=90, stretch=False)

    def clear(self):
        self._render_token += 1
        self._iid_to_row.clear()
        self.tree.delete(*self.tree.get_children())

    def render_rows_incremental(self, rows: List[Dict[str, Any]]):
        self.clear()
        token = self._render_token

        def _insert_chunk(start: int):
            if token != self._render_token:
                return
            end = min(start + RENDER_CHUNK, len(rows))
            for i in range(start, end):
                r = rows[i]
                iid = f"r{i}"
                self._iid_to_row[iid] = r
                vals = [_safe_str(r.get(c, "")) for c in self.columns]
                self.tree.insert("", "end", iid=iid, values=vals)
            if end < len(rows):
                self.after(1, lambda: _insert_chunk(end))

        self.after(0, lambda: _insert_chunk(0))

    # ---- legacy helpers used by Plugin(ttk.Frame) ----
    def set_rows(self, rows: List[Dict[str, Any]]):
        if not rows:
            self.clear()
            return
        if not self.columns:
            self.set_columns(list(rows[0].keys()))
        self.render_rows_incremental(rows)

    def get_selected_row(self) -> Optional[Dict[str, Any]]:
        sel = self.tree.selection()
        if not sel:
            return None
        return self._iid_to_row.get(sel[0])

    def _on_double_click(self, _evt):
        if not self.on_row_activate:
            return
        r = self.get_selected_row()
        if r is not None:
            self.on_row_activate(r)

def _now_stamp() -> str:
    # Same convention as EH/IH
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _export_il_xlsx(
    path: str,
    full_rows: List[Dict[str, Any]],
    survivors: List[Dict[str, Any]],
    base_header: List[str],
) -> None:
    """
    Writes a 2-sheet XLSX:
      - IL_FULL
      - IL_SURVIVORS
    Uses openpyxl (same approach as your EH/IH plugins).
    """
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required for XLSX export. Install it (pip install openpyxl)."
        ) from e

    def cell(v: Any) -> Any:
        if v is None:
            return ""
        if isinstance(v, (dict, list)):
            return json.dumps(v, ensure_ascii=False)
        return v

    wb = Workbook(write_only=True)

    # ---- headers (stable + tolerant) ----
    base = list(base_header or [])
    if "local_id" in base:
        base = ["local_id"] + [h for h in base if h != "local_id"]
    else:
        base = ["local_id"] + base

    el_cols = [
        "il_outcome",
        "il_failed_ids",
        "il_missing_ids",
        "il_met_ids",
        "il_uncertain_ids",
        "il_reason_summary",
        "il_evidence_json",
    ]

    # Full sheet header = base + IL cols + any extras present in data
    full_header = base + [c for c in el_cols if c not in base]
    if full_rows:
        extras = [k for k in full_rows[0].keys() if k not in full_header]
        full_header += extras

    # Survivors sheet header = base + any extras present in survivors (rare)
    surv_header = list(base)
    if survivors:
        extras2 = [k for k in survivors[0].keys() if k not in surv_header]
        surv_header += extras2

    # ---- sheet: IL_FULL ----
    ws1 = wb.create_sheet(title="IL_FULL")
    ws1.append(full_header)
    for r in full_rows:
        ws1.append([cell(r.get(h, "")) for h in full_header])

    # ---- sheet: IL_SURVIVORS ----
    ws2 = wb.create_sheet(title="IL_SURVIVORS")
    ws2.append(surv_header)
    for r in survivors:
        ws2.append([cell(r.get(h, "")) for h in surv_header])

    # Remove default sheet if present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)



def _find_bundle_member(zf: zipfile.ZipFile, root: str, rel: str) -> Optional[str]:
    """Return the member name for root+rel if present, else try suffix match case-insensitive."""
    exact = root + rel
    if exact in zf.namelist():
        return exact
    suffix = ("/" + rel).lower()
    for name in zf.namelist():
        if name.lower().endswith(suffix):
            return name
    return None


def _load_csv_rows_from_zip(zf: zipfile.ZipFile, member: str) -> Tuple[List[str], List[Dict[str, str]]]:
    b = _read_zip_bytes(zf, member)
    txt = _decode_bytes(b)
    return _csv_read(txt)


def _load_master_rows(zf: zipfile.ZipFile, root: str) -> Tuple[List[str], List[Dict[str, str]]]:
    """
    Best-effort: try to load the *original* A (all items) if present; else fall back to current.
    """
    candidates = [
        "data/A.csv",
        "data/original.csv",
        "data/aggregate.csv",
        "data/current.csv",
    ]
    for rel in candidates:
        mem = _find_bundle_member(zf, root, rel)
        if mem:
            try:
                return _load_csv_rows_from_zip(zf, mem)
            except Exception:
                continue
    return ([], [])


def _stage_prefix(stage: str) -> str:
    return stage.lower() + "_"


def _extract_contract_stage_row(stage: str, r: Dict[str, Any]) -> Dict[str, Any]:
    """
    Map a stage FULL row (plugin convention: {prefix}outcome, {prefix}*_ids, {prefix}reason_summary, {prefix}evidence_json)
    to contract v2 standardized columns.
    """
    pref = _stage_prefix(stage)
    a_id = _safe_str(r.get("a_id") or r.get("local_id") or r.get("id") or r.get("doi") or "")
    stage_outcome = _safe_str(r.get(pref + "outcome") or r.get("stage_outcome") or "")
    failed = _safe_str(r.get(pref + "failed_ids") or "")
    missing = _safe_str(r.get(pref + "missing_ids") or "")
    uncertain = _safe_str(r.get(pref + "uncertain_ids") or "")
    met = _safe_str(r.get(pref + "met_ids") or "")
    matched_evidence = _safe_str(r.get(pref + "evidence_json") or r.get(pref + "matched_evidence") or "")
    reason = _safe_str(r.get(pref + "reason_summary") or r.get("stage_reason_summary") or "")
    passed_to_next = "true" if stage_outcome and stage_outcome != "OUT" else "false"
    return {
        "a_id": a_id,
        "stage": stage,
        "stage_outcome": stage_outcome,
        "passed_to_next": passed_to_next,
        "failed_criteria_ids": failed,
        "missing_criteria_ids": missing,
        "uncertain_criteria_ids": uncertain,
        "met_criteria_ids": met,
        "matched_evidence": matched_evidence,
        "stage_reason_summary": reason,
        "history": "",
    }


def _compute_final_outcome(outcomes: Dict[str, str]) -> str:
    """
    Contract v2: FINAL outcome:
      - OUT if any stage outcome == OUT
      - PASS_CLEAN if ALL non-empty stage outcomes are PASS_CLEAN AND IL is PASS_CLEAN
      - REVIEW otherwise
    """
    for st in ("EH", "IH", "EL", "IL"):
        if outcomes.get(st) == "OUT":
            return "OUT"
    il_out = outcomes.get("IL", "")
    if il_out == "PASS_CLEAN" and all((outcomes.get(st, "") in {"", "PASS_CLEAN"} for st in ("EH", "IH", "EL", "IL"))):
        return "PASS_CLEAN"
    return "REVIEW"


def _build_final_report_xlsx_bytes(
    zf_in: zipfile.ZipFile,
    root: str,
    il_full_rows: List[Dict[str, Any]],
) -> bytes:
    """
    Build the contract v2 5-sheet Excel workbook:
      EH / IH / EL / IL / FINAL
    Uses stage FULL CSVs from the input bundle when available.
    """
    # Load prior stage FULL rows (best-effort)
    stage_full: Dict[str, List[Dict[str, str]]] = {}
    for st in ("EH", "IH", "EL"):
        mem = _find_bundle_member(zf_in, root, f"{REPORTS_DIR_REL}/{st}_FULL.csv")
        if mem:
            try:
                _, rows = _load_csv_rows_from_zip(zf_in, mem)
                stage_full[st] = rows
            except Exception:
                stage_full[st] = []
        else:
            stage_full[st] = []

    stage_full["IL"] = [{k: (_safe_str(v) if not isinstance(v, str) else v) for k, v in r.items()} for r in il_full_rows]

    # Master list (all items) for FINAL
    master_header, master_rows = _load_master_rows(zf_in, root)
    # Build meta lookup by a_id
    meta_by_id: Dict[str, Dict[str, str]] = {}
    for r in master_rows:
        a_id = _safe_str(r.get("a_id") or r.get("local_id") or r.get("id") or "")
        if a_id:
            meta_by_id[a_id] = r

    # Outcomes + reasons per stage
    stage_by_id: Dict[str, Dict[str, Dict[str, Any]]] = {st: {} for st in ("EH", "IH", "EL", "IL")}
    for st in ("EH", "IH", "EL", "IL"):
        for r in stage_full.get(st, []):
            a_id = _safe_str(r.get("a_id") or r.get("local_id") or r.get("id") or "")
            if not a_id:
                continue
            pref = _stage_prefix(st)
            stage_by_id[st][a_id] = {
                "outcome": _safe_str(r.get(pref + "outcome") or r.get("stage_outcome") or ""),
                "reason": _safe_str(r.get(pref + "reason_summary") or r.get("stage_reason_summary") or ""),
            }

    # Universe of ids
    all_ids = set(meta_by_id.keys())
    for st in ("EH", "IH", "EL", "IL"):
        all_ids.update(stage_by_id[st].keys())
    all_ids = sorted(all_ids, key=lambda x: (len(x), x))

    # Meta columns for FINAL (keep master header order, minus id columns)
    meta_cols = [c for c in master_header if c not in {"a_id", "local_id"}] if master_header else []

    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError("openpyxl is required to build the final report workbook.") from e

    wb = Workbook(write_only=True)

    # Stage sheets
    for st in ("EH", "IH", "EL", "IL"):
        ws = wb.create_sheet(title=st)
        ws.append(CONTRACT_STAGE_SHEET_COLS)
        for r in stage_full.get(st, []):
            row_obj = _extract_contract_stage_row(st, r)
            ws.append([row_obj.get(c, "") for c in CONTRACT_STAGE_SHEET_COLS])

    # FINAL sheet
    ws = wb.create_sheet(title="FINAL")
    final_cols = ["a_id"] + meta_cols + [
        "outcome_EH", "reason_EH",
        "outcome_IH", "reason_IH",
        "outcome_EL", "reason_EL",
        "outcome_IL", "reason_IL",
        "final_outcome", "final_reason_summary",
        "history",
    ]
    ws.append(final_cols)

    for a_id in all_ids:
        meta = meta_by_id.get(a_id, {})
        outcomes = {st: stage_by_id[st].get(a_id, {}).get("outcome", "") for st in ("EH", "IH", "EL", "IL")}
        reasons = {st: stage_by_id[st].get(a_id, {}).get("reason", "") for st in ("EH", "IH", "EL", "IL")}
        final_out = _compute_final_outcome(outcomes)
        if final_out == "OUT":
            # first stage that OUTed (in pipeline order)
            out_stage = next((st for st in ("EH", "IH", "EL", "IL") if outcomes.get(st) == "OUT"), "")
            final_reason = f"OUT at {out_stage}" if out_stage else "OUT"
        elif final_out == "PASS_CLEAN":
            final_reason = "PASS_CLEAN across all stages"
        else:
            final_reason = "Needs review after staged screening"

        row = [a_id]
        row.extend([meta.get(c, "") for c in meta_cols])
        row.extend([outcomes["EH"], reasons["EH"], outcomes["IH"], reasons["IH"], outcomes["EL"], reasons["EL"], outcomes["IL"], reasons["IL"]])
        row.extend([final_out, final_reason, ""])
        ws.append(row)

    # Ensure workbook has proper first sheet order (write_only creates default Sheet)
    # openpyxl write_only starts with a default sheet named "Sheet" sometimes; remove it if empty.
    # (In write_only mode, default may still exist.)
    try:
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 5:
            wb.remove(wb["Sheet"])
    except Exception:
        pass

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ----------------------------
# Main View
# ----------------------------

class ILView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.bundle_zip_path: Optional[str] = None
        self.bundle: Optional[BundleInfo] = None

        self.full_rows: List[Dict[str, Any]] = []
        self.survivors: List[Dict[str, str]] = []
        self.counts: Dict[str, int] = {}

        self.crit_impacts: Dict[str, Dict[str, int]] = {}
        self.row_eval_lists: List[Dict[str, List[str]]] = []

        self.cache_map: Dict[str, Dict[str, Any]] = {}

        self.active_criterion_id: Optional[str] = None

        self._worker: Optional[threading.Thread] = None
        self._cancel = threading.Event()

        self.sort_full: Tuple[Optional[str], bool] = (None, True)
        self.sort_surv: Tuple[Optional[str], bool] = (None, True)
        self.sort_crit: Tuple[Optional[str], bool] = (None, True)

        # Settings
        self.var_model = tk.StringVar(value=DEFAULT_MODEL)
        self.var_batch = tk.StringVar(value=str(DEFAULT_BATCH_SIZE))
        self.var_trunc = tk.StringVar(value=str(DEFAULT_TRUNC_CHARS))
        self.var_use_cache = tk.BooleanVar(value=DEFAULT_USE_CACHE)

        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        btn_b = ttk.Button(top, text="Load ScreenA bundle ZIP…", command=self._pick_bundle)
        btn_b.grid(row=0, column=0, padx=(0, 8), pady=2, sticky="w")

        self.lbl_bundle = ttk.Label(top, text="(no bundle loaded)")
        self.lbl_bundle.grid(row=0, column=1, sticky="w")

        self.lbl_bundle_meta = ttk.Label(top, text="")
        self.lbl_bundle_meta.grid(row=1, column=1, sticky="w")

        actions = ttk.Frame(top)
        actions.grid(row=0, column=2, rowspan=2, padx=(10, 0), sticky="e")

        self.btn_run = ttk.Button(actions, text="Run IL", command=self._run_clicked, state="disabled")
        self.btn_run.grid(row=0, column=0, padx=4, pady=2, sticky="e")

        self.btn_cancel = ttk.Button(actions, text="Cancel", command=self._cancel_run, state="disabled")
        self.btn_cancel.grid(row=1, column=0, padx=4, pady=2, sticky="e")

        self.btn_export = ttk.Button(actions, text="Export XLSX…", command=self._export_clicked, state="disabled")
        self.btn_export.grid(row=0, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_err = ttk.Button(actions, text="Export input_errors.csv…", command=self._export_errors_clicked, state="disabled")
        self.btn_export_err.grid(row=1, column=1, padx=4, pady=2, sticky="e")

        self.btn_export_bundle = ttk.Button(actions, text="Export next bundle ZIP…", command=self._export_bundle_clicked, state="disabled")
        self.btn_export_bundle.grid(row=0, column=2, padx=4, pady=2, sticky="e")

        self.btn_export_final = ttk.Button(actions, text="Export ScreenA_Report.xlsx…", command=self._export_final_clicked, state="disabled")
        self.btn_export_final.grid(row=0, column=3, padx=4, pady=2, sticky="e")

        # API key indicator (IL requires OpenAI)
        self.lbl_key = ttk.Label(actions, text="")
        self.lbl_key.grid(row=1, column=3, padx=4, pady=2, sticky="e")
        self._refresh_key_label()

        top.columnconfigure(1, weight=1)

        prog = ttk.Frame(self)
        prog.pack(fill="x", padx=10, pady=(0, 8))

        self.pbar = ttk.Progressbar(prog, orient="horizontal", mode="determinate")
        self.pbar.pack(fill="x", expand=True, side="left")

        self.lbl_status = ttk.Label(prog, text="Ready.")
        self.lbl_status.pack(side="left", padx=10)

        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        left = ttk.Frame(paned)
        right = ttk.Frame(paned)
        paned.add(left, weight=1)
        paned.add(right, weight=4)

        crit_box = ttk.Labelframe(left, text="IL Criteria (read-only)")
        crit_box.pack(fill="both", expand=True)

        self.criteria_table = DataTable(
            crit_box,
            on_sort=self._sort_criteria_table,
            on_row_activate=self._on_criterion_activated,
        )
        self.criteria_table.pack(fill="both", expand=True, padx=6, pady=6)

        cf = ttk.Frame(left)
        cf.pack(fill="x", pady=(6, 0))
        self.lbl_crit_filter = ttk.Label(cf, text="Criterion filter: (none)")
        self.lbl_crit_filter.pack(side="left")
        self.btn_clear_filter = ttk.Button(cf, text="Clear filter", command=self._clear_criterion_filter, state="disabled")
        self.btn_clear_filter.pack(side="right")

        settings = ttk.Labelframe(left, text="IL Settings")
        settings.pack(fill="x", pady=(6, 0))

        ttk.Label(settings, text="Model").grid(row=0, column=0, sticky="w", padx=6, pady=2)
        ttk.Entry(settings, textvariable=self.var_model, width=24).grid(row=0, column=1, sticky="ew", padx=6, pady=2)

        ttk.Label(settings, text="Batch size").grid(row=1, column=0, sticky="w", padx=6, pady=2)
        ttk.Entry(settings, textvariable=self.var_batch, width=10).grid(row=1, column=1, sticky="w", padx=6, pady=2)

        ttk.Label(settings, text="Trunc chars").grid(row=2, column=0, sticky="w", padx=6, pady=2)
        ttk.Entry(settings, textvariable=self.var_trunc, width=10).grid(row=2, column=1, sticky="w", padx=6, pady=2)

        ttk.Checkbutton(settings, text="Use cache (bundle cache/IL_cache.jsonl)", variable=self.var_use_cache).grid(row=3, column=0, columnspan=2, sticky="w", padx=6, pady=2)

        settings.columnconfigure(1, weight=1)

        warn_box = ttk.Labelframe(left, text="Notes / warnings")
        warn_box.pack(fill="both", expand=False, pady=(6, 0))

        self.txt_warn = tk.Text(warn_box, height=8, wrap="word")
        self.txt_warn.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_warn.configure(state="disabled")

        nb = ttk.Notebook(right)
        nb.pack(fill="both", expand=True)

        tab_full = ttk.Frame(nb)
        tab_surv = ttk.Frame(nb)
        tab_log = ttk.Frame(nb)
        nb.add(tab_full, text="IL Full report")
        nb.add(tab_surv, text="IL Survivors")
        nb.add(tab_log, text="Log")

        self.full_table = DataTable(tab_full, on_sort=self._sort_full_table, on_row_activate=self._open_row_detail_modal)
        self.full_table.pack(fill="both", expand=True, padx=6, pady=6)

        self.surv_table = DataTable(tab_surv, on_sort=self._sort_surv_table, on_row_activate=self._open_row_detail_modal)
        self.surv_table.pack(fill="both", expand=True, padx=6, pady=6)

        self.txt_log = tk.Text(tab_log, wrap="word")
        self.txt_log.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_log.configure(state="disabled")

        self.lbl_counts = ttk.Label(self, text="")
        self.lbl_counts.pack(fill="x", padx=10, pady=(0, 10))

    # -------- helpers --------

    def _refresh_key_label(self):
        self.lbl_key.configure(text=("OPENAI_API_KEY ✓" if _has_openai_key() else "OPENAI_API_KEY ✗"))

    def _log(self, msg: str) -> None:
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", msg)
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")

    def _set_warnings(self, lines: Sequence[str]) -> None:
        self.txt_warn.configure(state="normal")
        self.txt_warn.delete("1.0", "end")
        self.txt_warn.insert("end", "\n".join(lines) if lines else "(none)")
        self.txt_warn.configure(state="disabled")

    def _refresh_counts_label(self):
        if not self.bundle:
            self.lbl_counts.configure(text="")
            return
        pr = self.bundle.parse
        msg = f"Integral rows: {len(pr.rows)} | Skipped invalid: {len(pr.skipped)}"
        if self.counts:
            msg += (
                f" | OUT: {self.counts.get('OUT',0)}"
                f" | PASS_CLEAN: {self.counts.get('PASS_CLEAN',0)}"
                f" | REVIEW: {self.counts.get('REVIEW',0)}"
            )
        self.lbl_counts.configure(text=msg)

    def _set_controls_running(self, running: bool) -> None:
        self.btn_cancel.configure(state=("normal" if running else "disabled"))
        self.btn_run.configure(state=("disabled" if running else ("normal" if self.bundle_zip_path else "disabled")))
        self.btn_export.configure(state=("disabled" if running else ("normal" if self.full_rows else "disabled")))
        self.btn_export_err.configure(state=("disabled" if running else ("normal" if (self.bundle and self.bundle.parse.skipped) else "disabled")))
        self.btn_export_bundle.configure(state=("disabled" if running else ("normal" if self.full_rows else "disabled")))
        self.btn_export_final.configure(state=("disabled" if running else ("normal" if self.full_rows else "disabled")))

    def _sorted_rows(self, rows: List[Dict[str, Any]], col: str, asc: bool) -> List[Dict[str, Any]]:
        def _key(r: Dict[str, Any]):
            v = r.get(col, "")
            s = _safe_str(v)
            # numeric sort if it looks like a number
            try:
                if s.strip() == "":
                    return (1, 0.0, s)
                return (0, float(s), s)
            except Exception:
                return (0, 0.0, s.lower())
        return sorted(rows, key=_key, reverse=(not asc))

    # -------- bundle load --------

    def _pick_bundle(self):
        p = filedialog.askopenfilename(
            title="Select ScreenA bundle ZIP",
            filetypes=[("ZIP", "*.zip"), ("All files", "*.*")],
        )
        if not p:
            return
        self.bundle_zip_path = p
        self.lbl_bundle.configure(text=Path(p).name)
        self._load_bundle_inputs()

    def _load_bundle_inputs(self):
        if not self.bundle_zip_path:
            return

        warns: List[str] = []
        self.full_rows = []
        self.survivors = []
        self.counts = {}
        self.crit_impacts = {}
        self.row_eval_lists = []
        self.active_criterion_id = None
        self.lbl_crit_filter.configure(text="Criterion filter: (none)")
        self.btn_clear_filter.configure(state="disabled")

        self._refresh_key_label()

        try:
            self.bundle = _load_bundle(self.bundle_zip_path)
        except Exception as e:
            self.bundle = None
            messagebox.showerror("Bundle load failed", str(e))
            return

        # load cache if present
        self.cache_map = {}
        try:
            with zipfile.ZipFile(self.bundle_zip_path, "r") as zf:
                cache_member = self.bundle.root + IL_CACHE_REL
                if cache_member in zf.namelist():
                    self.cache_map = _load_cache_from_jsonl(_decode_bytes(_read_zip_bytes(zf, cache_member)))
        except Exception:
            self.cache_map = {}

        m = self.bundle.manifest
        schema = _safe_str(m.get("bundle_schema", m.get("schema", ""))).strip() or "unknown"
        created_at = _safe_str(m.get("created_at", "")).strip()
        created_by = _safe_str(m.get("created_by", "")).strip()
        stages = ((m.get("pipeline", {}) or {}).get("stages", None)) or ((m.get("pipeline_state", {}) or {}).get("stages", None)) or {}
        st_il = _safe_str(stages.get("IL", "")).strip() or "unknown"
        self.lbl_bundle_meta.configure(text=f"schema={schema} | created_at={created_at} | created_by={created_by} | IL={st_il}")

        # warnings: duplicates (they were moved to parse.skipped during load)
        dup_ids = []
        for e in (self.bundle.parse.skipped or []):
            if _safe_str(e.get("reason", "")).strip() == "duplicate local_id":
                lid = _safe_str((e.get("row") or {}).get("local_id", "")).strip()
                if lid:
                    dup_ids.append(lid)
        # unique sample
        sample = []
        seen = set()
        for x in dup_ids:
            if x not in seen:
                sample.append(x); seen.add(x)
            if len(sample) >= 10:
                break
        if dup_ids:
            warns.append(f"[data] duplicate local_id detected (n={len(set(dup_ids))}), sample={', '.join(sample)}")

        # refresh criteria table + notes
        self._refresh_criteria_table(pre_run=True)
        if self.bundle.criteria and self.bundle.criteria.warnings:
            warns.extend(self.bundle.criteria.warnings)
        self._set_warnings(warns)

        # clear report tables
        self.full_table.clear()
        self.surv_table.clear()
        self._refresh_counts_label()

        self.btn_run.configure(state="normal" if _has_openai_key() else "disabled")
        self.btn_export.configure(state="disabled")
        self.btn_export_bundle.configure(state="disabled")
        self.btn_export_err.configure(state=("normal" if self.bundle.parse.skipped else "disabled"))

        self.lbl_status.configure(text="Ready.")

    def _detect_duplicate_local_ids(self, rows: List[Dict[str, str]]) -> Tuple[int, List[str]]:
        seen = set()
        dups = []
        for r in rows:
            lid = _safe_str(r.get("local_id", "")).strip()
            if not lid:
                continue
            if lid in seen:
                dups.append(lid)
            else:
                seen.add(lid)
        uniq = []
        s2 = set()
        for x in dups:
            if x not in s2:
                uniq.append(x)
                s2.add(x)
        return len(uniq), uniq[:10]

    # -------- criteria table --------

    def _refresh_criteria_table(self, pre_run: bool):
        crits = self.bundle.criteria.criteria if (self.bundle and self.bundle.criteria) else []

        cols = ["id", "type", "targets", "operator", "what", "threshold", "status", "notes"]
        if not pre_run:
            cols += ["n_failed", "n_missing", "n_met", "n_uncertain"]

        rows: List[Dict[str, Any]] = []
        header_set = set(self.bundle.parse.header) if (self.bundle and self.bundle.parse) else set()
        header_set_l = {h.lower() for h in header_set}

        for c in crits:
            status = "OK"
            notes = ""

            if not c.targets:
                status = "WARNING"
                notes = "missing target -> treated as MISSING (REVIEW)"
            elif header_set and not any((t or "").lower() in header_set_l for t in c.targets):
                status = "WARNING"
                notes = f"missing column(s): {', '.join(c.targets)} -> treated as MISSING (REVIEW)"

            op = (c.operator or "").strip().lower()
            if op != "llm":
                status = "WARNING"
                notes = (notes + " | " if notes else "") + f"operator '{op}' treated as UNCERTAIN in IL"

            d: Dict[str, Any] = {
                "id": c.id,
                "type": c.ctype,
                "targets": ",".join(c.targets),
                "operator": c.operator,
                "what": c.what_raw,
                "threshold": _safe_str(c.threshold),
                "status": status,
                "notes": notes,
            }

            if not pre_run:
                imp = self.crit_impacts.get(c.id, {"failed": 0, "missing": 0, "met": 0, "uncertain": 0})
                d["n_failed"] = str(imp.get("failed", 0))
                d["n_missing"] = str(imp.get("missing", 0))
                d["n_met"] = str(imp.get("met", 0))
                d["n_uncertain"] = str(imp.get("uncertain", 0))

            rows.append(d)

        self.criteria_table.set_columns(cols)
        col, asc = self.sort_crit
        if col:
            rows = self._sorted_rows(rows, col, asc)
        self.criteria_table.render_rows_incremental(rows)

    def _on_criterion_activated(self, row: Dict[str, str]):
        cid = _safe_str(row.get("id", "")).strip()
        if not cid:
            return
        self.active_criterion_id = cid
        self.lbl_crit_filter.configure(text=f"Criterion filter: {cid}")
        self.btn_clear_filter.configure(state="normal")
        self._refresh_reports_view()

    def _clear_criterion_filter(self):
        self.active_criterion_id = None
        self.lbl_crit_filter.configure(text="Criterion filter: (none)")
        self.btn_clear_filter.configure(state="disabled")
        self._refresh_reports_view()

    def _sort_criteria_table(self, col: str):
        cur_col, asc = self.sort_crit
        if cur_col == col:
            asc = not asc
        else:
            asc = True
        self.sort_crit = (col, asc)
        self._refresh_criteria_table(pre_run=(not bool(self.full_rows)))

    # -------- reports --------

    def _refresh_reports_view(self):
        if not self.bundle:
            return

        full_cols = list(self.bundle.parse.header) + [
            "il_outcome", "il_failed_ids", "il_missing_ids", "il_met_ids", "il_uncertain_ids", "il_reason_summary"
        ]
        surv_cols = list(self.bundle.parse.header)

        full_view = self.full_rows
        if self.active_criterion_id:
            cid = self.active_criterion_id
            filtered = []
            for i, r in enumerate(self.full_rows):
                ev = self.row_eval_lists[i] if i < len(self.row_eval_lists) else {"failed": [], "missing": [], "met": [], "uncertain": []}
                if (cid in ev.get("failed", [])) or (cid in ev.get("missing", [])) or (cid in ev.get("met", [])) or (cid in ev.get("uncertain", [])):
                    filtered.append(r)
            full_view = filtered

        surv_view = self.survivors
        if self.active_criterion_id and self.full_rows:
            cid = self.active_criterion_id
            touched_survivor_ids = set()
            for i, r in enumerate(self.full_rows):
                if r.get("il_outcome") == "OUT":
                    continue
                ev = self.row_eval_lists[i]
                if (cid in ev.get("failed", [])) or (cid in ev.get("missing", [])) or (cid in ev.get("met", [])) or (cid in ev.get("uncertain", [])):
                    touched_survivor_ids.add(_safe_str(r.get("local_id", "")).strip())
            surv_view = [r for r in self.survivors if _safe_str(r.get("local_id", "")).strip() in touched_survivor_ids]

        col, asc = self.sort_full
        if col and full_view:
            full_view = self._sorted_rows(full_view, col, asc)

        col2, asc2 = self.sort_surv
        if col2 and surv_view:
            surv_view = self._sorted_rows(surv_view, col2, asc2)

        self.full_table.set_columns(full_cols)
        self.full_table.render_rows_incremental(full_view)

        self.surv_table.set_columns(surv_cols)
        self.surv_table.render_rows_incremental(surv_view)

    def _sort_full_table(self, col: str):
        cur_col, asc = self.sort_full
        if cur_col == col:
            asc = not asc
        else:
            asc = True
        self.sort_full = (col, asc)
        self._refresh_reports_view()

    def _sort_surv_table(self, col: str):
        cur_col, asc = self.sort_surv
        if cur_col == col:
            asc = not asc
        else:
            asc = True
        self.sort_surv = (col, asc)
        self._refresh_reports_view()

    # -------- Row detail modal --------

    def _open_row_detail_modal(self, row: Dict[str, Any]):
        if not self.bundle:
            return

        win = tk.Toplevel(self)
        win.title("IL Row details")
        win.geometry("950x650")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        lid = _safe_str(row.get("local_id", "")).strip()
        title = _safe_str(row.get("title", "")).strip() or _safe_str(row.get("Title", "")).strip()

        top = ttk.Frame(win)
        top.pack(fill="x", padx=10, pady=10)

        ttk.Label(top, text=f"local_id: {lid}").pack(anchor="w")
        if title:
            ttk.Label(top, text=f"title: {title[:250]}").pack(anchor="w")

        outcome = _safe_str(row.get("il_outcome", "")).strip()
        if outcome:
            ttk.Label(top, text=f"IL outcome: {outcome}").pack(anchor="w")
            rs = _safe_str(row.get("il_reason_summary", "")).strip()
            if rs:
                ttk.Label(top, text=f"summary: {rs}").pack(anchor="w")

        box = ttk.Labelframe(win, text="Per-criterion evidence (from il_evidence_json)")
        box.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        cols = ["cid", "type", "operator", "targets", "what", "threshold", "status", "decision", "confidence", "field", "quote_valid", "quote"]
        table = DataTable(box, on_sort=lambda _c: None, on_row_activate=None)
        table.pack(fill="both", expand=True, padx=6, pady=6)
        table.set_columns(cols)

        # parse evidence dict
        ev_raw = _safe_str(row.get("il_evidence_json", "{}"))
        try:
            evj = json.loads(ev_raw) if ev_raw.strip() else {}
        except Exception:
            evj = {}

        detail_rows: List[Dict[str, Any]] = []
        for c in self.bundle.criteria.criteria:
            obj = evj.get(c.id, {}) if isinstance(evj, dict) else {}
            status = _safe_str(obj.get("status", ""))
            decision = _safe_str(obj.get("decision", ""))
            confidence = _safe_str(obj.get("confidence", ""))
            field = _safe_str(obj.get("field", ""))
            quote = _safe_str(obj.get("quote", ""))
            qv = _safe_str(obj.get("quote_valid", obj.get("quote_valid", "")))
            if isinstance(obj.get("quote_valid", None), bool):
                qv = "True" if obj.get("quote_valid") else "False"

            detail_rows.append({
                "cid": c.id,
                "type": c.ctype,
                "operator": c.operator,
                "targets": ",".join(c.targets),
                "what": c.what_raw,
                "threshold": _safe_str(c.threshold),
                "status": status,
                "decision": decision,
                "confidence": confidence,
                "field": field,
                "quote_valid": qv,
                "quote": quote,
            })

        table.render_rows_incremental(detail_rows)

        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=10, pady=(0, 10))
        ttk.Button(btns, text="Close", command=win.destroy).pack(side="right")
        win.bind("<Escape>", lambda _e: win.destroy())

    # -------- run --------

    def _run_clicked(self):
        if not self.bundle or self._worker:
            return
        if not _has_openai_key():
            messagebox.showerror("Missing OPENAI_API_KEY", "IL uses the OpenAI API. Set OPENAI_API_KEY in your environment (.env).")
            return

        # parse settings
        model = (self.var_model.get() or DEFAULT_MODEL).strip()
        try:
            batch_size = int(self.var_batch.get())
        except Exception:
            batch_size = DEFAULT_BATCH_SIZE
        try:
            trunc_chars = int(self.var_trunc.get())
        except Exception:
            trunc_chars = DEFAULT_TRUNC_CHARS
        use_cache = bool(self.var_use_cache.get())

        self._cancel.clear()
        self.pbar.configure(value=0.0, maximum=100.0)
        self.lbl_status.configure(text="Running IL…")
        self._set_controls_running(True)

        def progress_cb(frac: float):
            self.after(0, lambda: self.pbar.configure(value=max(0.0, min(100.0, frac * 100.0))))

        def progress_evt(evt: Dict[str, Any]):
            kind = _safe_str(evt.get("kind", ""))
            sub = _safe_str(evt.get("sub", ""))
            if kind == "l_batch" and sub == "sending":
                bi = evt.get("batch_idx"); bt = evt.get("batch_total")
                ci = evt.get("crit_idx"); ct = evt.get("crit_total")
                self.after(0, lambda: self.lbl_status.configure(text=f"LLM: criterion {ci}/{ct} batch {bi}/{bt}…"))

        def work():
            try:
                full_rows, survivors, counts, crit_impacts, row_eval_lists, cache_out = run_il_screen(
                    self.bundle.parse,
                    self.bundle.criteria,
                    model=model,
                    trunc_chars=trunc_chars,
                    batch_size=batch_size,
                    use_cache=use_cache,
                    cache_in=self.cache_map if use_cache else {},
                    cancel_event=self._cancel,  # ✅ ILView uses self._cancel
                    log_cb=lambda s: self.after(0, (lambda ss=s: self._log(ss))),
                    progress_cb=progress_cb,
                    progress_evt=progress_evt,
                )

                if self._cancel.is_set():
                    self.after(0, lambda: self._log("\n[CANCELLED]\n"))
                    self.after(0, lambda: self.lbl_status.configure(text="Cancelled."))
                    return

                self.full_rows = full_rows
                self.survivors = survivors
                self.counts = counts
                self.crit_impacts = crit_impacts
                self.row_eval_lists = row_eval_lists
                self.cache_map = cache_out

                # ✅ refresh views + counts + criteria impacts
                self.after(0, self._refresh_reports_view)
                self.after(0, lambda: self._refresh_criteria_table(pre_run=False))
                self.after(0, self._refresh_counts_label)
                self.after(0, lambda: self.lbl_status.configure(text="IL done."))

            except Exception as e:
                self.after(0, lambda m=str(e): messagebox.showerror("IL run failed", m))
                self.after(0, lambda: self.lbl_status.configure(text="IL failed."))
            finally:
                self.after(0, lambda: self._set_controls_running(False))
                # clear worker flag so you can run again
                self._worker = None

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    def _cancel_run(self):
        self._cancel.set()
        self._log("\n[Cancel requested]\n")
        self.lbl_status.configure(text="Cancelling…")

    # -------- Export --------

    def _export_clicked(self):
        if not self.bundle:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        if not self.full_rows:
            messagebox.showwarning("Nothing to export", "Run IL first.")
            return

        default_name = f"{_now_stamp()}_IL_reports.xlsx"
        p = filedialog.asksaveasfilename(
            title="Save IL reports",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not p:
            return

        try:
            _export_il_xlsx(p, self.full_rows, self.survivors, self.bundle.parse.header)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_final_clicked(self):
        if not self.bundle or not self.full_rows:
            messagebox.showinfo("Not ready", "Run IL first.")
            return

        default_name = f"{_now_stamp()}_{FINAL_REPORT_NAME}"
        p = filedialog.asksaveasfilename(
            title="Save ScreenA_Report.xlsx",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not p:
            return

        try:
            with zipfile.ZipFile(self.bundle.zip_path, "r") as zf:
                wb_bytes = _build_final_report_xlsx_bytes(zf, self.bundle.root, self.full_rows)
            Path(p).write_bytes(wb_bytes)
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_errors_clicked(self):
        if not self.bundle:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        if not self.bundle.parse.skipped:
            messagebox.showinfo("No input errors", "No skipped/invalid rows were recorded for this bundle.")
            return

        default_name = f"{_now_stamp()}_input_errors.csv"
        p = filedialog.asksaveasfilename(
            title="Save input_errors.csv",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV", "*.csv")],
        )
        if not p:
            return

        try:
            with open(p, "w", encoding="utf-8", newline="") as f:
                w = csv.writer(f, lineterminator="\n")
                w.writerow(["reason", "row_json"])
                for e in self.bundle.parse.skipped:
                    w.writerow([_safe_str(e.get("reason", "")), json.dumps(e.get("row", {}), ensure_ascii=False)])
        except Exception as e:
            messagebox.showerror("Export failed", str(e))
            return

        self.lbl_status.configure(text=f"Exported: {Path(p).name}")

    def _export_bundle_clicked(self):
        if not self.bundle:
            messagebox.showwarning("Nothing to export", "Load a bundle first.")
            return
        if not self.full_rows:
            messagebox.showwarning("Nothing to export", "Run IL first.")
            return

        default_name = f"{_now_stamp()}_post_IL_bundle.zip"
        out_zip = filedialog.asksaveasfilename(
            title="Save next bundle ZIP (post-IL)",
            defaultextension=".zip",
            initialfile=default_name,
            filetypes=[("ZIP", "*.zip")],
        )
        if not out_zip:
            return

        try:
            self._build_next_bundle_zip(out_zip)
        except Exception as e:
            messagebox.showerror("Bundle export failed", str(e))
            return

        self.lbl_status.configure(text=f"Bundle exported: {Path(out_zip).name}")

    def _build_next_bundle_zip(self, out_zip: str) -> None:
        """Create a new bundle zip where data/current.csv becomes the IL survivors.

        Keeps other files from the input bundle, updates manifest pipeline stage, adds reports, carries input_errors,
        and writes cache/IL_cache.jsonl when enabled.
        """
        assert self.bundle is not None

        with zipfile.ZipFile(self.bundle.zip_path, "r") as zf_in:
            members = zf_in.namelist()
            root = self.bundle.root

            manifest = dict(self.bundle.manifest)

            def _set_stage(m: Dict[str, Any], key: str, value: str):
                if "pipeline" in m and isinstance(m.get("pipeline"), dict):
                    m["pipeline"].setdefault("stages", {})
                    m["pipeline"]["stages"][key] = value
                if "pipeline_state" in m and isinstance(m.get("pipeline_state"), dict):
                    m["pipeline_state"].setdefault("stages", {})
                    m["pipeline_state"]["stages"][key] = value

            _set_stage(manifest, "IL", "done")
            manifest["updated_at"] = datetime.utcnow().isoformat() + "Z"

            with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf_out:
                skip_exact = {
                    root + "data/current.csv",
                    root + f"{REPORTS_DIR_REL}/IL_FULL.csv",
                    root + f"{REPORTS_DIR_REL}/IL_SURVIVORS.csv",
                    root + FINAL_REPORT_REL,
                    root + "data/input_errors.csv",
                    root + IL_CACHE_REL,
                    root + "manifest.json",
                }

                for m in members:
                    if m.endswith("/"):
                        continue
                    if m in skip_exact:
                        continue
                    zf_out.writestr(m, _read_zip_bytes(zf_in, m))

                zf_out.writestr(root + "manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

                # survivors to data/current.csv
                header = list(self.bundle.parse.header)
                if "local_id" not in header:
                    header = ["local_id"] + header

                buf = io.StringIO()
                w = csv.DictWriter(buf, fieldnames=header, extrasaction="ignore")
                w.writeheader()
                for r in self.survivors:
                    w.writerow({k: r.get(k, "") for k in header})
                zf_out.writestr(root + "data/current.csv", buf.getvalue())

                # reports
                header_full = [
                    *header,
                    "il_outcome", "il_failed_ids", "il_missing_ids", "il_met_ids", "il_uncertain_ids", "il_reason_summary", "il_evidence_json",
                ]

                buf2 = io.StringIO()
                w2 = csv.DictWriter(buf2, fieldnames=header_full, extrasaction="ignore")
                w2.writeheader()
                for r in self.full_rows:
                    w2.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in header_full})
                zf_out.writestr(root + f"{REPORTS_DIR_REL}/IL_FULL.csv", buf2.getvalue())

                buf3 = io.StringIO()
                w3 = csv.DictWriter(buf3, fieldnames=header, extrasaction="ignore")
                w3.writeheader()
                for r in self.survivors:
                    w3.writerow({k: r.get(k, "") for k in header})
                zf_out.writestr(root + f"{REPORTS_DIR_REL}/IL_SURVIVORS.csv", buf3.getvalue())

                # FINAL contract workbook (EH/IH/EL/IL/FINAL)
                wb_bytes = _build_final_report_xlsx_bytes(zf_in, root, self.full_rows)
                zf_out.writestr(root + FINAL_REPORT_REL, wb_bytes)

                # input errors
                if self.bundle.parse.skipped:
                    buf4 = io.StringIO()
                    err_header = ["reason", "row_json"]
                    w4 = csv.DictWriter(buf4, fieldnames=err_header)
                    w4.writeheader()
                    for e in self.bundle.parse.skipped:
                        w4.writerow({"reason": _safe_str(e.get("reason", "")), "row_json": json.dumps(e.get("row", {}), ensure_ascii=False)})
                    zf_out.writestr(root + "data/input_errors.csv", buf4.getvalue())

                # cache
                if self.var_use_cache.get():
                    zf_out.writestr(root + IL_CACHE_REL, _dump_cache_to_jsonl(self.cache_map))


class StandaloneILPlugin(ttk.Frame):
    def __init__(self, parent: tk.Misc):
        super().__init__(parent)
        self.bundle_zip_path: Optional[str] = None
        self.bundle: Optional[BundleInfo] = None

        self.full_rows: List[Dict[str, Any]] = []
        self.survivors: List[Dict[str, str]] = []
        self.counts: Dict[str, int] = {}
        self.crit_impacts: Dict[str, Dict[str, int]] = {}
        self.row_eval_lists: List[Dict[str, List[str]]] = []
        self.cache_map: Dict[str, Dict[str, Any]] = {}

        self.cancel_event = threading.Event()
        self.worker: Optional[threading.Thread] = None

        # --- Top controls
        top = ttk.Frame(self)
        top.pack(fill="x", padx=8, pady=6)

        self.btn_load = ttk.Button(top, text="Load Bundle ZIP", command=self.on_load_bundle)
        self.btn_load.pack(side="left")

        ttk.Separator(top, orient="vertical").pack(side="left", padx=6, fill="y")

        ttk.Label(top, text="Model:").pack(side="left")
        self.var_model = tk.StringVar(value=DEFAULT_MODEL)
        ttk.Entry(top, textvariable=self.var_model, width=18).pack(side="left", padx=(2, 8))

        ttk.Label(top, text="Batch:").pack(side="left")
        self.var_batch = tk.IntVar(value=DEFAULT_BATCH_SIZE)
        ttk.Spinbox(top, from_=1, to=500, textvariable=self.var_batch, width=6).pack(side="left", padx=(2, 8))

        ttk.Label(top, text="Trunc:").pack(side="left")
        self.var_trunc = tk.IntVar(value=DEFAULT_TRUNC_CHARS)
        ttk.Spinbox(top, from_=200, to=5000, increment=50, textvariable=self.var_trunc, width=7).pack(side="left", padx=(2, 8))

        self.var_use_cache = tk.BooleanVar(value=DEFAULT_USE_CACHE)
        ttk.Checkbutton(top, text="Use cache", variable=self.var_use_cache).pack(side="left", padx=(2, 8))

        self.btn_run = ttk.Button(top, text="Run IL", command=self.on_run_il, state="disabled")
        self.btn_run.pack(side="left")

        self.btn_cancel = ttk.Button(top, text="Cancel", command=self.on_cancel, state="disabled")
        self.btn_cancel.pack(side="left", padx=(6, 0))

        ttk.Separator(top, orient="vertical").pack(side="left", padx=6, fill="y")

        self.btn_export_csv = ttk.Button(top, text="Export IL_FULL.csv", command=self.on_export_csv, state="disabled")
        self.btn_export_csv.pack(side="left")
        self.btn_export_xlsx = ttk.Button(top, text="Export IL_FULL.xlsx", command=self.on_export_xlsx, state="disabled")
        self.btn_export_xlsx.pack(side="left", padx=(6, 0))
        self.btn_next_bundle = ttk.Button(top, text="Build next bundle ZIP", command=self.on_build_next_bundle, state="disabled")
        self.btn_next_bundle.pack(side="left", padx=(6, 0))

        self.lbl_key = ttk.Label(top, text="")
        self.lbl_key.pack(side="right")
        self._refresh_key_label()

        # --- Meta + progress
        meta = ttk.Frame(self)
        meta.pack(fill="x", padx=8)
        self.lbl_bundle_meta = ttk.Label(meta, text="")
        self.lbl_bundle_meta.pack(side="left")

        self.lbl_counts = ttk.Label(meta, text="")
        self.lbl_counts.pack(side="right")

        self.pbar = ttk.Progressbar(self, orient="horizontal", mode="determinate")
        self.pbar.pack(fill="x", padx=8, pady=(4, 6))

        # --- Main split: left criteria, right table/log
        main = ttk.Panedwindow(self, orient="horizontal")
        main.pack(fill="both", expand=True, padx=8, pady=6)

        # left panel
        left = ttk.Frame(main, width=320)
        main.add(left, weight=0)

        ttk.Label(left, text="IL Criteria (enabled)").pack(anchor="w")
        self.lst_crit = tk.Listbox(left, height=14)
        self.lst_crit.pack(fill="both", expand=False, pady=(2, 6))
        self.lst_crit.bind("<Double-Button-1>", self.on_criterion_doubleclick)

        ttk.Label(left, text="Warnings").pack(anchor="w")
        self.txt_warn = tk.Text(left, height=6, wrap="word", state="disabled")
        self.txt_warn.pack(fill="both", expand=True, pady=(2, 6))

        # right panel
        right = ttk.Frame(main)
        main.add(right, weight=1)

        # table
        self.table_columns = [
            "local_id", "title", "year",
            "il_outcome", "il_failed_ids", "il_missing_ids", "il_uncertain_ids",
        ]
        self.table = DataTable(right, self.table_columns)
        self.table.pack(fill="both", expand=True)
        self.table.tree.bind("<Double-Button-1>", self.on_row_doubleclick)

        # log
        ttk.Label(right, text="Log").pack(anchor="w", pady=(6, 0))
        self.txt_log = tk.Text(right, height=8, wrap="word", state="disabled")
        self.txt_log.pack(fill="x", expand=False)

    # -------- UI helpers
    def _refresh_key_label(self):
        self.lbl_key.configure(text=("OPENAI_API_KEY ✓" if _has_openai_key() else "OPENAI_API_KEY ✗"))

    def _log(self, msg: str):
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", msg)
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")

    def _set_warn(self, lines: List[str]):
        self.txt_warn.configure(state="normal")
        self.txt_warn.delete("1.0", "end")
        self.txt_warn.insert("end", "\n".join(lines) if lines else "(none)")
        self.txt_warn.configure(state="disabled")

    def _refresh_counts_label(self):
        if not self.bundle:
            self.lbl_counts.configure(text="")
            return
        pr = self.bundle.parse
        msg = f"Rows: {len(pr.rows)} | Skipped: {len(pr.skipped)}"
        if self.counts:
            msg += (
                f" | OUT: {self.counts.get('OUT',0)}"
                f" | PASS_CLEAN: {self.counts.get('PASS_CLEAN',0)}"
                f" | REVIEW: {self.counts.get('REVIEW',0)}"
            )
        self.lbl_counts.configure(text=msg)

    def _set_controls_running(self, running: bool):
        self.btn_cancel.configure(state=("normal" if running else "disabled"))
        self.btn_run.configure(state=("disabled" if running else ("normal" if self.bundle else "disabled")))
        self.btn_load.configure(state=("disabled" if running else "normal"))
        self.btn_export_csv.configure(state=("disabled" if running else ("normal" if self.full_rows else "disabled")))
        self.btn_export_xlsx.configure(state=("disabled" if running else ("normal" if self.full_rows else "disabled")))
        self.btn_next_bundle.configure(state=("disabled" if running else ("normal" if self.full_rows else "disabled")))

    # -------- bundle load
    def on_load_bundle(self):
        p = filedialog.askopenfilename(
            title="Select ScreenA bundle ZIP",
            filetypes=[("ZIP files", "*.zip"), ("All files", "*.*")]
        )
        if not p:
            return
        self.bundle_zip_path = p
        self._refresh_key_label()

        try:
            self.bundle = _load_bundle(p)
        except Exception as e:
            self.bundle = None
            messagebox.showerror("Bundle load failed", str(e))
            return

        # read cache from bundle if present
        self.cache_map = {}
        try:
            with zipfile.ZipFile(p, "r") as zf:
                root = self.bundle.root
                mem = zf.namelist()
                cache_member = root + IL_CACHE_REL
                if cache_member in mem:
                    self.cache_map = _load_cache_from_jsonl(_decode_bytes(_read_zip_bytes(zf, cache_member)))
        except Exception:
            self.cache_map = {}

        # meta label
        m = self.bundle.manifest
        schema = _safe_str(m.get("bundle_schema", m.get("schema", ""))).strip() or "unknown"
        created_at = _safe_str(m.get("created_at", "")).strip()
        created_by = _safe_str(m.get("created_by", "")).strip()
        stages = ((m.get("pipeline", {}) or {}).get("stages", None)) or ((m.get("pipeline_state", {}) or {}).get("stages", None)) or {}
        st_il = _safe_str(stages.get("IL", "")).strip() or "unknown"
        self.lbl_bundle_meta.configure(text=f"schema={schema} | created_at={created_at} | created_by={created_by} | IL={st_il}")

        # criteria list
        self.lst_crit.delete(0, "end")
        crits = [c for c in self.bundle.criteria.criteria if c.enabled]
        for c in crits:
            lbl = f"{c.id}  thr={c.threshold:g}  → {c.source_text[:80]}"
            self.lst_crit.insert("end", lbl)

        self._set_warn(self.bundle.criteria.warnings)

        # reset results
        self.full_rows = []
        self.survivors = []
        self.counts = {}
        self.crit_impacts = {}
        self.row_eval_lists = []

        # set table to raw rows
        self.table.set_rows(self.bundle.parse.rows[:2000])  # show first 2000 pre-run for responsiveness
        self._refresh_counts_label()

        self.btn_run.configure(state="normal")

    # -------- run IL
    def on_run_il(self):
        if not self.bundle or self.worker:
            return

        model = self.var_model.get().strip() or DEFAULT_MODEL
        trunc_chars = int(self.var_trunc.get())
        batch_size = int(self.var_batch.get())
        use_cache = bool(self.var_use_cache.get())

        self.cancel_event.clear()
        self.pbar.configure(value=0.0, maximum=100.0)
        self._set_controls_running(True)

        def progress_cb(frac: float):
            # called from worker; route to UI thread
            self.after(0, lambda: self.pbar.configure(value=max(0.0, min(100.0, frac * 100.0))))

        def progress_evt(evt: Dict[str, Any]):
            # optional structured events for advanced UI; currently logs minimal
            kind = _safe_str(evt.get("kind",""))
            sub = _safe_str(evt.get("sub",""))
            if kind == "l_batch" and sub == "sending":
                bi = evt.get("batch_idx"); bt = evt.get("batch_total")
                ci = evt.get("crit_idx"); ct = evt.get("crit_total")
                self.after(0, lambda: self._log(f"[LLM] criterion {ci}/{ct} batch {bi}/{bt} sending...\n"))

        def work():
            try:
                full_rows, survivors, counts, crit_impacts, row_eval_lists, cache_out = run_il_screen(
                    self.bundle.parse,
                    self.bundle.criteria,
                    model=model,
                    trunc_chars=trunc_chars,
                    batch_size=batch_size,
                    use_cache=use_cache,
                    cache_in=self.cache_map if use_cache else {},
                    cancel_event=self.cancel_event,
                    log_cb=lambda s: self.after(0, lambda: self._log(s)),
                    progress_cb=progress_cb,
                    progress_evt=progress_evt,
                )
                if self.cancel_event.is_set():
                    self.after(0, lambda: self._log("\n[CANCELLED]\n"))
                    return
                self.full_rows = full_rows
                self.survivors = survivors
                self.counts = counts
                self.crit_impacts = crit_impacts
                self.row_eval_lists = row_eval_lists
                self.cache_map = cache_out

                # update table to full rows (show outcomes)
                self.after(0, lambda: self.table.set_rows(self.full_rows))
                self.after(0, self._refresh_counts_label)

                # enable exports
                self.after(0, lambda: self.btn_export_csv.configure(state="normal"))
                self.after(0, lambda: self.btn_export_xlsx.configure(state="normal"))
                self.after(0, lambda: self.btn_next_bundle.configure(state="normal"))

            except Exception as e:
                self.after(0, lambda m=str(e): messagebox.showerror("IL run failed", m))
            finally:
                self.after(0, lambda: self._set_controls_running(False))
                self.worker = None

        self.worker = threading.Thread(target=work, daemon=True)
        self.worker.start()

    def on_cancel(self):
        self.cancel_event.set()
        self._log("\n[Cancel requested]\n")

    # -------- exports
    def on_export_csv(self):
        if not self.full_rows:
            return
        p = filedialog.asksaveasfilename(
            title="Save IL_FULL.csv",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")]
        )
        if not p:
            return
        header = list(self.full_rows[0].keys())
        _write_csv(p, header, self.full_rows)
        messagebox.showinfo("Export", f"Saved:\n{p}")

    def on_export_xlsx(self):
        # Keep it standard-library only: write CSV next to requested xlsx name.
        if not self.full_rows:
            return
        p = filedialog.asksaveasfilename(
            title="Save IL_FULL.xlsx (fallback to CSV if no writer)",
            defaultextension=".xlsx",
            filetypes=[("XLSX", "*.xlsx"), ("CSV", "*.csv")]
        )
        if not p:
            return
        # If user asked CSV, do CSV.
        if p.lower().endswith(".csv"):
            header = list(self.full_rows[0].keys())
            _write_csv(p, header, self.full_rows)
            messagebox.showinfo("Export", f"Saved:\n{p}")
            return

        # Minimal XLSX writer without dependencies is non-trivial; provide a CSV next to it.
        csv_path = p[:-5] + ".csv"
        header = list(self.full_rows[0].keys())
        _write_csv(csv_path, header, self.full_rows)
        messagebox.showinfo("Export", f"No XLSX writer bundled.\nSaved CSV instead:\n{csv_path}")

    def on_build_next_bundle(self):
        if not self.bundle or not self.full_rows:
            return
        out_zip = filedialog.asksaveasfilename(
            title="Save next bundle ZIP (post-IL)",
            defaultextension=".zip",
            filetypes=[("ZIP", "*.zip")]
        )
        if not out_zip:
            return

        # Prepare output structure in-memory then zip.
        # - manifest.json updated
        # - data/current.csv replaced with survivors
        # - reports/IL_FULL.csv + reports/IL_SURVIVORS.csv
        # - input_errors.csv from skipped
        # - cache/IL_cache.jsonl if enabled
        try:
            with zipfile.ZipFile(self.bundle.zip_path, "r") as zf_in:
                members = zf_in.namelist()
                root = self.bundle.root

                # read original manifest
                manifest = dict(self.bundle.manifest)

                # update pipeline stage mark (tolerate both schemas)
                def _set_stage(m: Dict[str, Any], key: str, value: str):
                    if "pipeline" in m and isinstance(m.get("pipeline"), dict):
                        m["pipeline"].setdefault("stages", {})
                        m["pipeline"]["stages"][key] = value
                    if "pipeline_state" in m and isinstance(m.get("pipeline_state"), dict):
                        m["pipeline_state"].setdefault("stages", {})
                        m["pipeline_state"]["stages"][key] = value

                _set_stage(manifest, "IL", "done")
                manifest["updated_at"] = datetime.utcnow().isoformat() + "Z"

                # write new zip
                with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf_out:
                    # copy everything except data/current.csv and reports we replace
                    skip_prefixes = {
                        root + "data/current.csv",
                        root + f"{REPORTS_DIR_REL}/IL_FULL.csv",
                        root + f"{REPORTS_DIR_REL}/IL_SURVIVORS.csv",
                    root + FINAL_REPORT_REL,
                        root + "data/input_errors.csv",
                        root + IL_CACHE_REL,
                        root + "manifest.json",
                    }
                    for m in members:
                        if m in skip_prefixes:
                            continue
                        # also avoid directories
                        if m.endswith("/"):
                            continue
                        zf_out.writestr(m, _read_zip_bytes(zf_in, m))

                    # manifest
                    zf_out.writestr(root + "manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

                    # data/current.csv survivors
                    # Use original header order (minus EL columns)
                    header = list(self.bundle.parse.header)
                    # guarantee local_id present
                    if "local_id" not in header:
                        header = ["local_id"] + header
                    buf = io.StringIO()
                    w = csv.DictWriter(buf, fieldnames=header, extrasaction="ignore")
                    w.writeheader()
                    for r in self.survivors:
                        w.writerow({k: r.get(k, "") for k in header})
                    zf_out.writestr(root + "data/current.csv", buf.getvalue())

                    # reports
                    rep_full = root + f"{REPORTS_DIR_REL}/IL_FULL.csv"
                    rep_surv = root + f"{REPORTS_DIR_REL}/IL_SURVIVORS.csv"
                    # full
                    header_full = list(self.full_rows[0].keys())
                    buf2 = io.StringIO()
                    w2 = csv.DictWriter(buf2, fieldnames=header_full, extrasaction="ignore")
                    w2.writeheader()
                    for r in self.full_rows:
                        w2.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in header_full})
                    zf_out.writestr(rep_full, buf2.getvalue())

                    # survivors report = same schema as input current.csv
                    buf3 = io.StringIO()
                    w3 = csv.DictWriter(buf3, fieldnames=header, extrasaction="ignore")
                    w3.writeheader()
                    for r in self.survivors:
                        w3.writerow({k: r.get(k, "") for k in header})
                    zf_out.writestr(rep_surv, buf3.getvalue())

                    # input errors
                    if self.bundle.parse.skipped:
                        buf4 = io.StringIO()
                        # flatten skipped
                        err_header = ["reason", "row_json"]
                        w4 = csv.DictWriter(buf4, fieldnames=err_header)
                        w4.writeheader()
                        for e in self.bundle.parse.skipped:
                            w4.writerow({"reason": _safe_str(e.get("reason","")), "row_json": json.dumps(e.get("row", {}), ensure_ascii=False)})
                        zf_out.writestr(root + "data/input_errors.csv", buf4.getvalue())

                    # cache
                    if self.var_use_cache.get():
                        zf_out.writestr(root + IL_CACHE_REL, _dump_cache_to_jsonl(self.cache_map))

            messagebox.showinfo("Next bundle", f"Saved:\n{out_zip}")
        except Exception as e:
            messagebox.showerror("Next bundle failed", str(e))

    # -------- details
    def on_row_doubleclick(self, _evt=None):
        row = self.table.get_selected_row()
        if not row:
            return
        top = tk.Toplevel(self)
        top.title(f"Row {row.get('local_id','')}")
        top.geometry("850x600")

        txt = tk.Text(top, wrap="word")
        txt.pack(fill="both", expand=True)

        def add(title: str, content: str):
            txt.insert("end", f"\n=== {title} ===\n")
            txt.insert("end", content + "\n")

        add("local_id", _safe_str(row.get("local_id","")))
        add("title", _safe_str(row.get("title","")))
        add("abstract", _safe_str(row.get("abstract","")))
        add("keywords", _safe_str(row.get("keywords","")))
        add("IL outcome", _safe_str(row.get("il_outcome","")))
        add("EL summary", _safe_str(row.get("il_reason_summary","")))

        ev = _safe_str(row.get("il_evidence_json","{}"))
        try:
            evj = json.loads(ev)
            add("EL evidence (json)", json.dumps(evj, ensure_ascii=False, indent=2)[:15000])
        except Exception:
            add("EL evidence (raw)", ev[:15000])

        txt.configure(state="disabled")

    def on_criterion_doubleclick(self, _evt=None):
        # Filter table to rows touched by selected criterion (failed/missing/uncertain/met)
        if not self.full_rows or not self.bundle:
            return
        sel = self.lst_crit.curselection()
        if not sel:
            return
        # parse criterion id from listbox line
        line = self.lst_crit.get(sel[0])
        cid = line.split()[0].strip()

        touched: List[Dict[str, Any]] = []
        for r in self.full_rows:
            parts = ",".join([
                _safe_str(r.get("il_failed_ids","")),
                _safe_str(r.get("il_missing_ids","")),
                _safe_str(r.get("il_uncertain_ids","")),
                _safe_str(r.get("il_met_ids","")),
            ])
            if cid in {p.strip() for p in parts.split(",") if p.strip()}:
                touched.append(r)

        if touched:
            self._log(f"\n[Filter] Criterion {cid}: showing {len(touched)} rows touched.\n")
            self.table.set_rows(touched)
        else:
            self._log(f"\n[Filter] Criterion {cid}: no touched rows.\n")

# ------------------------------ plugin wrapper --------------------------------

class Plugin(BasePlugin):
    def __init__(self, app=None, meta: Optional[PluginMeta] = None):
        if meta is None:
            meta = PluginMeta(id=PLUGIN_ID, title=TAB_TITLE, version=PLUGIN_VERSION)
        super().__init__(app, meta)
        self.view: Optional[ILView] = None

    def build_tab(self, parent: ttk.Notebook) -> tk.Frame:
        frame = ttk.Frame(parent)
        self.view = ILView(frame)
        self.view.pack(fill="both", expand=True)
        return frame

    def on_close(self):
        try:
            if self.view:
                self.view.destroy()
        except Exception:
            pass
        self.view = None
