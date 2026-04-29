# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""
parser.py - Plugin 03 Harmoniser: criteria input parsing & utilities.

Concerns owned by this module:
  - Free-text IC/EC parsing (_parse_free_text_criteria)
  - Structured criteria-table loading (CSV/XLSX) and row normalization
  - Aggregate ("A") corpus-CSV header inspection and text-coverage stats
  - Target-name canonicalization against the corpus columns
  - Operand serialization for export (the `what` cell format)
  - Stage / operator vocabularies (STAGES, OPERATORS, TARGET_ALIASES)
  - Light shared utilities (_now_iso, _norm_space, _safe_str, _read_text_file,
    _rtf_to_text, _is_rtf_path)

This is the input-side / ingestion module of the Harmoniser pipeline; downstream
modules (inference, exporters, llm_refine, bundle, ui) consume what is produced
here. This module imports nothing from elsewhere in the package and has no GUI
or LLM dependencies - it is safe to reuse standalone.

Extracted from `plugin.py` in Conv 4 of the v3.1.0 refactor. All behavior is
preserved verbatim; the `tests/golden/criteria_harmonized_v3.1.0.csv`
regression test guards byte-identity of downstream output.
"""

import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple


# ============================
# Stage / operator vocabularies
# ============================

STAGES = ("EH", "IH", "EL", "IL")
OPERATORS = (
    "contains",
    "equals",
    "regex",
    "in_list",
    "not_in",
    "gte",
    "lte",
    "between",
    "llm",
)


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _safe_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    try:
        return str(x)
    except Exception:
        return ""


def _read_text_file(path: str) -> str:
    p = Path(path)
    b = p.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="ignore")


def _rtf_to_text(rtf: str) -> str:
    """Very lightweight RTF -> text."""
    if not rtf.lstrip().startswith(r"{\rtf"):
        return rtf

    rtf = re.sub(r"\\pict[\s\S]*?\}", "", rtf)  # drop pictures/blobs
    rtf = rtf.replace(r"\par", "\n").replace(r"\line", "\n")
    rtf = re.sub(r"\\[a-zA-Z]+-?\d* ?", "", rtf)  # control words
    rtf = rtf.replace("{", "").replace("}", "")

    def _hex(m: re.Match) -> str:
        try:
            return bytes.fromhex(m.group(1)).decode("cp1252", errors="ignore")
        except Exception:
            return ""

    rtf = re.sub(r"\\'([0-9a-fA-F]{2})", _hex, rtf)
    return rtf


def _is_rtf_path(path: str) -> bool:
    return Path(path).suffix.lower() == ".rtf"


def _load_a_header_and_stats(csv_path: str, sample_n: int = 200) -> Tuple[List[str], Dict[str, float]]:
    """Load A header and compute basic non-empty stats for common text fields."""
    cols: List[str] = []
    stats: Dict[str, float] = {}

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        try:
            cols = next(reader)
        except StopIteration:
            raise ValueError("A vector CSV is empty")

    fields = [c for c in ("title", "abstract", "keywords") if c in cols]
    if not fields:
        return cols, stats

    nonempty = {c: 0 for c in fields}
    total = 0

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            total += 1
            for c in fields:
                if _safe_str(row.get(c)).strip():
                    nonempty[c] += 1
            if total >= sample_n:
                break

    if total <= 0:
        return cols, stats

    for c in fields:
        stats[c] = nonempty[c] / total

    return cols, stats


def _get_best_text_targets(a_columns: List[str], text_stats: Dict[str, float]) -> str:
    """
    Select the best text field(s) based on availability and coverage.

    Policy:
      - Prefer among title/abstract/keywords when present.
      - If multiple are present and top coverage > 80%, use only the top one.
      - Otherwise, use all available text fields.
      - Fallback to first column.
    """
    available_fields: List[str] = []
    for field in ["title", "abstract", "keywords"]:
        if field in a_columns:
            available_fields.append(field)

    if not available_fields and a_columns:
        return a_columns[0]

    if text_stats and len(available_fields) > 1:
        sorted_fields = sorted(
            available_fields,
            key=lambda f: float(text_stats.get(f, 0.0)),
            reverse=True
        )
        if float(text_stats.get(sorted_fields[0], 0.0)) > 0.8:
            return sorted_fields[0]

    return ",".join(available_fields) if available_fields else (a_columns[0] if a_columns else "")


def _detect_id_column(a_columns: Sequence[str]) -> str:
    for cand in ("a_id", "local_id", "id", "record_id"):
        if cand in a_columns:
            return cand
    return a_columns[0] if a_columns else ""


# Aliases are applied ONLY when the alias exists in A columns.
TARGET_ALIASES = {
    "language": "lang",
    "langue": "lang",
    "publication_language": "lang",
    "type": "doc_type",
    "pub_type": "doc_type",
    "publication_type": "doc_type",
    "document_type": "doc_type",
    "journal": "venue",
    "source": "venue",
    "conference": "venue",
    "link": "url",
    "website": "url",
}


def _canonicalize_targets(target: str, a_columns: Sequence[str]) -> Tuple[str, List[str]]:
    """
    Return (canonical_target_str, unknown_targets).

    Safety rule:
      - Only apply alias mapping if the alias exists in A columns (case-insensitive).
      - Otherwise keep the original term if it exists.
    """
    unknown: List[str] = []
    parts = [p.strip() for p in (target or "").split(",") if p.strip()]
    canon_parts: List[str] = []

    def _find_ci(name: str) -> Optional[str]:
        for c in a_columns:
            if c == name:
                return c
        for c in a_columns:
            if c.lower() == name.lower():
                return c
        return None

    for p in parts:
        low = p.lower()
        mapped = TARGET_ALIASES.get(low, p)

        # Prefer mapped only if it exists; otherwise prefer original if it exists.
        mapped_ci = _find_ci(mapped) if mapped != p else None
        orig_ci = _find_ci(p)

        if mapped != p and mapped_ci is not None:
            canon_parts.append(mapped_ci)
            continue
        if orig_ci is not None:
            canon_parts.append(orig_ci)
            continue

        # last resort: try mapped CI even if same as p, else mark unknown
        if mapped != p:
            mapped_ci2 = _find_ci(mapped)
            if mapped_ci2 is not None:
                canon_parts.append(mapped_ci2)
                continue

        unknown.append(p)
        canon_parts.append(mapped)

    return ",".join(canon_parts), unknown


def _parse_what_cell(operator: str, raw: Any) -> List[str]:
    """Parse a 'what' cell from CSV/XLSX."""
    s = _safe_str(raw).strip()
    if not s:
        return []

    if s.startswith("[") and s.endswith("]"):
        try:
            arr = json.loads(s)
            if isinstance(arr, list):
                return [str(x) for x in arr if str(x).strip()]
        except Exception:
            pass

    if operator == "regex":
        return [s]

    if ";" in s:
        parts = [p.strip() for p in s.split(";")]
    elif "|" in s:
        parts = [p.strip() for p in s.split("|")]
    elif "," in s and operator in {"contains", "in_list", "not_in"}:
        parts = [p.strip() for p in s.split(",")]
    else:
        parts = [s]

    return [p for p in parts if p]


def _what_to_export(operator: str, what_list: List[str]) -> str:
    if operator == "between" and len(what_list) == 2:
        return f"{what_list[0]};{what_list[1]}"
    if operator in {"gte", "lte", "equals"} and len(what_list) == 1:
        return what_list[0]
    if operator == "regex" and what_list:
        return what_list[0]
    return ";".join([w for w in what_list if w is not None])


def _export_to_pipe_table(rows: List[Dict[str, Any]]) -> str:
    headers = [
        "stage",
        "id",
        "type",
        "scope",
        "label",
        "operator",
        "target",
        "what",
        "threshold",
        "enabled",
    ]

    def esc(s: str) -> str:
        return _safe_str(s).replace("|", "/").strip()

    out = []
    out.append("| " + " | ".join(headers) + " |")
    out.append("|" + "|".join(["---"] * len(headers)) + "|")

    for r in rows:
        line = [
            esc(r.get("stage", "")),
            esc(r.get("id", "")),
            esc(r.get("type", "")),
            esc(r.get("scope", "metadata")),
            esc(r.get("label", "")),
            esc(r.get("operator", "")),
            esc(r.get("target", "")),
            esc(_what_to_export(r.get("operator", ""), r.get("what", []) or [])),
            esc(r.get("threshold", "") if r.get("threshold", "") is not None else ""),
            esc("1" if bool(r.get("enabled", True)) else "0"),
        ]
        out.append("| " + " | ".join(line) + " |")

    return "\n".join(out) + "\n"


def _parse_free_text_criteria(text: str) -> List[Tuple[str, str, str, str]]:
    """
    Parse free-text criteria into tuples: (id, type, label, source_text).

    Supports:
      - IC-1: ... / EC-2 - ...
      - IC 1 ... / EC 3 ...
      - "Inclusion criteria" / "Exclusion criteria" headers + bullet lists
      - numbered bullets (1., 1), i., I., •, -, *)
      - mixed order and missing numbering (auto-numbering)
    """
    txt = (text or "").replace("–", "-").replace("—", "-")
    txt = txt.replace("\r\n", "\n").replace("\r", "\n")

    # normalize bullets
    txt = txt.replace("•", "- ").replace("●", "- ").replace("◦", "- ").replace("▪", "- ")

    lines = [l.strip() for l in txt.splitlines()]
    out: List[Tuple[str, str, str, str]] = []

    # explicit IC/EC forms
    explicit_pat = re.compile(
        r"^(IC|EC)\s*[- ]?\s*(\d+)\s*[:\-]?\s*(.+)$",
        re.IGNORECASE
    )
    # header detection
    inc_header = re.compile(r"^\s*(inclusion|include)\s*(criteria)?\s*:?\s*$", re.IGNORECASE)
    exc_header = re.compile(r"^\s*(exclusion|exclude)\s*(criteria)?\s*:?\s*$", re.IGNORECASE)

    # bullet / numbered item pattern (1., 1), i., I., -, *, etc.)
    bullet_pat = re.compile(
        r"^(\(?\d+\)?[.)]|[ivxlcdm]+[.)]|-|\*|\+)\s+(.+)$",
        re.IGNORECASE
    )

    mode: Optional[str] = None  # "include" | "exclude"
    inc_i = 0
    exc_i = 0

    for raw in lines:
        l = raw.strip()
        if not l:
            continue

        # headers
        if inc_header.match(l):
            mode = "include"
            continue
        if exc_header.match(l):
            mode = "exclude"
            continue

        # explicit IC/EC numbering
        m = explicit_pat.match(l)
        if m:
            prefix = m.group(1).upper()
            num = m.group(2)
            body = m.group(3).strip()
            crit_id = f"{prefix}-{num}"
            crit_type = "include" if prefix == "IC" else "exclude"
            out.append((crit_id, crit_type, body, raw))
            continue

        # bullet under mode
        m2 = bullet_pat.match(l)
        if m2 and mode in ("include", "exclude"):
            body = m2.group(2).strip()
            if mode == "include":
                inc_i += 1
                out.append((f"IC-{inc_i}", "include", body, raw))
            else:
                exc_i += 1
                out.append((f"EC-{exc_i}", "exclude", body, raw))
            continue

        # fallback: if line begins with "IC:" or "EC:" without numbering
        if l.lower().startswith("ic:") or l.lower().startswith("inclusion:"):
            inc_i += 1
            body = l.split(":", 1)[1].strip() if ":" in l else l
            out.append((f"IC-{inc_i}", "include", body, raw))
            continue
        if l.lower().startswith("ec:") or l.lower().startswith("exclusion:"):
            exc_i += 1
            body = l.split(":", 1)[1].strip() if ":" in l else l
            out.append((f"EC-{exc_i}", "exclude", body, raw))
            continue

        # permissive: if we are inside a mode, treat plain lines as criteria
        if mode in ("include", "exclude") and len(l) > 2:
            if mode == "include":
                inc_i += 1
                out.append((f"IC-{inc_i}", "include", l, raw))
            else:
                exc_i += 1
                out.append((f"EC-{exc_i}", "exclude", l, raw))

    return out


def _load_structured_criteria_table(path: str) -> Tuple[List[Dict[str, Any]], str]:
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return _load_structured_criteria_csv(path), "csv"
    if ext in {".xlsx", ".xlsm"}:
        return _load_structured_criteria_xlsx(path), "xlsx"
    raise ValueError("Unsupported criteria table format; use CSV or XLSX")


def _load_structured_criteria_csv(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = []
        for r in reader:
            rows.append({k.strip(): v for k, v in r.items() if k is not None})
        return rows


def _load_structured_criteria_xlsx(path: str) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(f"openpyxl required for XLSX: {e}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers: List[str] = []
    rows: List[Dict[str, Any]] = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [str(c).strip() if c is not None else "" for c in row]
            headers = [h for h in headers if h]
            continue
        if not any(row):
            continue
        d: Dict[str, Any] = {}
        for j, h in enumerate(headers):
            d[h] = row[j] if j < len(row) else ""
        rows.append(d)

    return rows


def _normalize_structured_row(r: Dict[str, Any]) -> Dict[str, Any]:
    target = r.get("target", r.get("field", ""))

    stage = _safe_str(r.get("stage", "")).strip().upper()
    if stage not in STAGES:
        stage = ""

    crit_id = _safe_str(r.get("id", "")).strip()
    crit_type = _safe_str(r.get("type", "")).strip().lower()
    if crit_type not in {"include", "exclude"}:
        if crit_type.upper().startswith("IC"):
            crit_type = "include"
        elif crit_type.upper().startswith("EC"):
            crit_type = "exclude"

    label = _safe_str(r.get("label", r.get("name", ""))).strip()
    operator = _safe_str(r.get("operator", "")).strip().lower()

    op_alias = {
        "eq": "equals",
        "==": "equals",
        "=": "equals",
        "in": "in_list",
        "not in": "not_in",
        "match": "regex",
        "re": "regex",
    }
    operator = op_alias.get(operator, operator)
    if operator not in OPERATORS:
        operator = ""

    what_list = _parse_what_cell(operator or "contains", r.get("what", ""))

    enabled_raw = r.get("enabled", True)
    if isinstance(enabled_raw, str):
        enabled = enabled_raw.strip().lower() not in {"0", "false", "no", "off"}
    else:
        enabled = bool(enabled_raw)

    scope = _safe_str(r.get("scope", "metadata")) or "metadata"
    threshold = _safe_str(r.get("threshold", "")).strip()
    source_text = _safe_str(r.get("source_text", "")).strip() or ""

    return {
        "stage": stage,
        "id": crit_id,
        "type": crit_type,
        "scope": scope,
        "label": label,
        "operator": operator,
        "target": _safe_str(target).strip(),
        "what": what_list,
        "threshold": threshold,
        "enabled": enabled,
        "source_text": source_text,
    }

