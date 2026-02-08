# -*- coding: utf-8 -*-
"""plugin.py — Harmoniser (Criteria) as a PRISMA Hub tab plugin

Single-file, self-contained (UI + engine).

What it does
- Requires BOTH:
  1) Criteria input: free-text IC/EC (TXT/RTF) or structured criteria table (CSV/XLSX)
  2) A vector: *_aggregate.csv
- Uses the A header to:
  - populate target pickers
  - apply alias mapping (language->lang, type->doc_type, journal->venue, ...)
  - validate that targets reference real columns
  - block export if invalid
- Harmonises criteria into stage-explicit rows for the split pipeline:
  EH / IH / EL / IL
- Optional LLM refinement (OpenAI) with strict guardrails.

Exports (defaults)
- Combined: criteria_harmonized.csv + criteria_harmonized.txt (pipe-table)
- Per-stage CSV: criteria_EH.csv / criteria_IH.csv / criteria_EL.csv / criteria_IL.csv

Notes
- This module does NOT screen articles.
- It only harmonises criteria.

"""

TAB_TITLE = "Harmoniser — Criteria"


import csv
import json
import os
import re
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from prisma_hub.plugin_api import PluginMeta, BasePlugin


# ============================
# Engine: parsing & utilities
# ============================

STAGES = ("EH", "IH", "EL", "IL")
OPERATORS = (
    "contains",
    "equals",
    "regex",
    "in_list",
    "not_in",
    "gte",
    "lte",
    "between",
    "llm",
)

DEFAULT_TEXT_TARGET = "title,abstract,keywords"
DEFAULT_THRESHOLD = 0.60


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _norm_space(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def _safe_str(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    try:
        return str(x)
    except Exception:
        return ""


def _read_text_file(path: str) -> str:
    p = Path(path)
    b = p.read_bytes()
    # try utf-8-sig first
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="ignore")


def _rtf_to_text(rtf: str) -> str:
    """Very lightweight RTF -> text.

    This is intentionally simple (no full RTF parsing), but works well for
    typical criteria files produced by Word/LibreOffice with basic formatting.
    """
    if not rtf.lstrip().startswith("{\\rtf"):
        return rtf

    # Remove binary blobs / pictures
    rtf = re.sub(r"\\pict[\s\S]*?\}", "", rtf)

    # Replace paragraph/line breaks
    rtf = rtf.replace("\\par", "\n")
    rtf = rtf.replace("\\line", "\n")

    # Remove RTF control words (e.g., \b0, \fs24, \u1234?)
    rtf = re.sub(r"\\[a-zA-Z]+-?\d* ?", "", rtf)

    # Remove groups/braces
    rtf = rtf.replace("{", "").replace("}", "")

    # Remove escaped hex like \'e9
    def _hex(m: re.Match) -> str:
        try:
            return bytes.fromhex(m.group(1)).decode("cp1252", errors="ignore")
        except Exception:
            return ""

    rtf = re.sub(r"\\'([0-9a-fA-F]{2})", _hex, rtf)
    return rtf


def _is_rtf_path(path: str) -> bool:
    return Path(path).suffix.lower() == ".rtf"


def _load_a_header_and_stats(csv_path: str, sample_n: int = 200) -> Tuple[List[str], Dict[str, float]]:
    """Load A header and compute basic non-empty stats for common text fields."""
    cols: List[str] = []
    stats: Dict[str, float] = {}

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        try:
            cols = next(reader)
        except StopIteration:
            raise ValueError("A vector CSV is empty")

    # Stats for common text fields (if present)
    fields = [c for c in ("title", "abstract", "keywords") if c in cols]
    if not fields:
        return cols, stats

    nonempty = {c: 0 for c in fields}
    total = 0

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            total += 1
            for c in fields:
                if _safe_str(row.get(c)).strip():
                    nonempty[c] += 1
            if total >= sample_n:
                break

    if total <= 0:
        return cols, stats

    for c in fields:
        stats[c] = nonempty[c] / total

    return cols, stats


def _detect_id_column(a_columns: Sequence[str]) -> str:
    # By user rule: use whatever A provides; detect best candidate.
    for cand in ("a_id", "local_id", "id", "record_id"):
        if cand in a_columns:
            return cand
    return a_columns[0] if a_columns else ""


TARGET_ALIASES = {
    "language": "lang",
    "langue": "lang",
    "publication_language": "lang",
    "type": "doc_type",
    "pub_type": "doc_type",
    "publication_type": "doc_type",
    "document_type": "doc_type",
    "journal": "venue",
    "source": "venue",
    "conference": "venue",
    "link": "url",
    "website": "url",
}


def _canonicalize_targets(target: str, a_columns: Sequence[str]) -> Tuple[str, List[str]]:
    """Return (canonical_target_str, unknown_targets).

    - Applies alias mapping (case-insensitive)
    - Preserves comma-separated multi-target
    """
    a_set = {c.strip() for c in a_columns}
    unknown: List[str] = []

    parts = [p.strip() for p in (target or "").split(",") if p.strip()]
    canon_parts: List[str] = []

    for p in parts:
        low = p.lower()
        mapped = TARGET_ALIASES.get(low, p)
        if mapped not in a_set:
            # Sometimes users provide weird casing
            # Try case-insensitive match
            ci = None
            for c in a_columns:
                if c.lower() == mapped.lower():
                    ci = c
                    break
            if ci:
                mapped = ci
            else:
                unknown.append(p)
        canon_parts.append(mapped)

    return ",".join(canon_parts), unknown


def _parse_what_cell(operator: str, raw: Any) -> List[str]:
    """Parse a 'what' cell from CSV/XLSX.

    Accepts:
    - JSON list string: ["a","b"]
    - Semicolon-separated: a; b; c
    - Pipe-separated: a | b | c
    - Comma-separated if no semicolons/pipes
    """
    s = _safe_str(raw).strip()
    if not s:
        return []

    # JSON list
    if s.startswith("[") and s.endswith("]"):
        try:
            arr = json.loads(s)
            if isinstance(arr, list):
                return [str(x) for x in arr if str(x).strip()]
        except Exception:
            pass

    # single regex is a single string
    if operator == "regex":
        return [s]

    # detect delimiter
    if ";" in s:
        parts = [p.strip() for p in s.split(";")]
    elif "|" in s:
        parts = [p.strip() for p in s.split("|")]
    elif "," in s and operator in {"contains", "in_list", "not_in"}:
        parts = [p.strip() for p in s.split(",")]
    else:
        parts = [s]

    return [p for p in parts if p]


def _what_to_export(operator: str, what_list: List[str]) -> str:
    if operator == "between" and len(what_list) == 2:
        return f"{what_list[0]};{what_list[1]}"
    if operator in {"gte", "lte", "equals"} and len(what_list) == 1:
        return what_list[0]
    if operator == "regex" and what_list:
        return what_list[0]
    return ";".join([w for w in what_list if w is not None])


def _export_to_pipe_table(rows: List[Dict[str, Any]]) -> str:
    headers = [
        "stage",
        "id",
        "type",
        "scope",
        "label",
        "operator",
        "target",
        "what",
        "threshold",
        "enabled",
    ]

    def esc(s: str) -> str:
        # Avoid breaking the pipe table
        return _safe_str(s).replace("|", "/").strip()

    out = []
    out.append("| " + " | ".join(headers) + " |")
    out.append("|" + "|".join(["---"] * len(headers)) + "|")

    for r in rows:
        line = [
            esc(r.get("stage", "")),
            esc(r.get("id", "")),
            esc(r.get("type", "")),
            esc(r.get("scope", "metadata")),
            esc(r.get("label", "")),
            esc(r.get("operator", "")),
            esc(r.get("target", "")),
            esc(_what_to_export(r.get("operator", ""), r.get("what", []) or [])),
            esc(r.get("threshold", "") if r.get("threshold", "") is not None else ""),
            esc("1" if bool(r.get("enabled", True)) else "0"),
        ]
        out.append("| " + " | ".join(line) + " |")

    return "\n".join(out) + "\n"


def _keywords_from_line(line: str) -> List[str]:
    """Best-effort keyword extraction for IH/contains defaults.

    Strategy:
    - Prefer quoted phrases
    - Else, collect multi-word phrases separated by commas/semicolons
    - Else, build a few bigrams/trigrams from content words

    This is intentionally conservative: if we can't extract anything useful,
    return [].
    """
    raw = _norm_space(line)

    # Quoted phrases
    quotes = re.findall(r"\"([^\"]{3,120})\"", raw)
    if quotes:
        return [_norm_space(q) for q in quotes if _norm_space(q)]

    # Parentheses phrases
    par = re.findall(r"\(([^\)]{3,120})\)", raw)
    par = [_norm_space(p) for p in par if _norm_space(p)]

    # After ':' or '—' often a list
    tail = raw
    m = re.split(r"[:—-]", raw, maxsplit=1)
    if len(m) == 2:
        tail = m[1]

    # Split tail by separators
    if any(sep in tail for sep in [";", ","]):
        parts = re.split(r"[;,]", tail)
        parts = [_norm_space(p) for p in parts if _norm_space(p)]
        parts = [p for p in parts if len(p) >= 3]
        # Combine parentheses as hints
        combined = parts + par
        # Filter out overly long chunks
        combined = [p for p in combined if 3 <= len(p) <= 80]
        # Deduplicate
        seen = set()
        out = []
        for p in combined:
            k = p.lower()
            if k not in seen:
                seen.add(k)
                out.append(p)
        return out[:10]

    # Token based
    stop = {
        "the", "a", "an", "and", "or", "of", "to", "in", "on", "for", "with", "as", "at",
        "from", "by", "is", "are", "be", "being", "been", "this", "that", "these", "those",
        "paper", "article", "study", "studies", "research", "focus", "primary", "aim", "aims",
        "include", "includes", "including", "excluding", "excludes", "must", "should", "will",
        "not", "no", "without", "within", "between", "after", "before", "since",
    }

    tokens = [t.lower() for t in re.findall(r"[A-Za-z][A-Za-z0-9_\-]{2,}", raw)]
    content = [t for t in tokens if t not in stop]
    if len(content) < 2:
        return par[:5] if par else []

    # Build bigrams/trigrams
    grams: List[str] = []
    for i in range(len(content) - 1):
        grams.append(f"{content[i]} {content[i+1]}")
    for i in range(len(content) - 2):
        grams.append(f"{content[i]} {content[i+1]} {content[i+2]}")

    # Prefer longer grams
    grams = sorted(set(grams), key=lambda s: (-len(s), s))
    return grams[:8]


def _infer_year_rule(text: str) -> Optional[Tuple[str, str, List[str]]]:
    """Return (stage, operator, what_list) if a year constraint is detected."""
    t = text.lower()

    # between YYYY and YYYY
    m = re.search(r"between\s+(19\d{2}|20\d{2})\s+(?:and|to)\s+(19\d{2}|20\d{2})", t)
    if m:
        y1, y2 = m.group(1), m.group(2)
        return "EH", "between", [y1, y2]

    # after/since YYYY
    m = re.search(r"(?:after|since|from)\s+(19\d{2}|20\d{2})", t)
    if m:
        y = m.group(1)
        return "EH", "gte", [y]

    # before/until YYYY
    m = re.search(r"(?:before|until|up\s+to)\s+(19\d{2}|20\d{2})", t)
    if m:
        y = m.group(1)
        return "EH", "lte", [y]

    # >= YYYY / <= YYYY
    m = re.search(r">=\s*(19\d{2}|20\d{2})", t)
    if m:
        return "EH", "gte", [m.group(1)]
    m = re.search(r"<=\s*(19\d{2}|20\d{2})", t)
    if m:
        return "EH", "lte", [m.group(1)]

    # standalone year mention can be ambiguous; ignore
    return None


def _infer_language_rule(text: str) -> Optional[Tuple[str, str, List[str]]]:
    t = text.lower()
    # very small mapping; users can edit
    langs = []
    if "english" in t or "anglais" in t:
        langs.append("en")
    if "french" in t or "français" in t or "francais" in t:
        langs.append("fr")
    if "spanish" in t or "espagnol" in t:
        langs.append("es")

    if not langs:
        return None

    # If explicitly says NOT in english/french
    neg = bool(re.search(r"\b(not|exclude|excluding|except)\b", t))
    if neg:
        return "EH", "not_in", langs

    # Else treat as equals if one, or in_list if many
    if len(langs) == 1:
        return "EH", "equals", [langs[0]]
    return "EH", "in_list", langs


def _infer_doctype_rule(text: str) -> Optional[Tuple[str, str, List[str]]]:
    t = text.lower()
    dtypes = []

    # conference/proceedings
    if "proceedings" in t or "conference" in t or "congress" in t or "colloque" in t:
        dtypes.append("conference")
    if "journal" in t or "article" in t:
        dtypes.append("journal")
    if "thesis" in t or "dissertation" in t or "mémoire" in t or "memoire" in t:
        dtypes.append("thesis")
    if "book" in t or "chapter" in t:
        dtypes.append("book")
    if "review" in t or "systematic review" in t or "meta-analysis" in t:
        dtypes.append("review")

    if not dtypes:
        return None

    # negation words
    neg = bool(re.search(r"\b(exclude|excluding|not|without|no)\b", t))
    if neg:
        return "EH", "not_in", dtypes

    if len(dtypes) == 1:
        return "EH", "equals", [dtypes[0]]
    return "EH", "in_list", dtypes


def _looks_semantic(text: str) -> bool:
    t = text.lower()
    return bool(
        re.search(
            r"\b(primary\s+focus|main\s+focus|focuses\s+on|focused\s+on|aims\s+to|addresses|investigates|examines|explores|concerns|is\s+about|rubber\s+hand|paradigm)\b",
            t,
        )
    )


def _auto_harmonize_line(
    crit_id: str,
    crit_type: str,
    label: str,
    source_text: str,
    a_columns: Sequence[str],
    text_field_stats: Dict[str, float],
) -> Dict[str, Any]:
    """No-LLM baseline harmonisation."""

    # Determine usable text targets from A availability
    text_targets = [c for c in ("title", "abstract", "keywords") if c in a_columns]
    if text_targets:
        # If a column is *almost always* empty, drop it from default targets
        # (still editable by user)
        usable = []
        for c in text_targets:
            if c not in text_field_stats:
                usable.append(c)
            else:
                if text_field_stats.get(c, 1.0) >= 0.05:
                    usable.append(c)
        if not usable:
            usable = text_targets[:1]
        default_text_target = ",".join(usable)
    else:
        default_text_target = ""

    line = _norm_space(source_text)

    # Start with safest defaults
    stage = "IL" if crit_type == "include" else "EL"
    operator = "llm"
    target = "abstract" if "abstract" in a_columns else (default_text_target or (a_columns[0] if a_columns else ""))
    what = [line] if line else []

    # language/year/doc_type rules
    yr = _infer_year_rule(line)
    if yr and "year" in a_columns:
        stage, operator, what = yr
        target = "year"

    lang = _infer_language_rule(line)
    if lang and "lang" in a_columns:
        stage, operator, what = lang
        target = "lang"

    dt = _infer_doctype_rule(line)
    if dt and "doc_type" in a_columns:
        stage, operator, what = dt
        target = "doc_type"

    # Heuristic keyword rule (IH) when not semantic
    if operator == "llm":
        if not _looks_semantic(line):
            kws = _keywords_from_line(line)
            if kws and default_text_target:
                stage = "IH"
                operator = "contains"
                target = default_text_target
                what = kws
        # Else keep as llm with EL/IL by type

    # Canonicalize targets against A
    target, _unk = _canonicalize_targets(target, a_columns)

    row: Dict[str, Any] = {
        "stage": stage,
        "id": crit_id,
        "type": crit_type,
        "scope": "metadata",
        "label": label,
        "operator": operator,
        "target": target,
        "what": what,
        "threshold": "" if stage in {"EH", "IH"} else f"{DEFAULT_THRESHOLD:.2f}",
        "enabled": True,
        "source_text": source_text,
    }

    # If EL/IL but operator not llm, keep threshold blank anyway? No: threshold only meaningful for llm scoring.
    # However user explicitly wants threshold for EL/IL (even if operator isn't llm), so we keep it for EL/IL.
    if row["stage"] in {"EH", "IH"}:
        row["threshold"] = ""

    return row


def _parse_free_text_criteria(text: str) -> List[Tuple[str, str, str, str]]:
    """Parse free-text criteria file into tuples:

    Returns list of (id, type, label, source_text)
    """
    # Normalize dashes
    txt = text.replace("–", "-").replace("—", "-")

    lines = [l.strip() for l in txt.splitlines()]
    out: List[Tuple[str, str, str, str]] = []

    # Accept patterns like:
    # IC-3 - ...
    # EC-4: ...
    # IC 3 ...
    pat = re.compile(r"^(IC|EC)\s*[- ]\s*(\d+)\s*[:\-]?\s*(.+)$", re.IGNORECASE)

    for l in lines:
        if not l:
            continue
        m = pat.match(l)
        if not m:
            continue
        prefix = m.group(1).upper()
        num = m.group(2)
        body = m.group(3).strip()
        crit_id = f"{prefix}-{num}"
        crit_type = "include" if prefix == "IC" else "exclude"
        label = body
        out.append((crit_id, crit_type, label, l))

    return out


def _load_structured_criteria_table(path: str) -> Tuple[List[Dict[str, Any]], str]:
    """Load criteria rows from CSV or XLSX.

    Returns (rows, source_kind)
    """
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        return _load_structured_criteria_csv(path), "csv"
    if ext in {".xlsx", ".xlsm"}:
        return _load_structured_criteria_xlsx(path), "xlsx"
    raise ValueError("Unsupported criteria table format; use CSV or XLSX")


def _load_structured_criteria_csv(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = []
        for r in reader:
            rows.append({k.strip(): v for k, v in r.items() if k is not None})
        return rows


def _load_structured_criteria_xlsx(path: str) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(f"openpyxl required for XLSX: {e}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers: List[str] = []
    rows: List[Dict[str, Any]] = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            headers = [str(c).strip() if c is not None else "" for c in row]
            headers = [h for h in headers if h]
            continue
        if not any(row):
            continue
        d: Dict[str, Any] = {}
        for j, h in enumerate(headers):
            d[h] = row[j] if j < len(row) else ""
        rows.append(d)

    return rows


def _normalize_structured_row(r: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize a row from a criteria table into the Harmoniser row schema."""

    # Accept both 'target' and legacy 'field'
    target = r.get("target", r.get("field", ""))

    stage = _safe_str(r.get("stage", "")).strip().upper()
    if stage not in STAGES:
        stage = ""

    crit_id = _safe_str(r.get("id", "")).strip()
    crit_type = _safe_str(r.get("type", "")).strip().lower()
    if crit_type not in {"include", "exclude"}:
        # Some tables may use IC/EC
        if crit_type.upper().startswith("IC"):
            crit_type = "include"
        elif crit_type.upper().startswith("EC"):
            crit_type = "exclude"

    label = _safe_str(r.get("label", r.get("name", ""))).strip()
    operator = _safe_str(r.get("operator", "")).strip().lower()

    # normalize operator aliases
    op_alias = {
        "eq": "equals",
        "==": "equals",
        "=": "equals",
        "in": "in_list",
        "not in": "not_in",
        "match": "regex",
        "re": "regex",
    }
    operator = op_alias.get(operator, operator)

    if operator not in OPERATORS:
        operator = ""

    what_list = _parse_what_cell(operator or "contains", r.get("what", ""))

    enabled_raw = r.get("enabled", True)
    enabled = True
    if isinstance(enabled_raw, str):
        enabled = enabled_raw.strip().lower() not in {"0", "false", "no", "off"}
    else:
        enabled = bool(enabled_raw)

    scope = _safe_str(r.get("scope", "metadata")) or "metadata"

    threshold = _safe_str(r.get("threshold", "")).strip()

    source_text = _safe_str(r.get("source_text", "")).strip() or ""

    return {
        "stage": stage,
        "id": crit_id,
        "type": crit_type,
        "scope": scope,
        "label": label,
        "operator": operator,
        "target": _safe_str(target).strip(),
        "what": what_list,
        "threshold": threshold,
        "enabled": enabled,
        "source_text": source_text,
    }


def _validate_row(row: Dict[str, Any], a_columns: Sequence[str]) -> Tuple[List[str], List[str]]:
    """Return (errors, warnings)."""
    errs: List[str] = []
    warns: List[str] = []

    stage = _safe_str(row.get("stage")).strip().upper()
    if stage not in STAGES:
        errs.append("Invalid stage")

    crit_id = _safe_str(row.get("id")).strip()
    if not crit_id:
        errs.append("Missing id")

    typ = _safe_str(row.get("type")).strip().lower()
    if typ not in {"include", "exclude"}:
        errs.append("Invalid type")

    op = _safe_str(row.get("operator")).strip().lower()
    if op not in OPERATORS:
        errs.append("Invalid operator")

    target = _safe_str(row.get("target")).strip()
    if not target:
        errs.append("Missing target")

    if a_columns:
        canon, unknown = _canonicalize_targets(target, a_columns)
        row["target"] = canon
        if unknown:
            errs.append(f"Unknown target(s): {', '.join(unknown)}")

    # what
    what_list = row.get("what")
    if not isinstance(what_list, list):
        warns.append("'what' was not a list; coerced")
        row["what"] = _parse_what_cell(op or "contains", what_list)

    what_list = row.get("what") or []
    if op == "between" and len(what_list) != 2:
        errs.append("between requires exactly 2 values")
    if op in {"gte", "lte", "equals"} and len(what_list) > 1:
        warns.append(f"{op} usually expects 1 value")
    if op == "llm":
        if len(what_list) != 1:
            errs.append("llm requires exactly 1 sentence in what")

    # threshold
    thr = _safe_str(row.get("threshold", "")).strip()
    if stage in {"EH", "IH"}:
        if thr:
            warns.append("threshold ignored for EH/IH; will be blanked")
            row["threshold"] = ""
    else:
        # EL/IL
        if not thr:
            row["threshold"] = f"{DEFAULT_THRESHOLD:.2f}"
        else:
            try:
                v = float(thr)
                if v < 0.0 or v > 1.0:
                    errs.append("threshold must be between 0 and 1")
            except Exception:
                errs.append("threshold must be a number")

    return errs, warns


def _export_csv(rows: List[Dict[str, Any]], path: str) -> None:
    cols = [
        "stage",
        "id",
        "type",
        "scope",
        "label",
        "operator",
        "target",
        "what",
        "threshold",
        "enabled",
        "source_text",
    ]

    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({
                "stage": r.get("stage", ""),
                "id": r.get("id", ""),
                "type": r.get("type", ""),
                "scope": r.get("scope", "metadata"),
                "label": r.get("label", ""),
                "operator": r.get("operator", ""),
                "target": r.get("target", ""),
                "what": _what_to_export(r.get("operator", ""), r.get("what", []) or []),
                "threshold": r.get("threshold", ""),
                "enabled": 1 if bool(r.get("enabled", True)) else 0,
                "source_text": r.get("source_text", ""),
            })


def _export_pipe(rows: List[Dict[str, Any]], path: str) -> None:
    txt = _export_to_pipe_table(rows)
    Path(path).write_text(txt, encoding="utf-8")


def _stage_filter(rows: List[Dict[str, Any]], stage: str) -> List[Dict[str, Any]]:
    s = stage.upper()
    return [r for r in rows if _safe_str(r.get("stage")).upper() == s]


def _call_openai_json(model: str, system: str, user: str, timeout_s: int = 120) -> Dict[str, Any]:
    """Best-effort OpenAI call returning JSON.

    Supports both new (OpenAI client) and legacy openai.ChatCompletion.
    """
    # New style
    try:
        from openai import OpenAI  # type: ignore

        client = OpenAI()
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=0,
        )
        txt = resp.choices[0].message.content or ""
        return json.loads(txt)
    except Exception:
        pass

    # Legacy style
    try:
        import openai  # type: ignore

        resp = openai.ChatCompletion.create(  # type: ignore[attr-defined]
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=0,
            request_timeout=timeout_s,
        )
        txt = resp["choices"][0]["message"]["content"] or ""
        return json.loads(txt)
    except Exception as e:
        raise RuntimeError(f"LLM call failed: {e}")


def _llm_available() -> bool:
    if not os.getenv("OPENAI_API_KEY"):
        return False
    try:
        import openai  # noqa: F401
        return True
    except Exception:
        try:
            from openai import OpenAI  # noqa: F401
            return True
        except Exception:
            return False


def _llm_refine(
    rows: List[Dict[str, Any]],
    full_criteria_text: str,
    a_columns: Sequence[str],
    model: str,
    log: Optional[callable] = None,
) -> List[Dict[str, Any]]:
    """LLM-assisted refinement.

    Guardrails:
    - Preserve ids and types
    - Preserve row count (1:1)
    - Allow edits to: stage/operator/target/what/label (label changes are allowed but not required)
    - Enforce targets are real columns
    - Enforce llm operator -> exactly one sentence in what
    """
    def _log(msg: str) -> None:
        if log:
            log(msg)

    _log("LLM refine: preparing prompt…")

    # Compact rows for the prompt
    compact = []
    for r in rows:
        compact.append({
            "id": r.get("id"),
            "type": r.get("type"),
            "stage": r.get("stage"),
            "label": r.get("label"),
            "operator": r.get("operator"),
            "target": r.get("target"),
            "what": r.get("what"),
            "threshold": r.get("threshold"),
            "enabled": r.get("enabled"),
            "source_text": r.get("source_text"),
        })

    system = (
        "You are a strict criteria harmoniser for PRISMA-like metadata screening.\n"
        "Return ONLY valid JSON. No extra text.\n"
        "You will receive a list of criteria rows. For each row, you may refine:\n"
        "- stage: EH/IH/EL/IL\n"
        "- operator: contains/equals/regex/in_list/not_in/gte/lte/between/llm\n"
        "- target: must be a comma-separated subset of allowed A columns\n"
        "- what: list of strings; for operator=llm MUST be exactly one short declarative sentence\n"
        "Rules:\n"
        "- Keep the SAME number of rows.\n"
        "- Do NOT change ids or types.\n"
        "- Do NOT invent targets not in the allowed A columns.\n"
        "- Be conservative: if you are unsure how to turn a criterion into heuristics, use operator=llm and keep stage EL (exclude) or IL (include).\n"
        "- Threshold: keep blank for EH/IH; for EL/IL ensure 0..1 string (default 0.60 if missing).\n"
        "Output schema: { \"rows\": [ {id,type,stage,label,operator,target,what,threshold,enabled} ... ] }\n"
    )

    user_payload = {
        "task": "Refine criteria rows based on the full criteria text and allowed A columns.",
        "allowed_a_columns": list(a_columns),
        "full_criteria_text": full_criteria_text[:8000],
        "rows": compact,
    }

    user = json.dumps(user_payload, ensure_ascii=False)

    _log(f"LLM refine: calling OpenAI model={model} …")
    d = _call_openai_json(model=model, system=system, user=user)

    if not isinstance(d, dict) or "rows" not in d or not isinstance(d["rows"], list):
        raise RuntimeError("LLM response missing 'rows' list")

    out_rows: List[Dict[str, Any]] = []
    got = d["rows"]

    if len(got) != len(rows):
        raise RuntimeError(f"LLM changed row count (expected {len(rows)}, got {len(got)})")

    # map expected ids/types
    expected = [(r.get("id"), r.get("type")) for r in rows]

    for i, rr in enumerate(got):
        if not isinstance(rr, dict):
            raise RuntimeError("LLM produced a non-object row")
        exp_id, exp_type = expected[i]
        if _safe_str(rr.get("id")).strip() != _safe_str(exp_id).strip():
            raise RuntimeError(f"LLM changed id at index {i}")
        if _safe_str(rr.get("type")).strip().lower() != _safe_str(exp_type).strip().lower():
            raise RuntimeError(f"LLM changed type at index {i}")

        nr = {
            "stage": _safe_str(rr.get("stage")).strip().upper(),
            "id": _safe_str(rr.get("id")).strip(),
            "type": _safe_str(rr.get("type")).strip().lower(),
            "scope": "metadata",
            "label": _safe_str(rr.get("label") or rows[i].get("label")).strip(),
            "operator": _safe_str(rr.get("operator")).strip().lower(),
            "target": _safe_str(rr.get("target")).strip(),
            "what": rr.get("what"),
            "threshold": _safe_str(rr.get("threshold", "")).strip(),
            "enabled": bool(rr.get("enabled", True)),
            "source_text": rows[i].get("source_text", ""),
        }

        # Normalize what
        if not isinstance(nr["what"], list):
            nr["what"] = _parse_what_cell(nr["operator"] or "contains", nr["what"])
        nr["what"] = [str(x) for x in nr["what"] if str(x).strip()]

        # Validate & canonicalize
        errs, warns = _validate_row(nr, a_columns)
        if errs:
            raise RuntimeError(f"LLM refined row invalid ({nr.get('id')}): {', '.join(errs)}")
        if warns:
            _log(f"LLM refined row warning ({nr.get('id')}): {', '.join(warns)}")

        out_rows.append(nr)

    _log("LLM refine: done.")
    return out_rows


# ============================
# UI
# ============================

@dataclass
class _UiState:
    criteria_path: str = ""
    criteria_kind: str = ""  # rtf/txt/csv/xlsx/paste
    a_path: str = ""
    a_columns: List[str] = None  # type: ignore
    a_id_col: str = ""
    text_stats: Dict[str, float] = None  # type: ignore
    criteria_text: str = ""
    rows: List[Dict[str, Any]] = None  # type: ignore

    def __post_init__(self):
        if self.a_columns is None:
            self.a_columns = []
        if self.text_stats is None:
            self.text_stats = {}
        if self.rows is None:
            self.rows = []


class HarmoniserView(ttk.Frame):
    def __init__(self, master: tk.Misc):
        super().__init__(master)
        self.state = _UiState()

        self._worker: Optional[threading.Thread] = None
        self._worker_err: Optional[str] = None
        self._worker_done: bool = False

        self._build_ui()
        self._refresh_buttons()

    # ---------------- UI build ----------------

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        # Top bar
        top = ttk.Frame(self)
        top.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))
        top.columnconfigure(1, weight=1)

        # Criteria controls
        crit_box = ttk.LabelFrame(top, text="1) Criteria (required)")
        crit_box.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        ttk.Button(crit_box, text="Load TXT/RTF…", command=self._load_criteria_text).grid(row=0, column=0, padx=6, pady=6, sticky="ew")
        ttk.Button(crit_box, text="Load table CSV/XLSX…", command=self._load_criteria_table).grid(row=0, column=1, padx=6, pady=6, sticky="ew")
        ttk.Button(crit_box, text="Clear", command=self._clear_criteria).grid(row=0, column=2, padx=6, pady=6, sticky="ew")

        self.lbl_crit = ttk.Label(crit_box, text="No criteria loaded")
        self.lbl_crit.grid(row=1, column=0, columnspan=3, padx=6, pady=(0, 6), sticky="w")

        # A vector controls
        a_box = ttk.LabelFrame(top, text="2) A vector (required)")
        a_box.grid(row=0, column=1, sticky="nsew")
        a_box.columnconfigure(0, weight=1)

        row0 = ttk.Frame(a_box)
        row0.grid(row=0, column=0, sticky="ew", padx=6, pady=6)
        row0.columnconfigure(1, weight=1)

        ttk.Button(row0, text="Load A CSV…", command=self._load_a_csv).grid(row=0, column=0, sticky="w")
        self.lbl_a = ttk.Label(row0, text="No A loaded")
        self.lbl_a.grid(row=0, column=1, sticky="w", padx=(8, 0))

        self.lbl_a_stats = ttk.Label(a_box, text="")
        self.lbl_a_stats.grid(row=1, column=0, sticky="w", padx=6, pady=(0, 6))

        # Main body: left criteria text + right table/logs
        body = ttk.Panedwindow(self, orient="horizontal")
        body.grid(row=1, column=0, sticky="nsew", padx=8, pady=(4, 8))

        # Left: criteria text editor
        left = ttk.Frame(body)
        left.columnconfigure(0, weight=1)
        left.rowconfigure(1, weight=1)

        left_top = ttk.Frame(left)
        left_top.grid(row=0, column=0, sticky="ew")
        left_top.columnconfigure(3, weight=1)

        ttk.Label(left_top, text="Criteria text (editable):").grid(row=0, column=0, sticky="w")

        self.var_llm = tk.BooleanVar(value=False)
        ttk.Checkbutton(left_top, text="LLM refine", variable=self.var_llm, command=self._refresh_buttons).grid(row=0, column=1, padx=(10, 0), sticky="w")

        ttk.Label(left_top, text="Model:").grid(row=0, column=2, padx=(10, 0), sticky="e")
        self.ent_model = ttk.Entry(left_top, width=18)
        self.ent_model.insert(0, "gpt-4o-mini")
        self.ent_model.grid(row=0, column=3, sticky="w")

        self.lbl_key = ttk.Label(left_top, text="API key: " + ("OK" if os.getenv("OPENAI_API_KEY") else "missing"))
        self.lbl_key.grid(row=0, column=4, padx=(10, 0), sticky="w")

        txt_frame = ttk.Frame(left)
        txt_frame.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
        txt_frame.rowconfigure(0, weight=1)
        txt_frame.columnconfigure(0, weight=1)

        self.txt_criteria = tk.Text(txt_frame, height=18, wrap="word")
        ysb = ttk.Scrollbar(txt_frame, orient="vertical", command=self.txt_criteria.yview)
        self.txt_criteria.configure(yscrollcommand=ysb.set)

        self.txt_criteria.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")

        body.add(left, weight=1)

        # Right: buttons + table + logs
        right = ttk.Frame(body)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)
        right.rowconfigure(3, weight=1)

        btns = ttk.Frame(right)
        btns.grid(row=0, column=0, sticky="ew")

        self.btn_harmonise = ttk.Button(btns, text="Harmonise (no-LLM)", command=self._harmonise_no_llm)
        self.btn_harmonise.grid(row=0, column=0, padx=(0, 6), pady=(0, 6), sticky="w")

        self.btn_harmonise_llm = ttk.Button(btns, text="Harmonise + LLM", command=self._harmonise_llm)
        self.btn_harmonise_llm.grid(row=0, column=1, padx=(0, 6), pady=(0, 6), sticky="w")

        self.btn_validate = ttk.Button(btns, text="Validate", command=self._validate)
        self.btn_validate.grid(row=0, column=2, padx=(0, 6), pady=(0, 6), sticky="w")

        self.btn_export = ttk.Button(btns, text="Export…", command=self._export)
        self.btn_export.grid(row=0, column=3, padx=(0, 6), pady=(0, 6), sticky="w")

        self.btn_pick_target = ttk.Button(btns, text="Pick target(s)…", command=self._pick_targets)
        self.btn_pick_target.grid(row=0, column=4, padx=(0, 6), pady=(0, 6), sticky="w")

        # Table
        table_frame = ttk.LabelFrame(right, text="Harmonised criteria")
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        cols = ("stage", "id", "type", "label", "operator", "target", "what", "threshold", "enabled")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="extended")
        for c in cols:
            self.tree.heading(c, text=c)
            width = 80
            if c == "label":
                width = 260
            elif c == "target":
                width = 220
            elif c == "what":
                width = 280
            elif c == "id":
                width = 90
            elif c == "enabled":
                width = 70
            self.tree.column(c, width=width, anchor="w")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # Row tags for highlighting
        self.tree.tag_configure("error", background="#ffe5e5")
        self.tree.tag_configure("warn", background="#fff6d5")

        self.tree.bind("<Double-1>", self._on_double_click)

        # Logs
        log_frame = ttk.LabelFrame(right, text="Log")
        log_frame.grid(row=3, column=0, sticky="nsew", pady=(6, 0))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.txt_log = tk.Text(log_frame, height=10, wrap="word", state="disabled")
        logsb = ttk.Scrollbar(log_frame, orient="vertical", command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=logsb.set)

        self.txt_log.grid(row=0, column=0, sticky="nsew")
        logsb.grid(row=0, column=1, sticky="ns")

        body.add(right, weight=2)

        # Editor widget holder
        self._edit_widget: Optional[tk.Widget] = None
        self._edit_item: Optional[str] = None
        self._edit_col: Optional[str] = None

    # ---------------- helpers ----------------

    def _log(self, msg: str) -> None:
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", f"[{_now_iso()}] {msg}\n")
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")

    def _refresh_buttons(self) -> None:
        has_crit = bool(self.state.criteria_text.strip()) or bool(self.state.rows)
        has_a = bool(self.state.a_path) and bool(self.state.a_columns)
        can_h = has_crit and has_a and (self._worker is None)

        self.btn_harmonise.configure(state=("normal" if can_h else "disabled"))

        llm_ok = _llm_available()
        self.lbl_key.configure(text="API key: " + ("OK" if llm_ok else "missing"))

        self.btn_harmonise_llm.configure(state=("normal" if (can_h and llm_ok) else "disabled"))
        self.btn_validate.configure(state=("normal" if (bool(self.state.rows) and has_a and self._worker is None) else "disabled"))
        self.btn_export.configure(state=("normal" if (bool(self.state.rows) and has_a and self._worker is None) else "disabled"))
        self.btn_pick_target.configure(state=("normal" if (bool(self.state.rows) and has_a and self._worker is None) else "disabled"))

    def _ensure_ready(self) -> bool:
        if not (self.state.a_path and self.state.a_columns):
            messagebox.showwarning("Missing A", "Please load the A vector CSV first.")
            return False
        # Either criteria_text is present (from text/paste) or rows exist (table loaded)
        if not (self.txt_criteria.get("1.0", "end").strip() or self.state.rows):
            messagebox.showwarning("Missing criteria", "Please load/paste criteria first.")
            return False
        return True

    # ---------------- load criteria ----------------

    def _load_criteria_text(self) -> None:
        path = filedialog.askopenfilename(
            title="Load criteria TXT/RTF",
            filetypes=[("Text/RTF", "*.txt *.rtf"), ("All", "*.*")],
        )
        if not path:
            return
        try:
            raw = _read_text_file(path)
            if _is_rtf_path(path):
                raw = _rtf_to_text(raw)
            raw = raw.replace("\r\n", "\n").replace("\r", "\n")
            self.state.criteria_path = path
            self.state.criteria_kind = "rtf" if _is_rtf_path(path) else "txt"
            self.state.criteria_text = raw
            self.txt_criteria.delete("1.0", "end")
            self.txt_criteria.insert("1.0", raw)
            self.state.rows = []
            self._clear_table()
            self.lbl_crit.configure(text=f"Loaded: {Path(path).name}")
            self._log(f"Criteria loaded ({self.state.criteria_kind}): {path}")
        except Exception as e:
            messagebox.showerror("Load failed", str(e))
        finally:
            self._refresh_buttons()

    def _load_criteria_table(self) -> None:
        path = filedialog.askopenfilename(
            title="Load criteria table",
            filetypes=[("CSV/XLSX", "*.csv *.xlsx *.xlsm"), ("All", "*.*")],
        )
        if not path:
            return
        try:
            raw_rows, kind = _load_structured_criteria_table(path)
            self.state.criteria_path = path
            self.state.criteria_kind = kind

            # Keep criteria text area as read-only-ish copy for LLM context
            self.txt_criteria.delete("1.0", "end")
            self.txt_criteria.insert("1.0", "\n".join([_safe_str(r) for r in raw_rows[:20]]) + ("\n…" if len(raw_rows) > 20 else ""))

            self.state.criteria_text = "\n".join([_safe_str(r) for r in raw_rows])

            # Normalize to schema but do NOT validate yet (A might not be loaded yet)
            self.state.rows = [_normalize_structured_row(r) for r in raw_rows]
            self._render_rows()
            self.lbl_crit.configure(text=f"Loaded: {Path(path).name} ({len(self.state.rows)} rows)")
            self._log(f"Criteria table loaded ({kind}): {path}")
        except Exception as e:
            messagebox.showerror("Load failed", str(e))
        finally:
            self._refresh_buttons()

    def _clear_criteria(self) -> None:
        self.state.criteria_path = ""
        self.state.criteria_kind = ""
        self.state.criteria_text = ""
        self.txt_criteria.delete("1.0", "end")
        self.state.rows = []
        self._clear_table()
        self.lbl_crit.configure(text="No criteria loaded")
        self._log("Criteria cleared")
        self._refresh_buttons()

    # ---------------- load A ----------------

    def _load_a_csv(self) -> None:
        path = filedialog.askopenfilename(
            title="Load A vector CSV",
            filetypes=[("CSV", "*.csv"), ("All", "*.*")],
        )
        if not path:
            return
        try:
            cols, stats = _load_a_header_and_stats(path)
            self.state.a_path = path
            self.state.a_columns = cols
            self.state.text_stats = stats
            self.state.a_id_col = _detect_id_column(cols)

            self.lbl_a.configure(text=f"Loaded: {Path(path).name} ({len(cols)} cols), id={self.state.a_id_col}")

            if stats:
                s = " / ".join([f"{k}:{stats[k]*100:.0f}%" for k in stats])
                self.lbl_a_stats.configure(text=f"Text coverage (sample): {s}")
            else:
                self.lbl_a_stats.configure(text="")

            self._log(f"A loaded: {path}")

            # If we already have rows, validate/canonicalize targets now
            if self.state.rows:
                self._validate(show_ok=False)

        except Exception as e:
            messagebox.showerror("Load failed", str(e))
        finally:
            self._refresh_buttons()

    # ---------------- harmonise ----------------

    def _harmonise_no_llm(self) -> None:
        if not self._ensure_ready():
            return

        # If criteria table already loaded, just validate/canonicalize with A and infer missing stage/operator if blank.
        if self.state.rows:
            self._log("Harmonise (no-LLM): normalizing existing rows…")
            for r in self.state.rows:
                # Fill missing stage/operator/target/threshold conservatively
                if not _safe_str(r.get("stage")).strip():
                    r["stage"] = "IL" if _safe_str(r.get("type")).lower() == "include" else "EL"
                if not _safe_str(r.get("operator")).strip():
                    r["operator"] = "llm"
                if not _safe_str(r.get("target")).strip():
                    # default to available text targets
                    target = DEFAULT_TEXT_TARGET
                    target, _ = _canonicalize_targets(target, self.state.a_columns)
                    r["target"] = target
                # Ensure threshold
                if _safe_str(r.get("stage")).upper() in {"EL", "IL"}:
                    if not _safe_str(r.get("threshold")).strip():
                        r["threshold"] = f"{DEFAULT_THRESHOLD:.2f}"
                else:
                    r["threshold"] = ""

            self._render_rows()
            self._validate(show_ok=True)
            return

        # Else parse free text
        text = self.txt_criteria.get("1.0", "end")
        parsed = _parse_free_text_criteria(text)
        if not parsed:
            messagebox.showerror("No criteria found", "No IC-/EC- lines detected. Check the formatting.")
            return

        self._log(f"Harmonise (no-LLM): parsing {len(parsed)} criteria…")

        rows = []
        for crit_id, crit_type, label, source_line in parsed:
            row = _auto_harmonize_line(
                crit_id=crit_id,
                crit_type=crit_type,
                label=label,
                source_text=source_line,
                a_columns=self.state.a_columns,
                text_field_stats=self.state.text_stats,
            )
            rows.append(row)

        self.state.rows = rows
        self._render_rows()
        self._validate(show_ok=True)

    def _harmonise_llm(self) -> None:
        if not self._ensure_ready():
            return
        if not _llm_available():
            messagebox.showwarning("LLM unavailable", "OPENAI_API_KEY missing or OpenAI package not available.")
            return

        # Ensure we have baseline rows first
        if not self.state.rows:
            self._harmonise_no_llm()
            if not self.state.rows:
                return

        model = self.ent_model.get().strip() or "gpt-4o-mini"
        full_text = self.txt_criteria.get("1.0", "end").strip() or self.state.criteria_text

        def worker():
            try:
                refined = _llm_refine(self.state.rows, full_text, self.state.a_columns, model=model, log=self._thread_log)
                self.state.rows = refined
                self._worker_err = None
            except Exception as e:
                self._worker_err = str(e)
            finally:
                self._worker_done = True

        self._start_worker(worker, "LLM harmonisation")

    # ---------------- worker helpers ----------------

    def _thread_log(self, msg: str) -> None:
        # Thread-safe log using after
        self.after(0, lambda: self._log(msg))

    def _start_worker(self, target, label: str) -> None:
        if self._worker is not None:
            return
        self._worker_done = False
        self._worker_err = None

        self._log(f"Starting: {label} …")
        self._worker = threading.Thread(target=target, daemon=True)
        self._worker.start()
        self._refresh_buttons()
        self.after(100, self._poll_worker)

    def _poll_worker(self) -> None:
        if self._worker is None:
            return
        if not self._worker_done:
            self.after(150, self._poll_worker)
            return

        # Done
        err = self._worker_err
        self._worker = None
        self._worker_done = False

        if err:
            self._log(f"Worker failed: {err}")
            messagebox.showerror("Operation failed", err)
        else:
            self._log("Worker finished successfully")
            self._render_rows()
            self._validate(show_ok=True)

        self._refresh_buttons()

    # ---------------- validate/export ----------------

    def _validate(self, show_ok: bool = True) -> bool:
        if not self.state.rows:
            return False
        if not self.state.a_columns:
            messagebox.showwarning("Missing A", "Load A vector first.")
            return False

        n_err = 0
        n_warn = 0

        for item in self.tree.get_children():
            self.tree.item(item, tags=())

        # Validate every row and tag items
        for item_id in self.tree.get_children():
            idx = int(self.tree.set(item_id, "__idx") or "-1") if "__idx" in self.tree["columns"] else -1
            # We don't store __idx as a visible column; fallback to mapping

        # We'll validate against state.rows and tag in render step
        for i, r in enumerate(self.state.rows):
            errs, warns = _validate_row(r, self.state.a_columns)
            if errs:
                n_err += 1
            if warns:
                n_warn += 1

        self._render_rows(with_validation=True)

        self._log(f"Validate: {len(self.state.rows)} rows, errors={n_err}, warnings={n_warn}")

        if n_err > 0:
            if show_ok:
                messagebox.showerror("Validation failed", f"{n_err} row(s) have errors. Fix them before export.")
            return False

        if show_ok:
            messagebox.showinfo("Validation OK", f"All good. Warnings: {n_warn}")
        return True

    def _export(self) -> None:
        if not self.state.rows:
            messagebox.showwarning("Nothing to export", "Harmonise criteria first.")
            return
        if not self._validate(show_ok=False):
            messagebox.showerror("Export blocked", "Fix validation errors before export.")
            return

        outdir = filedialog.askdirectory(title="Choose export folder")
        if not outdir:
            return

        base = "criteria_harmonized"
        try:
            # Combined
            csv_path = str(Path(outdir) / f"{base}.csv")
            txt_path = str(Path(outdir) / f"{base}.txt")
            _export_csv(self.state.rows, csv_path)
            _export_pipe(self.state.rows, txt_path)

            # Per-stage CSV
            for st in STAGES:
                st_rows = _stage_filter(self.state.rows, st)
                if not st_rows:
                    continue
                _export_csv(st_rows, str(Path(outdir) / f"criteria_{st}.csv"))

            self._log(f"Exported to: {outdir}")
            messagebox.showinfo("Export done", f"Exported combined CSV+TXT and per-stage CSV to:\n{outdir}")

        except Exception as e:
            messagebox.showerror("Export failed", str(e))

    # ---------------- target picker ----------------

    def _pick_targets(self) -> None:
        if not self.state.a_columns:
            messagebox.showwarning("Missing A", "Load A vector first.")
            return
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Pick targets", "Select one or more rows first.")
            return

        win = tk.Toplevel(self)
        win.title("Pick target columns")
        win.geometry("420x500")

        ttk.Label(win, text="Select one or more A columns (Ctrl/Shift):").pack(anchor="w", padx=10, pady=(10, 4))

        lb = tk.Listbox(win, selectmode="extended")
        lb.pack(fill="both", expand=True, padx=10, pady=6)

        for c in self.state.a_columns:
            lb.insert("end", c)

        btn_row = ttk.Frame(win)
        btn_row.pack(fill="x", padx=10, pady=10)

        def apply():
            picks = [lb.get(i) for i in lb.curselection()]
            if not picks:
                messagebox.showwarning("No selection", "Pick at least one column")
                return
            tgt = ",".join(picks)
            # Apply to selected rows
            for it in sel:
                rid = self.tree.set(it, "id")
                r = self._find_row_by_id(rid)
                if r is not None:
                    r["target"], _ = _canonicalize_targets(tgt, self.state.a_columns)
            self._render_rows(with_validation=True)
            win.destroy()

        ttk.Button(btn_row, text="Apply", command=apply).pack(side="left")
        ttk.Button(btn_row, text="Cancel", command=win.destroy).pack(side="right")

    # ---------------- table rendering/editing ----------------

    def _clear_table(self) -> None:
        for it in self.tree.get_children():
            self.tree.delete(it)

    def _find_row_by_id(self, rid: str) -> Optional[Dict[str, Any]]:
        rid = rid.strip()
        for r in self.state.rows:
            if _safe_str(r.get("id")).strip() == rid:
                return r
        return None

    def _render_rows(self, with_validation: bool = False) -> None:
        self._clear_table()
        if not self.state.rows:
            self._refresh_buttons()
            return

        for r in self.state.rows:
            # Ensure threshold lock for EH/IH
            st = _safe_str(r.get("stage")).upper()
            if st in {"EH", "IH"}:
                r["threshold"] = ""
            else:
                if not _safe_str(r.get("threshold")).strip():
                    r["threshold"] = f"{DEFAULT_THRESHOLD:.2f}"

        for r in self.state.rows:
            vals = (
                r.get("stage", ""),
                r.get("id", ""),
                r.get("type", ""),
                _safe_str(r.get("label", ""))[:200],
                r.get("operator", ""),
                r.get("target", ""),
                _what_to_export(r.get("operator", ""), r.get("what", []) or [])[:260],
                r.get("threshold", ""),
                "1" if bool(r.get("enabled", True)) else "0",
            )

            tags = ()
            if with_validation and self.state.a_columns:
                errs, warns = _validate_row(r, self.state.a_columns)
                if errs:
                    tags = ("error",)
                elif warns:
                    tags = ("warn",)

            self.tree.insert("", "end", values=vals, tags=tags)

        self._refresh_buttons()

    def _on_double_click(self, event) -> None:
        # Identify cell
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        item = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not item or not col:
            return

        col_index = int(col.replace("#", "")) - 1
        columns = list(self.tree["columns"])
        if col_index < 0 or col_index >= len(columns):
            return

        col_name = columns[col_index]
        if col_name not in {"stage", "label", "operator", "target", "what", "threshold", "enabled"}:
            return

        # enabled toggle
        if col_name == "enabled":
            rid = self.tree.set(item, "id")
            r = self._find_row_by_id(rid)
            if r is not None:
                r["enabled"] = not bool(r.get("enabled", True))
                self._render_rows(with_validation=True)
            return

        # Clean existing editor
        self._destroy_editor()

        bbox = self.tree.bbox(item, col)
        if not bbox:
            return
        x, y, w, h = bbox

        rid = self.tree.set(item, "id")
        row = self._find_row_by_id(rid)
        if row is None:
            return

        # threshold editable only for EL/IL
        if col_name == "threshold":
            st = _safe_str(row.get("stage")).upper()
            if st not in {"EL", "IL"}:
                return

        value = self.tree.set(item, col_name)

        if col_name == "stage":
            cb = ttk.Combobox(self.tree, values=list(STAGES), state="readonly")
            cb.set(value or "")
            cb.place(x=x, y=y, width=w, height=h)
            cb.focus_set()

            def save(_=None):
                v = cb.get().strip().upper()
                if v in STAGES:
                    row["stage"] = v
                    # manage threshold lock
                    if v in {"EH", "IH"}:
                        row["threshold"] = ""
                    else:
                        if not _safe_str(row.get("threshold")).strip():
                            row["threshold"] = f"{DEFAULT_THRESHOLD:.2f}"
                self._destroy_editor()
                self._render_rows(with_validation=True)

            cb.bind("<<ComboboxSelected>>", save)
            cb.bind("<Return>", save)
            cb.bind("<Escape>", lambda _=None: self._destroy_editor())

            self._edit_widget = cb
            return

        if col_name == "operator":
            cb = ttk.Combobox(self.tree, values=list(OPERATORS), state="readonly")
            cb.set(value or "")
            cb.place(x=x, y=y, width=w, height=h)
            cb.focus_set()

            def save(_=None):
                v = cb.get().strip().lower()
                if v in OPERATORS:
                    row["operator"] = v
                    # If operator=llm ensure single sentence what
                    if v == "llm":
                        what_list = row.get("what") or []
                        if isinstance(what_list, list):
                            if len(what_list) != 1:
                                row["what"] = [(_safe_str(row.get("label")) or _safe_str(row.get("source_text")) or "").strip()][:1]
                        else:
                            row["what"] = [str(what_list).strip()]
                self._destroy_editor()
                self._render_rows(with_validation=True)

            cb.bind("<<ComboboxSelected>>", save)
            cb.bind("<Return>", save)
            cb.bind("<Escape>", lambda _=None: self._destroy_editor())

            self._edit_widget = cb
            return

        # Entry editor
        ent = ttk.Entry(self.tree)
        ent.insert(0, value)
        ent.place(x=x, y=y, width=w, height=h)
        ent.focus_set()

        def save(_=None):
            v = ent.get().strip()
            if col_name == "label":
                row["label"] = v
            elif col_name == "target":
                # canonicalize against A
                canon, _unk = _canonicalize_targets(v, self.state.a_columns)
                row["target"] = canon
            elif col_name == "what":
                row["what"] = _parse_what_cell(row.get("operator", "contains"), v)
            elif col_name == "threshold":
                row["threshold"] = v
            self._destroy_editor()
            self._render_rows(with_validation=True)

        ent.bind("<Return>", save)
        ent.bind("<Escape>", lambda _=None: self._destroy_editor())
        ent.bind("<FocusOut>", save)

        self._edit_widget = ent

    def _destroy_editor(self) -> None:
        try:
            if self._edit_widget is not None:
                self._edit_widget.destroy()
        except Exception:
            pass
        finally:
            self._edit_widget = None


# ============================
# Hub plugin wrapper
# ============================

def create_plugin(app):
    return HarmoniserPlugin(app, PluginMeta(id="harmoniser", title="Harmoniser (Criteria)"))


class HarmoniserPlugin(BasePlugin):
    def __init__(self, app, meta: PluginMeta):
        super().__init__(app, meta)
        self.view: Optional[HarmoniserView] = None

    def build_tab(self, parent):
        frame = ttk.Frame(parent)
        self.view = HarmoniserView(frame)
        self.view.pack(fill="both", expand=True)
        return frame

    def on_close(self):
        # Best-effort cleanup
        try:
            if self.view:
                self.view.destroy()
        except Exception:
            pass
        self.view = None
