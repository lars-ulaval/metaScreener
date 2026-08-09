# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""tools/eval_ingest.py - ingest filled adjudication grids and compute agreement metrics.

Reads:
  - The partition manifest produced by tools/eval_grid_generator.py
    (defines which a_id was assigned to which rater, per stage).
  - The filled XLSX workbooks (one per rater) that the partition manifest
    says should exist.
  - The stage-filtered CSVs (post-EL bundle, post-IL bundle) whose
    *_evidence_json columns hold the LLM's per-criterion call.

Performs:
  - Schema validation: each filled workbook must contain exactly the rows
    the manifest assigned to that rater, with all decisions populated and
    drawn from the {include, exclude, uncertain} vocabulary. Empty cells,
    typos, and added/removed rows are reported with informative errors.
  - Long-format pivot: each (rater, stage, a_id, criterion) tuple gets one
    row recording human_decision and human_notes.
  - LLM join: each row is paired with the LLM's status for the same
    (stage, a_id, criterion) tuple, mapped to the canonical decision
    vocabulary (MET -> include, FAILED -> exclude, UNCERTAIN -> uncertain).
  - Cohen's kappa (human vs LLM): for every (stage, criterion) pair,
    computed on the union of (overlap-majority-vote rows) + (disjoint rows
    from each rater).
  - Fleiss' kappa (inter-human): for every (stage, criterion) pair,
    computed on the overlap subset only (the set of records every rater
    rated independently).
  - Per-criterion confusion matrices (human x LLM).

Writes:
  - eval_decisions_v1.csv   : every (rater, stage, a_id, criterion,
                              human_decision, human_notes) tuple.
  - eval_results_v1.csv     : same plus llm_status, llm_decision_canonical,
                              quote, quote_valid, agree flag.
  - eval_disagreements_v1.csv : the agree=False subset, sorted, for spot-check.
  - eval_summary_v1.txt     : human-readable kappa values, confusion matrices,
                              N's, limitations.

The tool is designed to be re-runnable. If you regenerate the grids with
a new seed and re-collect, the outputs simply overwrite. The C0 manifest
plus the filled XLSX files form a sufficient input set; nothing else from
the original generator run is required.

Usage:
    python tools/eval_ingest.py \\
        --manifest         docs/data/grids/partition_manifest.csv \\
        --criteria         tests/golden/criteria_harmonized_v3.1.0.csv \\
        --filled-grids-dir docs/data/grids/filled/ \\
        --el-filtered      tests/golden/el_filtered_v3.1.0.csv \\
        --il-filtered      tests/golden/il_filtered_v3.1.0.csv \\
        --output-dir       docs/data/

The tool exits non-zero on any schema-validation failure; the metrics step
runs only after every workbook validates cleanly.
"""
import argparse
import csv
import json
import math
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

# Canonical short codes used internally. The XLSX dropdowns (built by
# eval_grid_generator.py) present full sentences quoting the criterion
# text verbatim - "YES - this is true: <criterion>", "NO - this is not
# true: <criterion>", "I cannot tell from the abstract alone." The
# normalize_decision() function below collapses those long strings to
# the canonical codes here.
#
# These codes are deliberately distinct from any prior include/exclude/
# uncertain vocabulary in the codebase so that any stale code path
# fails loudly rather than silently miscoercing values.
CANONICAL_DECISIONS = ("yes", "no", "unsure")

# Map LLM evidence_json status -> canonical decision.
# The status field is the polarity-aware terminal call (already takes
# the criterion's threshold and inclusion/exclusion polarity into
# account), making it the correct join key for the human's yes/no/unsure
# rating, regardless of whether the criterion is IC- or EC-.
# --------------------------------------------------------------------------
# LLM status -> canonical decision: POLARITY-AWARE.
# --------------------------------------------------------------------------
#
# metaScreener's `status` field describes whether a record PASSES OR FAILS
# the screening RULE, not whether the criterion's claim holds of the paper.
# These coincide for inclusion criteria but invert for exclusion criteria.
#
#   For an INCLUSION criterion (e.g., "the paper considers HMD VR"):
#     status=MET    => paper passes the inclusion check
#                   => the criterion's claim IS true of the paper
#                   => canonical decision: "yes"
#     status=FAILED => paper fails the inclusion check
#                   => the criterion's claim is NOT true
#                   => canonical: "no"
#
#   For an EXCLUSION criterion (e.g., "primary focus is rubber hand illusion"):
#     status=MET    => paper passes the exclusion check (survives, not excluded)
#                   => the criterion's claim is NOT true of the paper
#                   => canonical: "no"
#     status=FAILED => paper fails the exclusion check (gets excluded)
#                   => the criterion's claim IS true of the paper
#                   => canonical: "yes"
#
# UNCERTAIN -> "unsure" regardless of polarity.
#
# The human rater always answers "is the criterion's claim true of this
# paper?" (yes/no/unsure). The polarity-aware mapping above ensures the
# LLM-side and human-side decisions are about the same question, so the
# resulting kappa measures actual agreement on the underlying judgement
# rather than artefacts of polarity translation.
_STATUS_MAP_INCLUDE = {"MET": "yes", "FAILED": "no", "UNCERTAIN": "unsure"}
_STATUS_MAP_EXCLUDE = {"MET": "no",  "FAILED": "yes", "UNCERTAIN": "unsure"}


def status_to_canonical(status: str, polarity: str) -> str:
    """Map an LLM evidence_json status to a canonical decision short code,
    taking the criterion's polarity into account.

    Args:
        status: One of "MET", "FAILED", "UNCERTAIN" (case-insensitive).
                Unknown statuses map to "unsure" (conservative default).
        polarity: One of "include", "exclude" (case-insensitive).
                  Unknown polarities default to "include" semantics.

    See module-level documentation for the full semantics.
    """
    s = str(status).strip().upper()
    p = str(polarity).strip().lower()
    table = _STATUS_MAP_EXCLUDE if p == "exclude" else _STATUS_MAP_INCLUDE
    return table.get(s, "unsure")


# Prefix patterns the rater dropdown options begin with. normalize_decision
# matches these case-insensitively. The criterion-specific tail of each
# YES/NO option is ignored - we only need the polarity prefix to canonicalize.
DROPDOWN_PREFIXES = {
    "yes":    ("yes -", "yes:", "yes,"),
    "no":     ("no -", "no:", "no,"),
    "unsure": ("i cannot tell", "cannot tell"),
}

STAGE_EVIDENCE_COL = {"EL": "el_evidence_json", "IL": "il_evidence_json"}
STAGE_SHEET_TITLE  = {"EL": "Decisions_EL", "IL": "Decisions_IL"}
STAGE_DISPLAY_ORDER = ("EL", "IL")

# Output row schema for eval_results_v1.csv.
RESULT_FIELDS = [
    "stage", "a_id", "criterion_id", "set", "rater_id",
    "human_decision", "human_notes",
    "llm_status", "llm_decision_canonical",
    "llm_quote", "llm_quote_valid", "llm_confidence",
    "agree",
]


# --------------------------------------------------------------------------
# Manifest + LLM evidence loaders
# --------------------------------------------------------------------------

def load_partition_manifest(path: Path) -> List[dict]:
    """Load the manifest written by eval_grid_generator.py.

    Returns the rows as-is (one per (stage, a_id, rater_id) tuple).
    """
    if not path.exists():
        raise FileNotFoundError(f"Manifest not found: {path}")
    with path.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        raise ValueError(f"Manifest is empty: {path}")
    required = {"stage", "a_id", "rater_id", "set"}
    missing = required - set(rows[0].keys())
    if missing:
        raise ValueError(f"Manifest missing columns: {sorted(missing)}")
    return rows


def index_manifest_by_rater(manifest: List[dict]) -> Dict[str, Dict[str, List[dict]]]:
    """Index manifest as {rater_id: {stage: [rows]}}."""
    out: Dict[str, Dict[str, List[dict]]] = defaultdict(lambda: defaultdict(list))
    for row in manifest:
        out[row["rater_id"]][row["stage"]].append(row)
    return {k: dict(v) for k, v in out.items()}


def load_llm_evidence(
    filtered_path: Path,
    stage: str,
) -> Dict[str, Dict[str, dict]]:
    """Load *_evidence_json from a stage-filtered CSV.

    Returns: {a_id: {criterion_id: {status, decision, confidence, quote,
                                    quote_valid, ...}}}
    """
    if not filtered_path.exists():
        raise FileNotFoundError(f"{stage} filtered CSV not found: {filtered_path}")
    ev_col = STAGE_EVIDENCE_COL[stage]
    out: Dict[str, Dict[str, dict]] = {}
    with filtered_path.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        raise ValueError(f"{stage} filtered CSV is empty: {filtered_path}")
    if ev_col not in rows[0]:
        raise ValueError(
            f"{stage} filtered CSV missing column {ev_col!r}: {filtered_path}"
        )
    for r in rows:
        a_id = (r.get("local_id") or "").strip()
        ev_str = r.get(ev_col, "") or ""
        if not ev_str:
            continue
        try:
            parsed = json.loads(ev_str)
        except json.JSONDecodeError as exc:
            raise ValueError(
                f"Malformed {ev_col} for a_id={a_id} in {filtered_path}: {exc}"
            ) from exc
        if not isinstance(parsed, dict):
            raise ValueError(
                f"{ev_col} for a_id={a_id} is not a JSON object"
            )
        out[a_id] = parsed
    return out


# --------------------------------------------------------------------------
# Workbook reader
# --------------------------------------------------------------------------

def normalize_decision(value) -> Optional[str]:
    """Coerce an Excel cell value to a canonical decision short code.

    The XLSX dropdowns present full sentences:
      "YES - this is true: <criterion text>"
      "NO - this is not true: <criterion text>"
      "I cannot tell from the abstract alone."

    This function strips whitespace, lowercases, and matches against the
    DROPDOWN_PREFIXES map to produce one of "yes" / "no" / "unsure".

    Returns None if the cell is empty or whitespace-only (caller decides
    whether that's an error). Raises ValueError if the cell contains
    text that does not match any canonical prefix - typically a sign
    that the rater typed a custom value or that the file was corrupted.

    Backwards-tolerant: also accepts the bare canonical codes ("yes",
    "no", "unsure") as direct values, in case a downstream tool feeds
    pre-normalized data through this function.
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    s_lower = s.lower()
    # Direct canonical match (e.g., synthetic test data, programmatic input).
    if s_lower in CANONICAL_DECISIONS:
        return s_lower
    # Prefix-based match for full-sentence dropdown values.
    for canonical, prefixes in DROPDOWN_PREFIXES.items():
        for prefix in prefixes:
            if s_lower.startswith(prefix):
                return canonical
    raise ValueError(
        f"Decision value {value!r} is not recognised. Expected a YES/NO/"
        "I-cannot-tell dropdown sentence, or one of "
        f"{CANONICAL_DECISIONS}."
    )


def read_filled_workbook(
    path: Path,
    rater_id: str,
    expected_per_stage: Dict[str, List[dict]],
) -> List[dict]:
    """Read and validate a filled rater workbook.

    expected_per_stage maps stage -> list of manifest rows for this rater.
    The workbook MUST contain Decisions_EL and Decisions_IL sheets matching
    those manifest rows exactly, with every decision cell populated and
    drawn from the canonical vocabulary.

    Returns long-format dict rows: each row has stage, a_id, criterion_id,
    rater_id, set, human_decision, human_notes.

    Raises ValueError with informative diagnostics on any schema problem.
    """
    if not path.exists():
        raise FileNotFoundError(f"Filled workbook not found: {path}")
    wb = load_workbook(path, data_only=True)

    out: List[dict] = []
    errors: List[str] = []

    for stage, manifest_rows in expected_per_stage.items():
        sheet_name = STAGE_SHEET_TITLE[stage]
        if sheet_name not in wb.sheetnames:
            errors.append(f"missing sheet {sheet_name!r}")
            continue
        ws = wb[sheet_name]

        # Header parsing.
        headers = [c.value for c in ws[1]]
        try:
            id_col_idx = headers.index("ID") + 1
        except ValueError:
            errors.append(f"sheet {sheet_name!r}: header row missing 'ID' column")
            continue

        decision_columns: List[Tuple[str, int, int]] = []
        for idx, h in enumerate(headers, start=1):
            if isinstance(h, str) and h.startswith("Decision: "):
                cid = h[len("Decision: "):].strip()
                # The companion notes column is the next one (Notes: <cid>).
                notes_idx = idx + 1
                notes_h = headers[idx] if idx < len(headers) else None  # 0-based
                if not (isinstance(notes_h, str) and notes_h.startswith("Notes: ")):
                    errors.append(
                        f"sheet {sheet_name!r}: 'Decision: {cid}' not "
                        f"followed by a 'Notes: {cid}' column"
                    )
                decision_columns.append((cid, idx, notes_idx))

        if not decision_columns:
            errors.append(f"sheet {sheet_name!r}: no 'Decision: <id>' columns found")
            continue

        # Build a_id index from data rows.
        rows_by_aid: Dict[str, int] = {}
        for r_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=r_idx, column=id_col_idx).value
            if cell is None:
                continue
            a_id = str(cell).strip()
            if not a_id:
                continue
            if a_id in rows_by_aid:
                errors.append(
                    f"sheet {sheet_name!r}: duplicate a_id {a_id} on rows "
                    f"{rows_by_aid[a_id]} and {r_idx}"
                )
                continue
            rows_by_aid[a_id] = r_idx

        # Validate the row set matches the manifest exactly.
        expected_aids = {row["a_id"] for row in manifest_rows}
        actual_aids = set(rows_by_aid.keys())
        spurious = actual_aids - expected_aids
        missing = expected_aids - actual_aids
        if spurious:
            errors.append(
                f"sheet {sheet_name!r}: workbook contains rows not in the "
                f"manifest for {rater_id}: {sorted(spurious)}"
            )
        if missing:
            errors.append(
                f"sheet {sheet_name!r}: manifest expects rows missing from "
                f"the workbook for {rater_id}: {sorted(missing)}"
            )

        # Build a fast lookup of (a_id, set) from the manifest.
        manifest_set: Dict[str, str] = {row["a_id"]: row["set"] for row in manifest_rows}

        # Collect per-row decisions.
        for a_id, r_idx in rows_by_aid.items():
            if a_id not in manifest_set:
                # already reported as spurious; skip.
                continue
            for cid, dec_idx, notes_idx in decision_columns:
                raw_dec = ws.cell(row=r_idx, column=dec_idx).value
                try:
                    dec = normalize_decision(raw_dec)
                except ValueError as exc:
                    errors.append(
                        f"sheet {sheet_name!r}, a_id={a_id}, criterion={cid}: {exc}"
                    )
                    continue
                if dec is None:
                    errors.append(
                        f"sheet {sheet_name!r}, a_id={a_id}, criterion={cid}: "
                        "decision is empty (every cell must be filled)"
                    )
                    continue
                notes_val = ws.cell(row=r_idx, column=notes_idx).value
                notes = "" if notes_val is None else str(notes_val).strip()
                out.append({
                    "stage": stage,
                    "a_id": a_id,
                    "criterion_id": cid,
                    "rater_id": rater_id,
                    "set": manifest_set[a_id],
                    "human_decision": dec,
                    "human_notes": notes,
                })

    if errors:
        msg = (
            f"\n{len(errors)} schema problem(s) in {path.name} "
            f"(rater={rater_id}):\n  - "
            + "\n  - ".join(errors)
        )
        raise ValueError(msg)

    return out


# --------------------------------------------------------------------------
# Join: human decisions x LLM evidence
# --------------------------------------------------------------------------

def join_with_llm(
    human_rows: List[dict],
    llm_evidence: Dict[str, Dict[str, Dict[str, dict]]],
    criteria_polarity: Dict[str, str],
) -> List[dict]:
    """Pair each human-decision row with the LLM's status for the same key.

    Args:
        human_rows: long-format rows from read_filled_workbook (one per
            (rater, stage, a_id, criterion) tuple).
        llm_evidence: {stage: {a_id: {criterion_id: evidence_dict}}}.
        criteria_polarity: {criterion_id: "include" | "exclude"}. Required
            for polarity-aware status mapping. Missing entries default to
            "include" semantics, which matches the legacy single-map
            behaviour for that one degenerate case.

    Returns rows with the full RESULT_FIELDS schema. Records where the
    LLM evidence is missing for a (stage, a_id, criterion) tuple get
    llm_status='MISSING' and agree=False (treated as disagreement so the
    metric reflects rather than masks the absence).
    """
    out: List[dict] = []
    for r in human_rows:
        stage = r["stage"]
        a_id = r["a_id"]
        cid = r["criterion_id"]
        polarity = criteria_polarity.get(cid, "include")
        ev = llm_evidence.get(stage, {}).get(a_id, {}).get(cid)
        if ev is None:
            llm_status = "MISSING"
            llm_can = "unsure"  # default sink; see comment above
            quote = ""
            quote_valid = ""
            confidence = ""
        else:
            llm_status = str(ev.get("status", "")).strip().upper()
            llm_can = status_to_canonical(llm_status, polarity)
            quote = str(ev.get("quote", "") or "")
            quote_valid_raw = ev.get("quote_valid", None)
            quote_valid = "" if quote_valid_raw is None else str(bool(quote_valid_raw)).lower()
            conf_raw = ev.get("confidence", None)
            confidence = "" if conf_raw is None else f"{float(conf_raw):.3f}"
        agree = (r["human_decision"] == llm_can) and (llm_status != "MISSING")
        out.append({
            "stage": stage,
            "a_id": a_id,
            "criterion_id": cid,
            "set": r["set"],
            "rater_id": r["rater_id"],
            "human_decision": r["human_decision"],
            "human_notes": r.get("human_notes", ""),
            "llm_status": llm_status,
            "llm_decision_canonical": llm_can,
            "llm_quote": quote,
            "llm_quote_valid": quote_valid,
            "llm_confidence": confidence,
            "agree": "True" if agree else "False",
        })
    return out


# --------------------------------------------------------------------------
# Agreement metrics (pure-python implementations)
# --------------------------------------------------------------------------

def _confusion_matrix(
    pairs: List[Tuple[str, str]],
    categories: Tuple[str, ...] = CANONICAL_DECISIONS,
) -> List[List[int]]:
    """Build a categories x categories confusion matrix from (rater_a, rater_b) pairs.

    Every label must be a member of `categories`. Anything else raises:
    this function used to skip unrecognised labels, which combined with
    cohen_kappa's n = len(pairs) to shift kappa silently (F-07). A
    published agreement statistic must not degrade quietly.
    """
    idx = {c: i for i, c in enumerate(categories)}
    n = len(categories)
    matrix = [[0] * n for _ in range(n)]
    for a, b in pairs:
        for label in (a, b):
            if label not in idx:
                raise ValueError(
                    "decision %r is not one of %s - a rater vocabulary or "
                    "status-mapping bug upstream would otherwise be absorbed "
                    "silently into the kappa" % (label, list(categories))
                )
        matrix[idx[a]][idx[b]] += 1
    return matrix


def cohen_kappa(
    pairs: List[Tuple[str, str]],
    categories: Tuple[str, ...] = CANONICAL_DECISIONS,
) -> Dict[str, float]:
    """Cohen's kappa for two raters over a list of (rater_a, rater_b) decisions.

    Returns {n, agreement_observed, agreement_expected, kappa}. Kappa is
    NaN when n=0 or expected agreement is 1.0 (i.e., one category dominates
    completely; kappa is undefined).
    """
    n = len(pairs)
    if n == 0:
        return {"n": 0, "agreement_observed": float("nan"),
                "agreement_expected": float("nan"), "kappa": float("nan")}
    cm = _confusion_matrix(pairs, categories)
    n_cat = len(categories)
    diag_sum = sum(cm[i][i] for i in range(n_cat))
    p_o = diag_sum / n
    row_totals = [sum(cm[i]) for i in range(n_cat)]
    col_totals = [sum(cm[i][j] for i in range(n_cat)) for j in range(n_cat)]
    p_e = sum((row_totals[i] / n) * (col_totals[i] / n) for i in range(n_cat))
    if math.isclose(p_e, 1.0):
        kappa = float("nan")
    else:
        kappa = (p_o - p_e) / (1.0 - p_e)
    return {
        "n": n,
        "agreement_observed": p_o,
        "agreement_expected": p_e,
        "kappa": kappa,
    }


def fleiss_kappa(
    rating_matrix: List[List[int]],
    categories: Tuple[str, ...] = CANONICAL_DECISIONS,
) -> Dict[str, float]:
    """Fleiss' kappa for N raters over M items, each rated into one of K categories.

    rating_matrix is M x K, where entry [i][k] = number of raters who put
    item i in category k. Every row must sum to the same value (the rater
    count per item).

    Returns {n_items, n_raters, agreement_observed, agreement_expected, kappa}.
    Kappa is NaN if n_items=0 or all items received the same category from
    all raters (perfect agreement degenerates the formula).
    """
    if not rating_matrix:
        return {"n_items": 0, "n_raters": 0,
                "agreement_observed": float("nan"),
                "agreement_expected": float("nan"),
                "kappa": float("nan")}
    n_items = len(rating_matrix)
    n_cat = len(categories)
    row_sums = [sum(row) for row in rating_matrix]
    if not all(s == row_sums[0] for s in row_sums):
        raise ValueError(
            f"Fleiss' kappa: every item must have the same rater count; got {row_sums}"
        )
    n_raters = row_sums[0]
    if n_raters < 2:
        raise ValueError(
            f"Fleiss' kappa: need at least 2 raters per item, got {n_raters}"
        )

    # Per-item agreement (P_i).
    P_i: List[float] = []
    for i in range(n_items):
        s = sum(rating_matrix[i][k] * (rating_matrix[i][k] - 1) for k in range(n_cat))
        P_i.append(s / (n_raters * (n_raters - 1)))
    P_bar = sum(P_i) / n_items

    # Category proportions across all ratings (p_j).
    total_ratings = n_items * n_raters
    p_j = [
        sum(rating_matrix[i][k] for i in range(n_items)) / total_ratings
        for k in range(n_cat)
    ]
    P_e = sum(p * p for p in p_j)

    if math.isclose(P_e, 1.0):
        kappa = float("nan")
    else:
        kappa = (P_bar - P_e) / (1.0 - P_e)
    return {
        "n_items": n_items,
        "n_raters": n_raters,
        "agreement_observed": P_bar,
        "agreement_expected": P_e,
        "kappa": kappa,
    }


# --------------------------------------------------------------------------
# Aggregation: build per-(stage, criterion) pairings + matrices
# --------------------------------------------------------------------------

def majority_vote(decisions: List[str]) -> str:
    """Return the most common decision; ties resolve to 'unsure' to be conservative.

    The return value is always a member of CANONICAL_DECISIONS. It used to
    be the literal "uncertain" on a tie or an empty input, which is not a
    canonical code: _confusion_matrix would then drop the resulting pair
    while cohen_kappa still counted it in n, silently deflating kappa.
    """
    if not decisions:
        return "unsure"
    counts = Counter(decisions)
    top = counts.most_common()
    if len(top) > 1 and top[0][1] == top[1][1]:
        return "unsure"
    return top[0][0]


def build_human_vs_llm_pairs(
    joined_rows: List[dict],
) -> Dict[Tuple[str, str], List[Tuple[str, str]]]:
    """Group rows by (stage, criterion_id) and produce (human, llm) pair lists.

    Overlap rows are collapsed by majority vote across raters BEFORE pairing
    with the LLM. Disjoint rows pair the single rater's decision directly.
    """
    pairs_by_key: Dict[Tuple[str, str], List[Tuple[str, str]]] = defaultdict(list)

    # Bucket by key + (a_id, set).
    overlap: Dict[Tuple[str, str, str], List[dict]] = defaultdict(list)
    disjoint: List[dict] = []
    for r in joined_rows:
        if r["set"] == "overlap":
            overlap[(r["stage"], r["criterion_id"], r["a_id"])].append(r)
        else:
            disjoint.append(r)

    for (stage, cid, a_id), rows in overlap.items():
        decs = [r["human_decision"] for r in rows]
        # All rows for the same (stage, cid, a_id) share the same llm_decision_canonical.
        llm = rows[0]["llm_decision_canonical"]
        if rows[0]["llm_status"] == "MISSING":
            continue
        human = majority_vote(decs)
        pairs_by_key[(stage, cid)].append((human, llm))

    for r in disjoint:
        if r["llm_status"] == "MISSING":
            continue
        pairs_by_key[(r["stage"], r["criterion_id"])].append(
            (r["human_decision"], r["llm_decision_canonical"])
        )

    return dict(pairs_by_key)


def build_inter_human_matrix(
    joined_rows: List[dict],
    raters: List[str],
) -> Dict[Tuple[str, str], List[List[int]]]:
    """Group overlap rows by (stage, criterion) and build Fleiss matrices.

    For each (stage, criterion) the matrix has one row per overlap a_id,
    with columns counting raters in each canonical decision.
    """
    n_raters = len(raters)
    out: Dict[Tuple[str, str], List[List[int]]] = defaultdict(list)
    by_key: Dict[Tuple[str, str], Dict[str, Dict[str, str]]] = defaultdict(
        lambda: defaultdict(dict)
    )
    for r in joined_rows:
        if r["set"] != "overlap":
            continue
        by_key[(r["stage"], r["criterion_id"])][r["a_id"]][r["rater_id"]] = \
            r["human_decision"]

    for key, by_aid in by_key.items():
        for a_id, by_rater in by_aid.items():
            if len(by_rater) != n_raters:
                # Defensive: schema validation should have caught this, but
                # skip this item rather than miscount the matrix.
                continue
            counts = [
                sum(1 for rater in raters if by_rater.get(rater) == cat)
                for cat in CANONICAL_DECISIONS
            ]
            out[key].append(counts)
    return dict(out)


def confusion_per_criterion(
    pairs_by_key: Dict[Tuple[str, str], List[Tuple[str, str]]],
) -> Dict[Tuple[str, str], List[List[int]]]:
    """Confusion matrix (rows: human, cols: llm) per (stage, criterion)."""
    return {k: _confusion_matrix(v) for k, v in pairs_by_key.items()}


# --------------------------------------------------------------------------
# Output writers
# --------------------------------------------------------------------------

def write_decisions_csv(path: Path, joined_rows: List[dict]) -> None:
    """Write eval_decisions_v1.csv (long-format human decisions only)."""
    fields = ["stage", "a_id", "criterion_id", "rater_id", "set",
              "human_decision", "human_notes"]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, lineterminator="\n")
        w.writeheader()
        for r in joined_rows:
            w.writerow({k: r.get(k, "") for k in fields})


def write_results_csv(path: Path, joined_rows: List[dict]) -> None:
    """Write eval_results_v1.csv (joined human+LLM evidence + agreement flag)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=RESULT_FIELDS, lineterminator="\n")
        w.writeheader()
        for r in joined_rows:
            w.writerow({k: r.get(k, "") for k in RESULT_FIELDS})


def write_disagreements_csv(path: Path, joined_rows: List[dict]) -> None:
    """Write eval_disagreements_v1.csv (only agree=False rows, sorted)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = [r for r in joined_rows if r["agree"] == "False"]
    rows.sort(key=lambda r: (r["stage"], r["criterion_id"], r["a_id"], r["rater_id"]))
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=RESULT_FIELDS, lineterminator="\n")
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in RESULT_FIELDS})


def write_summary_text(
    path: Path,
    cohen_results: Dict[Tuple[str, str], Dict[str, float]],
    fleiss_results: Dict[Tuple[str, str], Dict[str, float]],
    confusion: Dict[Tuple[str, str], List[List[int]]],
    n_raters: int,
) -> None:
    """Write eval_summary_v1.txt (human-readable kappa values + matrices)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []
    lines.append("metaScreener LLM-screening human validation: agreement metrics")
    lines.append("=" * 70)
    lines.append("")
    lines.append(f"Raters: {n_raters}")
    lines.append(f"Categories: {', '.join(CANONICAL_DECISIONS)}")
    lines.append("")
    lines.append("LLM-status -> canonical decision (polarity-aware):")
    lines.append("  For INCLUSION criteria (IC-*):")
    for k, v in _STATUS_MAP_INCLUDE.items():
        lines.append(f"    {k:<10} -> {v}")
    lines.append("  For EXCLUSION criteria (EC-*):")
    for k, v in _STATUS_MAP_EXCLUDE.items():
        lines.append(f"    {k:<10} -> {v}")
    lines.append("")

    # Iterate stages in display order, criteria sorted.
    keys = sorted(cohen_results.keys(), key=lambda k: (STAGE_DISPLAY_ORDER.index(k[0]), k[1]))
    for stage, cid in keys:
        lines.append("-" * 70)
        lines.append(f"Stage {stage}, criterion {cid}")
        lines.append("")
        c = cohen_results.get((stage, cid))
        if c is not None:
            lines.append(
                f"  Cohen's kappa (human vs LLM): {c['kappa']:.4f}  "
                f"(N={c['n']}, P_observed={c['agreement_observed']:.4f}, "
                f"P_expected={c['agreement_expected']:.4f})"
            )
        f = fleiss_results.get((stage, cid))
        if f is not None:
            lines.append(
                f"  Fleiss' kappa (inter-human, overlap subset): {f['kappa']:.4f}  "
                f"(items={f['n_items']}, raters={f['n_raters']}, "
                f"P_observed={f['agreement_observed']:.4f}, "
                f"P_expected={f['agreement_expected']:.4f})"
            )
        cm = confusion.get((stage, cid))
        if cm is not None:
            lines.append("")
            lines.append("  Confusion matrix (rows: human, cols: LLM):")
            header = "                " + "".join(f"{cat:>11}" for cat in CANONICAL_DECISIONS)
            lines.append(header)
            for i, cat in enumerate(CANONICAL_DECISIONS):
                row = f"    {cat:<12}" + "".join(f"{cm[i][j]:>11}" for j in range(len(CANONICAL_DECISIONS)))
                lines.append(row)
        lines.append("")

    lines.append("-" * 70)
    lines.append("Limitations:")
    lines.append("  - Validation N is bounded by the demonstration corpus universe;")
    lines.append("    extending kappa across multiple corpora is future work.")
    lines.append("  - Overlap subset is randomly selected with stratification on")
    lines.append("    stage_outcome; with N=15 per stage the inter-human kappa is")
    lines.append("    sensitive to small absolute disagreement counts.")
    lines.append("  - Human raters were the three paper authors. Independent")
    lines.append("    multi-disciplinary adjudication is not represented here.")
    lines.append("")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def parse_args(argv: List[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Ingest filled adjudication grids and compute "
                    "human-vs-LLM agreement metrics for metaScreener.",
    )
    p.add_argument("--manifest", required=True, type=Path,
                   help="Path to partition_manifest.csv from eval_grid_generator.py.")
    p.add_argument("--criteria", required=True, type=Path,
                   help="Path to criteria_harmonized CSV. Used to load each "
                        "criterion's polarity (include/exclude), which is "
                        "required for polarity-aware LLM-status mapping.")
    p.add_argument("--filled-grids-dir", required=True, type=Path,
                   help="Directory containing eval_grid_<rater>.xlsx (filled).")
    p.add_argument("--el-filtered", type=Path, default=None,
                   help="Path to el_filtered CSV (provides el_evidence_json).")
    p.add_argument("--il-filtered", type=Path, default=None,
                   help="Path to il_filtered CSV (provides il_evidence_json).")
    p.add_argument("--output-dir", required=True, type=Path,
                   help="Directory where evidence CSVs and summary text are written.")
    return p.parse_args(argv)


def load_criteria_polarity(path: Path) -> Dict[str, str]:
    """Load {criterion_id: polarity} from a harmonized criteria CSV.

    Polarity is one of "include" or "exclude" (the criterion's `type` field).
    Required by join_with_llm() so that LLM-status mapping is polarity-aware.
    """
    if not path.exists():
        raise FileNotFoundError(f"Criteria CSV not found: {path}")
    out: Dict[str, str] = {}
    with path.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            cid = (r.get("id") or "").strip()
            polarity = (r.get("type") or "").strip().lower()
            if cid:
                out[cid] = polarity
    return out


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])

    manifest = load_partition_manifest(args.manifest)
    by_rater = index_manifest_by_rater(manifest)
    raters = sorted(by_rater.keys())
    if len(raters) < 2:
        print(f"[error] need >=2 raters in manifest, got {raters}", file=sys.stderr)
        return 2

    # Load criteria polarities for polarity-aware LLM-status mapping.
    try:
        criteria_polarity = load_criteria_polarity(args.criteria)
    except (FileNotFoundError, ValueError) as exc:
        print(f"[error] {exc}", file=sys.stderr)
        return 2

    # Load LLM evidence per stage as required by the manifest.
    stages_in_manifest = {row["stage"] for row in manifest}
    stage_inputs = {"EL": args.el_filtered, "IL": args.il_filtered}
    llm_evidence: Dict[str, Dict[str, Dict[str, dict]]] = {}
    for stage in stages_in_manifest:
        path = stage_inputs.get(stage)
        if path is None:
            print(
                f"[error] manifest references stage {stage} but no "
                f"--{stage.lower()}-filtered CSV was supplied.",
                file=sys.stderr,
            )
            return 2
        llm_evidence[stage] = load_llm_evidence(path, stage)

    # Read each rater's workbook and produce long-form decision rows.
    all_human_rows: List[dict] = []
    for rater_id in raters:
        wb_path = args.filled_grids_dir / f"eval_grid_{rater_id}.xlsx"
        try:
            rater_rows = read_filled_workbook(
                path=wb_path,
                rater_id=rater_id,
                expected_per_stage=by_rater[rater_id],
            )
        except (FileNotFoundError, ValueError) as exc:
            print(f"[error] {exc}", file=sys.stderr)
            return 2
        all_human_rows.extend(rater_rows)
        print(f"[ok] read {wb_path.name}: {len(rater_rows)} decisions")

    # Join with LLM evidence (polarity-aware).
    joined = join_with_llm(all_human_rows, llm_evidence, criteria_polarity)

    # Write the three evidence CSVs.
    args.output_dir.mkdir(parents=True, exist_ok=True)
    decisions_path = args.output_dir / "eval_decisions_v1.csv"
    results_path = args.output_dir / "eval_results_v1.csv"
    disagreements_path = args.output_dir / "eval_disagreements_v1.csv"
    summary_path = args.output_dir / "eval_summary_v1.txt"

    write_decisions_csv(decisions_path, joined)
    write_results_csv(results_path, joined)
    write_disagreements_csv(disagreements_path, joined)

    # Compute metrics.
    pairs_by_key = build_human_vs_llm_pairs(joined)
    cohen_results = {key: cohen_kappa(pairs) for key, pairs in pairs_by_key.items()}
    matrices_by_key = build_inter_human_matrix(joined, raters)
    fleiss_results = {key: fleiss_kappa(m) for key, m in matrices_by_key.items()}
    confusion = confusion_per_criterion(pairs_by_key)

    write_summary_text(
        path=summary_path,
        cohen_results=cohen_results,
        fleiss_results=fleiss_results,
        confusion=confusion,
        n_raters=len(raters),
    )

    print(f"[ok] wrote {decisions_path.name} ({len(joined)} rows)")
    print(f"[ok] wrote {results_path.name} ({len(joined)} rows)")
    n_disagree = sum(1 for r in joined if r["agree"] == "False")
    print(f"[ok] wrote {disagreements_path.name} ({n_disagree} rows)")
    print(f"[ok] wrote {summary_path.name}")
    print()
    print("Agreement summary:")
    keys = sorted(cohen_results.keys(),
                  key=lambda k: (STAGE_DISPLAY_ORDER.index(k[0]), k[1]))
    for stage, cid in keys:
        c = cohen_results[(stage, cid)]
        f = fleiss_results.get((stage, cid))
        cohen_str = f"Cohen={c['kappa']:.3f} (N={c['n']})"
        fleiss_str = (f"Fleiss={f['kappa']:.3f} (items={f['n_items']})"
                      if f is not None else "Fleiss=n/a")
        print(f"  {stage} / {cid}:  {cohen_str}  {fleiss_str}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
