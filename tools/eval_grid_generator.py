# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""tools/eval_grid_generator.py — generate XLSX adjudication grids for N raters.

Reads:
  - One stage-filtered CSV per LLM-adjudicated stage (EL filtered, IL filtered).
  - The harmonized criteria CSV used by the demonstration.

Produces one XLSX file per rater. Each workbook has:
  1. "Read Me First"   — concise instructions for the co-author.
  2. "Decisions_EL"    — EL-stage records and EL/llm criteria (one decision
                          column per criterion).
  3. "Decisions_IL"    — IL-stage records and IL/llm criteria.
  4. "Reference"       — full text of every LLM-adjudicated criterion.

Per stage, a fixed-size overlap subset is rated by every rater; the remaining
records are split disjointly across raters with stratification by the stage's
outcome column. The two stages are partitioned independently with derived
seeds (stage-specific) so the IL overlap set and the EL overlap set are
independent — both Fleiss' kappas can be computed per stage.

LLM decisions present in the input filtered CSVs are NOT exposed in the
generated grids. Raters see only the bibliographic record and the criteria
text. Blind adjudication is the only methodologically defensible path: if
raters see the LLM's call, they anchor to it and the resulting kappa
becomes a measure of anchoring rather than independent agreement.

Usage:
    python tools/eval_grid_generator.py \\
        --el-filtered tests/golden/el_filtered_v3.1.0.csv \\
        --il-filtered tests/golden/il_filtered_v3.1.0.csv \\
        --criteria    tests/golden/criteria_harmonized_v3.1.0.csv \\
        --raters      AReyes JKiss JVoisin \\
        --output-dir  docs/data/grids/ \\
        --overlap     15 \\
        --seed        42

Outputs:
  - eval_grid_<rater>.xlsx (one per rater)
  - partition_manifest.csv (every (stage, a_id, rater_id, set) assignment)
  - partition_manifest.meta.txt (seed, raters, criteria — sidecar metadata)
"""
import argparse
import csv
import random
import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

DECISION_OPTIONS = ("include", "exclude", "uncertain")
DECISION_FORMULA = '"' + ",".join(DECISION_OPTIONS) + '"'

# Columns the rater needs to read in each Decisions sheet, in display order.
# Each tuple is (csv_key, display_label, column_width_chars).
RECORD_COLUMNS = [
    ("a_id",     "ID",       8),
    ("title",    "Title",    50),
    ("year",     "Year",     6),
    ("lang",     "Lang",     6),
    ("venue",    "Venue",    24),
    ("doi",      "DOI",      28),
    ("abstract", "Abstract", 80),
]

# Stage configuration: stage_id -> (outcome_column_in_csv, sheet_title).
STAGE_OUTCOME_COL = {"IL": "il_outcome", "EL": "el_outcome"}
STAGE_SHEET_TITLE = {"IL": "Decisions_IL", "EL": "Decisions_EL"}
STAGE_DISPLAY_ORDER = ("EL", "IL")  # raters see EL first (chronological in pipeline)

HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
OVERLAP_FILL = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
DECISION_FILL = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
NOTES_FILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="FFBFBFBF"),
    right=Side(style="thin", color="FFBFBFBF"),
    top=Side(style="thin", color="FFBFBFBF"),
    bottom=Side(style="thin", color="FFBFBFBF"),
)


# --------------------------------------------------------------------------
# Loading helpers
# --------------------------------------------------------------------------

def load_stage_records(path: Path, stage: str) -> List[dict]:
    """Load a stage-filtered CSV. Returns list of dict rows.

    The required outcome column depends on the stage (il_outcome for IL,
    el_outcome for EL). LLM decision columns (*_failed_ids, *_met_ids,
    *_uncertain_ids, *_evidence_json) are loaded but deliberately not
    propagated to the generated grids.
    """
    if stage not in STAGE_OUTCOME_COL:
        raise ValueError(f"Unknown stage: {stage!r}. Expected one of {list(STAGE_OUTCOME_COL)}.")
    if not path.exists():
        raise FileNotFoundError(f"{stage} filtered CSV not found: {path}")
    with path.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        raise ValueError(f"{stage} filtered CSV is empty: {path}")
    outcome_col = STAGE_OUTCOME_COL[stage]
    required = {"local_id", "title", outcome_col}
    missing = required - set(rows[0].keys())
    if missing:
        raise ValueError(
            f"{stage} filtered CSV missing required columns: {sorted(missing)}"
        )
    return rows


def load_llm_criteria(path: Path) -> Dict[str, List[dict]]:
    """Load harmonized criteria CSV, returning ALL llm-operator rows grouped by stage.

    Returns a dict keyed by stage ("IL", "EL") with the list of llm-operator
    criteria for that stage. In the published demonstration this is:
      - EL: EC-2 (spatial navigation), EC-3 (rubber hand illusion)
      - IL: IC-1 (HMD VR / immersive virtual reality)
    Stages with zero llm criteria are omitted from the output dict.
    """
    if not path.exists():
        raise FileNotFoundError(f"Criteria CSV not found: {path}")
    with path.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    by_stage: Dict[str, List[dict]] = defaultdict(list)
    for r in rows:
        stage = r.get("stage", "").strip()
        operator = r.get("operator", "").strip()
        enabled = str(r.get("enabled", "1")).strip()
        if stage in STAGE_OUTCOME_COL and operator == "llm" and enabled in ("1", "true", "True"):
            by_stage[stage].append(r)
    if not by_stage:
        raise ValueError(
            f"No llm-operator criteria found in {path} for stages "
            f"{list(STAGE_OUTCOME_COL)}; validation requires at least one."
        )
    return dict(by_stage)


# --------------------------------------------------------------------------
# Stratification + partition (single stage)
# --------------------------------------------------------------------------

def _stratify_by_outcome(records: List[dict], outcome_col: str) -> Dict[str, List[dict]]:
    """Group records by stage outcome column. Empty outcome -> 'UNKNOWN'."""
    buckets: Dict[str, List[dict]] = defaultdict(list)
    for r in records:
        key = (r.get(outcome_col) or "").strip() or "UNKNOWN"
        buckets[key].append(r)
    return dict(buckets)


def stratify_and_partition(
    records: List[dict],
    n_overlap: int,
    raters: List[str],
    seed: int,
    outcome_col: str = "il_outcome",
) -> Tuple[List[dict], Dict[str, List[dict]]]:
    """Partition records into (overlap_set, per_rater_disjoint_sets).

    - overlap_set contains exactly n_overlap records, drawn from each
      stratum (defined by outcome_col) proportionally so the overlap reflects
      the decision-distribution of the full corpus.
    - The remaining records are distributed disjointly across raters with
      cross-stratum load balancing (always assigning the next record from
      a stratum to the rater with the smallest current allotment), so the
      worst-case per-rater imbalance is at most 1.

    The seed pins the randomization so the partition is reproducible.

    Raises ValueError if n_overlap < 0 or > len(records), or if no raters.
    """
    if not raters:
        raise ValueError("At least one rater is required.")
    if n_overlap < 0:
        raise ValueError("n_overlap must be >= 0.")
    if n_overlap > len(records):
        raise ValueError(
            f"n_overlap ({n_overlap}) cannot exceed total records ({len(records)})."
        )

    rng = random.Random(seed)
    strata = _stratify_by_outcome(records, outcome_col)
    stratum_keys = sorted(strata.keys())

    # Step 1: draw overlap set proportionally from each stratum using
    # largest-remainder apportionment so totals sum to exactly n_overlap.
    total = len(records)
    raw = {k: n_overlap * len(strata[k]) / total for k in stratum_keys}
    floors = {k: int(raw[k]) for k in stratum_keys}
    leftover = n_overlap - sum(floors.values())
    fractional = sorted(
        stratum_keys,
        key=lambda k: (raw[k] - floors[k], len(strata[k])),
        reverse=True,
    )
    alloc = dict(floors)
    for i in range(leftover):
        alloc[fractional[i % len(fractional)]] += 1

    overlap_set: List[dict] = []
    remaining: Dict[str, List[dict]] = {}
    for k in stratum_keys:
        bucket = list(strata[k])
        rng.shuffle(bucket)
        take = alloc[k]
        overlap_set.extend(bucket[:take])
        remaining[k] = bucket[take:]

    # Step 2: distribute remaining records disjointly across raters, with
    # cross-stratum load balancing. Within each stratum we still iterate in
    # shuffled order, but we always assign to the rater with the currently
    # smallest allotment (ties broken by rater-list order). This guarantees
    # the worst-case per-rater imbalance is at most 1, even across multiple
    # strata of unequal size.
    per_rater: Dict[str, List[dict]] = {r: [] for r in raters}
    for k in stratum_keys:
        for rec in remaining[k]:
            target = min(raters, key=lambda r: (len(per_rater[r]), raters.index(r)))
            per_rater[target].append(rec)

    return overlap_set, per_rater


# --------------------------------------------------------------------------
# Multi-stage partition (drives final manifest + workbook contents)
# --------------------------------------------------------------------------

def partition_all_stages(
    stage_records: Dict[str, List[dict]],
    n_overlap: int,
    raters: List[str],
    seed: int,
) -> Dict[str, Tuple[List[dict], Dict[str, List[dict]]]]:
    """Run partition independently for each stage with derived seeds.

    Each stage's seed is derived from the base seed plus a stage-specific
    constant so that running both stages with `--seed 42` produces a fully
    reproducible joint partition that is also independent across stages.
    """
    stage_seed_offset = {"EL": 0, "IL": 1}  # arbitrary, fixed in spec
    out: Dict[str, Tuple[List[dict], Dict[str, List[dict]]]] = {}
    for stage, records in stage_records.items():
        outcome_col = STAGE_OUTCOME_COL[stage]
        offset = stage_seed_offset.get(stage, 0)
        out[stage] = stratify_and_partition(
            records=records,
            n_overlap=n_overlap,
            raters=raters,
            seed=seed + offset,
            outcome_col=outcome_col,
        )
    return out


# --------------------------------------------------------------------------
# Workbook construction
# --------------------------------------------------------------------------

def _set_header_row(ws, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def _build_readme_sheet(
    ws,
    rater_id: str,
    per_stage_totals: Dict[str, Tuple[int, int]],
    criteria_by_stage: Dict[str, List[dict]],
) -> None:
    """Single-page rater instructions. Plain English, no jargon.

    per_stage_totals maps stage -> (n_total_for_rater, n_overlap_for_stage).
    """
    ws.title = "Read Me First"
    ws.column_dimensions["A"].width = 110
    ws["A1"] = f"metaScreener - adjudication grid for {rater_id}"
    ws["A1"].font = Font(size=14, bold=True)

    n_records_total = sum(t for t, _ in per_stage_totals.values())
    n_decisions_total = sum(
        per_stage_totals[s][0] * len(criteria_by_stage.get(s, []))
        for s in per_stage_totals
    )

    lines = [
        "",
        f"You have been assigned {n_records_total} records across two pipeline "
        f"stages, totalling {n_decisions_total} decisions to make. The breakdown:",
    ]
    for stage in STAGE_DISPLAY_ORDER:
        if stage not in per_stage_totals:
            continue
        n_total, n_overlap = per_stage_totals[stage]
        n_crit = len(criteria_by_stage.get(stage, []))
        sheet_name = STAGE_SHEET_TITLE[stage]
        suffix = "a" if n_crit > 1 else ""
        lines.append(
            f"  - {sheet_name}: {n_total} records ({n_overlap} overlap + "
            f"{n_total - n_overlap} unique) x {n_crit} criterion{suffix} = "
            f"{n_total * n_crit} decisions."
        )
    lines += [
        "",
        "Records highlighted in pale yellow are the overlap subset - also "
        "rated independently by the other two co-authors. Please do NOT "
        "confer with the other raters while filling this in. The overlap "
        "subset measures inter-rater agreement, which only works if each "
        "rater decides independently.",
        "",
        "How to fill this in:",
        "  1. Open each Decisions sheet (Decisions_EL, then Decisions_IL).",
        "  2. For each row, read the Title and Abstract.",
        "  3. For each criterion column, pick include / exclude / uncertain "
        "from the dropdown. Add free-text notes if you want - optional.",
        "  4. Save the file (keep the same filename).",
        "  5. Send the file back to Alejandro (marec3@ulaval.ca).",
        "",
        "What the criteria mean (full text on the Reference sheet):",
    ]
    for stage in STAGE_DISPLAY_ORDER:
        for c in criteria_by_stage.get(stage, []):
            cid = c.get("id", "?").strip()
            label = (c.get("label") or "").strip()
            polarity = c.get("type", "").strip()
            lines.append(f"  - {cid} ({polarity}, stage {stage}): {label}")
    lines += [
        "",
        "Decision values (apply the SAME meaning regardless of polarity):",
        "  - include    - the abstract clearly satisfies the criterion as written.",
        "  - exclude    - the abstract clearly does NOT satisfy it.",
        "  - uncertain  - abstract is ambiguous; full text would be needed.",
        "",
        "Note: Inclusion criteria (IC-*) and exclusion criteria (EC-*) are "
        "rated the same way. 'include' means 'this criterion's condition is "
        "met by the record,' regardless of whether the criterion is itself "
        "an inclusion or exclusion rule. The validation analysis pairs your "
        "decisions with the LLM's per-criterion call and computes agreement.",
        "",
        "Estimated time: ~45 minutes if read carefully. Skim mode is fine "
        "for obvious yes/no calls; reserve careful reading for the genuinely "
        "uncertain ones.",
    ]
    for i, text in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if any(text.startswith(p) for p in (
            "How to fill this in:", "What the criteria",
            "Decision values", "Note:", "Records highlighted",
        )):
            cell.font = Font(bold=True)
    ws.sheet_view.showGridLines = False


def _build_decisions_sheet(
    ws,
    sheet_title: str,
    records: List[dict],
    criteria: List[dict],
    overlap_a_ids: set,
) -> None:
    """One row per record. Per-criterion decision dropdown + notes column."""
    ws.title = sheet_title

    # Build header: record cols + per-criterion (decision, notes) pairs.
    headers = [label for (_key, label, _w) in RECORD_COLUMNS]
    for c in criteria:
        cid = c.get("id", "?").strip()
        headers.append(f"Decision: {cid}")
        headers.append(f"Notes: {cid}")
    _set_header_row(ws, headers)

    # Column widths.
    for col_idx, (_key, _label, width) in enumerate(RECORD_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    next_col = len(RECORD_COLUMNS) + 1
    for _ in criteria:
        ws.column_dimensions[get_column_letter(next_col)].width = 14      # decision
        ws.column_dimensions[get_column_letter(next_col + 1)].width = 40  # notes
        next_col += 2

    # Data rows.
    for r_idx, rec in enumerate(records, start=2):
        a_id = (rec.get("local_id") or "").strip()
        is_overlap = a_id in overlap_a_ids
        row_fill = OVERLAP_FILL if is_overlap else None

        # Record columns.
        for c_idx, (key, _label, _w) in enumerate(RECORD_COLUMNS, start=1):
            if key == "a_id":
                value = a_id
            else:
                value = rec.get(key, "") or ""
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            if row_fill is not None:
                cell.fill = row_fill
            # Render DOI as hyperlink when present.
            if key == "doi" and value:
                url = value if str(value).startswith("http") else f"https://doi.org/{value}"
                cell.hyperlink = url
                cell.font = Font(color="FF0563C1", underline="single")

        # Per-criterion decision + notes columns.
        col = len(RECORD_COLUMNS) + 1
        for _c in criteria:
            dec_cell = ws.cell(row=r_idx, column=col, value=None)
            dec_cell.alignment = Alignment(horizontal="center", vertical="center")
            dec_cell.border = THIN_BORDER
            dec_cell.fill = OVERLAP_FILL if is_overlap else DECISION_FILL
            notes_cell = ws.cell(row=r_idx, column=col + 1, value=None)
            notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
            notes_cell.border = THIN_BORDER
            notes_cell.fill = OVERLAP_FILL if is_overlap else NOTES_FILL
            col += 2

        ws.row_dimensions[r_idx].height = 90

    # Freeze header + first column.
    ws.freeze_panes = "B2"

    # Data validation on each decision column.
    n_rows = len(records) + 1  # header + data
    col = len(RECORD_COLUMNS) + 1
    for _c in criteria:
        col_letter = get_column_letter(col)
        dv = DataValidation(
            type="list",
            formula1=DECISION_FORMULA,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid decision",
            error="Choose include, exclude, or uncertain from the dropdown.",
        )
        dv.add(f"{col_letter}2:{col_letter}{n_rows}")
        ws.add_data_validation(dv)
        col += 2


def _build_reference_sheet(ws, criteria_by_stage: Dict[str, List[dict]]) -> None:
    """Full criterion text for in-file lookup, grouped by stage."""
    ws.title = "Reference"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 100
    _set_header_row(ws, ["Stage", "ID", "Polarity", "Full text"])
    r_idx = 2
    for stage in STAGE_DISPLAY_ORDER:
        for c in criteria_by_stage.get(stage, []):
            ws.cell(row=r_idx, column=1, value=stage).alignment = \
                Alignment(vertical="top")
            ws.cell(row=r_idx, column=2, value=c.get("id", "")).alignment = \
                Alignment(vertical="top")
            ws.cell(row=r_idx, column=3, value=c.get("type", "")).alignment = \
                Alignment(vertical="top")
            cell = ws.cell(row=r_idx, column=4,
                           value=c.get("source_text", c.get("label", "")))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r_idx].height = 60
            r_idx += 1


def build_workbook(
    rater_id: str,
    per_stage_records: Dict[str, Tuple[List[dict], List[dict]]],
    criteria_by_stage: Dict[str, List[dict]],
) -> Workbook:
    """Build a complete rater workbook (Read Me First + N stage sheets + Reference).

    per_stage_records maps stage -> (overlap_records, unique_for_this_rater).
    The Decisions_<stage> sheet shows them concatenated (overlap rows first,
    then unique rows), with overlap rows shaded.
    """
    wb = Workbook()
    # Default sheet becomes Read Me First.
    readme_ws = wb.active

    per_stage_totals: Dict[str, Tuple[int, int]] = {}
    for stage in STAGE_DISPLAY_ORDER:
        if stage not in per_stage_records:
            continue
        overlap_records, unique_records = per_stage_records[stage]
        per_stage_totals[stage] = (
            len(overlap_records) + len(unique_records),
            len(overlap_records),
        )

    _build_readme_sheet(
        ws=readme_ws,
        rater_id=rater_id,
        per_stage_totals=per_stage_totals,
        criteria_by_stage=criteria_by_stage,
    )

    for stage in STAGE_DISPLAY_ORDER:
        if stage not in per_stage_records:
            continue
        overlap_records, unique_records = per_stage_records[stage]
        all_records = list(overlap_records) + list(unique_records)
        overlap_a_ids = {(r.get("local_id") or "").strip() for r in overlap_records}
        ws = wb.create_sheet(STAGE_SHEET_TITLE[stage])
        _build_decisions_sheet(
            ws=ws,
            sheet_title=STAGE_SHEET_TITLE[stage],
            records=all_records,
            criteria=criteria_by_stage.get(stage, []),
            overlap_a_ids=overlap_a_ids,
        )

    reference_ws = wb.create_sheet("Reference")
    _build_reference_sheet(reference_ws, criteria_by_stage)

    return wb


# --------------------------------------------------------------------------
# Manifest
# --------------------------------------------------------------------------

def write_partition_manifest(
    path: Path,
    partitions_by_stage: Dict[str, Tuple[List[dict], Dict[str, List[dict]]]],
    criteria_by_stage: Dict[str, List[dict]],
    raters: List[str],
    seed: int,
) -> None:
    """Record every (stage, a_id, rater_id, set) assignment plus run metadata.

    The manifest is what eval_ingest.py reads to validate that returned
    grids match the partition (no spurious additions, no missing rows, no
    rater swapping records). It also records the seed and per-stage criterion
    list so the partition is reproducible.
    """
    rows = []
    for stage, (overlap_records, per_rater) in partitions_by_stage.items():
        outcome_col = STAGE_OUTCOME_COL[stage]
        overlap_ids = {(r.get("local_id") or "").strip() for r in overlap_records}
        # Overlap rows first (one row per (a_id, rater) for each rater).
        for rec in overlap_records:
            a_id = (rec.get("local_id") or "").strip()
            for rater in raters:
                rows.append({
                    "stage": stage,
                    "a_id": a_id,
                    "rater_id": rater,
                    "set": "overlap",
                    "stage_outcome": (rec.get(outcome_col) or "").strip(),
                })
        # Disjoint rows.
        for rater, recs in per_rater.items():
            for rec in recs:
                a_id = (rec.get("local_id") or "").strip()
                if a_id in overlap_ids:
                    # defensive: should not occur given partition logic
                    continue
                rows.append({
                    "stage": stage,
                    "a_id": a_id,
                    "rater_id": rater,
                    "set": "disjoint",
                    "stage_outcome": (rec.get(outcome_col) or "").strip(),
                })

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=["stage", "a_id", "rater_id", "set", "stage_outcome"],
            lineterminator="\n",
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    # Sidecar metadata file (criteria list per stage + seed).
    meta_path = path.with_suffix(".meta.txt")
    with meta_path.open("w", encoding="utf-8", newline="") as fh:
        fh.write(f"seed={seed}\n")
        fh.write(f"raters={','.join(sorted(raters))}\n")
        for stage in STAGE_DISPLAY_ORDER:
            if stage in partitions_by_stage:
                overlap_records, _ = partitions_by_stage[stage]
                fh.write(f"{stage}_n_overlap={len(overlap_records)}\n")
                ids = ",".join(c.get("id", "") for c in criteria_by_stage.get(stage, []))
                fh.write(f"{stage}_criteria={ids}\n")


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def parse_args(argv: List[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Generate adjudication grids (XLSX) for the LLM-stage "
                    "validation of metaScreener (EL + IL).",
    )
    p.add_argument("--el-filtered", type=Path, default=None,
                   help="Path to el_filtered CSV (post-EL bundle records). "
                        "Required if EL has llm-operator criteria.")
    p.add_argument("--il-filtered", type=Path, default=None,
                   help="Path to il_filtered CSV (post-IL bundle records). "
                        "Required if IL has llm-operator criteria.")
    p.add_argument("--criteria", required=True, type=Path,
                   help="Path to criteria_harmonized CSV.")
    p.add_argument("--raters", required=True, nargs="+",
                   help="Rater identifiers (space-separated).")
    p.add_argument("--output-dir", required=True, type=Path,
                   help="Directory to write the XLSX grids and manifest.")
    p.add_argument("--overlap", type=int, default=15,
                   help="Number of records rated by every rater, per stage "
                        "(default: 15).")
    p.add_argument("--seed", type=int, default=42,
                   help="Base random seed for reproducible partition (default: 42).")
    return p.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])

    if len(set(args.raters)) != len(args.raters):
        print("[error] rater identifiers must be unique.", file=sys.stderr)
        return 2

    criteria_by_stage = load_llm_criteria(args.criteria)

    # Determine which stages need processing based on criteria + provided inputs.
    stage_inputs = {"EL": args.el_filtered, "IL": args.il_filtered}
    stage_records: Dict[str, List[dict]] = {}
    for stage, llm_criteria in criteria_by_stage.items():
        if not llm_criteria:
            continue
        path = stage_inputs.get(stage)
        if path is None:
            print(
                f"[error] stage {stage} has {len(llm_criteria)} llm criteria "
                f"but no --{stage.lower()}-filtered CSV was supplied.",
                file=sys.stderr,
            )
            return 2
        stage_records[stage] = load_stage_records(path, stage)

    if not stage_records:
        print("[error] no LLM-adjudicated stages to process.", file=sys.stderr)
        return 2

    partitions = partition_all_stages(
        stage_records=stage_records,
        n_overlap=args.overlap,
        raters=list(args.raters),
        seed=args.seed,
    )

    args.output_dir.mkdir(parents=True, exist_ok=True)

    # Build per-rater workbook contents from partitions.
    for rater_id in args.raters:
        per_stage_records: Dict[str, Tuple[List[dict], List[dict]]] = {}
        for stage, (overlap_records, per_rater) in partitions.items():
            per_stage_records[stage] = (overlap_records, per_rater[rater_id])
        wb = build_workbook(
            rater_id=rater_id,
            per_stage_records=per_stage_records,
            criteria_by_stage={
                s: criteria_by_stage.get(s, []) for s in partitions
            },
        )
        out_path = args.output_dir / f"eval_grid_{rater_id}.xlsx"
        wb.save(out_path)
        # Per-rater summary line.
        parts = []
        n_total_records = 0
        n_total_decisions = 0
        for stage, (overlap_records, per_rater) in partitions.items():
            n_unique = len(per_rater[rater_id])
            n_overlap_s = len(overlap_records)
            n_records = n_overlap_s + n_unique
            n_decisions = n_records * len(criteria_by_stage.get(stage, []))
            n_total_records += n_records
            n_total_decisions += n_decisions
            parts.append(f"{stage}={n_records} ({n_overlap_s}+{n_unique})")
        print(
            f"[ok] {out_path.name}: {n_total_records} records "
            f"[{', '.join(parts)}], {n_total_decisions} decisions"
        )

    manifest_path = args.output_dir / "partition_manifest.csv"
    write_partition_manifest(
        path=manifest_path,
        partitions_by_stage=partitions,
        criteria_by_stage=criteria_by_stage,
        raters=list(args.raters),
        seed=args.seed,
    )
    parts = []
    for stage, (overlap_records, per_rater) in partitions.items():
        n_disjoint = sum(len(v) for v in per_rater.values())
        parts.append(f"{stage} overlap={len(overlap_records)} disjoint={n_disjoint}")
    print(f"[ok] {manifest_path.name}: partition recorded "
          f"(seed={args.seed}, {'; '.join(parts)})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
