# -*- coding: utf-8 -*-

# SPDX-FileCopyrightText: 2026 Alejandro Reyes-Consuelo
# SPDX-License-Identifier: MIT

"""tools/eval_grid_filler_synthetic.py - test-only synthetic filler for adjudication grids.

This script is used by the test suite to produce deterministically-filled
copies of the empty XLSX grids that eval_grid_generator.py produces. It is
NOT intended for any real validation workflow; it exists so that
eval_ingest.py can be tested end-to-end without hand-filling spreadsheets.

A synthetic rater "profile" is parametrized by:
  - agree_with_llm_rate (0.0 - 1.0): probability the rater's decision
    matches the LLM's canonical decision for that (a_id, criterion).
  - uncertain_rate (0.0 - 1.0): probability the rater picks 'uncertain'
    instead of agreeing or disagreeing.
  - rater_id_seed_offset (int): added to the base seed so each rater
    produces a different but reproducible filling pattern.

The filler reads an empty workbook + the manifest + the LLM evidence,
computes the LLM canonical decision per (a_id, criterion), then for each
empty decision cell rolls a seeded random draw to assign a value.

This module is also importable as a library; eval_ingest tests use it
to produce filled fixtures inside tmp_path without leaving artifacts in
the repo.
"""
import argparse
import json
import random
import sys
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


CANONICAL_DECISIONS = ("yes", "no", "unsure")
LLM_STATUS_TO_DECISION = {"MET": "yes", "FAILED": "no", "UNCERTAIN": "unsure"}
STAGE_EVIDENCE_COL = {"EL": "el_evidence_json", "IL": "il_evidence_json"}
STAGE_SHEET_TITLE = {"EL": "Decisions_EL", "IL": "Decisions_IL"}
DROPDOWN_UNSURE_TEXT = "I cannot tell from the abstract alone."


def load_evidence_for_stage(filtered_csv: Path, stage: str) -> Dict[str, Dict[str, str]]:
    """Return {a_id: {criterion_id: canonical_decision_from_llm}}."""
    import csv
    out: Dict[str, Dict[str, str]] = {}
    ev_col = STAGE_EVIDENCE_COL[stage]
    with filtered_csv.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            a_id = (r.get("local_id") or "").strip()
            ev_str = r.get(ev_col, "") or ""
            if not ev_str:
                continue
            parsed = json.loads(ev_str)
            out[a_id] = {}
            for cid, v in parsed.items():
                if isinstance(v, dict):
                    status = str(v.get("status", "")).strip().upper()
                    out[a_id][cid] = LLM_STATUS_TO_DECISION.get(status, "unsure")
    return out


def load_criteria_labels(criteria_csv: Path) -> Dict[str, str]:
    """Return {criterion_id: label} for all criteria in the harmonized CSV.

    The label is the clean predicate text used to construct dropdown option
    sentences. Used by the synthetic filler so it can write the same
    full-sentence values that a real rater would pick from the dropdown.
    """
    import csv
    out: Dict[str, str] = {}
    with criteria_csv.open(encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            cid = (r.get("id") or "").strip()
            label = (r.get("label") or "").strip()
            if cid and label:
                out[cid] = label
    return out


def _option_for(canonical: str, criterion_label: str) -> str:
    """Render the dropdown sentence text matching one canonical decision."""
    if canonical == "yes":
        return f"YES - this is true: {criterion_label}"
    if canonical == "no":
        return f"NO - this is not true: {criterion_label}"
    return DROPDOWN_UNSURE_TEXT


def fill_workbook(
    src_path: Path,
    dst_path: Path,
    llm_canonical_by_stage: Dict[str, Dict[str, Dict[str, str]]],
    criterion_labels: Optional[Dict[str, str]] = None,
    agree_with_llm_rate: float = 0.9,
    uncertain_rate: float = 0.05,
    seed: int = 0,
) -> int:
    """Fill an empty grid with synthetic decisions.

    Cells are written with the full-sentence dropdown values that a real
    rater would pick. ``criterion_labels`` maps criterion_id -> label;
    if omitted, the labels are inferred from the empty grid's Reference
    sheet.

    Returns the number of cells filled.
    """
    if not 0.0 <= agree_with_llm_rate <= 1.0:
        raise ValueError("agree_with_llm_rate must be in [0,1]")
    if not 0.0 <= uncertain_rate <= 1.0:
        raise ValueError("uncertain_rate must be in [0,1]")
    if agree_with_llm_rate + uncertain_rate > 1.0:
        raise ValueError("agree_with_llm_rate + uncertain_rate must be <= 1.0")

    wb = load_workbook(src_path)
    rng = random.Random(seed)
    n_filled = 0

    if criterion_labels is None:
        criterion_labels = _infer_labels_from_reference_sheet(wb)

    for stage, sheet_title in STAGE_SHEET_TITLE.items():
        if sheet_title not in wb.sheetnames:
            continue
        ws = wb[sheet_title]
        headers = [c.value for c in ws[1]]
        try:
            id_col = headers.index("ID") + 1
        except ValueError:
            continue
        # Locate Decision: <cid> columns.
        decision_cols: List = []
        for idx, h in enumerate(headers, start=1):
            if isinstance(h, str) and h.startswith("Decision: "):
                cid = h[len("Decision: "):].strip()
                decision_cols.append((cid, idx))

        evidence = llm_canonical_by_stage.get(stage, {})
        for r_idx in range(2, ws.max_row + 1):
            a_id_val = ws.cell(row=r_idx, column=id_col).value
            if a_id_val is None:
                continue
            a_id = str(a_id_val).strip()
            llm_per_cid = evidence.get(a_id, {})
            for cid, dec_idx in decision_cols:
                llm_decision = llm_per_cid.get(cid, "unsure")
                roll = rng.random()
                if roll < uncertain_rate:
                    chosen = "unsure"
                elif roll < uncertain_rate + agree_with_llm_rate:
                    chosen = llm_decision
                else:
                    # Disagree: pick a different non-unsure decision.
                    alts = [d for d in CANONICAL_DECISIONS
                            if d != llm_decision and d != "unsure"]
                    chosen = rng.choice(alts) if alts else "unsure"
                # Write the full-sentence dropdown text the way a real rater
                # would have selected it. Falls back to bare canonical if the
                # criterion label is unknown (shouldn't happen in practice).
                criterion_label = criterion_labels.get(cid, cid)
                ws.cell(row=r_idx, column=dec_idx).value = _option_for(
                    chosen, criterion_label
                )
                # Notes column is the next one over.
                ws.cell(row=r_idx, column=dec_idx + 1).value = (
                    f"synthetic; profile=agree_{agree_with_llm_rate}/unc_{uncertain_rate}"
                )
                n_filled += 1

    dst_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst_path)
    return n_filled


def _infer_labels_from_reference_sheet(wb) -> Dict[str, str]:
    """Read the Reference sheet to extract {criterion_id: label}."""
    out: Dict[str, str] = {}
    if "Reference" not in wb.sheetnames:
        return out
    ws = wb["Reference"]
    headers = [c.value for c in ws[1]]
    try:
        id_col = headers.index("ID") + 1
        text_col = headers.index("Full text") + 1
    except ValueError:
        return out
    for r_idx in range(2, ws.max_row + 1):
        cid = ws.cell(row=r_idx, column=id_col).value
        text = ws.cell(row=r_idx, column=text_col).value
        if cid and text:
            # Strip "<cid> -" prefix if present (source_text format).
            cid_str = str(cid).strip()
            text_str = str(text).strip()
            prefix = f"{cid_str} -"
            if text_str.startswith(prefix):
                text_str = text_str[len(prefix):].strip()
            elif text_str.startswith(f"{cid_str} - "):
                text_str = text_str[len(f"{cid_str} - "):].strip()
            out[cid_str] = text_str
    return out


def parse_args(argv: List[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Synthetic filler for adjudication grids (test-only)."
    )
    p.add_argument("--src", required=True, type=Path, help="Empty XLSX path.")
    p.add_argument("--dst", required=True, type=Path, help="Filled XLSX output path.")
    p.add_argument("--el-filtered", type=Path, default=None,
                   help="Stage-EL filtered CSV (for evidence_json).")
    p.add_argument("--il-filtered", type=Path, default=None,
                   help="Stage-IL filtered CSV (for evidence_json).")
    p.add_argument("--agree", type=float, default=0.9)
    p.add_argument("--uncertain", type=float, default=0.05)
    p.add_argument("--seed", type=int, default=0)
    return p.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv if argv is not None else sys.argv[1:])
    by_stage: Dict[str, Dict[str, Dict[str, str]]] = {}
    if args.el_filtered:
        by_stage["EL"] = load_evidence_for_stage(args.el_filtered, "EL")
    if args.il_filtered:
        by_stage["IL"] = load_evidence_for_stage(args.il_filtered, "IL")
    n = fill_workbook(
        src_path=args.src, dst_path=args.dst,
        llm_canonical_by_stage=by_stage,
        agree_with_llm_rate=args.agree,
        uncertain_rate=args.uncertain,
        seed=args.seed,
    )
    print(f"[ok] filled {args.dst} with {n} synthetic decisions "
          f"(agree={args.agree}, uncertain={args.uncertain}, seed={args.seed})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
